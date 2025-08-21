# =============================================================================
# WORKLY - WORKFORCE MANAGEMENT ROUTES
# Organized by functional areas for better maintainability
# =============================================================================
#
# ROUTE ORGANIZATION:
# 1. Global Configuration & Utilities
# 2. Core Navigation Routes
# 3. Authentication Routes
# 4. Dashboard Routes
# 5. Attendance & Clock In/Out Routes
# 6. Shift Management Routes
# 7. Leave Management Routes
# 8. User Management Routes
# 9. Reports Routes
# 10. Holiday Management Routes
# 11. Reperibilità (On-Call) Routes
# 12. Expense Management Routes
# 13. Overtime Management Routes
# 14. Mileage Reimbursement Routes
# 15. Admin & System Management Routes
# 16. API Endpoints
#
# Total Routes: 169+
# =============================================================================

# Flask Core Imports
from flask import render_template, request, redirect, url_for, flash, jsonify, make_response
from flask_login import login_user, logout_user, login_required, current_user
from werkzeug.security import check_password_hash, generate_password_hash

# Standard Library Imports
from datetime import datetime, date, timedelta, time
from urllib.parse import urlparse, urljoin
from io import BytesIO, StringIO
import re
import qrcode
import base64
import json
from defusedcsv import csv

# Application Imports
from app import app, db, csrf
from config import get_config

# SQLAlchemy Imports
from sqlalchemy.orm import joinedload

# Model Imports
from models import (
    User, AttendanceEvent, LeaveRequest, LeaveType, Shift, ShiftTemplate, 
    ReperibilitaShift, ReperibilitaTemplate, ReperibilitaIntervention, Intervention,
    Sede, WorkSchedule, UserRole, PresidioCoverage, PresidioCoverageTemplate,
    ReperibilitaCoverage, Holiday, PasswordResetToken, OvertimeType, OvertimeRequest,
    ExpenseCategory, ExpenseReport, ACITable, MileageRequest,
    italian_now, get_active_presidio_templates, get_presidio_coverage_for_day
)

# Form Imports
from forms import (
    LoginForm, UserForm, UserProfileForm, AttendanceForm, LeaveRequestForm, LeaveTypeForm,
    ShiftForm, ShiftTemplateForm, SedeForm, WorkScheduleForm, RoleForm,
    PresidioCoverageTemplateForm, PresidioCoverageForm, PresidioCoverageSearchForm,
    ForgotPasswordForm, ResetPasswordForm, OvertimeTypeForm, OvertimeRequestForm,
    ApproveOvertimeForm, OvertimeFilterForm, ACIUploadForm, ACIRecordForm, ACIFilterForm,
    MileageRequestForm, ApproveMileageForm, MileageFilterForm
)

# Utility Imports
from utils import (
    get_user_statistics, get_team_statistics, format_hours, 
    check_user_schedule_with_permissions, send_overtime_request_message
)

# Blueprint registration will be handled at the end of this file

# =============================================================================
# GLOBAL CONFIGURATION AND UTILITY FUNCTIONS
# =============================================================================

@app.context_processor
def inject_config():
    """Inject configuration into all templates"""
    config = get_config()
    return dict(config=config)

def require_login(f):
    """Decorator to require login for routes"""
    from functools import wraps
    
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated:
            return redirect(url_for('auth.login'))
        return f(*args, **kwargs)
    
    return decorated_function

def is_safe_url(target):
    """Check if a URL is safe for redirect (same domain only)"""
    if not target:
        return False
    
    # Parse the target URL
    ref_url = urlparse(request.host_url)
    test_url = urlparse(urljoin(request.host_url, target))
    
    # Check if the scheme and netloc match (same domain)
    return test_url.scheme in ('http', 'https') and ref_url.netloc == test_url.netloc

# =============================================================================
# CORE NAVIGATION ROUTES
# =============================================================================

@app.route('/')
def index():
    """Main entry point - redirect to appropriate dashboard"""
    if current_user.is_authenticated:
        return redirect(url_for('dashboard.dashboard'))
    return redirect(url_for('auth.login'))

# =============================================================================
# AUTHENTICATION ROUTES - MOVED TO routes/auth.py BLUEPRINT
# =============================================================================
# Authentication routes now handled by auth_bp blueprint

# =============================================================================
# DASHBOARD ROUTES - MOVED TO blueprints/dashboard.py BLUEPRINT
# =============================================================================
# Dashboard routes now handled by dashboard_bp blueprint

# NEXT ROUTES TO MIGRATE: ATTENDANCE & CLOCK IN/OUT ROUTES

# =============================================================================
# ATTENDANCE & CLOCK IN/OUT ROUTES 
# =============================================================================

# API work_hours migrated to attendance blueprint

# =============================================================================
# ATTENDANCE & CLOCK IN/OUT ROUTES - MOVED TO blueprints/attendance.py BLUEPRINT  
# =============================================================================
# Attendance routes now handled by attendance_bp blueprint

# NEXT ROUTES TO MIGRATE: SHIFT MANAGEMENT ROUTES

# La prossima sezione con le vere Shift Management Routes inizia più avanti nel file
# Le routes Attendance sono state migrate al blueprint blueprints/attendance.py

# =============================================================================
# ROUTES NON ANCORA MIGRATE (da rimuovere quando migrazione completata)
# =============================================================================

# Prossime routes da migrare: Shift Management, Leave Management, etc.

# =============================================================================
# SHIFT MANAGEMENT ROUTES (non ancora migrati)  
# =============================================================================

# DUPLICATE FUNCTION REMOVED - La vera funzione turni_automatici() è alla riga ~941

# ATTENDANCE ROUTES COMPLETAMENTE MIGRATE A BLUEPRINT
# Il codice duplicato verrà rimosso sistematicamente

# =============================================================================
# FINE RIMOZIONE CODICE DUPLICATO
# =============================================================================
# Attendance route fragments completely removed - all functions migrated to attendance blueprint

# =============================================================================
# NEXT SECTIONS: REMAINING ROUTES TO MIGRATE
# =============================================================================

# check_shift_before_clock_out migrated to attendance blueprint

# clock_out migrated to attendance blueprint

# break_start migrated to attendance blueprint

# break_end migrated to attendance blueprint

# ATTENDANCE ROUTE MIGRATED TO blueprints/attendance.py - ALL LOGIC MOVED

# turni_automatici route MOVED TO shifts_bp blueprint

# =============================================================================
# SHIFT MANAGEMENT ROUTES
# =============================================================================

# get_shifts_for_template_api route MOVED TO shifts_bp blueprint

# visualizza_turni route MOVED TO shifts_bp blueprint

# genera_turni_da_template route MOVED TO shifts_bp blueprint

# create_shift route MOVED TO shifts_bp blueprint

# generate_shifts route MOVED TO shifts_bp blueprint

# regenerate_template route MOVED TO shifts_bp blueprint

# delete_template migrated to shifts module
@login_required
def delete_template(template_id):
    if not current_user.can_manage_shifts():
        flash('Non hai i permessi per eliminare template', 'danger')
        return redirect(url_for('manage_turni'))
    
    template = ShiftTemplate.query.get_or_404(template_id)
    
    # Delete associated shifts
    shifts_deleted = Shift.query.filter(
        Shift.date >= template.start_date,
        Shift.date <= template.end_date
    ).delete()
    
    # Delete template
    db.session.delete(template)
    db.session.commit()
    
    flash(f'Template "{template.name}" eliminato insieme a {shifts_deleted} turni associati', 'success')
    return redirect(url_for('manage_turni'))

# view_template migrated to shifts module
@login_required
def view_template(template_id):
    # All users can view templates, but only managers can manage them
    can_manage = current_user.can_manage_shifts()
    
    template = ShiftTemplate.query.get_or_404(template_id)
    
    # Get view mode from URL parameter
    view_mode = request.args.get('view', 'all')
    
    # Get shifts for this template period
    shifts_query = Shift.query.join(User, Shift.user_id == User.id).filter(
        Shift.date >= template.start_date,
        Shift.date <= template.end_date
    )
    
    # Apply view filter
    if view_mode == 'personal':
        shifts = shifts_query.filter(Shift.user_id == current_user.id).order_by(Shift.date.desc(), Shift.start_time).all()
    else:
        shifts = shifts_query.order_by(Shift.date.desc(), Shift.start_time).all()
    
    # Check for leave requests that overlap with each shift
    for shift in shifts:
        # Look for pending or approved leave requests that overlap with the shift date
        leave_request = LeaveRequest.query.filter(
            LeaveRequest.user_id == shift.user_id,
            LeaveRequest.start_date <= shift.date,
            LeaveRequest.end_date >= shift.date,
            LeaveRequest.status.in_(['Pending', 'Approved'])
        ).first()
        
        # Add leave request info to shift object
        shift.has_leave_request = leave_request is not None
        shift.leave_request = leave_request
    
    # Calculate statistics
    total_hours = sum(shift.get_duration_hours() for shift in shifts)
    future_shifts = len([s for s in shifts if s.date >= date.today()])
    unique_users = len(set(shift.user_id for shift in shifts))
    
    # Calculate hours per user
    user_hours = {}
    for shift in shifts:
        if shift.user_id not in user_hours:
            user_hours[shift.user_id] = {
                'user': shift.user,
                'total_hours': 0,
                'shift_count': 0,
                'shifts': []
            }
        hours = shift.get_duration_hours()
        user_hours[shift.user_id]['total_hours'] += hours
        user_hours[shift.user_id]['shift_count'] += 1
        user_hours[shift.user_id]['shifts'].append(shift)
    
    # Sort by total hours descending
    user_hours_list = sorted(user_hours.values(), key=lambda x: x['total_hours'], reverse=True)
    
    # Get forms only for managers
    if can_manage:
        shift_form = ShiftForm()
        template_form = ShiftTemplateForm()
        
        # Populate user choices for shift form - solo utenti con orario "Turni"
        # Escludi solo ruoli amministrativi (Amministratore)
        workers = User.query.join(WorkSchedule, User.work_schedule_id == WorkSchedule.id, isouter=True).filter(
            User.role != 'Amministratore',
            User.active.is_(True),
            WorkSchedule.name == 'Turni'
        ).all()
        shift_form.user_id.choices = [(u.id, u.get_full_name()) for u in workers]
        
        # Get all templates
        shift_templates = ShiftTemplate.query.order_by(ShiftTemplate.created_at.desc()).all()
    else:
        shift_form = None
        template_form = None
        shift_templates = []
    
    # Helper per giorni della settimana in italiano
    def get_italian_weekday(date_obj):
        giorni = ['Lunedì', 'Martedì', 'Mercoledì', 'Giovedì', 'Venerdì', 'Sabato', 'Domenica']
        return giorni[date_obj.weekday()]
    
    return render_template('shifts.html', 
                         shifts=shifts, 
                         shift_form=shift_form,
                         template_form=template_form,
                         shift_templates=shift_templates,
                         selected_template=template,
                         today=datetime.now().date(),
                         total_hours=round(total_hours, 1),
                         future_shifts=future_shifts,
                         unique_users=unique_users,
                         user_hours_list=user_hours_list,
                         get_italian_weekday=get_italian_weekday,
                         can_manage=can_manage,
                         view_mode=view_mode)

# =============================================================================
# LEAVE MANAGEMENT ROUTES
# =============================================================================

# leave_types migrated to leave module
@login_required
def leave_types():
    if not (current_user.can_manage_leave() or current_user.can_manage_leave_types()):
        flash('Non hai i permessi per gestire le tipologie di permesso', 'danger')
        return redirect(url_for('dashboard'))
    
    leave_types = LeaveType.query.order_by(LeaveType.name).all()
    return render_template('leave_types.html', leave_types=leave_types)

# leave_types/add migrated to leave module
@login_required
def add_leave_type_page():
    if not (current_user.can_manage_leave() or current_user.can_manage_leave_types()):
        flash('Non hai i permessi per aggiungere tipologie di permesso', 'danger')
        return redirect(url_for('leave_types'))
    
    if request.method == 'POST':
        try:
            name = request.form.get('name')
            description = request.form.get('description')
            requires_approval = 'requires_approval' in request.form
            active_status = 'active' in request.form
            
            # Verifica duplicati
            if LeaveType.query.filter_by(name=name).first():
                flash('Esiste già una tipologia con questo nome', 'warning')
                return render_template('add_leave_type.html')
            
            leave_type = LeaveType(
                name=name,
                description=description,
                requires_approval=requires_approval,
                active=active_status
            )
            
            db.session.add(leave_type)
            db.session.commit()
            flash(f'Tipologia "{name}" creata con successo', 'success')
            return redirect(url_for('leave_types'))
        except Exception as e:
            db.session.rollback()
            flash('Errore nella creazione della tipologia', 'danger')
            return render_template('add_leave_type.html')
    
    # GET request - mostra form di creazione
    return render_template('add_leave_type.html')

# leave_types/edit migrated to leave module
@login_required
def edit_leave_type_page(id):
    if not (current_user.can_manage_leave() or current_user.can_manage_leave_types()):
        flash('Non hai i permessi per modificare tipologie di permesso', 'danger')
        return redirect(url_for('leave_types'))
    
    leave_type = LeaveType.query.get_or_404(id)
    
    if request.method == 'POST':
        try:
            name = request.form.get('name')
            
            # Verifica duplicati (escludendo il record corrente)
            existing = LeaveType.query.filter(LeaveType.name == name, LeaveType.id != id).first()
            if existing:
                flash('Esiste già una tipologia con questo nome', 'warning')
                return render_template('edit_leave_type.html', leave_type=leave_type)
            
            leave_type.name = name
            leave_type.description = request.form.get('description')
            leave_type.requires_approval = 'requires_approval' in request.form
            leave_type.active = 'active' in request.form
            leave_type.updated_at = italian_now()
            
            db.session.commit()
            flash(f'Tipologia "{name}" aggiornata con successo', 'success')
            return redirect(url_for('leave_types'))
        except Exception as e:
            db.session.rollback()
            flash('Errore nell\'aggiornamento della tipologia', 'danger')
            return render_template('edit_leave_type.html', leave_type=leave_type)
    
    # GET request - mostra form di modifica
    return render_template('edit_leave_type.html', leave_type=leave_type)

# leave_types/delete migrated to leave module
@login_required
def delete_leave_type(id):
    if not (current_user.can_manage_leave() or current_user.can_manage_leave_types()):
        flash('Non hai i permessi per eliminare tipologie di permesso', 'danger')
        return redirect(url_for('leave_types'))
    
    leave_type = LeaveType.query.get_or_404(id)
    
    # Verifica che non ci siano richieste associate
    if leave_type.leave_requests.count() > 0:
        flash('Non è possibile eliminare una tipologia con richieste associate', 'warning')
        return redirect(url_for('leave_types'))
    
    try:
        name = leave_type.name
        db.session.delete(leave_type)
        db.session.commit()
        flash(f'Tipologia "{name}" eliminata con successo', 'success')
    except Exception as e:
        db.session.rollback()
        flash('Errore nell\'eliminazione della tipologia', 'danger')
    
    return redirect(url_for('leave_types'))

# ROUTE MOVED TO leave_bp blueprint
    # Ottieni il parametro view per determinare la modalità di visualizzazione
    view_mode = request.args.get('view', 'default')
    
    form = LeaveRequestForm()
    
    # Determina il comportamento in base al view_mode e ai permessi utente
    if view_mode == 'my':
        # Modalità "Le Mie Richieste" - sempre solo richieste dell'utente corrente
        requests = LeaveRequest.query.filter_by(
            user_id=current_user.id
        ).order_by(LeaveRequest.created_at.desc()).all()
        can_approve = False
        page_title = "Le Mie Richieste Ferie/Permessi"
        
    elif view_mode == 'approve' and (current_user.can_manage_leave() or current_user.can_approve_leave()):
        # Modalità approvazione - solo richieste pending da approvare
        if current_user.role == 'Responsabili':
            requests = LeaveRequest.query.join(User, LeaveRequest.user_id == User.id).filter(
                User.sede_id == current_user.sede_id,
                LeaveRequest.status == 'Pending'
            ).order_by(LeaveRequest.created_at.desc()).all()
        else:
            requests = LeaveRequest.query.join(User, LeaveRequest.user_id == User.id).filter(
                LeaveRequest.status == 'Pending'
            ).order_by(LeaveRequest.created_at.desc()).all()
        can_approve = True
        page_title = "Approva Richieste Ferie/Permessi"
        
    elif view_mode == 'view' and current_user.can_view_leave():
        # Modalità visualizzazione - tutte le richieste per reportistica
        if current_user.role == 'Responsabili':
            requests = LeaveRequest.query.join(User, LeaveRequest.user_id == User.id).filter(
                User.sede_id == current_user.sede_id
            ).order_by(LeaveRequest.created_at.desc()).all()
        else:
            requests = LeaveRequest.query.join(User, LeaveRequest.user_id == User.id).order_by(
                LeaveRequest.created_at.desc()
            ).all()
        can_approve = False
        page_title = "Visualizza Richieste Ferie/Permessi"
        
    elif current_user.can_manage_leave() or current_user.can_approve_leave():
        # Modalità gestione completa per manager
        if current_user.role == 'Responsabili':
            requests = LeaveRequest.query.join(User, LeaveRequest.user_id == User.id).filter(
                User.sede_id == current_user.sede_id
            ).order_by(LeaveRequest.created_at.desc()).all()
        else:
            requests = LeaveRequest.query.join(User, LeaveRequest.user_id == User.id).order_by(
                LeaveRequest.created_at.desc()
            ).all()
        can_approve = True
        page_title = "Gestione Richieste Ferie/Permessi"
        
    else:
        # Modalità utente normale - solo le proprie richieste
        requests = LeaveRequest.query.filter_by(
            user_id=current_user.id
        ).order_by(LeaveRequest.created_at.desc()).all()
        can_approve = False
        page_title = "Le Mie Richieste Ferie/Permessi"
    
    return render_template('leave_requests.html', 
                         requests=requests, 
                         form=form,
                         can_approve=can_approve,
                         page_title=page_title,
                         view_mode=view_mode,
                         today=datetime.now().date())

# ROUTE MOVED TO leave_bp blueprint
    if not current_user.can_request_leave():
        flash('Non hai i permessi per richiedere ferie/permessi', 'danger')
        return redirect(url_for('leave_requests'))
    
    form = LeaveRequestForm()
    
    if request.method == 'POST' and form.validate_on_submit():
        # Ottieni la tipologia di permesso selezionata
        leave_type = LeaveType.query.get(form.leave_type_id.data)
        if not leave_type or not leave_type.active:
            flash('Tipologia di permesso non valida', 'danger')
            return redirect(url_for('leave_requests'))
        
        # Determina la data di fine in base al tipo
        end_date = form.end_date.data if form.end_date.data else form.start_date.data
        
        # Check for overlapping requests
        overlapping = None
        
        # Controlla sovrapposizioni esistenti
        if form.start_time.data and form.end_time.data:
            # Per permessi orari, controlla sovrapposizione oraria nella stessa giornata
            existing_requests = LeaveRequest.query.filter(
                LeaveRequest.user_id == current_user.id,
                LeaveRequest.status.in_(['Pending', 'Approved']),
                LeaveRequest.start_date == form.start_date.data,  # Stessa giornata
                LeaveRequest.start_time.isnot(None),  # Solo permessi orari
                LeaveRequest.end_time.isnot(None)
            ).all()
            
            # Controlla sovrapposizione oraria
            for existing in existing_requests:
                if not (form.end_time.data <= existing.start_time or 
                       form.start_time.data >= existing.end_time):
                    overlapping = existing
                    break
        else:
            # Per ferie e permessi giornalieri, controllo sovrapposizioni per date
            overlapping = LeaveRequest.query.filter(
                LeaveRequest.user_id == current_user.id,
                LeaveRequest.status.in_(['Pending', 'Approved']),
                LeaveRequest.start_date <= end_date,
                LeaveRequest.end_date >= form.start_date.data
            ).first()
        
        if overlapping:
            if overlapping.start_time and overlapping.end_time:
                flash(f'Hai già un permesso sovrapposto dalle {overlapping.start_time.strftime("%H:%M")} alle {overlapping.end_time.strftime("%H:%M")} in questa giornata', 'warning')
            else:
                flash('Hai già una richiesta sovrapposta in questo periodo', 'warning')
        else:
            leave_request = LeaveRequest(
                user_id=current_user.id,
                leave_type_id=leave_type.id,
                start_date=form.start_date.data,
                end_date=end_date,
                reason=form.reason.data,
                leave_type=leave_type.name  # Manteniamo per retrocompatibilità
            )
            
            # Auto-approva se la tipologia non richiede autorizzazione o se l'utente può auto-approvarsi
            if not leave_type.requires_approval or current_user.can_approve_leave():
                leave_request.status = 'Approved'
                leave_request.approved_by = current_user.id  # Self-approved
                leave_request.approved_at = italian_now()
            else:
                leave_request.status = 'Pending'
            
            # Aggiungi orari se specificati
            if form.start_time.data and form.end_time.data:
                leave_request.start_time = form.start_time.data
                leave_request.end_time = form.end_time.data
            
            db.session.add(leave_request)
            db.session.commit()
            
            # Invia messaggi automatici
            from utils import send_leave_request_message
            send_leave_request_message(leave_request, 'created', current_user)
            
            # Messaggio di successo personalizzato
            if not leave_type.requires_approval:
                flash(f'Richiesta di {leave_type.name.lower()} approvata automaticamente', 'success')
            elif current_user.can_approve_leave():
                flash(f'Richiesta di {leave_type.name.lower()} approvata automaticamente', 'success')
            else:
                duration = leave_request.get_duration_display()
                flash(f'Richiesta di {leave_type.name.lower()} inviata con successo ({duration})', 'success')
            return redirect(url_for('leave_requests'))
    else:
        # Errori di validazione
        for field, errors in form.errors.items():
            for error in errors:
                flash(f'{field}: {error}', 'danger')
    
    # GET request - mostra form di creazione
    return render_template('create_leave_request.html', 
                         form=form,
                         today=datetime.now().date())

# approve_leave migrated to leave module
@login_required
def approve_leave(request_id):
    if not current_user.can_approve_leave():
        flash('Non hai i permessi per approvare richieste', 'danger')
        return redirect(url_for('leave_requests'))
    
    leave_request = LeaveRequest.query.get_or_404(request_id)
    leave_request.status = 'Approved'
    leave_request.approved_by = current_user.id
    leave_request.approved_at = datetime.utcnow()
    
    db.session.commit()
    
    # Invia messaggio di approvazione all'utente richiedente
    from utils import send_leave_request_message
    send_leave_request_message(leave_request, 'approved', current_user)
    
    flash('Richiesta approvata', 'success')
    return redirect(url_for('leave_requests'))

# reject_leave migrated to leave module
@login_required
def reject_leave(request_id):
    if not current_user.can_approve_leave():
        flash('Non hai i permessi per rifiutare richieste', 'danger')
        return redirect(url_for('leave_requests'))
    
    leave_request = LeaveRequest.query.get_or_404(request_id)
    leave_request.status = 'Rejected'
    leave_request.approved_by = current_user.id
    leave_request.approved_at = datetime.utcnow()
    
    db.session.commit()
    
    # Invia messaggio di rifiuto all'utente richiedente
    from utils import send_leave_request_message
    send_leave_request_message(leave_request, 'rejected', current_user)
    
    flash('Richiesta rifiutata', 'warning')
    return redirect(url_for('leave_requests'))

# delete_leave migrated to leave module
@login_required
def delete_leave(request_id):
    leave_request = LeaveRequest.query.get_or_404(request_id)
    
    # Verifica che sia l'utente proprietario della richiesta
    if leave_request.user_id != current_user.id:
        flash('Non puoi cancellare richieste di altri utenti', 'danger')
        return redirect(url_for('leave_requests'))
    
    # Verifica che la richiesta non sia già approvata E che non sia futura
    from zoneinfo import ZoneInfo
    italy_tz = ZoneInfo('Europe/Rome')
    today = datetime.now(italy_tz).date()
    
    if leave_request.status == 'Approved' and leave_request.start_date < today:
        flash('Non puoi cancellare richieste già approvate e iniziate', 'warning')
        return redirect(url_for('leave_requests'))
    
    # Invia messaggi di cancellazione prima di eliminare la richiesta
    from utils import send_leave_request_message
    send_leave_request_message(leave_request, 'cancelled', current_user)
    
    # Cancella la richiesta
    db.session.delete(leave_request)
    db.session.commit()
    flash('Richiesta cancellata con successo', 'success')
    
    # Determina dove reindirizzare in base al referer
    referer = request.headers.get('Referer', '')
    if 'dashboard' in referer or referer.endswith('/'):
        return redirect(url_for('dashboard'))
    else:
        return redirect(url_for('leave_requests'))

# =============================================================================
# USER MANAGEMENT ROUTES
# =============================================================================

# ROUTE MOVED TO user_management_bp blueprint
# users() function migrated to blueprints/user_management.py

# ROUTE MOVED TO user_management_bp blueprint
# new_user() function migrated to blueprints/user_management.py

# ROUTE MOVED TO user_management_bp blueprint
    if not (current_user.can_manage_users() or current_user.can_view_users()):
        flash('Non hai i permessi per accedere alla gestione utenti', 'danger')
        return redirect(url_for('dashboard'))
    
    # Applica filtro automatico per sede usando il metodo helper
    users_query = User.get_visible_users_query(current_user).options(joinedload(User.sede_obj))
    users = users_query.order_by(User.created_at.desc()).all()
    
    # Determina il nome della sede per il titolo
    sede_name = None if current_user.all_sedi else (current_user.sede_obj.name if current_user.sede_obj else None)
    form = UserForm(is_edit=False)
    
    # Aggiungi statistiche team per le statistiche utenti dinamiche
    team_stats = None
    if current_user.can_view_team_stats_widget():
        team_stats = get_team_statistics()
    
    return render_template('user_management.html', users=users, form=form, 
                         sede_name=sede_name, is_multi_sede=current_user.all_sedi,
                         team_stats=team_stats)

# ROUTE MOVED TO user_management_bp blueprint
# user_profile() function migrated to blueprints/user_management.py

# ROUTE MOVED TO user_management_bp blueprint
    if not current_user.can_manage_users():
        flash('Non hai i permessi per modificare gli utenti', 'danger')
        return redirect(url_for('user_management'))
    
    user = User.query.get_or_404(user_id)
    form = UserForm(original_username=user.username, is_edit=True, obj=user)
    
    if request.method == 'GET':
        # Popola i campi sede e all_sedi con i valori attuali
        form.all_sedi.data = user.all_sedi
        if user.sede_id:
            form.sede.data = user.sede_id
        
        if user.work_schedule_id:
            # Aggiungi l'orario corrente alle scelte se non già presente
            if user.work_schedule:
                schedule_choice = (user.work_schedule.id, f"{user.work_schedule.name} ({user.work_schedule.start_time.strftime('%H:%M') if user.work_schedule.start_time else ''}-{user.work_schedule.end_time.strftime('%H:%M') if user.work_schedule.end_time else ''})")
                if schedule_choice not in form.work_schedule.choices:
                    form.work_schedule.choices.append(schedule_choice)
            form.work_schedule.data = user.work_schedule_id
        else:
            # Se non ha un orario, imposta il valore di default
            form.work_schedule.data = ''
        
        # Gestione del veicolo ACI con campi progressivi
        if user.aci_vehicle_id and user.aci_vehicle:
            # Popola i campi progressivi basati sul veicolo esistente
            form.aci_vehicle_tipo.data = user.aci_vehicle.tipologia
            form.aci_vehicle_marca.data = user.aci_vehicle.marca
            form.aci_vehicle.data = user.aci_vehicle_id
            
            # Aggiorna le scelte per rendere i dropdown funzionali
            from models import ACITable
            aci_vehicles = ACITable.query.order_by(ACITable.tipologia, ACITable.marca, ACITable.modello).all()
            
            # Aggiorna le scelte delle marche per il tipo selezionato
            marche = list(set([v.marca for v in aci_vehicles if v.tipologia == user.aci_vehicle.tipologia and v.marca]))
            form.aci_vehicle_marca.choices = [('', 'Seleziona marca')] + [(marca, marca) for marca in sorted(marche)]
            
            # Aggiorna le scelte dei modelli per tipo e marca selezionati
            modelli = ACITable.query.filter(
                ACITable.tipologia == user.aci_vehicle.tipologia,
                ACITable.marca == user.aci_vehicle.marca
            ).order_by(ACITable.modello).all()
            form.aci_vehicle.choices = [('', 'Seleziona modello')] + [
                (v.id, f"{v.modello} (€{v.costo_km:.4f}/km)") for v in modelli
            ]
        else:
            form.aci_vehicle_tipo.data = ''
            form.aci_vehicle_marca.data = ''
            form.aci_vehicle.data = ''
    
    if form.validate_on_submit():
        # Impedisce la disattivazione dell'amministratore
        if user.role == 'Amministratore' and not form.active.data:
            flash('Non è possibile disattivare l\'utente amministratore', 'danger')
            return render_template('edit_user.html', form=form, user=user)
        
        user.username = form.username.data
        user.email = form.email.data
        user.role = form.role.data
        user.first_name = form.first_name.data
        user.last_name = form.last_name.data
        user.all_sedi = form.all_sedi.data
        user.sede_id = form.sede.data if not form.all_sedi.data else None
        user.work_schedule_id = form.work_schedule.data
        user.aci_vehicle_id = form.aci_vehicle.data if form.aci_vehicle.data and form.aci_vehicle.data != -1 else None
        user.part_time_percentage = form.get_part_time_percentage_as_float()
        user.active = form.active.data
        
        # Update password only if provided
        if form.password.data:
            user.password_hash = generate_password_hash(form.password.data)
        
        # Non c'è più gestione sedi multiple
        
        db.session.commit()
        flash(f'Utente {user.username} modificato con successo', 'success')
        return redirect(url_for('user_management'))
    else:
        # Populate the percentage field manually only on GET request
        form.part_time_percentage.data = str(user.part_time_percentage)
    
    return render_template('edit_user.html', form=form, user=user)

# ROUTE MOVED TO user_management_bp blueprint
# toggle_user() function migrated to blueprints/user_management.py

# =============================================================================
# REPORTS ROUTES
# =============================================================================

# MIGRATED TO BLUEPRINT: blueprints/reports.py
# @app.route('/reports')
# @login_required
# def reports():
    if not current_user.can_view_reports():
        flash('Non hai i permessi per visualizzare i report', 'danger')
        return redirect(url_for('dashboard'))
    
    # Get date range from request
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    if start_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    else:
        start_date = datetime.now().date() - timedelta(days=30)
    
    if end_date:
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    else:
        end_date = datetime.now().date()
    
    # Get team statistics with error handling
    try:
        team_stats = get_team_statistics(start_date, end_date)
    except Exception as e:
        pass  # Silent error handling
        team_stats = {
            'active_users': 0,
            'total_hours': 0,
            'shifts_this_period': 0,
            'avg_hours_per_user': 0
        }
    
    # Get user statistics for all active users (excluding Amministratore and Ospite)
    users = User.query.filter_by(active=True).filter(~User.role.in_(['Amministratore', 'Ospite'])).all()
    pass  # User count info
    
    user_stats = []
    chart_data = []  # Separate data for charts without User objects
    
    for user in users:
        try:
            stats = get_user_statistics(user.id, start_date, end_date)
            stats['user'] = user
            user_stats.append(stats)
            
            # Create chart-safe data
            chart_data.append({
                'user_id': user.id,
                'username': user.username,
                'full_name': user.get_full_name(),
                'role': user.role,
                'total_hours_worked': stats['total_hours_worked'],
                'days_worked': stats['days_worked'],
                'shifts_assigned': stats['shifts_assigned'],
                'shift_hours': stats['shift_hours'],
                'approved_leaves': stats['approved_leaves'],
                'pending_leaves': stats['pending_leaves']
            })
            
            pass  # User stats processed
        except Exception as e:
            pass  # Silent error handling
            # Continue without this user's stats
            continue
    
    pass  # Stats collected
    
    # Get interventions data for the table (entrambi i tipi)
    from models import Intervention, ReperibilitaIntervention
    start_datetime = datetime.combine(start_date, datetime.min.time())
    end_datetime = datetime.combine(end_date, datetime.max.time())
    
    try:
        # Interventi generici
        interventions = Intervention.query.filter(
            Intervention.start_datetime >= start_datetime,
            Intervention.start_datetime <= end_datetime
        ).order_by(Intervention.start_datetime.desc()).all()
        
        # Interventi reperibilità  
        reperibilita_interventions = ReperibilitaIntervention.query.filter(
            ReperibilitaIntervention.start_datetime >= start_datetime,
            ReperibilitaIntervention.start_datetime <= end_datetime
        ).order_by(ReperibilitaIntervention.start_datetime.desc()).all()
    except Exception as e:
        pass  # Silent error handling
        interventions = []
        reperibilita_interventions = []
    
    # Get attendance data for charts - calculate real data
    attendance_data = []
    current_date = start_date
    active_user_ids = [user.id for user in User.query.filter(User.active.is_(True)).filter(~User.role.in_(['Amministratore', 'Ospite'])).all()]
    
    while current_date <= end_date:
        # Calculate total hours and workers for this day
        daily_total_hours = 0
        workers_present = 0
        
        for user_id in active_user_ids:
            try:
                daily_hours = AttendanceEvent.get_daily_work_hours(user_id, current_date)
                if daily_hours and daily_hours > 0:
                    daily_total_hours += float(daily_hours)
                    workers_present += 1
            except Exception as e:
                pass  # Silent error handling
                # Continue to next user instead of stopping
                continue
        
        attendance_data.append({
            'date': current_date.strftime('%d/%m'),  # Format for display
            'hours': round(daily_total_hours, 1),
            'workers': workers_present
        })
        current_date += timedelta(days=1)
    
    return render_template('reports.html', 
                         team_stats=team_stats,
                         user_stats=user_stats,
                         chart_data=chart_data,
                         attendance_data=attendance_data,
                         interventions=interventions,
                         reperibilita_interventions=reperibilita_interventions,
                         start_date=start_date,
                         end_date=end_date)

# =============================================================================
# HOLIDAY MANAGEMENT ROUTES
# =============================================================================

# ROUTE MOVED TO holidays_bp blueprint - holidays

# ROUTE MOVED TO holidays_bp blueprint - add_holiday

# ROUTE MOVED TO holidays_bp blueprint - edit_holiday

# ROUTE MOVED TO holidays_bp blueprint - delete_holiday

# ROUTE MOVED TO holidays_bp blueprint - generate_holidays

# ROUTE MOVED TO auth_bp blueprint - change_password

# ROUTE MOVED TO auth_bp blueprint - forgot_password

# ROUTE MOVED TO auth_bp blueprint - reset_password

@app.errorhandler(404)
def not_found_error(error):
    return render_template('404.html'), 404

# edit_shift route migrated to shifts blueprint

# calculate_shift_presence function migrated to shifts blueprint

# team_shifts route migrated to shifts blueprint

# change_shift_user route migrated to shifts blueprint

@app.errorhandler(500)
def internal_error(error):
    db.session.rollback()
    return render_template('500.html'), 500

# =============================================================================
# REPERIBILITÀ (ON-CALL) ROUTES
# =============================================================================

# ROUTE MOVED TO reperibilita_bp blueprint - reperibilita_coverage
    """Lista coperture reperibilità"""
    if not current_user.can_access_reperibilita():
        flash('Non hai i permessi per visualizzare le coperture reperibilità', 'danger')
        return redirect(url_for('dashboard'))
    
    from models import ReperibilitaCoverage
    from collections import defaultdict
    
    # Raggruppa le coperture per periodo + sede per trattare duplicazioni come gruppi separati
    coverages = ReperibilitaCoverage.query.order_by(ReperibilitaCoverage.start_date.desc()).all()
    groups = defaultdict(lambda: {'coverages': [], 'start_date': None, 'end_date': None, 'creator': None, 'created_at': None})
    
    for coverage in coverages:
        # Include sede nel period_key per separare coperture duplicate con sedi diverse
        sede_ids = sorted(coverage.get_sedi_ids_list())
        sede_key = "_".join(map(str, sede_ids)) if sede_ids else "no_sede"
        period_key = f"{coverage.start_date.strftime('%Y-%m-%d')}_{coverage.end_date.strftime('%Y-%m-%d')}_{sede_key}"
        
        if not groups[period_key]['start_date']:
            groups[period_key]['start_date'] = coverage.start_date
            groups[period_key]['end_date'] = coverage.end_date
            groups[period_key]['creator'] = coverage.creator
            groups[period_key]['created_at'] = coverage.created_at
        groups[period_key]['coverages'].append(coverage)
    
    # Converte in oggetti simili ai presidi per il template
    reperibilita_groups = {}
    for period_key, data in groups.items():
        class ReperibilitaGroup:
            def __init__(self, coverages, start_date, end_date, creator, created_at):
                self.coverages = coverages
                self.start_date = start_date
                self.end_date = end_date
                self.creator = creator
                self.created_at = created_at
        
        reperibilita_groups[period_key] = ReperibilitaGroup(
            data['coverages'], data['start_date'], data['end_date'], 
            data['creator'], data['created_at']
        )
    
    return render_template('reperibilita_coverage.html', reperibilita_groups=reperibilita_groups)

# ROUTE MOVED TO reperibilita_bp blueprint - create_reperibilita_coverage
    """Crea nuova copertura reperibilità"""
    if not current_user.can_manage_reperibilita():
        flash('Non hai i permessi per creare coperture reperibilità', 'danger')
        return redirect(url_for('dashboard'))
    
    from forms import ReperibilitaCoverageForm
    from models import ReperibilitaCoverage
    
    form = ReperibilitaCoverageForm()
    
    if form.validate_on_submit():
        # Crea una copertura per ogni giorno selezionato
        import json
        success_count = 0
        
        for day_of_week in form.days_of_week.data:
            coverage = ReperibilitaCoverage()
            coverage.day_of_week = day_of_week
            coverage.start_time = form.start_time.data
            coverage.end_time = form.end_time.data
            coverage.set_required_roles_list(form.required_roles.data)
            coverage.set_sedi_ids_list(form.sedi.data)  # Aggiungi le sedi selezionate
            coverage.description = form.description.data
            coverage.active = form.active.data
            coverage.start_date = form.start_date.data
            coverage.end_date = form.end_date.data
            coverage.created_by = current_user.id
            
            db.session.add(coverage)
            success_count += 1
        
        try:
            db.session.commit()
            flash(f'Copertura reperibilità creata con successo per {success_count} giorni!', 'success')
            return redirect(url_for('reperibilita_coverage'))
        except Exception as e:
            db.session.rollback()
            flash(f'Errore durante la creazione: {str(e)}', 'error')
    
    return render_template('create_reperibilita_coverage.html', form=form)

# ROUTE MOVED TO reperibilita_bp blueprint - edit_reperibilita_coverage
    """Modifica copertura reperibilità"""
    if not current_user.can_manage_reperibilita():
        flash('Non hai i permessi per modificare coperture reperibilità', 'danger')
        return redirect(url_for('dashboard'))
    
    from forms import ReperibilitaCoverageForm
    from models import ReperibilitaCoverage
    
    coverage = ReperibilitaCoverage.query.get_or_404(coverage_id)
    form = ReperibilitaCoverageForm()
    
    if form.validate_on_submit():
        coverage.start_time = form.start_time.data
        coverage.end_time = form.end_time.data
        coverage.set_required_roles_list(form.required_roles.data)
        coverage.set_sedi_ids_list(form.sedi.data)  # Aggiungi le sedi selezionate
        coverage.description = form.description.data
        coverage.active = form.active.data
        coverage.start_date = form.start_date.data
        coverage.end_date = form.end_date.data
        
        try:
            db.session.commit()
            flash('Copertura reperibilità aggiornata con successo!', 'success')
            return redirect(url_for('reperibilita_coverage'))
        except Exception as e:
            db.session.rollback()
            flash(f'Errore durante l\'aggiornamento: {str(e)}', 'error')
    
    # Pre-popola il form con i dati esistenti
    if request.method == 'GET':
        form.start_time.data = coverage.start_time
        form.end_time.data = coverage.end_time
        form.required_roles.data = coverage.get_required_roles_list()
        form.description.data = coverage.description
        form.active.data = coverage.active
        form.start_date.data = coverage.start_date
        form.end_date.data = coverage.end_date
        form.days_of_week.data = [coverage.day_of_week]  # Single day for edit
    
    return render_template('edit_reperibilita_coverage.html', form=form, coverage=coverage)

# delete_reperibilita_coverage route MOVED TO reperibilita_bp blueprint

# view_reperibilita_coverage route MOVED TO reperibilita_bp blueprint

# delete_reperibilita_period route MOVED TO reperibilita_bp blueprint

# REPERIBILITA_SHIFTS ROUTE MIGRATED TO blueprints/reperibilita.py

# reperibilita_template_detail route MOVED TO reperibilita_bp blueprint

# reperibilita_replica route MOVED TO reperibilita_bp blueprint

# generate_reperibilita_shifts route MOVED TO reperibilita_bp blueprint

# regenerate_reperibilita_template route MOVED TO reperibilita_bp blueprint

# ROUTE MOVED TO reperibilita_bp blueprint - start_intervention
    """Inizia un intervento di reperibilità"""
    if current_user.role not in ['Management', 'Operatore', 'Redattore', 'Sviluppatore']:
        flash('Non hai i permessi per registrare interventi di reperibilità.', 'danger')
        return redirect(url_for('dashboard'))
    
    # Controlla se c'è già un intervento attivo
    active_intervention = ReperibilitaIntervention.query.filter_by(
        user_id=current_user.id,
        end_datetime=None
    ).first()
    
    if active_intervention:
        flash('Hai già un intervento di reperibilità in corso.', 'warning')
        return redirect(url_for('dashboard'))
    
    # Ottieni shift_id dal form se presente
    shift_id = request.form.get('shift_id')
    if shift_id:
        shift_id = int(shift_id)
    
    # Ottieni is_remote dal form (default True = remoto)
    is_remote = request.form.get('is_remote', 'true').lower() == 'true'
    
    # Ottieni priorità dal form (default Media)
    priority = request.form.get('priority', 'Media')
    if priority not in ['Bassa', 'Media', 'Alta']:
        priority = 'Media'
    
    # Crea nuovo intervento
    intervention = ReperibilitaIntervention(
        user_id=current_user.id,
        shift_id=shift_id,
        start_datetime=italian_now(),
        description=request.form.get('description', ''),
        priority=priority,
        is_remote=is_remote
    )
    
    db.session.add(intervention)
    db.session.commit()
    
    flash('Intervento di reperibilità iniziato con successo.', 'success')
    return redirect(url_for('reperibilita_shifts'))

# ROUTE MOVED TO reperibilita_bp blueprint - end_intervention
    """Termina un intervento di reperibilità"""
    if current_user.role not in ['Management', 'Operatore', 'Redattore', 'Sviluppatore']:
        flash('Non hai i permessi per registrare interventi di reperibilità.', 'danger')
        return redirect(url_for('dashboard'))
    
    # Trova l'intervento attivo
    active_intervention = ReperibilitaIntervention.query.filter_by(
        user_id=current_user.id,
        end_datetime=None
    ).first()
    
    if not active_intervention:
        flash('Nessun intervento di reperibilità attivo da terminare.', 'warning')
        return redirect(url_for('dashboard'))
    
    # Termina l'intervento
    active_intervention.end_datetime = italian_now()
    active_intervention.description = request.form.get('description', active_intervention.description)
    
    db.session.commit()
    
    flash('Intervento di reperibilità terminato con successo.', 'success')
    
    # Redirect PM to ente_home, others to reperibilita_shifts
    if current_user.role == 'Management':
        return redirect(url_for('ente_home'))
    else:
        return redirect(url_for('reperibilita_shifts'))

# ROUTE MOVED TO reperibilita_bp blueprint - delete_reperibilita_template
    """Elimina un template reperibilità e tutti i suoi turni"""
    if not current_user.can_manage_reperibilita():
        flash('Non hai i permessi per eliminare template di reperibilità', 'danger')
        return redirect(url_for('dashboard'))
    
    from models import ReperibilitaTemplate, ReperibilitaShift
    
    template = ReperibilitaTemplate.query.get_or_404(template_id)
    
    try:
        # Elimina tutti i turni del periodo del template
        shifts = ReperibilitaShift.query.filter(
            ReperibilitaShift.date >= template.start_date,
            ReperibilitaShift.date <= template.end_date
        ).all()
        
        for shift in shifts:
            db.session.delete(shift)
        
        # Elimina il template
        template_name = template.name
        db.session.delete(template)
        db.session.commit()
        
        flash(f'Template reperibilità "{template_name}" eliminato con successo', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Errore durante l\'eliminazione: {str(e)}', 'error')
    
    return redirect(url_for('reperibilita_shifts'))

# QR Code Authentication Routes
# QR ROUTE MIGRATED TO qr_bp blueprint - qr_login

# QR ROUTE MIGRATED TO qr_bp blueprint - qr_fresh

# QR ROUTE MIGRATED TO qr_bp blueprint - quick_attendance

# QR ROUTE MIGRATED TO qr_bp blueprint - generate_qr_codes

# EXPORT ROUTE MIGRATED TO export_bp blueprint - shifts/excel

# EXPORT ROUTE MIGRATED TO export_bp blueprint - shifts/pdf

# EXPORT ROUTE MIGRATED TO export_bp blueprint - attendance/excel

# INTERVENTION ROUTE MIGRATED TO interventions_bp blueprint - start_general_intervention

# INTERVENTION ROUTE MIGRATED TO interventions_bp blueprint - end_general_intervention

# INTERVENTION ROUTE MIGRATED TO interventions_bp blueprint - my_interventions

# EXPORT ROUTE MIGRATED TO export_bp blueprint - interventions/general/excel
# @app.route('/export_general_interventions_excel')
# @login_required
# def export_general_interventions_excel():
    """Export interventi generici in formato Excel"""
    if current_user.role == 'Admin':
        flash('Accesso non autorizzato.', 'danger')
        return redirect(url_for('dashboard'))
    
    from zoneinfo import ZoneInfo
    italy_tz = ZoneInfo('Europe/Rome')
    today = datetime.now(italy_tz).date()
    
    # Gestisci filtri data
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')
    
    # Default: primo del mese corrente - oggi
    if not start_date_str:
        start_date = today.replace(day=1)
    else:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
    
    if not end_date_str:
        end_date = today
    else:
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    
    # Converti le date in datetime per il filtro
    start_datetime = datetime.combine(start_date, datetime.min.time())
    end_datetime = datetime.combine(end_date, datetime.max.time())
    
    # PM ed Ente vedono tutti gli interventi, altri utenti solo i propri
    if current_user.role in ['Management', 'Ente']:
        general_interventions = Intervention.query.join(User).filter(
            Intervention.start_datetime >= start_datetime,
            Intervention.start_datetime <= end_datetime
        ).order_by(Intervention.start_datetime.desc()).all()
    else:
        general_interventions = Intervention.query.filter(
            Intervention.user_id == current_user.id,
            Intervention.start_datetime >= start_datetime,
            Intervention.start_datetime <= end_datetime
        ).order_by(Intervention.start_datetime.desc()).all()
    
    # Crea Excel in memoria usando openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    import tempfile
    import os
    
    wb = Workbook()
    ws = wb.active  
    ws.title = "Interventi Generici"
    
    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Header
    if current_user.role in ['Management', 'Ente']:
        headers = ['Utente', 'Nome', 'Cognome', 'Ruolo', 'Data Inizio', 'Ora Inizio', 'Data Fine', 'Ora Fine', 
                  'Durata (minuti)', 'Priorità', 'Tipologia', 'Descrizione', 'Stato']
    else:
        headers = ['Data Inizio', 'Ora Inizio', 'Data Fine', 'Ora Fine', 
                  'Durata (minuti)', 'Priorità', 'Tipologia', 'Descrizione', 'Stato']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    # Dati
    for row_idx, intervention in enumerate(general_interventions, 2):
        row_data = []
        
        if current_user.role in ['Management', 'Ente']:
            row_data.extend([
                intervention.user.username,
                intervention.user.first_name,
                intervention.user.last_name,
                intervention.user.role
            ])
        
        row_data.extend([
            intervention.start_datetime.strftime('%d/%m/%Y'),
            intervention.start_datetime.strftime('%H:%M'),
            intervention.end_datetime.strftime('%d/%m/%Y') if intervention.end_datetime else 'In corso',
            intervention.end_datetime.strftime('%H:%M') if intervention.end_datetime else '',
            round(intervention.duration_minutes, 1) if intervention.end_datetime else '',
            intervention.priority or '',
            'In presenza',
            intervention.description or '',
            'Completato' if intervention.end_datetime else 'In corso'
        ])
        
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border
            if col <= 4 or col in [5, 7]:  # Date e orari
                cell.alignment = Alignment(horizontal='center')
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to temporary file
    temp_dir = tempfile.mkdtemp()
    filename = f'interventi_generici_{start_date.strftime("%Y%m%d")}_{end_date.strftime("%Y%m%d")}.xlsx'
    excel_path = os.path.join(temp_dir, filename)
    wb.save(excel_path)
    
    # Read file for response
    with open(excel_path, 'rb') as f:
        excel_data = f.read()
    
    # Cleanup
    os.remove(excel_path)
    os.rmdir(temp_dir)
    
    response = make_response(excel_data)
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response

# EXPORT ROUTE MIGRATED TO export_bp blueprint - interventions/reperibilita/excel
# @app.route('/export_reperibilita_interventions_excel')
# @login_required
# def export_reperibilita_interventions_excel():
    """Export interventi reperibilità in formato Excel"""
    if current_user.role == 'Admin':
        flash('Accesso non autorizzato.', 'danger')
        return redirect(url_for('dashboard'))
    
    from zoneinfo import ZoneInfo
    italy_tz = ZoneInfo('Europe/Rome')
    today = datetime.now(italy_tz).date()
    
    # Gestisci filtri data
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')
    
    # Default: primo del mese corrente - oggi
    if not start_date_str:
        start_date = today.replace(day=1)
    else:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
    
    if not end_date_str:
        end_date = today
    else:
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    
    # Converti le date in datetime per il filtro
    start_datetime = datetime.combine(start_date, datetime.min.time())
    end_datetime = datetime.combine(end_date, datetime.max.time())
    
    # PM ed Ente vedono tutti gli interventi, altri utenti solo i propri
    if current_user.role in ['Management', 'Ente']:
        reperibilita_interventions = ReperibilitaIntervention.query.join(User).filter(
            ReperibilitaIntervention.start_datetime >= start_datetime,
            ReperibilitaIntervention.start_datetime <= end_datetime
        ).order_by(ReperibilitaIntervention.start_datetime.desc()).all()
    else:
        reperibilita_interventions = ReperibilitaIntervention.query.filter(
            ReperibilitaIntervention.user_id == current_user.id,
            ReperibilitaIntervention.start_datetime >= start_datetime,
            ReperibilitaIntervention.start_datetime <= end_datetime
        ).order_by(ReperibilitaIntervention.start_datetime.desc()).all()
    
    # Crea Excel in memoria usando openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    import tempfile
    import os
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Interventi Reperibilità"
    
    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Header
    if current_user.role in ['Management', 'Ente']:
        headers = ['Utente', 'Nome', 'Cognome', 'Ruolo', 'Data Inizio', 'Ora Inizio', 'Data Fine', 'Ora Fine', 
                  'Durata (minuti)', 'Priorità', 'Tipologia', 'Data Turno', 'Ora Inizio Turno', 'Ora Fine Turno', 
                  'Descrizione', 'Stato']
    else:
        headers = ['Data Inizio', 'Ora Inizio', 'Data Fine', 'Ora Fine', 
                  'Durata (minuti)', 'Priorità', 'Tipologia', 'Data Turno', 'Ora Inizio Turno', 'Ora Fine Turno', 
                  'Descrizione', 'Stato']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    # Dati
    for row_idx, intervention in enumerate(reperibilita_interventions, 2):
        row_data = []
        
        if current_user.role in ['Management', 'Ente']:
            row_data.extend([
                intervention.user.username,
                intervention.user.first_name,
                intervention.user.last_name,
                intervention.user.role
            ])
        
        row_data.extend([
            intervention.start_datetime.strftime('%d/%m/%Y'),
            intervention.start_datetime.strftime('%H:%M'),
            intervention.end_datetime.strftime('%d/%m/%Y') if intervention.end_datetime else 'In corso',
            intervention.end_datetime.strftime('%H:%M') if intervention.end_datetime else '',
            round(intervention.duration_minutes, 1) if intervention.end_datetime else '',
            intervention.priority or '',
            'Remoto' if intervention.is_remote else 'In presenza',
            intervention.shift.date.strftime('%d/%m/%Y') if intervention.shift else '',
            intervention.shift.start_time.strftime('%H:%M') if intervention.shift else '',
            intervention.shift.end_time.strftime('%H:%M') if intervention.shift else '',
            intervention.description or '',
            'Completato' if intervention.end_datetime else 'In corso'
        ])
        
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border
            if col <= 4 or col in [5, 7, 12]:  # Date e orari
                cell.alignment = Alignment(horizontal='center')
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to temporary file
    temp_dir = tempfile.mkdtemp()
    filename = f'interventi_reperibilita_{start_date.strftime("%Y%m%d")}_{end_date.strftime("%Y%m%d")}.xlsx'
    excel_path = os.path.join(temp_dir, filename)
    wb.save(excel_path)
    
    # Read file for response
    with open(excel_path, 'rb') as f:
        excel_data = f.read()
    
    # Cleanup
    os.remove(excel_path)
    os.rmdir(temp_dir)
    
    response = make_response(excel_data)
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response

# QR ROUTE MIGRATED TO qr_bp blueprint - qr_page

# =============================================================================
# ADMIN & SYSTEM MANAGEMENT ROUTES
# =============================================================================

# ROUTE MOVED TO user_management_bp blueprint - admin_qr_codes

# ROUTE MOVED TO user_management_bp blueprint - view_qr_codes

# ROUTE MOVED TO user_management_bp blueprint - admin_generate_static_qr

# EXPORT ROUTE MIGRATED TO export_bp blueprint - leave/excel
# @app.route('/export_leave_requests_excel')
# @login_required
# def export_leave_requests_excel():
    """Export delle richieste di ferie/permessi in formato Excel"""
    if not current_user.can_view_leave_requests() and not current_user.can_request_leave():
        flash('Non hai i permessi per esportare le richieste', 'danger')
        return redirect(url_for('dashboard'))
    
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    
    # Determina se l'utente può vedere tutte le richieste o solo le proprie
    can_approve = current_user.can_approve_leave()
    
    if can_approve:
        # Admin può vedere tutte le richieste
        requests = LeaveRequest.query.order_by(LeaveRequest.start_date.desc()).all()
        filename = f"richieste_ferie_permessi_{date.today().strftime('%Y%m%d')}.xlsx"
    else:
        # Utente normale vede solo le proprie
        requests = LeaveRequest.query.filter_by(user_id=current_user.id).order_by(LeaveRequest.start_date.desc()).all()
        filename = f"mie_richieste_ferie_permessi_{date.today().strftime('%Y%m%d')}.xlsx"
    
    # Crea il workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Richieste Ferie e Permessi"
    
    # Definisci gli header
    if can_approve:
        headers = ['Utente', 'Ruolo', 'Periodo', 'Durata', 'Tipo', 'Motivo', 'Stato', 'Data Richiesta', 'Approvato da', 'Data Approvazione']
    else:
        headers = ['Periodo', 'Durata', 'Tipo', 'Motivo', 'Stato', 'Data Richiesta', 'Approvato da', 'Data Approvazione']
    
    # Scrive gli header
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Scrive i dati
    for row_idx, request in enumerate(requests, 2):
        col = 1
        
        if can_approve:
            # Utente
            ws.cell(row=row_idx, column=col, value=request.user.get_full_name())
            col += 1
            
            # Ruolo
            ws.cell(row=row_idx, column=col, value=request.user.role)
            col += 1
        
        # Periodo
        if request.leave_type == 'Permesso' and request.is_time_based():
            periodo = f"{request.start_date.strftime('%d/%m/%Y')} {request.start_time.strftime('%H:%M')}-{request.end_time.strftime('%H:%M')}"
        elif request.start_date != request.end_date:
            periodo = f"{request.start_date.strftime('%d/%m/%Y')} - {request.end_date.strftime('%d/%m/%Y')}"
        else:
            periodo = request.start_date.strftime('%d/%m/%Y')
        ws.cell(row=row_idx, column=col, value=periodo)
        col += 1
        
        # Durata
        if request.leave_type == 'Permesso' and request.is_time_based():
            durata = f"{request.duration_hours}h"
        else:
            durata = f"{request.duration_days} giorni"
        ws.cell(row=row_idx, column=col, value=durata)
        col += 1
        
        # Tipo
        ws.cell(row=row_idx, column=col, value=request.leave_type)
        col += 1
        
        # Motivo
        ws.cell(row=row_idx, column=col, value=request.reason or '-')
        col += 1
        
        # Stato
        status_cell = ws.cell(row=row_idx, column=col, value=request.status)
        if request.status == 'Approved':
            status_cell.fill = PatternFill(start_color="D4F8D4", end_color="D4F8D4", fill_type="solid")
        elif request.status == 'Rejected':
            status_cell.fill = PatternFill(start_color="F8D4D4", end_color="F8D4D4", fill_type="solid")
        elif request.status == 'Pending':
            status_cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        col += 1
        
        # Data Richiesta
        ws.cell(row=row_idx, column=col, value=request.created_at.strftime('%d/%m/%Y %H:%M'))
        col += 1
        
        # Approvato da
        ws.cell(row=row_idx, column=col, value=request.approved_by.get_full_name() if request.approved_by else '-')
        col += 1
        
        # Data Approvazione
        ws.cell(row=row_idx, column=col, value=request.approved_at.strftime('%d/%m/%Y %H:%M') if request.approved_at else '-')
        col += 1
    
    # Ajusta la larghezza delle colonne
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    # Prepara la risposta
    response = make_response()
    
    # Salva in un buffer temporaneo
    from io import BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    response.data = buffer.getvalue()
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename={filename}'
    
    return response

# EXPORT ROUTE MIGRATED TO export_bp blueprint - expense/excel
# @app.route('/export_expense_reports_excel')
# @login_required
# def export_expense_reports_excel():
    """Export delle note spese in formato Excel"""
    if not current_user.can_view_expense_reports() and not current_user.can_create_expense_reports():
        flash('Non hai i permessi per esportare le note spese', 'danger')
        return redirect(url_for('dashboard'))
    
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from models import ExpenseReport, ExpenseCategory, User
    from sqlalchemy.orm import joinedload
    
    # Determina se l'utente può vedere tutte le note spese o solo le proprie
    can_manage = current_user.can_manage_expense_reports() or current_user.can_approve_expense_reports()
    
    if can_manage:
        # Admin può vedere tutte le note spese
        expenses = ExpenseReport.query.options(
            joinedload(ExpenseReport.employee),
            joinedload(ExpenseReport.category)
        ).order_by(ExpenseReport.expense_date.desc()).all()
        filename = f"note_spese_{date.today().strftime('%Y%m%d')}.xlsx"
    else:
        # Utente normale vede solo le proprie
        expenses = ExpenseReport.query.filter_by(employee_id=current_user.id).options(
            joinedload(ExpenseReport.employee),
            joinedload(ExpenseReport.category)
        ).order_by(ExpenseReport.expense_date.desc()).all()
        filename = f"mie_note_spese_{date.today().strftime('%Y%m%d')}.xlsx"
    
    # Crea il workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Note Spese"
    
    # Definisci gli header
    if can_manage:
        headers = ['Dipendente', 'Data Spesa', 'Categoria', 'Descrizione', 'Importo', 'Stato', 'Data Creazione', 'Approvato da', 'Data Approvazione', 'Note']
    else:
        headers = ['Data Spesa', 'Categoria', 'Descrizione', 'Importo', 'Stato', 'Data Creazione', 'Approvato da', 'Data Approvazione', 'Note']
    
    # Scrive gli header
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Scrive i dati
    for row_idx, expense in enumerate(expenses, 2):
        col = 1
        
        if can_manage:
            # Dipendente
            ws.cell(row=row_idx, column=col, value=expense.employee.get_full_name())
            col += 1
        
        # Data Spesa
        ws.cell(row=row_idx, column=col, value=expense.expense_date.strftime('%d/%m/%Y'))
        col += 1
        
        # Categoria
        ws.cell(row=row_idx, column=col, value=expense.category.name if expense.category else '-')
        col += 1
        
        # Descrizione
        ws.cell(row=row_idx, column=col, value=expense.description)
        col += 1
        
        # Importo (come numero per calcoli Excel)
        ws.cell(row=row_idx, column=col, value=float(expense.amount))
        col += 1
        
        # Stato
        status_text = {
            'pending': 'In attesa',
            'approved': 'Approvata',
            'rejected': 'Rifiutata'
        }.get(expense.status, expense.status)
        
        status_cell = ws.cell(row=row_idx, column=col, value=status_text)
        if expense.status == 'approved':
            status_cell.fill = PatternFill(start_color="D4F8D4", end_color="D4F8D4", fill_type="solid")
        elif expense.status == 'rejected':
            status_cell.fill = PatternFill(start_color="F8D4D4", end_color="F8D4D4", fill_type="solid")
        elif expense.status == 'pending':
            status_cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        col += 1
        
        # Data Creazione
        ws.cell(row=row_idx, column=col, value=expense.created_at.strftime('%d/%m/%Y %H:%M'))
        col += 1
        
        # Approvato da
        approved_by_user = User.query.get(expense.approved_by) if expense.approved_by else None
        ws.cell(row=row_idx, column=col, value=approved_by_user.get_full_name() if approved_by_user else '-')
        col += 1
        
        # Data Approvazione
        ws.cell(row=row_idx, column=col, value=expense.approved_at.strftime('%d/%m/%Y %H:%M') if expense.approved_at else '-')
        col += 1
        
        # Note (campo approval_comment del modello)
        ws.cell(row=row_idx, column=col, value=expense.approval_comment or '-')
        col += 1
    
    # Ajusta la larghezza delle colonne
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    # Prepara la risposta
    response = make_response()
    
    # Salva in un buffer temporaneo
    from io import BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    response.data = buffer.getvalue()
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename={filename}'
    
    return response

# ===============================
# GESTIONE TURNI PER SEDI
# ===============================

# ROUTE MOVED TO user_management_bp blueprint - admin/turni (manage_turni)
    """Gestione turni per sedi di tipo 'Turni'"""
    if not current_user.can_access_turni():
        flash('Non hai i permessi per accedere alla gestione turni', 'danger')
        return redirect(url_for('dashboard'))
    
    # Ottieni sedi turni accessibili dall'utente
    sedi_turni = current_user.get_turni_sedi()
    
    # Per ogni sede, calcola statistiche sui turni esistenti
    sede_stats = {}
    for sede in sedi_turni:
        from models import Shift, ReperibilitaShift
        
        turni_count = db.session.query(Shift).join(User, Shift.user_id == User.id).filter(
            User.sede_id == sede.id,
            User.active == True
        ).count()
        
        # Conta coperture attive per questa sede (usando PresidioCoverage temporaneamente)
        from models import PresidioCoverage
        coperture_count = PresidioCoverage.query.filter(
            PresidioCoverage.active == True
        ).count()
        
        sede_stats[sede.id] = {
            'turni_count': turni_count,
            'coperture_count': coperture_count,
            'users_count': len([u for u in sede.users if u.active])
        }
    
    return render_template('manage_turni.html', 
                         sedi_turni=sedi_turni, 
                         sede_stats=sede_stats,
                         can_manage_all=(current_user.can_manage_shifts()))

# ADMIN TURNI ROUTE MIGRATED TO shifts_bp blueprint

# ADMIN TURNI ROUTE MIGRATED TO shifts_bp blueprint
# def view_turni_coverage():
    """Visualizza le coperture create per una sede specifica"""
    if not current_user.can_access_turni():
        flash('Non hai i permessi per visualizzare le coperture', 'danger')
        return redirect(url_for('dashboard'))
    
    sede_id = request.args.get('sede', type=int)
    if not sede_id:
        # Se l'utente non è Admin, usa la sua sede per default
        if current_user.role != 'Admin' and current_user.sede_obj and current_user.sede_obj.is_turni_mode():
            sede_id = current_user.sede_obj.id
        else:
            flash('ID sede non specificato. Seleziona una sede dalla pagina Gestione Turni.', 'warning')
            return redirect(url_for('manage_turni'))
    
    sede = Sede.query.get_or_404(sede_id)
    
    # Verifica permessi sulla sede - supporta utenti multi-sede
    if not current_user.can_manage_shifts() and not current_user.can_view_shifts():
        flash('Non hai i permessi per visualizzare le coperture', 'danger')
        return redirect(url_for('dashboard'))
    
    # Per utenti non-admin, verifica accesso alla sede specifica
    if not current_user.all_sedi and current_user.sede_obj and current_user.sede_obj.id != sede_id:
        flash('Non hai accesso a questa sede specifica', 'danger')
        return redirect(url_for('manage_turni'))
    
    if not sede.is_turni_mode():
        flash('La sede selezionata non è configurata per la modalità turni', 'warning')
        return redirect(url_for('manage_turni'))
    
    # Ottieni le coperture create per questa sede
    # Per ora prendiamo tutte le coperture attive - in futuro potremmo aggiungere un campo sede_id
    from models import PresidioCoverage
    coperture = PresidioCoverage.query.filter_by(active=True).order_by(
        PresidioCoverage.start_date.desc(),
        PresidioCoverage.day_of_week,
        PresidioCoverage.start_time
    ).all()
    
    # Raggruppa coperture per periodo di validità (evita duplicati)
    coperture_grouped = {}
    coperture_ids_seen = set()
    for copertura in coperture:
        # Evita duplicati
        if copertura.id in coperture_ids_seen:
            continue
        coperture_ids_seen.add(copertura.id)
        
        period_key = f"{copertura.start_date.strftime('%Y-%m-%d')} - {copertura.end_date.strftime('%Y-%m-%d')}"
        if period_key not in coperture_grouped:
            coperture_grouped[period_key] = {
                'start_date': copertura.start_date,
                'end_date': copertura.end_date,
                'coperture': [],
                'active_status': copertura.active and copertura.end_date >= date.today()
            }
        coperture_grouped[period_key]['coperture'].append(copertura)
    
    # Statistiche
    total_coperture = len(coperture)
    active_coperture = len([c for c in coperture if c.is_valid_for_date(date.today())])
    
    return render_template('view_turni_coverage.html',
                         sede=sede,
                         coperture_grouped=coperture_grouped,
                         total_coperture=total_coperture,
                         active_coperture=active_coperture,
                         today=datetime.now().date(),
                         is_admin=(current_user.role == 'Admin'))

# ADMIN TURNI ROUTE MIGRATED TO shifts_bp blueprint  
# def generate_turni_from_coverage():
    """Pagina per generare turni basati sulle coperture create"""
    if not current_user.can_access_turni():
        flash('Non hai i permessi per generare turni', 'danger')
        return redirect(url_for('dashboard'))
    
    sede_id = request.args.get('sede', type=int)
    if not sede_id:
        # Se l'utente non è Admin, usa la sua sede per default
        if current_user.role != 'Admin' and current_user.sede_obj and current_user.sede_obj.is_turni_mode():
            sede_id = current_user.sede_obj.id
        else:
            flash('ID sede non specificato. Seleziona una sede dalla pagina Genera Turni.', 'warning')
            return redirect(url_for('generate_turnazioni'))
    
    sede = Sede.query.get_or_404(sede_id)
    
    # Verifica permessi sulla sede - supporta utenti multi-sede  
    if not current_user.can_manage_shifts():
        flash('Non hai i permessi per generare turni', 'danger')
        return redirect(url_for('dashboard'))
        
    # Per utenti non-admin, verifica accesso alla sede specifica
    if not current_user.all_sedi and current_user.sede_obj and current_user.sede_obj.id != sede_id:
        flash('Non hai accesso a questa sede specifica', 'danger')
        return redirect(url_for('generate_turnazioni'))
    
    if not sede.is_turni_mode():
        flash('La sede selezionata non è configurata per la modalità turni', 'warning')
        return redirect(url_for('generate_turnazioni'))
    
    # Ottieni le coperture attive per questa sede
    from models import PresidioCoverage
    coperture = PresidioCoverage.query.filter_by(active=True).order_by(
        PresidioCoverage.start_date.desc(),
        PresidioCoverage.day_of_week,
        PresidioCoverage.start_time
    ).all()
    
    # Raggruppa coperture per periodo di validità (evita duplicati con ID univoci)
    coperture_grouped = {}
    coperture_ids_seen = set()
    for copertura in coperture:
        # Evita duplicati
        if copertura.id in coperture_ids_seen:
            continue
        coperture_ids_seen.add(copertura.id)
        
        period_key = f"{copertura.start_date.strftime('%Y-%m-%d')} - {copertura.end_date.strftime('%Y-%m-%d')}"
        if period_key not in coperture_grouped:
            coperture_grouped[period_key] = {
                'start_date': copertura.start_date,
                'end_date': copertura.end_date,
                'coperture': [],
                'active_status': copertura.active and copertura.end_date >= date.today(),
                'period_id': f"{copertura.start_date.strftime('%Y%m%d')}-{copertura.end_date.strftime('%Y%m%d')}"
            }
        coperture_grouped[period_key]['coperture'].append(copertura)
    
    # Statistiche
    total_coperture = len(coperture)
    active_coperture = len([c for c in coperture if c.is_valid_for_date(date.today())])
    
    return render_template('generate_turni_from_coverage.html',
                         sede=sede,
                         coperture_grouped=coperture_grouped,
                         total_coperture=total_coperture,
                         active_coperture=active_coperture,
                         today=datetime.now().date(),
                         is_admin=(current_user.role == 'Admin'))

# ADMIN TURNI ROUTE MIGRATED TO shifts_bp blueprint
# def process_generate_turni_from_coverage():
    """Processa la generazione dei turni basata sulle coperture"""
    if not current_user.can_access_turni():
        flash('Non hai i permessi per generare turni', 'danger')
        return redirect(url_for('dashboard'))
    
    sede_id = request.form.get('sede_id', type=int)
    coverage_period_id = request.form.get('coverage_period_id')
    use_coverage_dates = 'use_coverage_dates' in request.form
    replace_existing = 'replace_existing' in request.form
    confirm_overwrite = 'confirm_overwrite' in request.form
    
    # Debug parametri ricevuti

    
    if not sede_id or not coverage_period_id or coverage_period_id.strip() == '':
        flash(f'Dati mancanti per la generazione turni (sede_id: {sede_id}, coverage_period_id: \'{coverage_period_id}\')', 'danger')
        return redirect(url_for('generate_turnazioni'))
    
    sede = Sede.query.get_or_404(sede_id)
    
    # Verifica permessi - supporta utenti multi-sede
    if not current_user.can_manage_shifts():
        flash('Non hai i permessi per generare turni', 'danger')
        return redirect(url_for('dashboard'))
        
    # Per utenti non-admin, verifica accesso alla sede specifica
    if not current_user.all_sedi and current_user.sede_obj and current_user.sede_obj.id != sede_id:
        flash('Non hai accesso per generare turni per questa sede', 'danger')
        return redirect(url_for('generate_turnazioni'))
    
    try:
        # Decodifica period_id per ottenere le date della copertura
        start_str, end_str = coverage_period_id.split('-')
        start_date = datetime.strptime(start_str, '%Y%m%d').date()
        end_date = datetime.strptime(end_str, '%Y%m%d').date()
        
        # Importa i modelli necessari
        from models import PresidioCoverage, Shift
        coperture = PresidioCoverage.query.filter(
            PresidioCoverage.start_date <= end_date,
            PresidioCoverage.end_date >= start_date,
            PresidioCoverage.active == True
        ).all()
        
        if not coperture:
            flash('Nessuna copertura trovata per il periodo specificato', 'warning')
            return redirect(url_for('generate_turnazioni'))
        
        # Controlla se esistono già turni nel periodo prima di procedere
        existing_shifts = Shift.query.join(User, Shift.user_id == User.id).filter(
            User.sede_id == sede_id,
            Shift.date >= start_date,
            Shift.date <= end_date
        ).all()
        
        # Se esistono turni e l'utente non ha scelto di sostituirli, chiedi conferma
        if existing_shifts and not replace_existing and not confirm_overwrite:
            turni_count = len(existing_shifts)
            date_range = f"{start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')}"
            
            # Renderizza template di conferma con informazioni sui turni esistenti
            return render_template('confirm_overwrite_shifts.html',
                                 sede=sede,
                                 period_id=coverage_period_id,
                                 start_date=start_date,
                                 end_date=end_date,
                                 date_range=date_range,
                                 existing_shifts_count=turni_count,
                                 use_coverage_dates=use_coverage_dates,
                                 replace_existing=replace_existing)
        
        # Implementa la generazione turni reale basata sulle coperture
        turni_creati = 0
        turni_sostituiti = 0
        current_date = start_date
        
        while current_date <= end_date:
            # Trova le coperture per questo giorno della settimana
            day_of_week = current_date.weekday()  # 0=Lunedì, 6=Domenica
            
            coperture_giorno = [c for c in coperture if c.day_of_week == day_of_week and c.is_valid_for_date(current_date)]
            
            for copertura in coperture_giorno:
                # Verifica se esiste già un turno per questa data e orario
                existing_shift = Shift.query.filter_by(
                    date=current_date,
                    start_time=copertura.start_time,
                    end_time=copertura.end_time
                ).first()
                
                if existing_shift and not replace_existing:
                    continue  # Salta se esiste già e non si vuole sostituire
                elif existing_shift and replace_existing:
                    db.session.delete(existing_shift)
                    turni_sostituiti += 1
                
                # Trova utenti disponibili per i ruoli richiesti con numerosità
                required_roles_dict = copertura.get_required_roles_dict()
                
                # Per ogni ruolo e numerosità richiesta
                for role, count_needed in required_roles_dict.items():
                    available_users = User.query.filter(
                        User.sede_id == sede_id,
                        User.active == True,
                        User.role == role
                    ).all()
                    
                    if len(available_users) >= count_needed:
                        # Assegna il numero richiesto di utenti per questo ruolo
                        for i in range(count_needed):
                            user_index = (current_date.day + copertura.id + i) % len(available_users)
                            assigned_user = available_users[user_index]
                            
                            # Crea il turno
                            new_shift = Shift(
                                user_id=assigned_user.id,
                                date=current_date,
                                start_time=copertura.start_time,
                                end_time=copertura.end_time,
                                shift_type='Normale',
                                created_by=current_user.id
                            )
                            db.session.add(new_shift)
                            turni_creati += 1
            
            current_date += timedelta(days=1)
        
        db.session.commit()
        
        if turni_creati > 0 or turni_sostituiti > 0:
            message_parts = []
            if turni_creati > 0:
                message_parts.append(f'{turni_creati} turni creati')
            if turni_sostituiti > 0:
                message_parts.append(f'{turni_sostituiti} turni sostituiti')
            
            flash(f'Generazione completata! {" e ".join(message_parts)} per {sede.name} dal {start_date.strftime("%d/%m/%Y")} al {end_date.strftime("%d/%m/%Y")}', 'success')
        else:
            flash(f'Nessun turno generato - potrebbero già esistere turni per il periodo o non ci sono utenti disponibili', 'warning')
        
    except (ValueError, AttributeError) as e:
        flash('ID periodo non valido', 'danger')
        return redirect(url_for('generate_turnazioni'))
    
    return redirect(url_for('generate_turnazioni'))

# ADMIN TURNI ROUTE MIGRATED TO shifts_bp blueprint
# def view_generated_shifts():
    """Visualizza i turni generati per una specifica copertura"""
    if not current_user.can_access_turni():
        flash('Non hai i permessi per visualizzare i turni', 'danger')
        return redirect(url_for('dashboard'))
    
    sede_id = request.args.get('sede', type=int)
    period_id = request.args.get('period') or request.args.get('coverage_period')
    
    if not all([sede_id, period_id]):
        flash('Parametri mancanti per la visualizzazione turni', 'danger')
        return redirect(url_for('generate_turnazioni'))
    
    sede = Sede.query.get_or_404(sede_id)
    
    # Verifica permessi sulla sede - supporta utenti multi-sede
    if not current_user.can_view_shifts() and not current_user.can_manage_shifts():
        flash('Non hai i permessi per visualizzare i turni', 'danger')
        return redirect(url_for('dashboard'))
        
    # Per utenti non-admin, verifica accesso alla sede specifica
    if not current_user.all_sedi and current_user.sede_obj and current_user.sede_obj.id != sede_id:
        flash('Non hai accesso per visualizzare i turni di questa sede', 'danger')
        return redirect(url_for('generate_turnazioni'))
    
    # Decodifica period_id per ottenere le date
    try:
        start_str, end_str = period_id.split('-')
        coverage_start_date = datetime.strptime(start_str, '%Y%m%d').date()
        coverage_end_date = datetime.strptime(end_str, '%Y%m%d').date()
    except (ValueError, AttributeError):
        flash('Periodo non valido specificato', 'danger')
        return redirect(url_for('generate_turnazioni'))
    
    # Ottieni i turni generati nel periodo delle coperture
    shifts = Shift.query.filter(
        Shift.user.has(sede_id=sede_id),
        Shift.date >= coverage_start_date,
        Shift.date <= coverage_end_date
    ).order_by(Shift.date, Shift.start_time).all()
    
    # Raggruppa turni per data
    shifts_by_date = {}
    for shift in shifts:
        date_str = shift.date.strftime('%Y-%m-%d')
        if date_str not in shifts_by_date:
            shifts_by_date[date_str] = {
                'date': shift.date,
                'date_display': shift.date.strftime('%d/%m/%Y'),
                'day_name': ['Lunedì', 'Martedì', 'Mercoledì', 'Giovedì', 'Venerdì', 'Sabato', 'Domenica'][shift.date.weekday()],
                'shifts': []
            }
        shifts_by_date[date_str]['shifts'].append(shift)
    
    # Ottieni le coperture di riferimento per il confronto
    from models import PresidioCoverage
    reference_coverages = PresidioCoverage.query.filter(
        PresidioCoverage.start_date <= coverage_end_date,
        PresidioCoverage.end_date >= coverage_start_date,
        PresidioCoverage.active == True
    ).all()
    
    total_shifts = len(shifts)
    dates_with_shifts = len(shifts_by_date)
    period_days = (coverage_end_date - coverage_start_date).days + 1
    
    # Calcola utenti unici coinvolti
    unique_users = set()
    for shift in shifts:
        if shift.user:
            unique_users.add(shift.user.id)
    unique_users_count = len(unique_users)
    
    return render_template('view_generated_shifts.html',
                         sede=sede,
                         coverage_start_date=coverage_start_date,
                         coverage_end_date=coverage_end_date,
                         shifts_by_date=shifts_by_date,
                         reference_coverages=reference_coverages,
                         total_shifts=total_shifts,
                         dates_with_shifts=dates_with_shifts,
                         period_days=period_days,
                         unique_users_count=unique_users_count,
                         today=datetime.now().date(),
                         is_admin=(current_user.role == 'Admin'))

# ADMIN TURNI ROUTE MIGRATED TO shifts_bp blueprint
# def regenerate_turni_from_coverage():
    """Rigenera i turni eliminando quelli esistenti da oggi in poi e creandone di nuovi"""
    if not current_user.can_access_turni():
        flash('Non hai i permessi per rigenerare turni', 'danger')
        return redirect(url_for('dashboard'))
    
    sede_id = request.form.get('sede_id', type=int)
    coverage_period_id = request.form.get('coverage_period_id')
    
    if not all([sede_id, coverage_period_id]):
        flash('Dati mancanti per la rigenerazione turni', 'danger')
        return redirect(url_for('generate_turnazioni'))
    
    sede = Sede.query.get_or_404(sede_id)
    
    # Verifica permessi
    if current_user.role != 'Admin':
        if not current_user.sede_obj or current_user.sede_obj.id != sede_id:
            flash('Non hai i permessi per rigenerare turni per questa sede', 'danger')
            return redirect(url_for('generate_turnazioni'))
    
    try:
        # Decodifica period_id per ottenere le date della copertura
        start_str, end_str = coverage_period_id.split('-')
        start_date = datetime.strptime(start_str, '%Y%m%d').date()
        end_date = datetime.strptime(end_str, '%Y%m%d').date()
        
        # Data di inizio per l'eliminazione (da oggi in poi)
        from datetime import date, timedelta
        today = date.today()
        delete_from_date = max(start_date, today)
        
        # Elimina turni esistenti da oggi in poi
        from models import Shift
        shifts_to_delete = Shift.query.join(User, Shift.user_id == User.id).filter(
            User.sede_id == sede_id,
            Shift.date >= delete_from_date,
            Shift.date <= end_date
        ).all()
        
        deleted_count = len(shifts_to_delete)
        for shift in shifts_to_delete:
            db.session.delete(shift)
        
        db.session.commit()
        
        # Rigenera turni per tutto il periodo originale
        from models import PresidioCoverage
        
        coperture = PresidioCoverage.query.filter(
            PresidioCoverage.start_date <= end_date,
            PresidioCoverage.end_date >= start_date,
            PresidioCoverage.active == True
        ).all()
        
        if not coperture:
            flash('Nessuna copertura trovata per il periodo specificato', 'warning')
            return redirect(url_for('generate_turnazioni'))
        
        # Genera i nuovi turni
        new_shifts_count = 0
        current_date = start_date
        
        while current_date <= end_date:
            # Trova le coperture per questo giorno della settimana
            day_of_week = current_date.weekday()  # 0=Lunedì, 6=Domenica
            
            day_coverages = [c for c in coperture if 
                           c.start_date <= current_date <= c.end_date and
                           c.day_of_week == day_of_week]
            
            for coverage in day_coverages:
                # Ottieni utenti disponibili per questa sede e copertura
                available_users = User.query.filter(
                    User.sede_id == sede_id,
                    User.active == True,
                    User.role.in_(['Operatore', 'Sviluppatore', 'Redattore', 'Management'])
                ).all()
                
                # Calcola il numero totale di staff richiesto dai ruoli
                roles_dict = coverage.get_required_roles_dict()
                total_required_staff = sum(roles_dict.values()) if roles_dict else 1
                
                if available_users and total_required_staff > 0:
                    # Seleziona utenti per questa copertura (logica semplificata)
                    selected_users = available_users[:total_required_staff]
                    
                    for user in selected_users:
                        new_shift = Shift(
                            user_id=user.id,
                            date=current_date,
                            start_time=coverage.start_time,
                            end_time=coverage.end_time,
                            shift_type='Turno',
                            created_by=current_user.id
                        )
                        db.session.add(new_shift)
                        new_shifts_count += 1
            
            current_date += timedelta(days=1)
        
        db.session.commit()
        
        # Messaggio di successo
        if deleted_count > 0:
            flash(f'Turni rigenerati con successo! Eliminati {deleted_count} turni esistenti, creati {new_shifts_count} nuovi turni.', 'success')
        else:
            flash(f'Turni generati con successo! Creati {new_shifts_count} nuovi turni.', 'success')
        
        # Reindirizza alla visualizzazione dei turni generati
        return redirect(url_for('view_generated_shifts', sede=sede_id, period=coverage_period_id))
        
    except Exception as e:
        db.session.rollback()
        flash(f'Errore durante la rigenerazione turni: {str(e)}', 'danger')
        return redirect(url_for('generate_turnazioni'))

# delete_shift route migrated to shifts blueprint

# ADMIN TURNI ROUTE MIGRATED TO shifts_bp blueprint
# def delete_turni_period():
    """Elimina tutti i turni di un periodo da oggi in poi"""
    if not current_user.can_access_turni():
        flash('Non hai i permessi per eliminare turni', 'danger')
        return redirect(url_for('dashboard'))
    
    sede_id = request.form.get('sede_id', type=int)
    coverage_period_id = request.form.get('coverage_period_id')
    
    if not all([sede_id, coverage_period_id]):
        flash('Dati mancanti per l\'eliminazione turni', 'danger')
        return redirect(url_for('generate_turnazioni'))
    
    sede = Sede.query.get_or_404(sede_id)
    
    # Verifica permessi
    if current_user.role != 'Admin':
        if not current_user.sede_obj or current_user.sede_obj.id != sede_id:
            flash('Non hai i permessi per eliminare turni per questa sede', 'danger')
            return redirect(url_for('generate_turnazioni'))
    
    try:
        # Decodifica period_id per ottenere le date della copertura
        start_str, end_str = coverage_period_id.split('-')
        start_date = datetime.strptime(start_str, '%Y%m%d').date()
        end_date = datetime.strptime(end_str, '%Y%m%d').date()
        
        # Data di inizio per l'eliminazione (da oggi in poi)
        from datetime import date
        today = date.today()
        delete_from_date = max(start_date, today)
        
        # Elimina turni esistenti da oggi in poi
        from models import Shift
        shifts_to_delete = Shift.query.join(User, Shift.user_id == User.id).filter(
            User.sede_id == sede_id,
            Shift.date >= delete_from_date,
            Shift.date <= end_date
        ).all()
        
        deleted_count = len(shifts_to_delete)
        for shift in shifts_to_delete:
            db.session.delete(shift)
        
        db.session.commit()
        
        # Messaggio di successo
        if deleted_count > 0:
            preserved_days = (today - start_date).days if today > start_date else 0
            if preserved_days > 0:
                flash(f'Eliminati {deleted_count} turni dal {delete_from_date.strftime("%d/%m/%Y")} al {end_date.strftime("%d/%m/%Y")} (preservati {preserved_days} giorni già lavorati)', 'success')
            else:
                flash(f'Eliminati {deleted_count} turni dal {delete_from_date.strftime("%d/%m/%Y")} al {end_date.strftime("%d/%m/%Y")}', 'success')
        else:
            flash('Nessun turno da eliminare nel periodo specificato', 'info')
        
        # Reindirizza alla visualizzazione dei turni generati
        return redirect(url_for('view_generated_shifts', sede=sede_id, period=coverage_period_id))
        
    except Exception as e:
        db.session.rollback()
        flash(f'Errore durante l\'eliminazione turni: {str(e)}', 'danger')
        return redirect(url_for('generate_turnazioni'))

# ADMIN TURNI ROUTE MIGRATED TO shifts_bp blueprint
# def generate_turnazioni():
    """Generazione automatica turnazioni con visualizzazione coperture inline"""
    if not current_user.can_access_turni():
        flash('Non hai i permessi per generare turnazioni', 'danger')
        return redirect(url_for('dashboard'))
    
    # Ottieni sedi turni accessibili dall'utente
    sedi_turni = current_user.get_turni_sedi()
    
    # Per ogni sede, carica le coperture direttamente server-side
    from models import PresidioCoverage
    sedi_with_coverage = []
    
    for sede in sedi_turni:
        # Ottieni le coperture attive per questa sede
        coperture = PresidioCoverage.query.filter_by(active=True).order_by(
            PresidioCoverage.start_date.desc(),
            PresidioCoverage.day_of_week,
            PresidioCoverage.start_time
        ).all()
        
        # Raggruppa coperture per periodo di validità
        coperture_grouped = {}
        for copertura in coperture:
            period_key = f"{copertura.start_date.strftime('%d/%m/%Y')} - {copertura.end_date.strftime('%d/%m/%Y')}"
            if period_key not in coperture_grouped:
                coperture_grouped[period_key] = {
                    'start_date': copertura.start_date,
                    'end_date': copertura.end_date,
                    'coperture': [],
                    'active_status': copertura.active and copertura.end_date >= date.today(),
                    'period_id': f"{copertura.start_date.strftime('%Y%m%d')}-{copertura.end_date.strftime('%Y%m%d')}"
                }
            coperture_grouped[period_key]['coperture'].append(copertura)
        
        # Statistiche
        total_coperture = len(coperture)
        active_coperture = len([c for c in coperture if c.is_valid_for_date(date.today())])
        
        sedi_with_coverage.append({
            'sede': sede,
            'coperture_grouped': coperture_grouped,
            'total_coperture': total_coperture,
            'active_coperture': active_coperture
        })
    
    return render_template('generate_turnazioni.html', 
                         sedi_with_coverage=sedi_with_coverage,
                         today=datetime.now().date(),
                         is_admin=(current_user.role == 'Admin'))

# =============================================================================
# API ENDPOINTS
# =============================================================================

# API ROUTE MIGRATED TO api_bp blueprint - get_sede_users

# API ROUTE MIGRATED TO api_bp blueprint - api_sede_work_schedules
# API ROUTE MIGRATED TO api_bp blueprint - api_sede_work_schedules (exception handler migrated)

# API ROUTE MIGRATED TO api_bp blueprint - api_roles

# ===============================
# GESTIONE SEDI E ORARI DI LAVORO
# ===============================

# ROUTE MOVED TO admin_bp blueprint - manage_sedi

# ROUTE MOVED TO admin_bp blueprint - create_sede

# ROUTE MOVED TO admin_bp blueprint - edit_sede

# ROUTE MOVED TO admin_bp blueprint - toggle_sede (FIXED PERMISSION BUG)

# ROUTE MOVED TO admin_bp blueprint - manage_work_schedules

# ROUTE MOVED TO admin_bp blueprint - create_work_schedule

# ROUTE MOVED TO admin_bp blueprint - edit_work_schedule

# ROUTE MOVED TO admin_bp blueprint - toggle_work_schedule

# ROUTE MOVED TO admin_bp blueprint - delete_work_schedule

# GESTIONE RUOLI DINAMICI

# ROUTE MOVED TO admin_bp blueprint - manage_roles

# ROUTE MOVED TO admin_bp blueprint - create_role

# ROUTE MOVED TO admin_bp blueprint - edit_role

# ROUTE MOVED TO admin_bp blueprint - toggle_role

# ROUTE MOVED TO admin_bp blueprint - delete_role

# ROUTE MOVED TO messages_bp blueprint - internal_messages

# ROUTE MOVED TO messages_bp blueprint - mark_message_read

# ROUTE MOVED TO messages_bp blueprint - delete_message

# ROUTE MOVED TO messages_bp blueprint - mark_all_messages_read

# ROUTE MOVED TO messages_bp blueprint - send_message

# =====================================
# NUOVO SISTEMA TURNI - 3 FUNZIONALITÀ
# =====================================

# ROUTE MOVED TO presidio_bp blueprint - manage_coverage

# ROUTE MOVED TO presidio_bp blueprint - view_presidio_coverage

# ROUTE MOVED TO presidio_bp blueprint - edit_presidio_coverage

# ROUTE MOVED TO presidio_bp blueprint - create_presidio_coverage

# ROUTE MOVED TO presidio_bp blueprint - generate_turnazioni_coverage

# VIEW COVERAGE TEMPLATES ROUTE MIGRATED TO shifts_bp blueprint
# def view_coverage_templates():

# VIEW TURNI FOR PERIOD ROUTE MIGRATED TO shifts_bp blueprint
# def view_turni_for_period():

# =====================================
# NUOVO SISTEMA PRESIDIO - PACCHETTO COMPLETO
# =====================================

# PRESIDIO COVERAGE ROUTE MIGRATED TO shifts_bp blueprint
# def presidio_coverage():
    """Pagina principale per gestione copertura presidio - Sistema completo"""
    if not current_user.can_manage_shifts():
        flash('Non hai i permessi per gestire coperture presidio', 'danger')
        return redirect(url_for('dashboard'))
    
    # Ottieni tutti i template di copertura presidio ordinati per data creazione
    templates = get_active_presidio_templates()
    
    # Form per nuovo template
    form = PresidioCoverageTemplateForm()
    search_form = PresidioCoverageSearchForm()
    current_template = None
    
    # Applica filtri di ricerca se presenti
    if request.args.get('search'):
        query = PresidioCoverageTemplate.query.filter_by(active=True)
        
        template_name = request.args.get('template_name')
        if template_name:
            query = query.filter(PresidioCoverageTemplate.name.ilike(f"%{template_name}%"))
        
        date_from = request.args.get('date_from')
        if date_from:
            try:
                date_from = datetime.strptime(date_from, '%Y-%m-%d').date()
                query = query.filter(PresidioCoverageTemplate.start_date >= date_from)
            except ValueError:
                pass
        
        date_to = request.args.get('date_to')
        if date_to:
            try:
                date_to = datetime.strptime(date_to, '%Y-%m-%d').date()
                query = query.filter(PresidioCoverageTemplate.end_date <= date_to)
            except ValueError:
                pass
        
        active_arg = request.args.get('active')
        if active_arg:
            active_bool = active_arg == 'true'
            query = query.filter(PresidioCoverageTemplate.active == active_bool)
        
        templates = query.order_by(PresidioCoverageTemplate.created_at.desc()).all()
    
    # Gestisci creazione/modifica template
    if request.method == 'POST':
        action = request.form.get('action', 'create')
        
        if action == 'create' and form.validate_on_submit():
            template = PresidioCoverageTemplate(
                name=form.name.data,
                start_date=form.start_date.data,
                end_date=form.end_date.data,
                description=form.description.data,
                sede_id=form.sede_id.data,
                created_by=current_user.id
            )
            db.session.add(template)
            db.session.commit()
            flash(f'Template "{template.name}" creato con successo', 'success')
            return redirect(url_for('presidio_coverage_edit', template_id=template.id))
    
    return render_template('presidio_coverage.html', 
                         templates=templates,
                         form=form,
                         search_form=search_form,
                         current_template=current_template)

# PRESIDIO COVERAGE ROUTE MIGRATED TO shifts_bp blueprint
# def presidio_coverage_edit(template_id):
    """Modifica template esistente - Sistema completo"""
    if not current_user.can_manage_shifts():
        flash('Non hai i permessi per gestire coperture presidio', 'danger')
        return redirect(url_for('dashboard'))
    
    current_template = PresidioCoverageTemplate.query.get_or_404(template_id)
    templates = get_active_presidio_templates()
    
    # Pre-popola form con dati template
    form = PresidioCoverageTemplateForm()
    coverage_form = PresidioCoverageForm()
    
    if request.method == 'GET':
        form.name.data = current_template.name
        form.start_date.data = current_template.start_date
        form.end_date.data = current_template.end_date
        form.description.data = current_template.description
        form.sede_id.data = current_template.sede_id
    
    if request.method == 'POST':
        action = request.form.get('action', 'update')
        
        if action == 'update' and form.validate_on_submit():
            current_template.name = form.name.data
            current_template.start_date = form.start_date.data
            current_template.end_date = form.end_date.data
            current_template.description = form.description.data
            current_template.sede_id = form.sede_id.data
            db.session.commit()
            flash(f'Template "{current_template.name}" aggiornato con successo', 'success')
            return redirect(url_for('presidio_coverage_edit', template_id=template_id))
        
        elif action == 'add_coverage' and coverage_form.validate_on_submit():
            # Aggiungi nuove coperture per i giorni selezionati
            success_count = 0
            error_count = 0
            
            for day_of_week in coverage_form.days_of_week.data:
                
                # Crea nuova copertura
                new_coverage = PresidioCoverage(
                    template_id=template_id,
                    day_of_week=day_of_week,
                    start_time=coverage_form.start_time.data,
                    end_time=coverage_form.end_time.data,
                    required_roles=json.dumps(coverage_form.required_roles.data),
                    role_count=coverage_form.role_count.data,
                    description=coverage_form.description.data,
                    active=coverage_form.active.data,
                    start_date=current_template.start_date,
                    end_date=current_template.end_date,
                    created_by=current_user.id
                )
                
                # Gestione pause opzionali
                if coverage_form.break_start.data and coverage_form.break_end.data:
                    try:
                        new_coverage.break_start = datetime.strptime(coverage_form.break_start.data, '%H:%M').time()
                        new_coverage.break_end = datetime.strptime(coverage_form.break_end.data, '%H:%M').time()
                    except ValueError:
                        pass  # Ignora errori di parsing per campi opzionali
                
                db.session.add(new_coverage)
                success_count += 1
            
            db.session.commit()
            
            if success_count > 0:
                flash(f'{success_count} coperture aggiunte con successo', 'success')
            if error_count > 0:
                flash(f'{error_count} coperture non aggiunte per sovrapposizioni orarie', 'warning')
            
            return redirect(url_for('presidio_coverage_edit', template_id=template_id))
    
    return render_template('presidio_coverage.html', 
                         templates=templates,
                         form=form,
                         coverage_form=coverage_form,
                         current_template=current_template)

# PRESIDIO COVERAGE ROUTE MIGRATED TO shifts_bp blueprint
# def presidio_detail(template_id):
    """Visualizza dettagli di un template di copertura presidio"""
    if not (current_user.can_manage_shifts() or current_user.can_view_shifts()):
        flash('Non hai i permessi per visualizzare le coperture presidio', 'danger')
        return redirect(url_for('dashboard'))
    
    template = PresidioCoverageTemplate.query.get_or_404(template_id)
    
    # Organizza le coperture per giorno della settimana
    coverages_by_day = {}
    for coverage in template.coverages.filter_by(active=True).order_by(PresidioCoverage.start_time):
        day = coverage.day_of_week
        if day not in coverages_by_day:
            coverages_by_day[day] = []
        coverages_by_day[day].append(coverage)
    
    return render_template('presidio_detail.html', 
                         template=template,
                         coverages_by_day=coverages_by_day)

# PRESIDIO COVERAGE ROUTE MIGRATED TO shifts_bp blueprint
# def view_presidi():
    """Visualizzazione sola lettura dei presidi configurati"""
    if not current_user.can_view_shifts():
        flash('Non hai i permessi per visualizzare i presidi', 'warning')
        return redirect(url_for('dashboard'))
    
    templates = PresidioCoverageTemplate.query.filter_by(active=True).order_by(PresidioCoverageTemplate.start_date.desc()).all()
    return render_template('view_presidi.html', templates=templates)

# API ROUTE MIGRATED TO api_bp blueprint - api_presidio_coverage

# PRESIDIO COVERAGE ROUTE MIGRATED TO shifts_bp blueprint
# def toggle_presidio_template_status(template_id):
    """Attiva/disattiva template presidio"""
    if not current_user.can_manage_shifts():
        return jsonify({'success': False, 'message': 'Non autorizzato'}), 403
    
    template = PresidioCoverageTemplate.query.get_or_404(template_id)
    new_status = request.json.get('active', not template.active)
    
    template.active = new_status
    template.updated_at = italian_now()
    
    # Aggiorna anche tutte le coperture associate
    for coverage in template.coverages:
        coverage.active = new_status
    
    db.session.commit()
    
    return jsonify({
        'success': True, 
        'message': f'Template {"attivato" if new_status else "disattivato"} con successo'
    })

# ROUTE MOVED TO presidio_bp blueprint - delete_presidio_template

# ROUTE MOVED TO presidio_bp blueprint - duplicate_presidio_template

# ============ FUNZIONI UTILITÀ PRESIDIO ============

# FUNCTION MOVED TO presidio_bp blueprint - get_presidio_coverage_for_period

# FUNCTION MOVED TO presidio_bp blueprint - get_required_roles_for_day_time

# FUNCTION MOVED TO presidio_bp blueprint - get_active_presidio_templates

# FUNCTION MOVED TO presidio_bp blueprint - create_presidio_shift_from_template

# ROUTE MOVED TO presidio_bp blueprint - delete_presidio_coverage

# =============================================================================
# EXPENSE MANAGEMENT ROUTES
# =============================================================================

# MIGRATED TO BLUEPRINT: blueprints/expense.py
# @app.route('/expenses', methods=['GET', 'POST'])
# @login_required
# def expense_reports():
    """Visualizza elenco note spese"""
    if not current_user.can_access_expense_reports_menu():
        flash('Non hai i permessi per accedere alle note spese', 'danger')
        return redirect(url_for('dashboard'))
    
    from models import ExpenseReport, ExpenseCategory
    from forms import ExpenseFilterForm
    
    filter_form = ExpenseFilterForm(current_user=current_user)
    
    # Query base
    query = ExpenseReport.query
    
    # Check view mode from URL parameter
    view_mode = request.args.get('view', 'all')
    
    # Usa i nuovi permessi espliciti per determinare cosa può vedere l'utente
    if view_mode == 'my' or current_user.can_view_my_expense_reports() and not current_user.can_view_expense_reports():
        # Mostra solo le note spese dell'utente corrente
        query = query.filter(ExpenseReport.employee_id == current_user.id)
        page_title = "Le Mie Note Spese"
    elif current_user.can_view_expense_reports() or current_user.can_approve_expense_reports():
        # Utente può vedere tutte le note spese (eventualmente filtrate per sede)
        if not current_user.all_sedi and current_user.sede_id:
            # Filtra per sede se non ha accesso globale
            from models import User
            sede_users = User.query.filter(User.sede_id == current_user.sede_id).with_entities(User.id).all()
            sede_user_ids = [u.id for u in sede_users]
            query = query.filter(ExpenseReport.employee_id.in_(sede_user_ids))
        page_title = "Note Spese"
    else:
        # Fallback: mostra solo le proprie
        query = query.filter(ExpenseReport.employee_id == current_user.id)
        page_title = "Le Mie Note Spese"
    
    # Applica filtri se presenti
    if filter_form.validate_on_submit():
        if filter_form.employee_id.data:
            query = query.filter(ExpenseReport.employee_id == filter_form.employee_id.data)
        if filter_form.category_id.data:
            query = query.filter(ExpenseReport.category_id == filter_form.category_id.data)
        if filter_form.status.data:
            query = query.filter(ExpenseReport.status == filter_form.status.data)
        if filter_form.date_from.data:
            query = query.filter(ExpenseReport.expense_date >= filter_form.date_from.data)
        if filter_form.date_to.data:
            query = query.filter(ExpenseReport.expense_date <= filter_form.date_to.data)
    
    # Ordina per data più recente
    expenses = query.order_by(ExpenseReport.expense_date.desc(), ExpenseReport.created_at.desc()).all()
    
    return render_template('expense_reports.html', 
                         expenses=expenses, 
                         filter_form=filter_form,
                         page_title=page_title,
                         view_mode=view_mode)

# ROUTE MOVED TO expense_bp blueprint - create_expense_report
    """Crea nuova nota spese"""
    if not current_user.can_create_expense_reports():
        flash('Non hai i permessi per creare note spese', 'danger')
        return redirect(url_for('expense_reports'))
    
    from models import ExpenseReport, ExpenseCategory
    from forms import ExpenseReportForm
    import os
    from werkzeug.utils import secure_filename
    
    form = ExpenseReportForm()
    
    if form.validate_on_submit():
        # Gestione upload file
        receipt_filename = None
        if form.receipt_file.data:
            file = form.receipt_file.data
            filename = secure_filename(file.filename)
            
            # Crea nome file unico
            import uuid
            file_extension = filename.rsplit('.', 1)[1].lower() if '.' in filename else ''
            unique_filename = f"{uuid.uuid4().hex}.{file_extension}"
            
            # Crea directory uploads se non esiste
            upload_dir = os.path.join(app.root_path, 'static', 'uploads', 'expenses')
            os.makedirs(upload_dir, exist_ok=True)
            
            file_path = os.path.join(upload_dir, unique_filename)
            file.save(file_path)
            receipt_filename = unique_filename
        
        # Crea nota spese
        expense = ExpenseReport(
            employee_id=current_user.id,
            expense_date=form.expense_date.data,
            description=form.description.data,
            amount=form.amount.data,
            category_id=form.category_id.data,
            receipt_filename=receipt_filename
        )
        
        db.session.add(expense)
        db.session.commit()
        
        flash('Nota spese creata con successo', 'success')
        return redirect(url_for('expense_reports'))
    
    return render_template('create_expense_report.html', form=form)

# ROUTE MOVED TO expense_bp blueprint - edit_expense_report
    """Modifica nota spese esistente"""
    from models import ExpenseReport
    from forms import ExpenseReportForm
    import os
    from werkzeug.utils import secure_filename
    
    expense = ExpenseReport.query.get_or_404(expense_id)
    
    # Verifica permessi
    if expense.employee_id != current_user.id and not current_user.can_manage_expense_reports():
        flash('Non hai i permessi per modificare questa nota spese', 'danger')
        return redirect(url_for('expense_reports'))
    
    # Verifica se modificabile
    if not expense.can_be_edited():
        flash('Questa nota spese non può più essere modificata', 'warning')
        return redirect(url_for('expense_reports'))
    
    form = ExpenseReportForm()
    
    if form.validate_on_submit():
        # Gestione upload file
        if form.receipt_file.data:
            # Elimina vecchio file se esiste
            if expense.receipt_filename:
                old_file_path = os.path.join(app.root_path, 'static', 'uploads', 'expenses', expense.receipt_filename)
                if os.path.exists(old_file_path):
                    os.remove(old_file_path)
            
            file = form.receipt_file.data
            filename = secure_filename(file.filename)
            
            # Crea nome file unico
            import uuid
            file_extension = filename.rsplit('.', 1)[1].lower() if '.' in filename else ''
            unique_filename = f"{uuid.uuid4().hex}.{file_extension}"
            
            # Crea directory uploads se non esiste
            upload_dir = os.path.join(app.root_path, 'static', 'uploads', 'expenses')
            os.makedirs(upload_dir, exist_ok=True)
            
            file_path = os.path.join(upload_dir, unique_filename)
            file.save(file_path)
            expense.receipt_filename = unique_filename
        
        # Aggiorna dati
        expense.expense_date = form.expense_date.data
        expense.description = form.description.data
        expense.amount = form.amount.data
        expense.category_id = form.category_id.data
        
        db.session.commit()
        flash('Nota spese aggiornata con successo', 'success')
        return redirect(url_for('expense_reports'))
    
    # Precompila form con dati esistenti
    form.expense_date.data = expense.expense_date
    form.description.data = expense.description
    form.amount.data = expense.amount
    form.category_id.data = expense.category_id
    
    return render_template('edit_expense_report.html', form=form, expense=expense)

# ROUTE MOVED TO expense_bp blueprint - approve_expense_report
    """Approva/rifiuta nota spese"""
    from models import ExpenseReport
    from forms import ExpenseApprovalForm
    
    expense = ExpenseReport.query.get_or_404(expense_id)
    
    # Verifica permessi
    if not expense.can_be_approved_by(current_user):
        flash('Non hai i permessi per approvare questa nota spese', 'danger')
        return redirect(url_for('expense_reports'))
    
    if expense.status != 'pending':
        flash('Questa nota spese è già stata processata', 'warning')
        return redirect(url_for('expense_reports'))
    
    form = ExpenseApprovalForm()
    
    if form.validate_on_submit():
        if form.action.data == 'approve':
            expense.approve(current_user, form.comment.data)
            flash('Nota spese approvata con successo', 'success')
        else:
            expense.reject(current_user, form.comment.data)
            flash('Nota spese rifiutata', 'info')
        
        db.session.commit()
        return redirect(url_for('expense_reports'))
    
    return render_template('approve_expense_report.html', form=form, expense=expense)

# ROUTE MOVED TO expense_bp blueprint - download_expense_receipt
    """Download ricevuta allegata"""
    from models import ExpenseReport
    from flask import send_file
    import os
    
    expense = ExpenseReport.query.get_or_404(expense_id)
    
    # Verifica permessi
    if (expense.employee_id != current_user.id and 
        not current_user.can_view_expense_reports() and 
        not current_user.can_approve_expense_reports()):
        flash('Non hai i permessi per scaricare questo documento', 'danger')
        return redirect(url_for('expense_reports'))
    
    if not expense.receipt_filename:
        flash('Nessun documento allegato a questa nota spese', 'warning')
        return redirect(url_for('expense_reports'))
    
    file_path = os.path.join(app.root_path, 'static', 'uploads', 'expenses', expense.receipt_filename)
    
    if not os.path.exists(file_path):
        flash('File non trovato', 'danger')
        return redirect(url_for('expense_reports'))
    
    return send_file(file_path, as_attachment=True, 
                    download_name=f"ricevuta_{expense.id}_{expense.expense_date.strftime('%Y%m%d')}.{expense.receipt_filename.split('.')[-1]}")

# ROUTE MOVED TO expense_bp blueprint - expense_categories
    """Gestisci categorie note spese"""
    if not current_user.can_manage_expense_reports():
        flash('Non hai i permessi per gestire le categorie', 'danger')
        return redirect(url_for('expense_reports'))
    
    from models import ExpenseCategory
    categories = ExpenseCategory.query.order_by(ExpenseCategory.name).all()
    
    return render_template('expense_categories.html', categories=categories)

# ROUTE MOVED TO expense_bp blueprint - create_expense_category
    """Crea nuova categoria"""
    if not current_user.can_manage_expense_reports():
        flash('Non hai i permessi per creare categorie', 'danger')
        return redirect(url_for('expense_reports'))
    
    from models import ExpenseCategory
    from forms import ExpenseCategoryForm
    
    form = ExpenseCategoryForm()
    
    if form.validate_on_submit():
        category = ExpenseCategory(
            name=form.name.data,
            description=form.description.data,
            active=form.active.data,
            created_by=current_user.id
        )
        
        db.session.add(category)
        
        try:
            db.session.commit()
            flash('Categoria creata con successo', 'success')
            return redirect(url_for('expense_categories'))
        except:
            db.session.rollback()
            flash('Errore: categoria già esistente', 'danger')
    
    return render_template('create_expense_category.html', form=form)

# ROUTE MOVED TO expense_bp blueprint - edit_expense_category
    """Modifica categoria nota spese"""
    if not current_user.can_manage_expense_reports():
        flash('Non hai i permessi per modificare le categorie', 'danger')
        return redirect(url_for('expense_categories'))
    
    from models import ExpenseCategory
    from forms import ExpenseCategoryForm
    
    category = ExpenseCategory.query.get_or_404(category_id)
    form = ExpenseCategoryForm(obj=category)
    
    if form.validate_on_submit():
        category.name = form.name.data
        category.description = form.description.data
        category.active = form.active.data
        
        try:
            db.session.commit()
            flash('Categoria modificata con successo', 'success')
            return redirect(url_for('expense_categories'))
        except:
            db.session.rollback()
            flash('Errore: nome categoria già esistente', 'danger')
    
    return render_template('edit_expense_category.html', form=form, category=category)

# ROUTE MOVED TO expense_bp blueprint - delete_expense_category
    """Elimina categoria nota spese"""
    if not current_user.can_manage_expense_reports():
        flash('Non hai i permessi per eliminare le categorie', 'danger')
        return redirect(url_for('expense_categories'))
    
    from models import ExpenseCategory
    
    category = ExpenseCategory.query.get_or_404(category_id)
    
    # Verifica se ci sono note spese associate
    if category.expense_reports and len(category.expense_reports) > 0:
        flash('Non è possibile eliminare una categoria con note spese associate', 'warning')
        return redirect(url_for('expense_categories'))
    
    try:
        name = category.name
        db.session.delete(category)
        db.session.commit()
        flash(f'Categoria "{name}" eliminata con successo', 'success')
    except:
        db.session.rollback()
        flash('Errore nell\'eliminazione della categoria', 'danger')
    
    return redirect(url_for('expense_categories'))

# ROUTE MOVED TO expense_bp blueprint - delete_expense_report
    """Elimina nota spese"""
    from models import ExpenseReport
    import os
    
    expense = ExpenseReport.query.get_or_404(expense_id)
    
    # Verifica permessi
    if expense.employee_id != current_user.id and not current_user.can_manage_expense_reports():
        flash('Non hai i permessi per eliminare questa nota spese', 'danger')
        return redirect(url_for('expense_reports'))
    
    # Solo note in attesa possono essere eliminate
    if expense.status != 'pending':
        flash('Solo le note spese in attesa possono essere eliminate', 'warning')
        return redirect(url_for('expense_reports'))
    
    # Elimina file allegato se esiste
    if expense.receipt_filename:
        file_path = os.path.join(app.root_path, 'static', 'uploads', 'expenses', expense.receipt_filename)
        if os.path.exists(file_path):
            os.remove(file_path)
    
    db.session.delete(expense)
    db.session.commit()
    
    flash('Nota spese eliminata con successo', 'success')
    return redirect(url_for('expense_reports'))

# =============================================================================
# OVERTIME MANAGEMENT ROUTES
# =============================================================================

# ROUTE MOVED TO expense_bp blueprint - overtime_types
    """Visualizzazione e gestione tipologie straordinari"""
    if not (current_user.can_manage_overtime_types() or current_user.can_view_overtime_types()):
        flash('Non hai i permessi per visualizzare le tipologie di straordinario.', 'warning')
        return redirect(url_for('dashboard'))
    
    types = OvertimeType.query.all()
    return render_template('overtime_types.html', types=types)

# ROUTE MOVED TO expense_bp blueprint - create_overtime_type
    """Creazione nuova tipologia straordinario"""
    if not (current_user.can_manage_overtime_types() or current_user.can_create_overtime_types()):
        flash('Non hai i permessi per creare tipologie di straordinario.', 'warning')
        return redirect(url_for('overtime_types'))
    
    form = OvertimeTypeForm()
    if form.validate_on_submit():
        overtime_type = OvertimeType(
            name=form.name.data,
            description=form.description.data,
            hourly_rate_multiplier=form.hourly_rate_multiplier.data,
            active=form.active.data
        )
        db.session.add(overtime_type)
        db.session.commit()
        flash('Tipologia straordinario creata con successo!', 'success')
        return redirect(url_for('overtime_types'))
    
    return render_template('create_overtime_type.html', form=form)

# ROUTE MOVED TO expense_bp blueprint - overtime_requests_management
    """Visualizzazione richieste straordinari"""
    if not current_user.can_view_overtime_requests():
        flash('Non hai i permessi per visualizzare le richieste di straordinario.', 'warning')
        return redirect(url_for('dashboard'))
    
    # Filtro per sede con join esplicito per evitare ambiguità
    if current_user.all_sedi:
        requests = OvertimeRequest.query.options(
            joinedload(OvertimeRequest.employee),
            joinedload(OvertimeRequest.overtime_type)
        ).order_by(OvertimeRequest.created_at.desc()).all()
    else:
        requests = OvertimeRequest.query.join(
            User, OvertimeRequest.employee_id == User.id
        ).filter(
            User.sede_id == current_user.sede_id
        ).options(
            joinedload(OvertimeRequest.employee),
            joinedload(OvertimeRequest.overtime_type)
        ).order_by(OvertimeRequest.created_at.desc()).all()
    
    # Crea form per filtri
    form = OvertimeFilterForm()
    return render_template('overtime_requests.html', requests=requests, form=form)

# ROUTE MOVED TO expense_bp blueprint - create_overtime_request
    """Creazione richiesta straordinario"""
    if not current_user.can_create_overtime_requests():
        flash('Non hai i permessi per creare richieste di straordinario.', 'warning')
        return redirect(url_for('dashboard'))
    
    form = OvertimeRequestForm()
    if form.validate_on_submit():
        # Calcola le ore di straordinario
        start_datetime = datetime.combine(form.overtime_date.data, form.start_time.data)
        end_datetime = datetime.combine(form.overtime_date.data, form.end_time.data)
        hours = (end_datetime - start_datetime).total_seconds() / 3600
        
        overtime_request = OvertimeRequest(
            employee_id=current_user.id,
            overtime_date=form.overtime_date.data,
            start_time=form.start_time.data,
            end_time=form.end_time.data,

            motivation=form.motivation.data,
            overtime_type_id=form.overtime_type_id.data,
            status='pending'
        )
        db.session.add(overtime_request)
        db.session.commit()
        
        # Invia notifica automatica agli approvatori
        send_overtime_request_message(overtime_request, 'created', current_user)
        
        flash('Richiesta straordinario inviata con successo!', 'success')
        return redirect(url_for('my_overtime_requests'))
    
    return render_template('create_overtime_request.html', form=form)

# ROUTE MOVED TO expense_bp blueprint - my_overtime_requests
    """Le mie richieste straordinari"""
    if not current_user.can_view_my_overtime_requests():
        flash('Non hai i permessi per visualizzare le tue richieste di straordinario.', 'warning')
        return redirect(url_for('dashboard'))
    
    requests = OvertimeRequest.query.filter_by(employee_id=current_user.id).options(
        joinedload(OvertimeRequest.overtime_type)
    ).order_by(OvertimeRequest.created_at.desc()).all()
    
    return render_template('my_overtime_requests.html', requests=requests)

# ROUTE MOVED TO expense_bp blueprint - approve_overtime_request

# ROUTE MOVED TO expense_bp blueprint - reject_overtime_request

# ROUTE MOVED TO expense_bp blueprint - delete_overtime_request

# ROUTE MOVED TO expense_bp blueprint - overtime_requests_excel

# ROUTE MOVED TO expense_bp blueprint - edit_overtime_type

# ROUTE MOVED TO expense_bp blueprint - delete_overtime_type

# =============================================================================
# MILEAGE REIMBURSEMENT ROUTES
# =============================================================================

# ROUTE MOVED TO expense_bp blueprint - mileage_requests

# ROUTE MOVED TO expense_bp blueprint - create_mileage_request

# ROUTE MOVED TO expense_bp blueprint - my_mileage_requests

# ROUTE MOVED TO expense_bp blueprint - approve_mileage_request

# ROUTE MOVED TO expense_bp blueprint - delete_mileage_request

# ROUTE MOVED TO expense_bp blueprint - calculate_distance API

def calculate_approximate_distance(start_address, end_address):
    """Calcola distanza approssimativa tra due indirizzi italiani"""
    
    # Database semplificato di coordinate delle principali città italiane
    city_coords = {
        'roma': (41.9028, 12.4964),
        'milano': (45.4642, 9.1900),
        'napoli': (40.8518, 14.2681),
        'torino': (45.0703, 7.6869),
        'firenze': (43.7696, 11.2558),
        'bologna': (44.4949, 11.3426),
        'genova': (44.4056, 8.9463),
        'palermo': (38.1157, 13.3613),
        'bari': (41.1171, 16.8719),
        'catania': (37.5079, 15.0830),
        'venezia': (45.4408, 12.3155),
        'verona': (45.4384, 10.9916),
        'messina': (38.1938, 15.5540),
        'padova': (45.4064, 11.8768),
        'trieste': (45.6495, 13.7768),
        'brescia': (45.5416, 10.2118),
        'parma': (44.8015, 10.3279),
        'modena': (44.6471, 10.9252),
        'reggio calabria': (38.1059, 15.6219),
        'perugia': (43.1122, 12.3888)
    }
    
    def extract_city(address):
        """Estrae il nome della città dall'indirizzo"""
        address = address.lower().strip()
        
        # Cerca la città nell'indirizzo
        for city in city_coords.keys():
            if city in address:
                return city
        
        # Se non trova la città, prova a estrarre la prima parola
        first_word = address.split(',')[0].strip()
        if first_word in city_coords:
            return first_word
            
        return None
    
    def haversine_distance(coord1, coord2):
        """Calcola la distanza in km tra due coordinate usando la formula di Haversine"""
        import math
        
        lat1, lon1 = coord1
        lat2, lon2 = coord2
        
        # Converti in radianti
        lat1, lon1, lat2, lon2 = map(math.radians, [lat1, lon1, lat2, lon2])
        
        # Formula di Haversine
        dlat = lat2 - lat1
        dlon = lon2 - lon1
        a = math.sin(dlat/2)**2 + math.cos(lat1) * math.cos(lat2) * math.sin(dlon/2)**2
        c = 2 * math.asin(math.sqrt(a))
        
        # Raggio della Terra in km
        r = 6371
        
        return c * r
    
    # Estrai le città dagli indirizzi
    start_city = extract_city(start_address)
    end_city = extract_city(end_address)
    
    if start_city and end_city and start_city in city_coords and end_city in city_coords:
        # Calcola la distanza reale tra le città
        distance = haversine_distance(city_coords[start_city], city_coords[end_city])
        # Aggiungi un fattore di correzione per le strade (circa 1.3x la distanza in linea d'aria)
        return distance * 1.3
    else:
        # Fallback: distanza approssimativa basata sulla lunghezza degli indirizzi
        return max(20, min(100, len(start_address) + len(end_address)))

# ROUTE MOVED TO expense_bp blueprint - export_mileage_requests

# =============================================
# SISTEMA ACI - BACK OFFICE AMMINISTRATORE
# =============================================

# ACI HELPER FUNCTION MIGRATED TO aci_bp blueprint - admin_required

# ACI ROUTE MIGRATED TO aci_bp blueprint - aci_tables
    """Visualizza tabelle ACI con caricamento lazy - record caricati solo dopo filtro"""
    form = ACIFilterForm()
    tables = []
    total_records = ACITable.query.count()
    
    # LAZY LOADING: carica record solo se è stato applicato un filtro (POST)
    if request.method == "POST" and form.validate_on_submit():
        query = ACITable.query
        
        # Applica filtri selezionati
        filters_applied = False
        if form.tipologia.data:
            query = query.filter(ACITable.tipologia == form.tipologia.data)
            filters_applied = True
        if form.marca.data:
            query = query.filter(ACITable.marca == form.marca.data)
            filters_applied = True
        if form.modello.data:
            query = query.filter(ACITable.modello == form.modello.data)
            filters_applied = True
            
        # Carica risultati solo se almeno un filtro è applicato
        if filters_applied:
            tables = query.order_by(ACITable.tipologia, ACITable.marca, ACITable.modello).all()
            import logging
            logging.info(f"ACI Tables: Caricati {len(tables)} record con filtri applicati")
        else:
            # Se nessun filtro è selezionato ma form è stato inviato, mostra messaggio
            flash("⚠️ Seleziona almeno un filtro prima di cercare", "warning")
    
    return render_template("aci_tables.html", 
                         tables=tables, 
                         form=form, 
                         total_records=total_records,
                         show_results=(request.method == "POST"))

# ACI API ROUTE MIGRATED TO aci_bp blueprint - api_aci_marcas
    """API per ottenere le marche filtrate per tipologia"""
    tipologia = request.args.get('tipologia')
    
    query = db.session.query(ACITable.marca).distinct()
    
    if tipologia:
        query = query.filter(ACITable.tipologia == tipologia)
    
    marcas = [row.marca for row in query.order_by(ACITable.marca).all()]
    return jsonify(marcas)

# ACI API ROUTE MIGRATED TO aci_bp blueprint - api_aci_modelos
    """API per ottenere i modelli filtrati per tipologia e marca"""
    tipologia = request.args.get('tipologia')
    marca = request.args.get('marca')
    
    query = db.session.query(ACITable.modello).distinct()
    
    if tipologia:
        query = query.filter(ACITable.tipologia == tipologia)
    if marca:
        query = query.filter(ACITable.marca == marca)
    
    modelos = [row.modello for row in query.order_by(ACITable.modello).all()]
    return jsonify(modelos)

# ACI ROUTE MIGRATED TO aci_bp blueprint - aci_upload
    """Upload e importazione file Excel ACI"""
    form = ACIUploadForm()
    
    if form.validate_on_submit():
        file = form.excel_file.data
        tipologia = form.tipologia.data
        
        try:
            # Salva file temporaneo
            import tempfile
            import os
            import pandas as pd
            
            with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
                file.save(tmp_file.name)
                
                # Leggi file Excel con ottimizzazioni per file grandi
                df = pd.read_excel(tmp_file.name, engine='openpyxl', 
                                 usecols=[0, 1, 2],  # Leggi solo colonne A, B, C
                                 dtype={0: 'str', 1: 'str', 2: 'float64'},  # Forza tipi di dato
                                 na_filter=True)
                
                # Pulisci file temporaneo
                os.unlink(tmp_file.name)
            
            # Verifica struttura file Excel - richiede almeno 3 colonne (A, B, C)
            if len(df.columns) < 3:
                flash(f"Errore: Il file Excel deve avere almeno 3 colonne (MARCA, MODELLO, COSTO KM). Trovate {len(df.columns)} colonne.", "danger")
                return render_template("aci_upload.html", form=form)
            
            # Log struttura per debug
            import logging
            logging.info(f"Colonne Excel trovate: {list(df.columns)}")
            logging.info(f"Utilizzo solo colonne A, B, C - ignorando tutte le altre")
            
            # Usa il nome del file come tipologia se non specificata esplicitamente
            if not tipologia.strip():
                # Estrae nome file senza estensione
                import os
                filename = file.filename or "Excel_File"
                tipologia = os.path.splitext(filename)[0]
            
            # Processa dati Excel - SOLO COLONNE A, B, C con ottimizzazione per file grandi
            imported_count = 0
            skipped_count = 0
            batch_size = 100  # Processa a lotti per evitare timeout
            
            # Converti DataFrame in lista per processamento più veloce
            rows_data = []
            for index, row in df.iterrows():
                # UTILIZZO SOLO LE PRIME 3 COLONNE (A, B, C)
                marca = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ""
                modello = str(row.iloc[1]).strip() if pd.notna(row.iloc[1]) else ""
                costo_km = row.iloc[2] if pd.notna(row.iloc[2]) else None
                
                # Skip righe vuote o incomplete
                if not marca or not modello or modello == 'nan' or costo_km is None:
                    skipped_count += 1
                    continue
                    
                rows_data.append((marca, modello, costo_km))
            
            logging.info(f"Processando {len(rows_data)} righe valide in batch da {batch_size}")
            
            # Processa in batch per evitare timeout
            for i in range(0, len(rows_data), batch_size):
                batch = rows_data[i:i + batch_size]
                
                for marca, modello, costo_km in batch:
                    # Verifica duplicati esistenti
                    existing = ACITable.query.filter_by(
                        tipologia=tipologia,
                        marca=marca,
                        modello=modello
                    ).first()
                    
                    if existing:
                        # Aggiorna record esistente - SOLO COSTO KM
                        existing.costo_km = float(costo_km)
                        imported_count += 1
                    else:
                        # Crea nuovo record ACI
                        aci_record = ACITable(
                            tipologia=tipologia,
                            marca=marca,
                            modello=modello,
                            costo_km=float(costo_km)
                        )
                        db.session.add(aci_record)
                        imported_count += 1
                
                # Commit ogni batch per liberare memoria
                try:
                    db.session.commit()
                    logging.info(f"Batch {i//batch_size + 1} completato: {len(batch)} record")
                except Exception as batch_error:
                    db.session.rollback()
                    logging.error(f"Errore batch {i//batch_size + 1}: {batch_error}")
                    raise batch_error
            
            # Messaggi di feedback dettagliati
            if imported_count > 0:
                flash(f"✅ File Excel importato con successo!", "success")
                flash(f"📊 {imported_count} record processati (nuovi o aggiornati)", "info")
                if skipped_count > 0:
                    flash(f"⏭️ {skipped_count} righe saltate (intestazioni o righe vuote)", "info")
                flash(f"🚫 Importate solo colonne A, B, C (MARCA, MODELLO, COSTO KM) - tutte le altre colonne ignorate", "info")
            else:
                flash("⚠️ Nessun record valido trovato nel file Excel", "warning")
                
            return redirect(url_for("aci_tables"))
            
        except Exception as e:
            db.session.rollback()
            flash(f"Errore durante l'importazione del file: {str(e)}", "danger")
    
    return render_template("aci_upload.html", form=form)

# ACI ROUTE MIGRATED TO aci_bp blueprint - aci_create
    """Crea nuovo record ACI manualmente"""
    form = ACIRecordForm()
    
    if form.validate_on_submit():
        try:
            aci_record = ACITable(
                tipologia=form.tipologia.data,
                marca=form.marca.data,
                modello=form.modello.data,
                costo_km=form.costo_km.data
            )
            
            db.session.add(aci_record)
            db.session.commit()
            flash("Record ACI creato con successo!", "success")
            return redirect(url_for("aci_tables"))
            
        except Exception as e:
            db.session.rollback()
            flash(f"Errore durante la creazione del record: {str(e)}", "danger")
    
    return render_template("aci_create.html", form=form)

# ACI ROUTE MIGRATED TO aci_bp blueprint - aci_edit
    """Modifica record ACI esistente"""
    aci_record = ACITable.query.get_or_404(record_id)
    form = ACIRecordForm(obj=aci_record)
    
    if form.validate_on_submit():
        try:
            aci_record.tipologia = form.tipologia.data
            aci_record.marca = form.marca.data
            aci_record.modello = form.modello.data
            aci_record.costo_km = form.costo_km.data
            
            db.session.commit()
            flash("Record ACI aggiornato con successo!", "success")
            return redirect(url_for("aci_tables"))
            
        except Exception as e:
            db.session.rollback()
            flash(f"Errore durante l'aggiornamento del record: {str(e)}", "danger")
    
    return render_template("aci_edit.html", form=form, record=aci_record)

# ACI ROUTE MIGRATED TO aci_bp blueprint - aci_delete
    """Cancella record ACI"""
    try:
        aci_record = ACITable.query.get_or_404(record_id)
        db.session.delete(aci_record)
        db.session.commit()
        flash("Record ACI cancellato con successo!", "success")
        
    except Exception as e:
        db.session.rollback()
        flash(f"Errore durante la cancellazione: {str(e)}", "danger")
    
    return redirect(url_for("aci_tables"))

# ACI ROUTE MIGRATED TO aci_bp blueprint - aci_export
    """Export Excel delle tabelle ACI"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from io import BytesIO
        
        # Recupera tutti i record ACI
        records = ACITable.query.order_by(ACITable.tipologia, ACITable.tipo, ACITable.marca, ACITable.modello).all()
        
        # Crea workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Tabelle ACI"
        
        # Intestazioni
        headers = [
            "ID", "Tipologia", "Tipo", "Marca", "Modello", 
            "Costo KM", "Fringe Benefit 10%", "Fringe Benefit 25%", 
            "Fringe Benefit 30%", "Fringe Benefit 50%", 
            "Creato il", "Aggiornato il"
        ]
        
        # Stile intestazione
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Dati
        for row, record in enumerate(records, 2):
            ws.cell(row=row, column=1, value=record.id)
            ws.cell(row=row, column=2, value=record.tipologia)
            ws.cell(row=row, column=3, value=record.tipo)
            ws.cell(row=row, column=4, value=record.marca)
            ws.cell(row=row, column=5, value=record.modello)
            ws.cell(row=row, column=6, value=float(record.costo_km) if record.costo_km else None)
            ws.cell(row=row, column=7, value=float(record.fringe_benefit_10) if record.fringe_benefit_10 else None)
            ws.cell(row=row, column=8, value=float(record.fringe_benefit_25) if record.fringe_benefit_25 else None)
            ws.cell(row=row, column=9, value=float(record.fringe_benefit_30) if record.fringe_benefit_30 else None)
            ws.cell(row=row, column=10, value=float(record.fringe_benefit_50) if record.fringe_benefit_50 else None)
            ws.cell(row=row, column=11, value=record.created_at.strftime('%d/%m/%Y %H:%M') if record.created_at else "")
            ws.cell(row=row, column=12, value=record.updated_at.strftime('%d/%m/%Y %H:%M') if record.updated_at else "")
        
        # Autofit colonne
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Salva in BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Response
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename=tabelle_aci_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        return response
        
    except Exception as e:
        flash(f"Errore durante l'export: {str(e)}", "danger")
        return redirect(url_for("aci_tables"))

# ACI ROUTE MIGRATED TO aci_bp blueprint - aci_bulk_delete
    """Cancellazione in massa per tipologia"""
    tipologia = request.form.get('tipologia')
    if not tipologia:
        flash("Tipologia non specificata.", "warning")
        return redirect(url_for("aci_tables"))
    
    try:
        # Prima conta i record da cancellare
        records_to_delete = ACITable.query.filter_by(tipologia=tipologia).count()
        
        if records_to_delete == 0:
            flash("Nessun record trovato per la tipologia specificata.", "info")
            return redirect(url_for("aci_tables"))
        
        # Esegui la cancellazione
        deleted_count = ACITable.query.filter_by(tipologia=tipologia).delete()
        db.session.commit()
        
        flash(f"✅ Cancellazione completata con successo!", "success")
        flash(f"📊 {deleted_count} record eliminati dalla tipologia '{tipologia}'", "info")
        
        import logging
        logging.info(f"Admin {current_user.username} ha cancellato {deleted_count} record ACI tipologia '{tipologia}'")
        
    except Exception as e:
        db.session.rollback()
        flash(f"❌ Errore durante la cancellazione in massa: {str(e)}", "danger")
        import logging
        logging.error(f"Errore cancellazione bulk ACI: {str(e)}")
    
    return redirect(url_for("aci_tables"))

# =============================================================================
# BLUEPRINT REGISTRATION
# =============================================================================

# Blueprint registration moved to main.py to avoid circular imports
