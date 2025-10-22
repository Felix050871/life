# =============================================================================
# DASHBOARD ROUTES BLUEPRINT
# =============================================================================

from flask import Blueprint, render_template, request, redirect, url_for, flash, make_response
from flask_login import login_required, current_user
from datetime import datetime, date, timedelta, time

# Application Imports
from app import db
from models import (
    User, AttendanceEvent, LeaveRequest, Shift, PresidioCoverageTemplate, 
    ReperibilitaShift, Intervention, ExpenseReport, OvertimeRequest, MileageRequest,
    Sede, italian_now
)
from utils import get_user_statistics, get_team_statistics, format_hours
from utils_tenant import filter_by_company

# Create Blueprint
dashboard_bp = Blueprint('dashboard', __name__)

@dashboard_bp.route('/')
def index():
    """Main entry point - redirect to appropriate dashboard"""
    if current_user.is_authenticated:
        return redirect(url_for('dashboard.dashboard'))
    return redirect(url_for('auth.login'))

@dashboard_bp.route('/dashboard')
@login_required
def dashboard():
    # Verifica permessi di accesso alla dashboard
    if not current_user.can_access_dashboard():
        flash('Non hai i permessi per accedere alla dashboard.', 'danger')
        return redirect(url_for('auth.login'))
    
    stats = get_user_statistics(current_user.id)
    
    # Widget statistics team (solo per utenti autorizzati)
    team_stats = None
    if current_user.can_view_team_stats_widget():
        team_stats = get_team_statistics()
    
    # Get today's attendance events
    today_events_check = filter_by_company(AttendanceEvent.query).filter(
        AttendanceEvent.user_id == current_user.id,
        AttendanceEvent.date == date.today()
    ).first()
    today_attendance = today_events_check  # Per compatibilità con il template
    
    # Get daily status variables for the status section
    today_date = date.today()
    
    # Initialize variables for all users
    user_status = 'out'
    today_work_hours = 0
    today_events = []
    
    # Get current user's status and today's events (for all users who can view their own attendance)
    if current_user.can_view_my_attendance() or current_user.can_view_attendance():
        user_status, _ = AttendanceEvent.get_user_status(current_user.id, today_date)
        today_events = AttendanceEvent.get_daily_events(current_user.id, today_date)
        today_work_hours = AttendanceEvent.get_daily_work_hours(current_user.id, today_date)
    
    # Get presidio coverage templates for shifts widget
    upcoming_shifts = []
    if current_user.can_view_shifts():
        # Show active presidio coverage templates
        upcoming_shifts = filter_by_company(PresidioCoverageTemplate.query).filter_by(active=True).order_by(PresidioCoverageTemplate.start_date.desc()).all()
    
    # Widget team management data (per Management e Responsabile)
    team_management_data = []
    if current_user.can_view_team_management_widget():
        users_to_check = []
        
        if current_user.role == 'Amministratore':
            # Admin vede tutti gli utenti attivi
            users_to_check = filter_by_company(User.query).filter(
                User.active.is_(True),
                ~User.role.in_(['Admin', 'Staff'])  # Escludi admin e staff
            ).order_by(User.last_name, User.first_name).all()
        elif current_user.role == 'Management' and current_user.all_sedi:
            # Management multi-sede vede tutti gli utenti
            users_to_check = filter_by_company(User.query).filter(
                User.active.is_(True),
                ~User.role.in_(['Admin', 'Staff'])
            ).order_by(User.last_name, User.first_name).all()
        elif current_user.sede_id:
            # Responsabile e Management vede utenti della propria sede
            users_to_check = filter_by_company(User.query).filter(
                User.sede_id == current_user.sede_id,
                User.active.is_(True),
                ~User.role.in_(['Admin', 'Staff'])
            ).order_by(User.last_name, User.first_name).all()
        
        for user in users_to_check:
            user_today_status, last_event = AttendanceEvent.get_user_status(user.id, today_date)
            team_management_data.append({
                'user': user,
                'status': user_today_status,
                'last_event': last_event
            })
    
    # Widget reperibilità (on-call) shifts
    upcoming_reperibilita_shifts = []
    active_intervention = None
    active_general_intervention = None
    recent_interventions = []
    
    if current_user.can_view_my_reperibilita():
        # My upcoming reperibilità shifts (next 7 days)
        seven_days_from_now = date.today() + timedelta(days=7)
        upcoming_reperibilita_shifts = filter_by_company(ReperibilitaShift.query).filter(
            ReperibilitaShift.user_id == current_user.id,
            ReperibilitaShift.date >= date.today(),
            ReperibilitaShift.date <= seven_days_from_now
        ).order_by(ReperibilitaShift.date, ReperibilitaShift.start_time).all()
        
        # Active intervention for current user
        active_intervention = filter_by_company(Intervention.query).filter(
            Intervention.user_id == current_user.id,
            Intervention.end_datetime.is_(None)
        ).first()
        
        # Recent interventions (last 7 days)
        seven_days_ago = datetime.now() - timedelta(days=7)
        recent_interventions = filter_by_company(Intervention.query).filter(
            Intervention.user_id == current_user.id,
            Intervention.start_datetime >= seven_days_ago
        ).order_by(Intervention.start_datetime.desc()).limit(5).all()
    
    # Active general intervention (per tutti gli utenti)
    active_general_intervention = filter_by_company(Intervention.query).filter(
        Intervention.user_id.is_(None),  # Intervento generale
        Intervention.end_datetime.is_(None)
    ).first()
    
    # Widget leave requests
    recent_leaves = []
    if current_user.can_view_my_leave():
        # Recent leave requests (last 30 days)
        thirty_days_ago = date.today() - timedelta(days=30)
        recent_leaves = filter_by_company(LeaveRequest.query).filter(
            LeaveRequest.user_id == current_user.id,
            LeaveRequest.created_at >= thirty_days_ago
        ).order_by(LeaveRequest.created_at.desc()).limit(5).all()
    
    # Widget calendar view (next 7 days with shifts and events)
    weekly_calendar = []
    if current_user.can_view_shifts() or current_user.can_view_my_reperibilita():
        for i in range(7):
            day = date.today() + timedelta(days=i)
            day_data = {'date': day, 'shifts': [], 'reperibilita': []}
            
            # Get shifts for this day
            if current_user.can_view_shifts():
                day_shifts = filter_by_company(Shift.query).filter(
                    Shift.user_id == current_user.id,
                    Shift.date == day
                ).all()
                day_data['shifts'] = day_shifts
            
            # Get reperibilità for this day
            if current_user.can_view_my_reperibilita():
                day_reperibilita = filter_by_company(ReperibilitaShift.query).filter(
                    ReperibilitaShift.user_id == current_user.id,
                    ReperibilitaShift.date == day
                ).all()
                day_data['reperibilita'] = day_reperibilita
            
            weekly_calendar.append(day_data)
    
    # Widget for daily attendance data (by sede for managers)
    daily_attendance_data = {}
    if current_user.can_view_daily_attendance_widget():
        # Get users visible to current user based on role and sede access
        if current_user.role == 'Amministratore' or current_user.all_sedi:
            # Amministratori e utenti multi-sede vedono tutti gli utenti di tutte le sedi
            visible_users = filter_by_company(User.query).filter(
                User.active.is_(True),
                ~User.role.in_(['Admin', 'Staff'])
            ).all()
        elif current_user.role in ['Responsabile', 'Management']:
            # Responsabili e Management vedono solo utenti della propria sede
            visible_users = filter_by_company(User.query).filter(
                User.sede_id == current_user.sede_id,
                User.active.is_(True),
                ~User.role.in_(['Admin', 'Staff'])
            ).all()
        else:
            # Altri utenti vedono solo utenti della propria sede se specificata
            if current_user.sede_id:
                visible_users = filter_by_company(User.query).filter(
                    User.sede_id == current_user.sede_id,
                    User.active.is_(True),
                    ~User.role.in_(['Admin', 'Staff'])
                ).all()
            else:
                visible_users = []
        
        # Group users by sede and get their attendance status for today
        sede_groups = {}
        for user in visible_users:
            sede_name = user.sede_obj.name if user.sede_obj else 'Sede Non Specificata'
            if sede_name not in sede_groups:
                sede_groups[sede_name] = {
                    'total_users': 0,
                    'present_users': [],
                    'coverage_rate': 0
                }
            
            sede_groups[sede_name]['total_users'] += 1
            
            # Check if user is present today
            user_status, _ = AttendanceEvent.get_user_status(user.id, today_date)
            if user_status in ['in', 'on_break']:
                sede_groups[sede_name]['present_users'].append(user)
        
        # Calculate coverage rates
        for sede_name, data in sede_groups.items():
            if data['total_users'] > 0:
                data['coverage_rate'] = int((len(data['present_users']) / data['total_users']) * 100)
            else:
                data['coverage_rate'] = 0
        
        daily_attendance_data = sede_groups
    
    # Widget shifts coverage alerts
    shifts_coverage_alerts = []
    if current_user.can_view_shifts() and current_user.can_view_team_management_widget():
        # Check for shifts without coverage in next 7 days
        end_date = date.today() + timedelta(days=7)
        templates_to_check = filter_by_company(PresidioCoverageTemplate.query).filter(
            PresidioCoverageTemplate.active.is_(True),
            PresidioCoverageTemplate.start_date <= end_date
        ).all()
        
        for template in templates_to_check:
            template_start = max(template.start_date, date.today())
            template_end = min(template.end_date, end_date) if template.end_date else end_date
            
            current_date = template_start
            while current_date <= template_end:
                # Check if this day needs coverage and is missing
                # This is a simplified check - in real implementation you'd check against actual coverage
                shifts_coverage_alerts.append({
                    'date': current_date,
                    'template': template,
                    'missing_roles': []  # Would be calculated based on actual coverage
                })
                current_date += timedelta(days=1)
    
    # Widget my leave requests (pending and recent)
    my_leave_requests = []
    if current_user.can_view_my_leave():
        my_leave_requests = filter_by_company(LeaveRequest.query).filter(
            LeaveRequest.user_id == current_user.id
        ).order_by(LeaveRequest.created_at.desc()).limit(5).all()
    
    # Widget my shifts (next 7 days)
    my_shifts = []
    if current_user.can_view_shifts():
        seven_days_from_now = date.today() + timedelta(days=7)
        my_shifts = filter_by_company(Shift.query).filter(
            Shift.user_id == current_user.id,
            Shift.date >= date.today(),
            Shift.date <= seven_days_from_now
        ).order_by(Shift.date, Shift.start_time).all()
    
    # Widget my reperibilità (next 7 days)
    my_reperibilita = []
    if current_user.can_view_my_reperibilita():
        seven_days_from_now = date.today() + timedelta(days=7)
        my_reperibilita = filter_by_company(ReperibilitaShift.query).filter(
            ReperibilitaShift.user_id == current_user.id,
            ReperibilitaShift.date >= date.today(),
            ReperibilitaShift.date <= seven_days_from_now
        ).order_by(ReperibilitaShift.date, ReperibilitaShift.start_time).all()
    
    # Widget expense reports data
    expense_reports_data = []
    if current_user.can_view_my_expense_reports():
        recent_reports = filter_by_company(ExpenseReport.query).filter(
            ExpenseReport.employee_id == current_user.id
        ).order_by(ExpenseReport.created_at.desc()).limit(3).all()
        expense_reports_data = recent_reports
    
    # Widget overtime requests
    recent_overtime_requests = []
    my_overtime_requests = []
    if current_user.can_view_my_overtime_requests():
        my_overtime_requests = filter_by_company(OvertimeRequest.query).filter(
            OvertimeRequest.employee_id == current_user.id
        ).order_by(OvertimeRequest.created_at.desc()).limit(5).all()
    
    if current_user.can_approve_overtime_requests():
        recent_overtime_requests = filter_by_company(OvertimeRequest.query).filter(
            OvertimeRequest.status == 'pending'
        ).order_by(OvertimeRequest.created_at.desc()).limit(5).all()
    
    # Widget mileage requests
    recent_mileage_requests = []
    my_mileage_requests = []
    if current_user.can_view_my_mileage_requests():
        my_mileage_requests = filter_by_company(MileageRequest.query).filter(
            MileageRequest.user_id == current_user.id
        ).order_by(MileageRequest.created_at.desc()).limit(5).all()
    
    if current_user.can_approve_mileage_requests():
        recent_mileage_requests = filter_by_company(MileageRequest.query).filter(
            MileageRequest.status == 'pending'
        ).order_by(MileageRequest.created_at.desc()).limit(5).all()
    
    # Widget banca ore wallet
    banca_ore_wallet = None
    if current_user.can_view_my_banca_ore_widget():
        # Usa la funzione dedicata per calcolare il wallet banca ore
        from blueprints.banca_ore import calculate_banca_ore_balance
        banca_ore_wallet = calculate_banca_ore_balance(current_user.id)
    
    # Get current time for dashboard display
    today = italian_now()
    current_time = today.time()
    
    # Get all sedi for multi-sede users
    all_sedi_list = []
    if current_user.all_sedi:
        all_sedi_list = filter_by_company(Sede.query).filter_by(active=True).all()

    return render_template('dashboard.html', 
                         stats=stats, 
                         team_stats=team_stats,
                         today_attendance=today_attendance,
                         upcoming_shifts=upcoming_shifts,
                         team_management_data=team_management_data,
                         upcoming_reperibilita_shifts=upcoming_reperibilita_shifts,
                         active_intervention=active_intervention,
                         active_general_intervention=active_general_intervention,
                         recent_interventions=recent_interventions,
                         recent_leaves=recent_leaves,
                         weekly_calendar=weekly_calendar,
                         all_sedi_list=all_sedi_list,
                         today=today,
                         today_date=today_date,
                         current_time=current_time,
                         user_status=user_status,
                         today_events=today_events,
                         today_work_hours=today_work_hours,
                         daily_attendance_data=daily_attendance_data,
                         shifts_coverage_alerts=shifts_coverage_alerts,
                         my_leave_requests=my_leave_requests,
                         my_shifts=my_shifts,
                         my_reperibilita=my_reperibilita,
                         expense_reports_data=expense_reports_data,
                         recent_overtime_requests=recent_overtime_requests,
                         my_overtime_requests=my_overtime_requests,
                         recent_mileage_requests=recent_mileage_requests,
                         my_mileage_requests=my_mileage_requests,
                         banca_ore_wallet=banca_ore_wallet,
                         format_hours=format_hours)

@dashboard_bp.route('/dashboard_team')
@login_required
def dashboard_team():
    """Dashboard per visualizzare le presenze di tutte le sedi - per Management"""
    if not current_user.can_view_all_attendance():
        flash('Non hai i permessi per visualizzare questo contenuto.', 'danger')
        return redirect(url_for('dashboard.dashboard'))
    
    # Get users visible to current user based on role and sede access - consistent with attendance logic
    if current_user.role == 'Amministratore' or current_user.all_sedi:
        # Amministratori e utenti multi-sede vedono tutti gli utenti di tutte le sedi
        all_users = filter_by_company(User.query).filter(
            User.active.is_(True),
            ~User.role.in_(['Admin', 'Staff'])
        ).all()
    elif current_user.role in ['Responsabile', 'Management']:
        # Responsabili e Management vedono solo utenti della propria sede
        all_users = filter_by_company(User.query).filter(
            User.sede_id == current_user.sede_id,
            User.active.is_(True),
            ~User.role.in_(['Admin', 'Staff'])
        ).all()
    else:
        # Altri utenti vedono solo utenti della propria sede se specificata
        if current_user.sede_id:
            all_users = filter_by_company(User.query).filter(
                User.sede_id == current_user.sede_id,
                User.active.is_(True),
                ~User.role.in_(['Admin', 'Staff'])
            ).all()
        else:
            all_users = []
    
    # Parse date range from query parameters or use default FIRST
    today = date.today()
    
    # Check for custom date range in query parameters
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')
    
    if start_date_str and end_date_str:
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            # Generate period label
            if start_date == end_date:
                period_label = start_date.strftime('%d/%m/%Y')
            else:
                period_label = f"{start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')}"
        except ValueError:
            # Invalid date format, use default
            start_date = today - timedelta(days=30)
            end_date = today
            period_label = "Ultimi 30 giorni"
    else:
        # No custom range specified, use default (last 30 days)
        start_date = today - timedelta(days=30)
        end_date = today
        period_label = "Ultimi 30 giorni"
    
    # Build user data with attendance status for ALL dates in range
    users_data = []
    
    # Generate list of all dates in range
    current_date = start_date
    date_list = []
    while current_date <= end_date:
        date_list.append(current_date)
        current_date += timedelta(days=1)
    
    # OPTIMIZATION: Single query to fetch ALL events for ALL users in the date range
    user_ids = [u.id for u in all_users]
    
    if user_ids:
        all_events = filter_by_company(AttendanceEvent.query).filter(
            AttendanceEvent.user_id.in_(user_ids),
            AttendanceEvent.timestamp >= datetime.combine(start_date, time.min),
            AttendanceEvent.timestamp <= datetime.combine(end_date, time.max)
        ).order_by(AttendanceEvent.timestamp).all()
        
        # Build a map: (user_id, date) -> [events]
        events_by_user_date = {}
        for event in all_events:
            event_date = event.timestamp.date()
            key = (event.user_id, event_date)
            if key not in events_by_user_date:
                events_by_user_date[key] = []
            events_by_user_date[key].append(event)
        
        # Process each user/date combination using the pre-fetched data
        for user in all_users:
            for check_date in date_list:
                key = (user.id, check_date)
                daily_events = events_by_user_date.get(key, [])
                
                # Calculate status and work hours from daily_events
                user_status = 'out'
                last_event = None
                daily_work_hours = 0
                
                if daily_events:
                    last_event = daily_events[-1]
                    
                    # Determine status from last event
                    if last_event.event_type == 'clock_in':
                        user_status = 'in'
                    elif last_event.event_type == 'break_start':
                        user_status = 'on_break'
                    elif last_event.event_type in ['clock_out', 'break_end']:
                        user_status = 'out'
                    
                    # Calculate work hours
                    total_seconds = 0
                    clock_in_time = None
                    break_start_time = None
                    
                    for event in daily_events:
                        if event.event_type == 'clock_in':
                            clock_in_time = event.timestamp
                        elif event.event_type == 'clock_out' and clock_in_time:
                            total_seconds += (event.timestamp - clock_in_time).total_seconds()
                            clock_in_time = None
                        elif event.event_type == 'break_start' and clock_in_time:
                            total_seconds += (event.timestamp - clock_in_time).total_seconds()
                            break_start_time = event.timestamp
                            clock_in_time = None
                        elif event.event_type == 'break_end':
                            clock_in_time = event.timestamp
                            break_start_time = None
                    
                    # If still clocked in, count until end of day or now
                    if clock_in_time:
                        end_time = min(datetime.now(), datetime.combine(check_date, time.max))
                        if clock_in_time.date() == check_date:
                            total_seconds += (end_time - clock_in_time).total_seconds()
                    
                    daily_work_hours = total_seconds / 3600
                
                users_data.append({
                    'user': user,
                    'date': check_date,
                    'status': user_status,
                    'last_event': last_event,
                    'daily_events': daily_events,
                    'daily_work_hours': daily_work_hours
                })
    
    # Sort by date (most recent first), then sede, then name
    users_data.sort(key=lambda x: (
        -x['date'].toordinal(),  # Negative for reverse chronological
        x['user'].sede_obj.name if x['user'].sede_obj else 'ZZZ',
        x['user'].last_name,
        x['user'].first_name
    ))
    
    # Get all sedi for filtering
    all_sedi = []
    if current_user.role == 'Amministratore' or current_user.all_sedi:
        all_sedi = filter_by_company(Sede.query).filter_by(active=True).order_by(Sede.name).all()
    elif current_user.sede_id:
        all_sedi = [current_user.sede]
    
    # Check if export to Excel is requested
    if request.args.get('export') == 'excel':
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from io import BytesIO
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Dashboard Team"
        
        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # Header
        headers = ['Data', 'Utente', 'Ruolo', 'Sede', 'Stato', 'Entrata', 'Uscita', 'Ore Lavorate', 'Note']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        
        # Dati
        for row_idx, user_data in enumerate(users_data, 2):
            # Determina orario entrata
            entrata = '-'
            if user_data['daily_events']:
                first_in = next((e for e in user_data['daily_events'] if e.event_type == 'clock_in'), None)
                if first_in:
                    entrata = first_in.timestamp.strftime('%H:%M')
            
            # Determina orario uscita
            uscita = '-'
            if user_data['last_event'] and user_data['last_event'].event_type == 'clock_out':
                uscita = user_data['last_event'].timestamp.strftime('%H:%M')
            
            # Determina stato
            stato = {
                'in': 'Presente',
                'on_break': 'In Pausa',
                'out': 'Uscito' if user_data['daily_work_hours'] > 0 else 'Assente'
            }.get(user_data['status'], 'Assente')
            
            # Note
            note = ''
            if user_data['last_event'] and user_data['last_event'].notes:
                note = user_data['last_event'].notes
            
            row_data = [
                user_data['date'].strftime('%d/%m/%Y'),
                user_data['user'].get_full_name(),
                user_data['user'].role,
                user_data['user'].sede_obj.name if user_data['user'].sede_obj else 'Nessuna Sede',
                stato,
                entrata,
                uscita,
                f"{user_data['daily_work_hours']:.2f}h" if user_data['daily_work_hours'] > 0 else '0h',
                note
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = thin_border
                if col in [1, 6, 7]:  # Data e Orari
                    cell.alignment = Alignment(horizontal='center')
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        # Generate filename
        filename = f"dashboard_team_{start_date.strftime('%Y-%m-%d')}_{end_date.strftime('%Y-%m-%d')}.xlsx"
        
        # Create response
        response = make_response(excel_file.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename={filename}'
        
        return response
    
    return render_template('dashboard_team.html', 
                         users_data=users_data,
                         all_sedi=all_sedi,
                         today=today,
                         start_date=start_date,
                         end_date=end_date,
                         period_label=period_label,
                         format_hours=format_hours)

@dashboard_bp.route('/dashboard_sede')
@login_required
def dashboard_sede():
    """Dashboard specifica per sede - mostra dati aggregati per sede"""
    if not current_user.can_view_all_attendance():
        flash('Non hai i permessi per visualizzare questo contenuto.', 'danger')
        return redirect(url_for('dashboard.dashboard'))
    
    # Get accessible sedi for current user
    if current_user.role == 'Amministratore' or current_user.all_sedi:
        available_sedi = filter_by_company(Sede.query).filter_by(active=True).order_by(Sede.name).all()
    elif current_user.sede_id:
        available_sedi = [current_user.sede]
    else:
        available_sedi = []
    
    # Get selected sede from query parameter
    selected_sede_id = request.args.get('sede_id', type=int)
    selected_sede = None
    
    if selected_sede_id:
        selected_sede = filter_by_company(Sede.query).get(selected_sede_id)
        # Verify user has access to this sede
        if selected_sede and not (
            current_user.role == 'Amministratore' or 
            current_user.all_sedi or 
            current_user.sede_id == selected_sede_id
        ):
            selected_sede = None
            flash('Non hai accesso ai dati di questa sede.', 'warning')
    
    # If no valid sede selected, default to user's sede
    if not selected_sede and current_user.sede_id:
        selected_sede = current_user.sede
    
    sede_data = {}
    if selected_sede:
        # Get users from selected sede
        sede_users = filter_by_company(User.query).filter(
            User.sede_id == selected_sede.id,
            User.active.is_(True),
            ~User.role.in_(['Admin', 'Staff'])
        ).all()
        
        # Get today's attendance data for sede users
        today = date.today()
        attendance_summary = {
            'total_users': len(sede_users),
            'users_in': 0,
            'users_out': 0,
            'users_on_break': 0,
            'total_work_hours': 0
        }
        
        user_details = []
        for user in sede_users:
            user_status, last_event = AttendanceEvent.get_user_status(user.id, today)
            today_work_hours = AttendanceEvent.get_daily_work_hours(user.id, today)
            
            # Update summary counts
            if user_status == 'in':
                attendance_summary['users_in'] += 1
            elif user_status == 'on_break':
                attendance_summary['users_on_break'] += 1
            else:
                attendance_summary['users_out'] += 1
            
            attendance_summary['total_work_hours'] += today_work_hours
            
            user_details.append({
                'user': user,
                'status': user_status,
                'last_event': last_event,
                'today_work_hours': today_work_hours
            })
        
        # Sort user details by name
        user_details.sort(key=lambda x: (x['user'].last_name, x['user'].first_name))
        
        sede_data = {
            'sede': selected_sede,
            'attendance_summary': attendance_summary,
            'user_details': user_details
        }
    
    return render_template('dashboard_sede.html',
                         available_sedi=available_sedi,
                         selected_sede=selected_sede,
                         sede_data=sede_data,
                         format_hours=format_hours)

@dashboard_bp.route('/ente-home')
@login_required
def ente_home():
    """Home page speciale per enti/organizzazioni - dashboard semplificata"""
    # Questa è una dashboard alternativa più semplice per certi tipi di utenti
    
    # Get basic user stats
    stats = get_user_statistics(current_user.id)
    
    # Get today's attendance status
    today_date = date.today()
    user_status = 'out'
    today_events = []
    today_work_hours = 0
    
    if current_user.can_view_my_attendance() or current_user.can_view_attendance():
        user_status, _ = AttendanceEvent.get_user_status(current_user.id, today_date)
        today_events = AttendanceEvent.get_daily_events(current_user.id, today_date)
        today_work_hours = AttendanceEvent.get_daily_work_hours(current_user.id, today_date)
    
    # Get recent leave requests
    recent_leaves = []
    if current_user.can_view_my_leave():
        recent_leaves = filter_by_company(LeaveRequest.query).filter(
            LeaveRequest.user_id == current_user.id
        ).order_by(LeaveRequest.created_at.desc()).limit(3).all()
    
    # Get upcoming shifts
    upcoming_shifts = []
    if current_user.can_view_shifts():
        seven_days_from_now = date.today() + timedelta(days=7)
        upcoming_shifts = filter_by_company(Shift.query).filter(
            Shift.user_id == current_user.id,
            Shift.date >= date.today(),
            Shift.date <= seven_days_from_now
        ).order_by(Shift.date, Shift.start_time).limit(5).all()
    
    return render_template('ente_home.html',
                         stats=stats,
                         today_date=today_date,
                         user_status=user_status,
                         today_events=today_events,
                         today_work_hours=today_work_hours,
                         recent_leaves=recent_leaves,
                         upcoming_shifts=upcoming_shifts,
                         format_hours=format_hours)