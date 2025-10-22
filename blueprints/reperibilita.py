# =============================================================================
# REPERIBILITÀ BLUEPRINT - Modulo gestione turni reperibilità
# =============================================================================
#
# ROUTES INCLUSE:
# 1. reperibilita_coverage (GET) - Visualizzazione coperture reperibilità
# 2. reperibilita_shifts (GET) - Gestione turni reperibilità
# 3. api/get_reperibilita_data (GET) - API dati reperibilità
# 4. generate_reperibilita (POST) - Generazione automatica turni
# 5. my_reperibilita (GET) - Le mie reperibilità
# 6. interventions/start (POST) - Inizia intervento generico
# 7. interventions/end (POST) - Termina intervento generico
# 8. interventions/my (GET) - Visualizza miei interventi
# 9. interventions/export/general/excel (GET) - Export interventi generici Excel
# 10. interventions/export/reperibilita/excel (GET) - Export interventi reperibilità Excel
# + altre route di gestione coperture, template, replica
#
# Total routes: 20+ reperibilità + interventions management routes
# =============================================================================

from flask import Blueprint, request, jsonify, render_template, redirect, url_for, flash, make_response
from flask_login import login_required, current_user
from datetime import datetime, date, timedelta
from functools import wraps
from app import db
from models import User, Sede, ReperibilitaShift, ReperibilitaCoverage, ReperibilitaIntervention, ReperibilitaTemplate, Intervention, AttendanceEvent, italian_now
from forms import ReperibilitaCoverageForm, ReperibilitaReplicaForm
from collections import defaultdict
from utils import generate_reperibilita_shifts
from utils_tenant import filter_by_company, set_company_on_create
from io import BytesIO, StringIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from defusedcsv import csv

# Create blueprint
reperibilita_bp = Blueprint('reperibilita', __name__, url_prefix='/reperibilita')

# Helper functions
def require_reperibilita_permissions(f):
    """Decorator to require reperibilità permissions for routes"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated:
            return redirect(url_for('auth.login'))
        if not current_user.can_access_reperibilita_menu():
            flash('Non hai i permessi per accedere alla reperibilità', 'danger')
            return redirect(url_for('dashboard.dashboard'))
        return f(*args, **kwargs)
    return decorated_function

# =============================================================================
# REPERIBILITÀ MANAGEMENT ROUTES
# =============================================================================

@reperibilita_bp.route('/coverage')
@login_required
@require_reperibilita_permissions
def reperibilita_coverage():
    """Lista coperture reperibilità"""
    if not current_user.can_access_reperibilita():
        flash('Non hai i permessi per visualizzare le coperture reperibilità', 'danger')
        return redirect(url_for('dashboard.dashboard'))
    
    # Raggruppa le coperture per periodo + sede per trattare duplicazioni come gruppi separati
    coverages = filter_by_company(ReperibilitaCoverage.query).order_by(ReperibilitaCoverage.start_date.desc()).all()
    groups = defaultdict(lambda: {'coverages': [], 'start_date': None, 'end_date': None, 'creator': None, 'created_at': None})
    
    for coverage in coverages:
        # Include sede nel period_key per separare coperture duplicate con sedi diverse
        sede_ids = sorted(coverage.get_sedi_ids_list())
        sede_key = "_".join(map(str, sede_ids)) if sede_ids else "no_sede"
        period_key = f"{coverage.start_date.strftime('%Y-%m-%d')}_{coverage.end_date.strftime('%Y-%m-%d')}_{sede_key}"
        
        if not groups[period_key]['start_date']:
            groups[period_key]['start_date'] = coverage.start_date
            groups[period_key]['end_date'] = coverage.end_date
            groups[period_key]['creator'] = coverage.creator
            groups[period_key]['created_at'] = coverage.created_at
        if 'coverages' in groups[period_key] and groups[period_key]['coverages'] is not None:
            groups[period_key]['coverages'].append(coverage)
    
    # Converte in oggetti simili ai presidi per il template
    reperibilita_groups = {}
    for period_key, data in groups.items():
        class ReperibilitaGroup:
            def __init__(self, coverages, start_date, end_date, creator, created_at):
                self.coverages = coverages
                self.start_date = start_date
                self.end_date = end_date
                self.creator = creator
                self.created_at = created_at
        
        reperibilita_groups[period_key] = ReperibilitaGroup(
            data['coverages'], data['start_date'], data['end_date'], 
            data['creator'], data['created_at']
        )
    
    return render_template('reperibilita_coverage.html', reperibilita_groups=reperibilita_groups)

@reperibilita_bp.route('/reperibilita_shifts')
@login_required
@require_reperibilita_permissions
def reperibilita_shifts():
    """Gestione turni reperibilità"""
    if not (current_user.can_manage_reperibilita() or current_user.can_view_reperibilita()):
        flash('Non hai i permessi per gestire i turni reperibilità', 'danger')
        return redirect(url_for('dashboard.dashboard'))
    
    # Filtri dalla query string
    month_filter = request.args.get('month')
    user_filter = request.args.get('user', 'all')
    
    # Query base
    query = filter_by_company(ReperibilitaShift.query)
    
    # Filtro mese
    if month_filter:
        try:
            year, month = month_filter.split('-')
            query = query.filter(
                db.extract('year', ReperibilitaShift.date) == int(year),
                db.extract('month', ReperibilitaShift.date) == int(month)
            )
        except ValueError:
            pass
    else:
        # Mese corrente di default
        now = italian_now()
        query = query.filter(
            db.extract('year', ReperibilitaShift.date) == now.year,
            db.extract('month', ReperibilitaShift.date) == now.month
        )
    
    # Filtro utente
    if user_filter != 'all':
        try:
            user_id = int(user_filter)
            query = query.filter(ReperibilitaShift.user_id == user_id)
        except ValueError:
            pass
    
    # Controllo sede
    if not current_user.all_sedi and current_user.sede_obj:
        sede_users = User.query.filter_by(sede_id=current_user.sede_obj.id).all()
        user_ids = [u.id for u in sede_users]
        query = query.filter(ReperibilitaShift.user_id.in_(user_ids))
    
    # Esecuzione query
    shifts = query.join(User, ReperibilitaShift.user_id == User.id).order_by(
        ReperibilitaShift.date.desc(),
        ReperibilitaShift.start_time
    ).all()
    
    # Lista utenti per filtro
    available_users = User.query.filter_by(active=True).order_by(User.last_name, User.first_name).all()
    if not current_user.all_sedi and current_user.sede_obj:
        available_users = [u for u in available_users if u.sede_id == current_user.sede_obj.id]
    
    # Parametri per navigation 
    period_mode = request.args.get('period', 'month')
    view_mode = request.args.get('view', 'calendar')
    display_mode = request.args.get('display', 'calendar')
    
    # Crea oggetto navigation per il template
    now = italian_now()
        
    # Calcola navigation date
    if month_filter:
        try:
            year, month = month_filter.split('-')
            current_month = datetime(int(year), int(month), 1)
        except ValueError:
            current_month = datetime(now.year, now.month, 1)
    else:
        current_month = datetime(now.year, now.month, 1)
    
    # Previous/Next month
    prev_month = current_month.replace(day=1) - timedelta(days=1)
    prev_month = prev_month.replace(day=1)
    
    if current_month.month == 12:
        next_month = current_month.replace(year=current_month.year + 1, month=1)
    else:
        next_month = current_month.replace(month=current_month.month + 1)
    
    # Crea oggetto navigation
    navigation = {
        'prev_date': prev_month,
        'next_date': next_month,
        'current_period': current_month.strftime('%B %Y')
    }

    return render_template('reperibilita_shifts.html',
                         shifts=shifts,
                         available_users=available_users,
                         selected_month=month_filter,
                         selected_user=user_filter,
                         navigation=navigation,
                         period_mode=period_mode,
                         view_mode=view_mode,
                         display_mode=display_mode,
                         can_manage=current_user.can_manage_reperibilita())

@reperibilita_bp.route('/my_reperibilita')
@login_required
@require_reperibilita_permissions
def my_reperibilita():
    """Le mie reperibilità"""
    if not current_user.can_view_my_reperibilita():
        flash('Non hai i permessi per visualizzare le tue reperibilità', 'danger')
        return redirect(url_for('dashboard.dashboard'))
    
    # Filtri dalla query string
    month_filter = request.args.get('month')
    status_filter = request.args.get('status', 'all')
    
    # Base query - solo le proprie reperibilità
    query = filter_by_company(ReperibilitaShift.query).filter_by(user_id=current_user.id)
    
    # Filtro mese
    if month_filter:
        try:
            year, month = month_filter.split('-')
            query = query.filter(
                db.extract('year', ReperibilitaShift.date) == int(year),
                db.extract('month', ReperibilitaShift.date) == int(month)
            )
        except ValueError:
            pass
    
    # Ordinamento
    shifts = query.order_by(ReperibilitaShift.date.desc()).all()
    
    # Statistiche personali
    stats = {
        'total_shifts': len(shifts),
        'upcoming_shifts': len([s for s in shifts if s.date >= italian_now().date()]),
        'past_shifts': len([s for s in shifts if s.date < italian_now().date()]),
    }
    
    return render_template('my_reperibilita.html',
                         shifts=shifts,
                         stats=stats,
                         selected_month=month_filter,
                         selected_status=status_filter)

@reperibilita_bp.route('/generate_shifts', methods=['GET', 'POST'])
@login_required
@require_reperibilita_permissions
def generate_shifts():
    """Generazione turni reperibilità (placeholder)"""
    if not current_user.can_manage_reperibilita():
        flash('Non hai i permessi per generare turni', 'danger')
        return redirect(url_for('dashboard.dashboard'))
    
    flash('Funzione generazione automatica turni in sviluppo', 'info')
    return redirect(url_for('reperibilita.reperibilita_coverage'))

@reperibilita_bp.route('/api/get_reperibilita_data')
@login_required
def get_reperibilita_data():
    """API per ottenere dati reperibilità"""
    try:
        start_date_str = request.args.get('start_date')
        end_date_str = request.args.get('end_date')
        
        if not start_date_str or not end_date_str:
            return jsonify({'error': 'Date mancanti'}), 400
        
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        
        # Query shifts nel periodo
        query = filter_by_company(ReperibilitaShift.query).filter(
            ReperibilitaShift.date >= start_date,
            ReperibilitaShift.date <= end_date
        )
        
        # Controllo sede
        if not current_user.all_sedi and current_user.sede_obj:
            sede_users = User.query.filter_by(sede_id=current_user.sede_obj.id).all()
            user_ids = [u.id for u in sede_users]
            query = query.filter(ReperibilitaShift.user_id.in_(user_ids))
        
        shifts = query.all()
        
        # Formatta dati per risposta JSON
        shifts_data = []
        for shift in shifts:
            shifts_data.append({
                'id': shift.id,
                'user_name': shift.user.get_full_name(),
                'date': shift.date.isoformat(),
                'start_time': shift.start_time.strftime('%H:%M'),
                'end_time': shift.end_time.strftime('%H:%M'),
                'type': getattr(shift, 'shift_type', 'Standard')
            })
        
        return jsonify({
            'success': True,
            'data': shifts_data,
            'period': f"{start_date.isoformat()} - {end_date.isoformat()}"
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@reperibilita_bp.route('/coverage/create', methods=['GET', 'POST'])
@login_required
@require_reperibilita_permissions
def create_reperibilita_coverage():
    """Crea nuova copertura reperibilità"""
    if not current_user.can_manage_reperibilita():
        flash('Non hai i permessi per creare coperture reperibilità', 'danger')
        return redirect(url_for('dashboard.dashboard'))
    
    form = ReperibilitaCoverageForm()
    
    if form.validate_on_submit():
        # Crea una copertura per ogni giorno selezionato
        import json
        success_count = 0
        
        for day_of_week in (form.days_of_week.data or []):
            coverage = ReperibilitaCoverage()
            coverage.day_of_week = day_of_week
            coverage.start_time = form.start_time.data
            coverage.end_time = form.end_time.data
            coverage.set_required_roles_list(form.required_roles.data)
            coverage.set_sedi_ids_list(form.sedi.data)  # Aggiungi le sedi selezionate
            coverage.description = form.description.data
            coverage.active = form.active.data
            coverage.start_date = form.start_date.data
            coverage.end_date = form.end_date.data
            coverage.created_by = current_user.id
            
            set_company_on_create(coverage)
            db.session.add(coverage)
            success_count += 1
        
        try:
            db.session.commit()
            flash(f'Copertura reperibilità creata con successo per {success_count} giorni!', 'success')
            return redirect(url_for('reperibilita.reperibilita_coverage'))
        except Exception as e:
            db.session.rollback()
            flash(f'Errore durante la creazione: {str(e)}', 'error')
    
    return render_template('create_reperibilita_coverage.html', form=form)

@reperibilita_bp.route('/coverage/edit/<int:coverage_id>', methods=['GET', 'POST'])
@login_required
@require_reperibilita_permissions
def edit_reperibilita_coverage(coverage_id):
    """Modifica copertura reperibilità"""
    if not current_user.can_manage_reperibilita():
        flash('Non hai i permessi per modificare coperture reperibilità', 'danger')
        return redirect(url_for('dashboard.dashboard'))
    
    coverage = filter_by_company(ReperibilitaCoverage.query).filter_by(id=coverage_id).first_or_404()
    form = ReperibilitaCoverageForm()
    
    if form.validate_on_submit():
        coverage.start_time = form.start_time.data
        coverage.end_time = form.end_time.data
        coverage.set_required_roles_list(form.required_roles.data)
        coverage.set_sedi_ids_list(form.sedi.data)  # Aggiungi le sedi selezionate
        coverage.description = form.description.data
        coverage.active = form.active.data
        coverage.start_date = form.start_date.data
        coverage.end_date = form.end_date.data
        
        try:
            db.session.commit()
            flash('Copertura reperibilità aggiornata con successo!', 'success')
            return redirect(url_for('reperibilita.reperibilita_coverage'))
        except Exception as e:
            db.session.rollback()
            flash(f'Errore durante l\'aggiornamento: {str(e)}', 'error')
    
    # Pre-popola il form con i dati esistenti
    if request.method == 'GET':
        form.start_time.data = coverage.start_time
        form.end_time.data = coverage.end_time
        form.required_roles.data = coverage.get_required_roles_list()
        form.description.data = coverage.description
        form.active.data = coverage.active
        form.start_date.data = coverage.start_date
        form.end_date.data = coverage.end_date
        form.days_of_week.data = [coverage.day_of_week]  # Single day for edit
    
    return render_template('edit_reperibilita_coverage.html', form=form, coverage=coverage)

@reperibilita_bp.route('/coverage/delete/<int:coverage_id>', methods=['GET'])
@login_required
@require_reperibilita_permissions
def delete_reperibilita_coverage(coverage_id):
    """Elimina copertura reperibilità"""
    if not current_user.can_manage_reperibilita():
        flash('Non hai i permessi per eliminare coperture reperibilità', 'danger')
        return redirect(url_for('dashboard.dashboard'))
    
    coverage = filter_by_company(ReperibilitaCoverage.query).filter_by(id=coverage_id).first_or_404()
    
    try:
        db.session.delete(coverage)
        db.session.commit()
        flash('Copertura reperibilità eliminata con successo!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Errore durante l\'eliminazione: {str(e)}', 'error')
    
    return redirect(url_for('reperibilita.reperibilita_coverage'))

@reperibilita_bp.route('/coverage/view/<period_key>')
@login_required
@require_reperibilita_permissions
def view_reperibilita_coverage(period_key):
    """Visualizza dettagli coperture reperibilità per un periodo"""
    if not current_user.can_access_reperibilita():
        flash('Non hai i permessi per visualizzare coperture reperibilità', 'danger')
        return redirect(url_for('dashboard.dashboard'))
    
    # Decodifica period_key
    start_date_str, end_date_str = period_key.split('_')
    start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
    end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    
    # Trova tutte le coperture per questo periodo
    coverages = filter_by_company(ReperibilitaCoverage.query).filter(
        ReperibilitaCoverage.start_date == start_date,
        ReperibilitaCoverage.end_date == end_date
    ).order_by(ReperibilitaCoverage.day_of_week, ReperibilitaCoverage.start_time).all()
    
    if not coverages:
        flash('Periodo di copertura reperibilità non trovato', 'error')
        return redirect(url_for('reperibilita.reperibilita_coverage'))
    
    return render_template('view_reperibilita_coverage.html', 
                         coverages=coverages, 
                         start_date=start_date, 
                         end_date=end_date,
                         period_key=period_key)

@reperibilita_bp.route('/coverage/delete_period/<period_key>')
@login_required
@require_reperibilita_permissions  
def delete_reperibilita_period(period_key):
    """Elimina tutte le coperture reperibilità di un periodo"""
    if not current_user.can_manage_reperibilita():
        flash('Non hai i permessi per eliminare periodi reperibilità', 'danger')
        return redirect(url_for('dashboard.dashboard'))
    
    # Decodifica period_key
    start_date_str, end_date_str = period_key.split('_')
    start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
    end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    
    # Trova tutte le coperture per questo periodo
    coverages = filter_by_company(ReperibilitaCoverage.query).filter(
        ReperibilitaCoverage.start_date == start_date,
        ReperibilitaCoverage.end_date == end_date
    ).all()
    
    # Trova anche tutti i turni generati per questo periodo
    shifts = filter_by_company(ReperibilitaShift.query).filter(
        ReperibilitaShift.date >= start_date,
        ReperibilitaShift.date <= end_date
    ).all()
    
    try:
        coverage_count = len(coverages)
        shift_count = len(shifts)
        
        # Elimina prima i turni, poi le coperture
        for shift in shifts:
            db.session.delete(shift)
        for coverage in coverages:
            db.session.delete(coverage)
            
        db.session.commit()
        flash(f'Eliminate {coverage_count} coperture reperibilità e {shift_count} turni del periodo {start_date.strftime("%d/%m/%Y")} - {end_date.strftime("%d/%m/%Y")}', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Errore durante l\'eliminazione: {str(e)}', 'error')
    
    return redirect(url_for('reperibilita.reperibilita_shifts'))

@reperibilita_bp.route('/template/<start_date>/<end_date>')
@login_required
@require_reperibilita_permissions
def reperibilita_template_detail(start_date, end_date):
    """Mostra dettaglio template reperibilità (come shift_template_detail)"""
    from collections import defaultdict
    
    # Parse delle date
    start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    
    # Trova tutti i turni di reperibilità per questo periodo
    shifts = filter_by_company(ReperibilitaShift.query).filter(
        ReperibilitaShift.date >= start_date,
        ReperibilitaShift.date <= end_date
    ).order_by(ReperibilitaShift.date, ReperibilitaShift.start_time).all()
    
    # Organizza per giorno della settimana per la vista calendario
    shifts_by_day = defaultdict(list)
    for shift in shifts:
        shifts_by_day[shift.date].append(shift)
    
    # Genera calendario giorni
    calendar_days = []
    current_date = start_date
    weekdays = ['Lunedì', 'Martedì', 'Mercoledì', 'Giovedì', 'Venerdì', 'Sabato', 'Domenica']
    
    while current_date <= end_date:
        calendar_days.append({
            'date': current_date,
            'weekday': weekdays[current_date.weekday()],
            'is_today': current_date == italian_now().date()
        })
        current_date += timedelta(days=1)
    
    period_key = f"{start_date.strftime('%Y-%m-%d')}_{end_date.strftime('%Y-%m-%d')}"
    
    return render_template('reperibilita_template_detail.html', 
                         shifts=shifts,
                         shifts_by_day=shifts_by_day,
                         calendar_days=calendar_days,
                         start_date=start_date,
                         end_date=end_date,
                         period_key=period_key)

@reperibilita_bp.route('/replica/<period_key>', methods=['GET', 'POST'])
@login_required
@require_reperibilita_permissions
def reperibilita_replica(period_key):
    """Replica template reperibilità"""
    if not current_user.can_manage_reperibilita():
        flash('Non hai i permessi per replicare i template di reperibilità', 'danger')
        return redirect(url_for('dashboard.dashboard'))
    
    from forms import ReperibilitaReplicaForm
    
    # Decodifica period_key
    start_date_str, end_date_str = period_key.split('_')
    start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
    end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    
    form = ReperibilitaReplicaForm()
    
    if form.validate_on_submit():
        
        # Ottieni mappatura ruoli dal form
        role_mapping = form.get_role_mapping_dict()
        
        # Trova le coperture originali
        original_coverages = filter_by_company(ReperibilitaCoverage.query).filter(
            ReperibilitaCoverage.start_date == start_date,
            ReperibilitaCoverage.end_date == end_date
        ).all()
        
        if not original_coverages:
            flash('Template di copertura originale non trovato', 'error')
            return redirect(url_for('reperibilita.reperibilita_coverage'))
        
        # Verifica se esistono già coperture per informazione (non blocca la creazione)
        existing_coverages = filter_by_company(ReperibilitaCoverage.query).filter(
            ReperibilitaCoverage.start_date == form.start_date.data,
            ReperibilitaCoverage.end_date == form.end_date.data
        ).all()
        
        # Replica le coperture con nuove date e ruoli modificati
        new_coverages_count = 0
        for original_coverage in original_coverages:
            new_coverage = ReperibilitaCoverage()
            new_coverage.day_of_week = original_coverage.day_of_week
            new_coverage.start_time = original_coverage.start_time
            new_coverage.end_time = original_coverage.end_time
            new_coverage.description = original_coverage.description
            new_coverage.active = original_coverage.active
            new_coverage.start_date = form.start_date.data
            new_coverage.end_date = form.end_date.data
            new_coverage.created_by = current_user.id
            
            # Applica mappatura ruoli se specificata
            original_roles = original_coverage.get_required_roles_list()
            if role_mapping:
                # Sostituisce i ruoli secondo la mappatura
                new_roles = []
                for role in original_roles:
                    if role in role_mapping:
                        new_roles.append(role_mapping[role])
                    else:
                        new_roles.append(role)  # Mantiene il ruolo originale se non mappato
                new_coverage.set_required_roles_list(new_roles)
            else:
                # Mantiene i ruoli originali
                new_coverage.set_required_roles_list(original_roles)
            
            # Gestisce il cambio di sede se specificato
            if form.sede_id.data:
                # Assegna la nuova sede specificata
                new_coverage.set_sedi_ids_list([int(form.sede_id.data)])
            else:
                # Mantiene le sedi originali
                new_coverage.set_sedi_ids_list(original_coverage.get_sedi_ids_list())
            
            set_company_on_create(new_coverage)
            db.session.add(new_coverage)
            new_coverages_count += 1
        
        try:
            db.session.commit()
            
            success_msg = f'Template reperibilità replicato con successo. Coperture create: {new_coverages_count}.'
            if role_mapping:
                success_msg += f' Ruoli sostituiti: {len(role_mapping)}.'
            if form.sede_id.data:
                from models import Sede
                sede = Sede.query.get(int(form.sede_id.data))
                if sede:
                    success_msg += f' Sede cambiata in: {sede.name}.'
            if existing_coverages:
                success_msg += f' Aggiunte a {len(existing_coverages)} coperture esistenti.'
            
            flash(success_msg, 'success')
            return redirect(url_for('reperibilita.reperibilita_coverage'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Errore durante la replica: {str(e)}', 'error')
    
    # Pre-popola le date originali come suggerimento
    if request.method == 'GET':
        form.start_date.data = start_date
        form.end_date.data = end_date
    
    # Trova le coperture originali per mostrare informazioni nel template
    original_coverages = filter_by_company(ReperibilitaCoverage.query).filter(
        ReperibilitaCoverage.start_date == start_date,
        ReperibilitaCoverage.end_date == end_date
    ).order_by(ReperibilitaCoverage.day_of_week, ReperibilitaCoverage.start_time).all()
    
    return render_template('reperibilita_replica.html', 
                         form=form,
                         original_coverages=original_coverages,
                         start_date=start_date,
                         end_date=end_date)

# =============================================================================
# INTERVENTION MANAGEMENT ROUTES
# =============================================================================

@reperibilita_bp.route('/start-intervention', methods=['POST'])
@login_required
@require_reperibilita_permissions
def start_intervention():
    """Inizia un intervento di reperibilità"""
    if current_user.role not in ['Management', 'Operatore', 'Redattore', 'Sviluppatore']:
        flash('Non hai i permessi per registrare interventi di reperibilità.', 'danger')
        return redirect(url_for('dashboard.dashboard'))
    
    # Controlla se c'è già un intervento attivo
    from models import ReperibilitaIntervention
    active_intervention = filter_by_company(ReperibilitaIntervention.query).filter_by(
        user_id=current_user.id,
        end_datetime=None
    ).first()
    
    if active_intervention:
        flash('Hai già un intervento di reperibilità in corso.', 'warning')
        return redirect(url_for('dashboard.dashboard'))
    
    # Ottieni shift_id dal form se presente
    shift_id = request.form.get('shift_id')
    if shift_id:
        shift_id = int(shift_id)
    
    # Ottieni is_remote dal form (default True = remoto)
    is_remote = request.form.get('is_remote', 'true').lower() == 'true'
    
    # Ottieni priorità dal form (default Media)
    priority = request.form.get('priority', 'Media')
    if priority not in ['Bassa', 'Media', 'Alta']:
        priority = 'Media'
    
    # Crea nuovo intervento
    intervention = ReperibilitaIntervention()
    intervention.user_id = current_user.id
    intervention.shift_id = shift_id
    intervention.start_datetime = italian_now()
    intervention.description = request.form.get('description', '')
    intervention.priority = priority
    intervention.is_remote = is_remote
    
    set_company_on_create(intervention)
    db.session.add(intervention)
    db.session.commit()
    
    flash('Intervento di reperibilità iniziato con successo.', 'success')
    return redirect(url_for('reperibilita.reperibilita_shifts'))

@reperibilita_bp.route('/end-intervention', methods=['POST'])
@login_required
@require_reperibilita_permissions
def end_intervention():
    """Termina un intervento di reperibilità"""
    if current_user.role not in ['Management', 'Operatore', 'Redattore', 'Sviluppatore']:
        flash('Non hai i permessi per registrare interventi di reperibilità.', 'danger')
        return redirect(url_for('dashboard.dashboard'))
    
    # Trova l'intervento attivo
    active_intervention = filter_by_company(ReperibilitaIntervention.query).filter_by(
        user_id=current_user.id,
        end_datetime=None
    ).first()
    
    if not active_intervention:
        flash('Nessun intervento di reperibilità attivo da terminare.', 'warning')
        return redirect(url_for('dashboard.dashboard'))
    
    # Termina l'intervento
    active_intervention.end_datetime = italian_now()
    
    active_intervention.description = request.form.get('description', active_intervention.description)
    
    db.session.commit()
    
    flash('Intervento di reperibilità terminato con successo.', 'success')
    
    # Redirect to appropriate page
    if current_user.role == 'Management':
        return redirect(url_for('dashboard.ente_home'))
    else:
        return redirect(url_for('reperibilita.reperibilita_shifts'))

@reperibilita_bp.route('/shifts/regenerate/<int:template_id>', methods=['GET'])
@login_required
@require_reperibilita_permissions
def regenerate_reperibilita_template(template_id):
    """Rigenera turni reperibilità da template esistente"""
    if not current_user.can_manage_reperibilita():
        flash('Non hai i permessi per rigenerare turni di reperibilità', 'danger')
        return redirect(url_for('dashboard.dashboard'))
    
    from models import ReperibilitaTemplate
    from utils import generate_reperibilita_shifts
    
    # Trova il template esistente
    template = filter_by_company(ReperibilitaTemplate.query).filter_by(id=template_id).first_or_404()
    
    try:
        # Elimina turni esistenti nel periodo del template
        existing_shifts = filter_by_company(ReperibilitaShift.query).filter(
            ReperibilitaShift.date >= template.start_date,
            ReperibilitaShift.date <= template.end_date
        ).all()
        
        for shift in existing_shifts:
            db.session.delete(shift)
        
        # Rigenera turni con gli stessi parametri del template
        shifts_created, warnings = generate_reperibilita_shifts(
            template.start_date,
            template.end_date,
            current_user.id
        )
        
        # Aggiorna la data di creazione del template
        template.created_at = italian_now()
        
        db.session.commit()
        
        # Costruisci messaggio di successo
        success_msg = f'Template "{template.name}" rigenerato con successo. Turni reperibilità generati: {shifts_created}.'
        
        if warnings:
            if len(warnings) <= 3:
                warning_text = " Attenzione: " + "; ".join(warnings)
            else:
                warning_text = f" Attenzione: {warnings[0]}; {warnings[1]}; {warnings[2]} e altri {len(warnings) - 3} avvisi."
            success_msg += warning_text
        
        flash(success_msg, 'success' if not warnings else 'warning')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Errore durante la rigenerazione: {str(e)}', 'error')
    
    return redirect(url_for('reperibilita.reperibilita_shifts'))

@reperibilita_bp.route('/template/delete/<template_id>')
@login_required
@require_reperibilita_permissions
def delete_reperibilita_template(template_id):
    """Elimina template reperibilità"""
    if not current_user.can_manage_reperibilita():
        flash('Non hai i permessi per eliminare template di reperibilità', 'danger')
        return redirect(url_for('dashboard.dashboard'))
    
    from models import ReperibilitaTemplate
    
    template = filter_by_company(ReperibilitaTemplate.query).filter_by(id=template_id).first_or_404()
    
    try:
        # Elimina tutti i turni del periodo del template
        shifts = filter_by_company(ReperibilitaShift.query).filter(
            ReperibilitaShift.date >= template.start_date,
            ReperibilitaShift.date <= template.end_date
        ).all()
        
        for shift in shifts:
            db.session.delete(shift)
        
        # Elimina il template
        template_name = template.name
        db.session.delete(template)
        db.session.commit()
        
        flash(f'Template reperibilità "{template_name}" eliminato con successo', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Errore durante l\'eliminazione: {str(e)}', 'error')
    
    return redirect(url_for('reperibilita.reperibilita_shifts'))

# =============================================================================
# INTERVENTIONS MANAGEMENT ROUTES
# =============================================================================

@reperibilita_bp.route('/interventions/start', methods=['POST'])
@login_required
@require_reperibilita_permissions
def start_general_intervention():
    """Inizia un nuovo intervento generico"""
    if not current_user.can_manage_interventions():
        return jsonify({
            'success': False,
            'message': 'Non hai i permessi per gestire interventi'
        }), 403
    
    # Controlla se l'utente è presente
    user_status, _ = AttendanceEvent.get_user_status(current_user.id)
    if user_status != 'in':
        flash('Devi essere presente per iniziare un intervento.', 'warning')
        return redirect(url_for('dashboard.dashboard'))
    
    # Controlla se c'è già un intervento attivo
    active_intervention = Intervention.query.filter_by(
        user_id=current_user.id,
        end_datetime=None
    ).first()
    
    if active_intervention:
        flash('Hai già un intervento attivo. Terminalo prima di iniziarne un altro.', 'warning')
        return redirect(url_for('dashboard.dashboard'))
    
    # Ottieni i dati dal form
    description = request.form.get('description', '')
    priority = request.form.get('priority', 'Media')
    is_remote = request.form.get('is_remote', 'false').lower() == 'true'
    
    # Crea nuovo intervento
    now = italian_now()
    
    intervention = Intervention(
        user_id=current_user.id,
        start_datetime=now,
        description=description,
        priority=priority,
        is_remote=is_remote
    )
    
    try:
        db.session.add(intervention)
        db.session.commit()
        flash(f'Intervento iniziato alle {now.strftime("%H:%M")}', 'success')
    except Exception as e:
        db.session.rollback()
        flash('Errore nel salvare l\'intervento', 'danger')
    
    return redirect(url_for('dashboard.dashboard'))

@reperibilita_bp.route('/interventions/end', methods=['POST'])
@login_required
@require_reperibilita_permissions
def end_general_intervention():
    """Termina un intervento generico attivo"""
    if not current_user.can_manage_interventions():
        return jsonify({
            'success': False,
            'message': 'Non hai i permessi per gestire interventi'
        }), 403
    
    # Trova l'intervento attivo
    active_intervention = Intervention.query.filter_by(
        user_id=current_user.id,
        end_datetime=None
    ).first()
    
    if not active_intervention:
        flash('Nessun intervento attivo trovato.', 'warning')
        return redirect(url_for('dashboard.dashboard'))
    
    # Termina l'intervento
    now = italian_now()
    active_intervention.end_datetime = now
    
    # Gestisci la descrizione finale
    end_description = request.form.get('end_description', '').strip()
    if end_description:
        # Combina descrizione iniziale e finale
        initial_desc = active_intervention.description or ''
        if initial_desc and end_description:
            active_intervention.description = f"{initial_desc}\n\n--- Risoluzione ---\n{end_description}"
        elif end_description:
            active_intervention.description = end_description
    
    try:
        db.session.commit()
        duration = active_intervention.duration_minutes
        flash(f'Intervento terminato alle {now.strftime("%H:%M")} (durata: {duration:.1f} minuti)', 'success')
    except Exception as e:
        db.session.rollback()
        flash('Errore nel terminare l\'intervento', 'danger')
    
    return redirect(url_for('dashboard.dashboard'))

@reperibilita_bp.route('/interventions/my')
@login_required
@require_reperibilita_permissions  
def my_interventions():
    """Pagina per visualizzare gli interventi - tutti per PM/Management, solo propri per altri utenti"""
    if not current_user.can_view_my_interventions():
        flash('Non hai i permessi per visualizzare gli interventi', 'danger')
        return redirect(url_for('dashboard.dashboard'))
    
    # Solo Admin non può accedere a questa pagina (non ha interventi)
    if current_user.role == 'Admin':
        flash('Accesso non autorizzato.', 'danger')
        return redirect(url_for('dashboard.dashboard'))
    
    now = italian_now()
    today = now.date()
    
    # Gestisci filtri data
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')
    
    # Default: primo del mese corrente - oggi
    if not start_date_str:
        start_date = today.replace(day=1)
    else:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
    
    if not end_date_str:
        end_date = today
    else:
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    
    # Converti le date in datetime per il filtro
    start_datetime = datetime.combine(start_date, datetime.min.time())
    end_datetime = datetime.combine(end_date, datetime.max.time())
    
    # PM/Management vedono tutti gli interventi, altri utenti solo i propri
    if current_user.can_view_interventions():
        # Ottieni tutti gli interventi di reperibilità filtrati per data
        reperibilita_interventions = filter_by_company(ReperibilitaIntervention.query).join(User).filter(
            ReperibilitaIntervention.start_datetime >= start_datetime,
            ReperibilitaIntervention.start_datetime <= end_datetime
        ).order_by(ReperibilitaIntervention.start_datetime.desc()).all()
        
        # Ottieni tutti gli interventi generici filtrati per data
        general_interventions = Intervention.query.join(User).filter(
            Intervention.start_datetime >= start_datetime,
            Intervention.start_datetime <= end_datetime
        ).order_by(Intervention.start_datetime.desc()).all()
    else:
        # Ottieni solo gli interventi dell'utente corrente filtrati per data
        reperibilita_interventions = filter_by_company(ReperibilitaIntervention.query).filter(
            ReperibilitaIntervention.user_id == current_user.id,
            ReperibilitaIntervention.start_datetime >= start_datetime,
            ReperibilitaIntervention.start_datetime <= end_datetime
        ).order_by(ReperibilitaIntervention.start_datetime.desc()).all()
        
        general_interventions = Intervention.query.filter(
            Intervention.user_id == current_user.id,
            Intervention.start_datetime >= start_datetime,
            Intervention.start_datetime <= end_datetime
        ).order_by(Intervention.start_datetime.desc()).all()
    
    return render_template('my_interventions.html',
                         reperibilita_interventions=reperibilita_interventions,
                         general_interventions=general_interventions,
                         start_date=start_date,
                         end_date=end_date)

@reperibilita_bp.route('/interventions/export/general/excel')
@login_required
@require_reperibilita_permissions
def export_general_interventions_excel():
    """Export interventi generici in formato Excel"""
    if not current_user.can_view_my_interventions():
        return jsonify({'error': 'Non autorizzato'}), 403
    
    if current_user.role == 'Admin':
        flash('Accesso non autorizzato.', 'danger')
        return redirect(url_for('dashboard.dashboard'))
    
    now = italian_now()
    today = now.date()
    
    # Gestisci filtri data
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')
    
    # Default: primo del mese corrente - oggi
    if not start_date_str:
        start_date = today.replace(day=1)
    else:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
    
    if not end_date_str:
        end_date = today
    else:
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    
    # Converti le date in datetime per il filtro
    start_datetime = datetime.combine(start_date, datetime.min.time())
    end_datetime = datetime.combine(end_date, datetime.max.time())
    
    # PM/Management vedono tutti gli interventi, altri utenti solo i propri
    if current_user.can_view_interventions():
        general_interventions = Intervention.query.join(User).filter(
            Intervention.start_datetime >= start_datetime,
            Intervention.start_datetime <= end_datetime
        ).order_by(Intervention.start_datetime.desc()).all()
    else:
        general_interventions = Intervention.query.filter(
            Intervention.user_id == current_user.id,
            Intervention.start_datetime >= start_datetime,
            Intervention.start_datetime <= end_datetime
        ).order_by(Intervention.start_datetime.desc()).all()
    
    # Crea workbook Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Interventi Generici"
    
    # Stili
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Headers
    headers = ['Data Inizio', 'Ora Inizio', 'Data Fine', 'Ora Fine', 'Durata (min)', 
               'Utente', 'Descrizione', 'Priorità', 'Remoto']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Dati
    for row, intervention in enumerate(general_interventions, 2):
        ws.cell(row=row, column=1, value=intervention.start_datetime.strftime('%d/%m/%Y') if intervention.start_datetime else '')
        ws.cell(row=row, column=2, value=intervention.start_datetime.strftime('%H:%M') if intervention.start_datetime else '')
        ws.cell(row=row, column=3, value=intervention.end_datetime.strftime('%d/%m/%Y') if intervention.end_datetime else 'In corso')
        ws.cell(row=row, column=4, value=intervention.end_datetime.strftime('%H:%M') if intervention.end_datetime else 'In corso')
        ws.cell(row=row, column=5, value=f'{intervention.duration_minutes:.1f}' if intervention.end_datetime else 'In corso')
        ws.cell(row=row, column=6, value=f'{intervention.user.first_name} {intervention.user.last_name}' if intervention.user else 'N/A')
        ws.cell(row=row, column=7, value=intervention.description or '')
        ws.cell(row=row, column=8, value=intervention.priority or 'Media')
        ws.cell(row=row, column=9, value='Sì' if intervention.is_remote else 'No')
    
    # Regola larghezza colonne
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Prepara response
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"interventi_generici_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx"
    
    response = make_response(output.getvalue())
    response.headers['Content-Disposition'] = f'attachment; filename={filename}'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    
    return response

@reperibilita_bp.route('/interventions/export/reperibilita/excel')
@login_required
@require_reperibilita_permissions
def export_reperibilita_interventions_excel():
    """Export interventi di reperibilità in formato Excel"""
    if not current_user.can_view_my_interventions():
        return jsonify({'error': 'Non autorizzato'}), 403
    
    if current_user.role == 'Admin':
        flash('Accesso non autorizzato.', 'danger')
        return redirect(url_for('dashboard.dashboard'))
    
    now = italian_now()
    today = now.date()
    
    # Gestisci filtri data
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')
    
    # Default: primo del mese corrente - oggi
    if not start_date_str:
        start_date = today.replace(day=1)
    else:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
    
    if not end_date_str:
        end_date = today
    else:
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    
    # Converti le date in datetime per il filtro
    start_datetime = datetime.combine(start_date, datetime.min.time())
    end_datetime = datetime.combine(end_date, datetime.max.time())
    
    # PM/Management vedono tutti gli interventi, altri utenti solo i propri
    if current_user.can_view_interventions():
        reperibilita_interventions = filter_by_company(ReperibilitaIntervention.query).join(User).filter(
            ReperibilitaIntervention.start_datetime >= start_datetime,
            ReperibilitaIntervention.start_datetime <= end_datetime
        ).order_by(ReperibilitaIntervention.start_datetime.desc()).all()
    else:
        reperibilita_interventions = filter_by_company(ReperibilitaIntervention.query).filter(
            ReperibilitaIntervention.user_id == current_user.id,
            ReperibilitaIntervention.start_datetime >= start_datetime,
            ReperibilitaIntervention.start_datetime <= end_datetime
        ).order_by(ReperibilitaIntervention.start_datetime.desc()).all()
    
    # Crea workbook Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Interventi Reperibilità"
    
    # Stili
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Headers
    headers = ['Data Inizio', 'Ora Inizio', 'Data Fine', 'Ora Fine', 'Durata (min)', 
               'Utente', 'Descrizione', 'Priorità', 'Tipo', 'Cliente']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Dati
    for row, intervention in enumerate(reperibilita_interventions, 2):
        ws.cell(row=row, column=1, value=intervention.start_datetime.strftime('%d/%m/%Y') if intervention.start_datetime else '')
        ws.cell(row=row, column=2, value=intervention.start_datetime.strftime('%H:%M') if intervention.start_datetime else '')
        ws.cell(row=row, column=3, value=intervention.end_datetime.strftime('%d/%m/%Y') if intervention.end_datetime else 'In corso')
        ws.cell(row=row, column=4, value=intervention.end_datetime.strftime('%H:%M') if intervention.end_datetime else 'In corso')
        ws.cell(row=row, column=5, value=f'{intervention.duration_minutes:.1f}' if intervention.end_datetime else 'In corso')
        ws.cell(row=row, column=6, value=f'{intervention.user.first_name} {intervention.user.last_name}' if intervention.user else 'N/A')
        ws.cell(row=row, column=7, value=intervention.description or '')
        ws.cell(row=row, column=8, value=intervention.priority or 'Media')
        ws.cell(row=row, column=9, value=intervention.intervention_type or '')
        ws.cell(row=row, column=10, value=intervention.client or '')
    
    # Regola larghezza colonne
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Prepara response
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"interventi_reperibilita_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx"
    
    response = make_response(output.getvalue())
    response.headers['Content-Disposition'] = f'attachment; filename={filename}'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    
    return response