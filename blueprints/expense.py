# =============================================================================
# EXPENSE & FINANCIAL MANAGEMENT BLUEPRINT
# =============================================================================
# Blueprint for managing expense reports, overtime requests, mileage reimbursements
# and all related financial operations including categories and approvals
# =============================================================================

from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, make_response, send_file
from flask_login import login_required, current_user
from datetime import datetime, date, timedelta
from io import BytesIO
from defusedcsv import csv
import os

# Local imports - Add as needed during migration
from models import User, ExpenseReport, ExpenseCategory, OvertimeRequest, OvertimeType, MileageRequest
from forms import ExpenseFilterForm, ExpenseReportForm, OvertimeRequestForm, MileageRequestForm, MileageFilterForm, ExpenseCategoryForm, OvertimeTypeForm, ExpenseApprovalForm
from app import db, app
from sqlalchemy.orm import joinedload
from werkzeug.utils import secure_filename
from utils_tenant import filter_by_company, set_company_on_create, get_user_company_id
import uuid

# =============================================================================
# BLUEPRINT CONFIGURATION
# =============================================================================

expense_bp = Blueprint(
    'expense', 
    __name__, 
    url_prefix='/expense',
    template_folder='../templates',
    static_folder='../static'
)

# =============================================================================
# PERMISSION DECORATORS
# =============================================================================

def require_expense_permission(f):
    """Decorator to check expense viewing permission"""
    from functools import wraps
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.can_view_expense_reports():
            flash('Non hai i permessi per visualizzare le note spese', 'danger')
            return redirect(url_for('dashboard.dashboard'))
        return f(*args, **kwargs)
    return decorated_function

def require_manage_expense_permission(f):
    """Decorator to check expense management permission"""
    from functools import wraps
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.can_manage_expense_reports():
            flash('Non hai i permessi per gestire le note spese', 'danger')
            return redirect(url_for('dashboard.dashboard'))
        return f(*args, **kwargs)
    return decorated_function

def require_overtime_permission(f):
    """Decorator to check overtime permission"""
    from functools import wraps
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.can_view_overtime_requests():
            flash('Non hai i permessi per visualizzare gli straordinari', 'danger')
            return redirect(url_for('dashboard.dashboard'))
        return f(*args, **kwargs)
    return decorated_function

def require_mileage_permission(f):
    """Decorator to check mileage permission"""
    from functools import wraps
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.can_view_mileage_requests():
            flash('Non hai i permessi per visualizzare i rimborsi chilometrici', 'danger')
            return redirect(url_for('dashboard.dashboard'))
        return f(*args, **kwargs)
    return decorated_function

# =============================================================================
# EXPENSE REPORTS ROUTES
# =============================================================================

@expense_bp.route('/reports', methods=['GET', 'POST'])
@login_required
@require_expense_permission
def expense_reports():
    """Visualizza elenco note spese"""
    filter_form = ExpenseFilterForm(current_user=current_user)
    
    # Query base (with company filter)
    query = filter_by_company(ExpenseReport.query)
    
    # Check view mode from URL parameter
    view_mode = request.args.get('view', 'all')
    
    # Usa i nuovi permessi espliciti per determinare cosa può vedere l'utente
    if view_mode == 'my' or current_user.can_view_my_expense_reports() and not current_user.can_view_expense_reports():
        # Mostra solo le note spese dell'utente corrente
        query = query.filter(ExpenseReport.employee_id == current_user.id)
        page_title = "Le Mie Note Spese"
    elif current_user.can_view_expense_reports() or current_user.can_approve_expense_reports():
        # Utente può vedere tutte le note spese (eventualmente filtrate per sede)
        if not current_user.all_sedi and current_user.sede_id:
            # Filtra per sede se non ha accesso globale
            sede_users = User.query.filter(User.sede_id == current_user.sede_id).with_entities(User.id).all()
            sede_user_ids = [u.id for u in sede_users]
            query = query.filter(ExpenseReport.employee_id.in_(sede_user_ids))
        page_title = "Note Spese"
    else:
        # Fallback: mostra solo le proprie
        query = query.filter(ExpenseReport.employee_id == current_user.id)
        page_title = "Le Mie Note Spese"
    
    # Applica filtri se presenti
    if filter_form.validate_on_submit():
        if filter_form.employee_id.data:
            query = query.filter(ExpenseReport.employee_id == filter_form.employee_id.data)
        if filter_form.category_id.data:
            query = query.filter(ExpenseReport.category_id == filter_form.category_id.data)
        if filter_form.status.data:
            query = query.filter(ExpenseReport.status == filter_form.status.data)
        if filter_form.date_from.data:
            query = query.filter(ExpenseReport.expense_date >= filter_form.date_from.data)
        if filter_form.date_to.data:
            query = query.filter(ExpenseReport.expense_date <= filter_form.date_to.data)
    
    # Ordina per data più recente
    expenses = query.order_by(ExpenseReport.expense_date.desc(), ExpenseReport.created_at.desc()).all()
    
    return render_template('expense_reports.html', 
                         expenses=expenses, 
                         filter_form=filter_form,
                         page_title=page_title,
                         view_mode=view_mode)

@expense_bp.route('/reports/create', methods=['GET', 'POST'])
@login_required
def create_expense_report():
    """Crea nuova nota spese"""
    if not current_user.can_create_expense_reports():
        flash('Non hai i permessi per creare note spese', 'danger')
        return redirect(url_for('expense.expense_reports'))
    
    form = ExpenseReportForm()
    
    if form.validate_on_submit():
        # Gestione upload file
        receipt_filename = None
        if form.receipt_file.data:
            file = form.receipt_file.data
            filename = secure_filename(file.filename)
            
            # Crea nome file unico
            file_extension = filename.rsplit('.', 1)[1].lower() if '.' in filename else ''
            unique_filename = f"{uuid.uuid4().hex}.{file_extension}"
            
            # Crea directory uploads se non esiste
            upload_dir = os.path.join(app.root_path, 'static', 'uploads', 'expenses')
            os.makedirs(upload_dir, exist_ok=True)
            
            file_path = os.path.join(upload_dir, unique_filename)
            file.save(file_path)
            receipt_filename = unique_filename
        
        # Crea nota spese
        expense = ExpenseReport(
            employee_id=current_user.id,
            expense_date=form.expense_date.data,
            description=form.description.data,
            amount=form.amount.data,
            category_id=form.category_id.data,
            receipt_filename=receipt_filename
        )
        set_company_on_create(expense)
        
        db.session.add(expense)
        db.session.commit()
        
        flash('Nota spese creata con successo', 'success')
        return redirect(url_for('expense.expense_reports'))
    
    return render_template('create_expense_report.html', form=form)

@expense_bp.route('/reports/edit/<int:expense_id>', methods=['GET', 'POST'])
@login_required
def edit_expense_report(expense_id):
    """Modifica nota spese esistente"""
    import os
    from werkzeug.utils import secure_filename
    import uuid
    
    expense = filter_by_company(ExpenseReport.query).filter_by(id=expense_id).first_or_404()
    
    # Verifica permessi
    if expense.employee_id != current_user.id and not current_user.can_manage_expense_reports():
        flash('Non hai i permessi per modificare questa nota spese', 'danger')
        return redirect(url_for('expense.expense_reports'))
    
    # Verifica se modificabile
    if not expense.can_be_edited():
        flash('Questa nota spese non può più essere modificata', 'warning')
        return redirect(url_for('expense.expense_reports'))
    
    form = ExpenseReportForm()
    
    if form.validate_on_submit():
        # Gestione upload file
        if form.receipt_file.data:
            # Elimina vecchio file se esiste
            if expense.receipt_filename:
                old_file_path = os.path.join(app.root_path, 'static', 'uploads', 'expenses', expense.receipt_filename)
                if os.path.exists(old_file_path):
                    os.remove(old_file_path)
            
            file = form.receipt_file.data
            filename = secure_filename(file.filename)
            
            # Crea nome file unico
            file_extension = filename.rsplit('.', 1)[1].lower() if '.' in filename else ''
            unique_filename = f"{uuid.uuid4().hex}.{file_extension}"
            
            # Crea directory uploads se non esiste
            upload_dir = os.path.join(app.root_path, 'static', 'uploads', 'expenses')
            os.makedirs(upload_dir, exist_ok=True)
            
            file_path = os.path.join(upload_dir, unique_filename)
            file.save(file_path)
            expense.receipt_filename = unique_filename
        
        # Aggiorna dati
        expense.expense_date = form.expense_date.data
        expense.description = form.description.data
        expense.amount = form.amount.data
        expense.category_id = form.category_id.data
        
        db.session.commit()
        
        flash('Nota spese aggiornata con successo', 'success')
        return redirect(url_for('expense.expense_reports'))
    
    # Pre-popola il form
    if request.method == 'GET':
        form.expense_date.data = expense.expense_date
        form.description.data = expense.description
        form.amount.data = expense.amount
        form.category_id.data = expense.category_id
    
    return render_template('edit_expense_report.html', form=form, expense=expense)

@expense_bp.route('/reports/approve/<int:expense_id>', methods=['GET', 'POST'])
@login_required
@require_manage_expense_permission
def approve_expense_report(expense_id):
    """Approva/rifiuta nota spese"""
    from forms import ExpenseApprovalForm
    
    expense = filter_by_company(ExpenseReport.query).filter_by(id=expense_id).first_or_404()
    
    # Verifica permessi
    if not expense.can_be_approved_by(current_user):
        flash('Non hai i permessi per approvare questa nota spese', 'danger')
        return redirect(url_for('expense.expense_reports'))
    
    if expense.status != 'pending':
        flash('Questa nota spese è già stata processata', 'warning')
        return redirect(url_for('expense.expense_reports'))
    
    form = ExpenseApprovalForm()
    
    if form.validate_on_submit():
        if form.action.data == 'approve':
            expense.approve(current_user, form.comment.data)
            flash('Nota spese approvata con successo', 'success')
        else:
            expense.reject(current_user, form.comment.data)
            flash('Nota spese rifiutata', 'info')
        
        db.session.commit()
        return redirect(url_for('expense.expense_reports'))
    
    return render_template('approve_expense_report.html', form=form, expense=expense)

@expense_bp.route('/reports/download/<int:expense_id>')
@login_required
def download_expense_receipt(expense_id):
    """Download ricevuta allegata"""
    import os
    
    expense = filter_by_company(ExpenseReport.query).filter_by(id=expense_id).first_or_404()
    
    # Verifica permessi
    if (expense.employee_id != current_user.id and 
        not current_user.can_view_expense_reports() and 
        not current_user.can_approve_expense_reports()):
        flash('Non hai i permessi per scaricare questo documento', 'danger')
        return redirect(url_for('expense.expense_reports'))
    
    if not expense.receipt_filename:
        flash('Nessun documento allegato a questa nota spese', 'warning')
        return redirect(url_for('expense.expense_reports'))
    
    file_path = os.path.join(app.root_path, 'static', 'uploads', 'expenses', expense.receipt_filename)
    
    if not os.path.exists(file_path):
        flash('File non trovato', 'danger')
        return redirect(url_for('expense.expense_reports'))
    
    return send_file(file_path, as_attachment=True, 
                    download_name=f"ricevuta_{expense.id}_{expense.expense_date.strftime('%Y%m%d')}.{expense.receipt_filename.split('.')[-1]}")

@expense_bp.route('/reports/delete/<int:expense_id>', methods=['POST'])
@login_required
def delete_expense_report(expense_id):
    """Elimina nota spese"""
    import os
    
    expense = filter_by_company(ExpenseReport.query).filter_by(id=expense_id).first_or_404()
    
    # Verifica permessi
    if expense.employee_id != current_user.id and not current_user.can_manage_expense_reports():
        flash('Non hai i permessi per eliminare questa nota spese', 'danger')
        return redirect(url_for('expense.expense_reports'))
    
    # Solo note in attesa possono essere eliminate
    if expense.status != 'pending':
        flash('Solo le note spese in attesa possono essere eliminate', 'warning')
        return redirect(url_for('expense.expense_reports'))
    
    # Elimina file allegato se esiste
    if expense.receipt_filename:
        file_path = os.path.join(app.root_path, 'static', 'uploads', 'expenses', expense.receipt_filename)
        if os.path.exists(file_path):
            os.remove(file_path)
    
    db.session.delete(expense)
    db.session.commit()
    
    flash('Nota spese eliminata con successo', 'success')
    return redirect(url_for('expense.expense_reports'))

# =============================================================================
# EXPENSE CATEGORIES ROUTES
# =============================================================================

@expense_bp.route('/categories')
@login_required
@require_expense_permission
def expense_categories():
    """Gestisci categorie note spese"""
    if not current_user.can_manage_expense_reports():
        flash('Non hai i permessi per gestire le categorie', 'danger')
        return redirect(url_for('expense.expense_reports'))
    
    from models import ExpenseCategory
    categories = filter_by_company(ExpenseCategory.query).order_by(ExpenseCategory.name).all()
    
    return render_template('expense_categories.html', categories=categories)

@expense_bp.route('/categories/create', methods=['GET', 'POST'])
@login_required
@require_manage_expense_permission
def create_expense_category():
    """Crea nuova categoria"""
    if not current_user.can_manage_expense_reports():
        flash('Non hai i permessi per creare categorie', 'danger')
        return redirect(url_for('expense.expense_reports'))
    
    from models import ExpenseCategory
    from forms import ExpenseCategoryForm
    
    form = ExpenseCategoryForm()
    
    if form.validate_on_submit():
        category = ExpenseCategory(
            name=form.name.data,
            description=form.description.data,
            active=form.active.data,
            created_by=current_user.id
        )
        set_company_on_create(category)
        
        db.session.add(category)
        
        try:
            db.session.commit()
            flash('Categoria creata con successo', 'success')
            return redirect(url_for('expense.expense_categories'))
        except:
            db.session.rollback()
            flash('Errore nella creazione della categoria', 'danger')
    
    return render_template('create_expense_category.html', form=form)

@expense_bp.route('/categories/<int:category_id>/edit', methods=['GET', 'POST'])
@login_required
@require_manage_expense_permission
def edit_expense_category(category_id):
    """Modifica categoria nota spese"""
    if not current_user.can_manage_expense_reports():
        flash('Non hai i permessi per modificare le categorie', 'danger')
        return redirect(url_for('expense.expense_categories'))
    
    from models import ExpenseCategory
    
    category = filter_by_company(ExpenseCategory.query).filter_by(id=category_id).first_or_404()
    form = ExpenseCategoryForm(obj=category)
    
    if form.validate_on_submit():
        category.name = form.name.data
        category.description = form.description.data
        category.active = form.active.data
        
        try:
            db.session.commit()
            flash('Categoria modificata con successo', 'success')
            return redirect(url_for('expense.expense_categories'))
        except:
            db.session.rollback()
            flash('Errore: nome categoria già esistente', 'danger')
    
    return render_template('edit_expense_category.html', form=form, category=category)

@expense_bp.route('/categories/<int:category_id>/delete', methods=['POST'])
@login_required
@require_manage_expense_permission
def delete_expense_category(category_id):
    """Elimina categoria nota spese"""
    if not current_user.can_manage_expense_reports():
        flash('Non hai i permessi per eliminare le categorie', 'danger')
        return redirect(url_for('expense.expense_categories'))
    
    from models import ExpenseCategory
    
    category = filter_by_company(ExpenseCategory.query).filter_by(id=category_id).first_or_404()
    
    # Verifica se ci sono note spese associate
    if category.expense_reports and len(category.expense_reports) > 0:
        flash('Non è possibile eliminare una categoria con note spese associate', 'warning')
        return redirect(url_for('expense.expense_categories'))
    
    try:
        name = category.name
        db.session.delete(category)
        db.session.commit()
        flash(f'Categoria "{name}" eliminata con successo', 'success')
    except:
        db.session.rollback()
        flash('Errore nell\'eliminazione della categoria', 'danger')
    
    return redirect(url_for('expense.expense_categories'))

# =============================================================================
# OVERTIME MANAGEMENT ROUTES
# =============================================================================

@expense_bp.route('/overtime/types')
@login_required
@require_overtime_permission
def overtime_types():
    """Visualizzazione e gestione tipologie straordinari"""
    if not (current_user.can_manage_overtime_types() or current_user.can_view_overtime_types()):
        flash('Non hai i permessi per visualizzare le tipologie di straordinario.', 'warning')
        return redirect(url_for('dashboard.dashboard'))
    
    from models import OvertimeType
    types = filter_by_company(OvertimeType.query).all()
    return render_template('overtime_types.html', types=types)

@expense_bp.route('/overtime/types/create', methods=['GET', 'POST'])
@login_required
def create_overtime_type():
    """Creazione nuova tipologia straordinario"""
    if not (current_user.can_manage_overtime_types() or current_user.can_create_overtime_types()):
        flash('Non hai i permessi per creare tipologie di straordinario.', 'warning')
        return redirect(url_for('expense.overtime_types'))
    
    from models import OvertimeType
    
    form = OvertimeTypeForm()
    if form.validate_on_submit():
        overtime_type = OvertimeType(
            name=form.name.data,
            description=form.description.data,
            hourly_rate_multiplier=form.hourly_rate_multiplier.data,
            active=form.active.data
        )
        set_company_on_create(overtime_type)
        db.session.add(overtime_type)
        db.session.commit()
        flash('Tipologia straordinario creata con successo!', 'success')
        return redirect(url_for('expense.overtime_types'))
    
    return render_template('create_overtime_type.html', form=form)

@expense_bp.route('/overtime/types/<int:type_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_overtime_type(type_id):
    """Modifica tipologia straordinario"""
    if not current_user.can_manage_overtime_types():
        flash("Non hai i permessi per modificare le tipologie di straordinario.", "warning")
        return redirect(url_for("expense.overtime_types"))
    
    from models import OvertimeType
    
    overtime_type = filter_by_company(OvertimeType.query).filter_by(id=type_id).first_or_404()
    form = OvertimeTypeForm(obj=overtime_type)
    
    if form.validate_on_submit():
        overtime_type.name = form.name.data
        overtime_type.description = form.description.data
        overtime_type.hourly_rate_multiplier = form.hourly_rate_multiplier.data
        overtime_type.active = form.active.data
        
        db.session.commit()
        flash("Tipologia straordinario aggiornata con successo!", "success")
        return redirect(url_for("expense.overtime_types"))
    
    return render_template("edit_overtime_type.html", form=form, overtime_type=overtime_type)

@expense_bp.route('/overtime/types/<int:type_id>/delete', methods=['POST'])
@login_required
def delete_overtime_type(type_id):
    """Cancella tipologia straordinario"""
    if not current_user.can_manage_overtime_types():
        flash("Non hai i permessi per cancellare le tipologie di straordinario.", "warning")
        return redirect(url_for("expense.overtime_types"))
    
    from models import OvertimeType, OvertimeRequest
    
    overtime_type = filter_by_company(OvertimeType.query).filter_by(id=type_id).first_or_404()
    
    # Controlla se ci sono richieste associate
    requests_count = filter_by_company(OvertimeRequest.query).filter_by(overtime_type_id=type_id).count()
    if requests_count > 0:
        flash(f"Impossibile cancellare: ci sono {requests_count} richieste associate a questa tipologia.", "warning")
        return redirect(url_for("expense.overtime_types"))
    
    db.session.delete(overtime_type)
    db.session.commit()
    flash("Tipologia straordinario cancellata.", "info")
    return redirect(url_for("expense.overtime_types"))

@expense_bp.route('/overtime/requests')
@login_required
@require_overtime_permission
def overtime_requests_management():
    """Gestione richieste straordinario"""
    from models import OvertimeRequest, OvertimeType
    from forms import OvertimeFilterForm
    
    filter_form = OvertimeFilterForm()
    
    # Query base (with company filter)
    query = filter_by_company(OvertimeRequest.query)
    
    # Filtra per sede se l'utente non ha accesso globale
    if not current_user.all_sedi and current_user.sede_id:
        sede_users = User.query.filter(User.sede_id == current_user.sede_id).with_entities(User.id).all()
        sede_user_ids = [u.id for u in sede_users]
        query = query.filter(OvertimeRequest.employee_id.in_(sede_user_ids))
    
    # Applica filtri
    if filter_form.validate_on_submit():
        if filter_form.user_id.data:
            query = query.filter(OvertimeRequest.employee_id == filter_form.user_id.data)
        if filter_form.status.data:
            query = query.filter(OvertimeRequest.status == filter_form.status.data)
        if filter_form.overtime_type_id.data:
            query = query.filter(OvertimeRequest.overtime_type_id == filter_form.overtime_type_id.data)
        if filter_form.date_from.data:
            query = query.filter(OvertimeRequest.overtime_date >= filter_form.date_from.data)
        if filter_form.date_to.data:
            query = query.filter(OvertimeRequest.overtime_date <= filter_form.date_to.data)
    
    # Ordina per data più recente
    requests = query.order_by(OvertimeRequest.overtime_date.desc()).all()
    
    return render_template('overtime_requests.html', 
                         requests=requests, 
                         form=filter_form)

@expense_bp.route('/overtime/requests/create', methods=['GET', 'POST'])
@login_required
def create_overtime_request():
    """Creazione richiesta straordinario"""
    if not current_user.can_create_overtime_requests():
        flash('Non hai i permessi per creare richieste di straordinario.', 'warning')
        return redirect(url_for('dashboard.dashboard'))
    
    form = OvertimeRequestForm()
    if form.validate_on_submit():
        # Calcola le ore di straordinario
        start_datetime = datetime.combine(form.overtime_date.data, form.start_time.data)
        end_datetime = datetime.combine(form.overtime_date.data, form.end_time.data)
        hours = (end_datetime - start_datetime).total_seconds() / 3600
        
        overtime_request = OvertimeRequest(
            employee_id=current_user.id,
            overtime_date=form.overtime_date.data,
            start_time=form.start_time.data,
            end_time=form.end_time.data,
            motivation=form.motivation.data,
            overtime_type_id=form.overtime_type_id.data,
            status='Pending'
        )
        set_company_on_create(overtime_request)
        db.session.add(overtime_request)
        db.session.commit()
        
        # Invia notifica automatica agli approvatori (se la funzione esiste)
        try:
            from utils import send_overtime_request_message
            send_overtime_request_message(overtime_request, 'created', current_user)
        except ImportError:
            pass  # Funzione di notifica non disponibile
        
        flash('Richiesta straordinario inviata con successo!', 'success')
        return redirect(url_for('expense.my_overtime_requests'))
    
    return render_template('create_overtime_request.html', form=form)

@expense_bp.route('/overtime/requests/my')
@login_required
def my_overtime_requests():
    """Le mie richieste straordinari"""
    if not current_user.can_view_my_overtime_requests():
        flash('Non hai i permessi per visualizzare le tue richieste di straordinario.', 'warning')
        return redirect(url_for('dashboard.dashboard'))
    
    requests = filter_by_company(OvertimeRequest.query).filter_by(employee_id=current_user.id).options(
        joinedload(OvertimeRequest.overtime_type)
    ).order_by(OvertimeRequest.created_at.desc()).all()
    
    return render_template('my_overtime_requests.html', requests=requests)

@expense_bp.route('/overtime/requests/<int:request_id>/approve', methods=['POST'])
@login_required
def approve_overtime_request(request_id):
    """Approva richiesta straordinario"""
    if not current_user.can_approve_overtime_requests():
        flash('Non hai i permessi per approvare richieste di straordinario.', 'warning')
        return redirect(url_for('expense.overtime_requests_management'))
    
    overtime_request = filter_by_company(OvertimeRequest.query).filter_by(id=request_id).first_or_404()
    
    if not current_user.all_sedi and overtime_request.employee.sede_id != current_user.sede_id:
        flash('Non puoi approvare richieste di altre sedi.', 'warning')
        return redirect(url_for('expense.overtime_requests_management'))
    
    overtime_request.status = 'Approved'
    overtime_request.approval_comment = request.form.get('approval_comment', '')
    overtime_request.approved_by = current_user.id
    try:
        from utils import italian_now
        overtime_request.approved_at = italian_now()
    except ImportError:
        overtime_request.approved_at = datetime.now()
    
    db.session.commit()
    
    # Invia notifica automatica all'utente (se la funzione esiste)
    try:
        from utils import send_overtime_request_message
        send_overtime_request_message(overtime_request, 'approved', current_user)
    except ImportError:
        pass  # Funzione di notifica non disponibile
    
    flash('Richiesta straordinario approvata con successo!', 'success')
    return redirect(url_for('expense.overtime_requests_management'))

@expense_bp.route('/overtime/requests/<int:request_id>/reject', methods=['POST'])
@login_required
def reject_overtime_request(request_id):
    """Rifiuta richiesta straordinario"""
    if not current_user.can_approve_overtime_requests():
        flash('Non hai i permessi per rifiutare richieste di straordinario.', 'warning')
        return redirect(url_for('expense.overtime_requests_management'))
    
    overtime_request = filter_by_company(OvertimeRequest.query).filter_by(id=request_id).first_or_404()
    
    if not current_user.all_sedi and overtime_request.employee.sede_id != current_user.sede_id:
        flash('Non puoi rifiutare richieste di altre sedi.', 'warning')
        return redirect(url_for('expense.overtime_requests_management'))
    
    overtime_request.status = 'Rejected'
    overtime_request.approval_comment = request.form.get('approval_comment', '')
    overtime_request.approved_by = current_user.id
    try:
        from utils import italian_now
        overtime_request.approved_at = italian_now()
    except ImportError:
        overtime_request.approved_at = datetime.now()
    
    db.session.commit()
    
    # Invia notifica automatica all'utente (se la funzione esiste)
    try:
        from utils import send_overtime_request_message
        send_overtime_request_message(overtime_request, 'rejected', current_user)
    except ImportError:
        pass  # Funzione di notifica non disponibile
    
    flash('Richiesta straordinario rifiutata.', 'info')
    return redirect(url_for('expense.overtime_requests_management'))

@expense_bp.route('/overtime/requests/<int:request_id>/delete', methods=['POST'])
@login_required
def delete_overtime_request(request_id):
    """Cancella richiesta straordinario"""
    overtime_request = filter_by_company(OvertimeRequest.query).filter_by(id=request_id).first_or_404()
    
    # Solo il proprietario può cancellare se in stato pending
    if overtime_request.employee_id != current_user.id:
        flash('Non puoi cancellare richieste di altri utenti.', 'warning')
        return redirect(url_for('expense.my_overtime_requests'))
    
    if overtime_request.status != 'pending':
        flash('Non puoi cancellare richieste già approvate o rifiutate.', 'warning')
        return redirect(url_for('expense.my_overtime_requests'))
    
    db.session.delete(overtime_request)
    db.session.commit()
    flash('Richiesta straordinario cancellata.', 'info')
    return redirect(url_for('expense.my_overtime_requests'))

@expense_bp.route('/overtime/requests/export')
@login_required
def overtime_requests_excel():
    """Export Excel richieste straordinari"""
    if not current_user.can_view_overtime_requests():
        flash('Non hai i permessi per esportare le richieste di straordinario.', 'warning')
        return redirect(url_for('dashboard.dashboard'))
    
    # Query con filtri sede
    if current_user.all_sedi:
        requests = filter_by_company(OvertimeRequest.query).options(
            joinedload(OvertimeRequest.employee),
            joinedload(OvertimeRequest.overtime_type)
        ).order_by(OvertimeRequest.created_at.desc()).all()
    else:
        requests = filter_by_company(OvertimeRequest.query).join(User).filter(
            User.sede_id == current_user.sede_id
        ).options(
            joinedload(OvertimeRequest.employee),
            joinedload(OvertimeRequest.overtime_type)
        ).order_by(OvertimeRequest.created_at.desc()).all()
    
    # Creazione Excel
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from io import BytesIO
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Richieste Straordinari"
    
    # Headers
    headers = ['Utente', 'Data', 'Orario Inizio', 'Orario Fine', 'Ore', 'Tipologia', 'Moltiplicatore', 'Stato', 'Motivazione', 'Creata il']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Dati
    for row, req in enumerate(requests, 2):
        ws.cell(row=row, column=1, value=req.employee.get_full_name())
        ws.cell(row=row, column=2, value=req.overtime_date.strftime('%d/%m/%Y'))
        ws.cell(row=row, column=3, value=req.start_time.strftime('%H:%M'))
        ws.cell(row=row, column=4, value=req.end_time.strftime('%H:%M'))
        ws.cell(row=row, column=5, value=f"{req.hours:.1f}")
        ws.cell(row=row, column=6, value=req.overtime_type.name)
        ws.cell(row=row, column=7, value=f"x{req.overtime_type.hourly_rate_multiplier:.1f}")
        ws.cell(row=row, column=8, value=req.status.upper())
        ws.cell(row=row, column=9, value=req.motivation[:100])
        ws.cell(row=row, column=10, value=req.created_at.strftime('%d/%m/%Y %H:%M'))
    
    # Auto-adjust columns
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Salva in memoria
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    try:
        from utils import italian_now
        today = italian_now().strftime("%Y%m%d")
    except ImportError:
        today = datetime.now().strftime("%Y%m%d")
    
    filename = f"richieste_straordinari_{today}.xlsx"
    
    response = make_response(output.read())
    response.headers['Content-Disposition'] = f'attachment; filename={filename}'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    
    return response

# =============================================================================
# MILEAGE MANAGEMENT ROUTES
# =============================================================================

@expense_bp.route('/mileage/requests')
@login_required
@require_mileage_permission
def mileage_requests():
    """Visualizza le richieste di rimborso chilometrico"""
    try:
        if not (current_user.can_view_mileage_requests() or current_user.can_manage_mileage_requests()):
            flash('Non hai i permessi per visualizzare le richieste di rimborso chilometrico.', 'warning')
            return redirect(url_for('dashboard.dashboard'))
        
        # Filtri
        filter_form = MileageFilterForm(current_user=current_user)
        
        # Base query (with company filter)
        query = filter_by_company(MileageRequest.query).options(
            joinedload(MileageRequest.user),
            joinedload(MileageRequest.approver),  
            joinedload(MileageRequest.vehicle)
        )
        
        # Filtri per sede (se l'utente non ha accesso globale)
        if not current_user.all_sedi and current_user.sede_id:
            query = query.join(User, MileageRequest.user_id == User.id).filter(User.sede_id == current_user.sede_id)
        
        # Applica filtri dal form
        if request.method == 'POST' and filter_form.validate_on_submit():
            if filter_form.status.data:
                query = query.filter(MileageRequest.status == filter_form.status.data)
            if filter_form.user_id.data:
                query = query.filter(MileageRequest.user_id == filter_form.user_id.data)
            if filter_form.date_from.data:
                query = query.filter(MileageRequest.travel_date >= filter_form.date_from.data)
            if filter_form.date_to.data:
                query = query.filter(MileageRequest.travel_date <= filter_form.date_to.data)
        
        # Ordina per data più recente
        requests = query.order_by(MileageRequest.created_at.desc()).all()
        
        return render_template('mileage_requests.html', requests=requests, filter_form=filter_form)
    except Exception as e:
        flash('Errore nel caricamento delle richieste di rimborso.', 'danger')
        return redirect(url_for('dashboard.dashboard'))

@expense_bp.route('/mileage/requests/create', methods=['GET', 'POST'])
@login_required
def create_mileage_request():
    """Crea una nuova richiesta di rimborso chilometrico"""
    if not current_user.can_create_mileage_requests():
        flash('Non hai i permessi per creare richieste di rimborso chilometrico.', 'warning')
        return redirect(url_for('dashboard.dashboard'))
    
    # Controlla se l'utente ha un veicolo ACI assegnato
    if not current_user.aci_vehicle_id:
        flash('Non puoi creare richieste di rimborso chilometrico senza avere un veicolo ACI assegnato. Contatta l\'amministratore per associare un veicolo al tuo profilo.', 'warning')
        return redirect(url_for('expense.my_mileage_requests'))
    
    form = MileageRequestForm(user=current_user)
    
    if form.validate_on_submit():
        try:
            # Converti route_addresses da stringa a array
            route_addresses_list = []
            if form.route_addresses.data:
                # Dividi per righe e filtra righe vuote
                route_addresses_list = [addr.strip() for addr in form.route_addresses.data.split('\n') if addr.strip()]
            
            # Crea la richiesta
            mileage_request = MileageRequest(
                user_id=current_user.id,
                travel_date=form.travel_date.data,
                route_addresses=route_addresses_list,  # Array invece di stringa
                total_km=form.total_km.data,
                is_km_manual=form.is_km_manual.data,
                purpose=form.purpose.data,
                notes=form.notes.data,
                vehicle_id=form.vehicle_id.data if form.vehicle_id.data else None,
                vehicle_description=form.vehicle_description.data if form.vehicle_description.data else None
            )
            
            # Calcola l'importo del rimborso
            mileage_request.calculate_reimbursement_amount()
            set_company_on_create(mileage_request)
            
            db.session.add(mileage_request)
            db.session.commit()
            
            flash('Richiesta di rimborso chilometrico inviata con successo!', 'success')
            return redirect(url_for('expense.my_mileage_requests'))
            
        except Exception as e:
            db.session.rollback()
            flash('Errore nella creazione della richiesta di rimborso. Riprova più tardi.', 'danger')
            return render_template('create_mileage_request.html', form=form)
    
    return render_template('create_mileage_request.html', form=form)

@expense_bp.route('/mileage/requests/my')
@login_required
def my_mileage_requests():
    """Visualizza le richieste di rimborso chilometrico dell'utente corrente"""
    try:
        if not current_user.can_view_my_mileage_requests():
            flash('Non hai i permessi per visualizzare le tue richieste di rimborso chilometrico.', 'warning')
            return redirect(url_for('dashboard.dashboard'))
        
        requests = filter_by_company(MileageRequest.query).filter_by(user_id=current_user.id)\
                                      .options(joinedload(MileageRequest.approver),
                                              joinedload(MileageRequest.vehicle))\
                                      .order_by(MileageRequest.created_at.desc()).all()
        
        return render_template('my_mileage_requests.html', requests=requests)
    except Exception as e:
        flash('Errore nel caricamento delle richieste di rimborso.', 'danger')
        return redirect(url_for('dashboard.dashboard'))

@expense_bp.route('/mileage/requests/<int:request_id>/approve', methods=['POST'])
@login_required
def approve_mileage_request(request_id):
    """Approva una richiesta di rimborso chilometrico"""
    if not current_user.can_approve_mileage_requests():
        flash('Non hai i permessi per approvare richieste di rimborso chilometrico.', 'warning')
        return redirect(url_for('expense.mileage_requests'))
    
    mileage_request = filter_by_company(MileageRequest.query).filter_by(id=request_id).first_or_404()
    
    # Controllo per sede se necessario
    if not current_user.all_sedi and current_user.sede_id:
        if mileage_request.user.sede_id != current_user.sede_id:
            flash('Non puoi approvare richieste di utenti di altre sedi.', 'warning')
            return redirect(url_for('expense.mileage_requests'))
    
    action = request.form.get('action')
    comment = request.form.get('comment', '')
    
    if action == 'approve':
        mileage_request.status = 'Approved'
        mileage_request.approval_comment = comment
        mileage_request.approved_by = current_user.id
        try:
            from utils import italian_now
            mileage_request.approved_at = italian_now()
        except ImportError:
            mileage_request.approved_at = datetime.now()
        
        flash('Richiesta di rimborso chilometrico approvata con successo!', 'success')
    elif action == 'reject':
        if not comment:
            flash('Il commento è obbligatorio per rifiutare una richiesta.', 'warning')
            return redirect(url_for('expense.mileage_requests'))
        
        mileage_request.status = 'Rejected'
        mileage_request.approval_comment = comment
        mileage_request.approved_by = current_user.id
        try:
            from utils import italian_now
            mileage_request.approved_at = italian_now()
        except ImportError:
            mileage_request.approved_at = datetime.now()
        
        flash('Richiesta di rimborso chilometrico rifiutata.', 'success')
    else:
        flash('Azione non valida.', 'danger')
        return redirect(url_for('expense.mileage_requests'))
    
    db.session.commit()
    
    # Invia notifica automatica all'utente (se la funzione esiste)
    try:
        from utils import send_mileage_request_message
        send_mileage_request_message(mileage_request, action, current_user)
    except ImportError:
        pass  # Funzione di notifica non disponibile
    
    return redirect(url_for('expense.mileage_requests'))

@expense_bp.route('/mileage/requests/<int:request_id>/delete', methods=['POST'])
@login_required
def delete_mileage_request(request_id):
    """Cancella una richiesta di rimborso chilometrico"""
    mileage_request = filter_by_company(MileageRequest.query).filter_by(id=request_id).first_or_404()
    
    # Solo l'autore o un manager può cancellare
    can_delete = (mileage_request.user_id == current_user.id and 
                  current_user.can_view_my_mileage_requests()) or \
                 current_user.can_manage_mileage_requests()
    
    if not can_delete:
        flash('Non hai i permessi per cancellare questa richiesta.', 'warning')
        return redirect(url_for('expense.mileage_requests'))
    
    # Non permettere cancellazione se già approvata
    if mileage_request.status == 'approved':
        flash('Non è possibile cancellare una richiesta già approvata.', 'warning')
        return redirect(url_for('expense.mileage_requests'))
    
    db.session.delete(mileage_request)
    db.session.commit()
    
    flash('Richiesta di rimborso chilometrico cancellata con successo.', 'success')
    return redirect(url_for('expense.mileage_requests'))

# Duplicate create_mileage_request function removed - already defined earlier in file

@expense_bp.route('/api/calculate_distance', methods=['POST'])
@login_required
def calculate_distance():
    """Calcola la distanza tra indirizzi usando API di geocoding"""
    try:
        data = request.get_json()
        addresses = data.get('addresses', [])
        
        if len(addresses) < 2:
            return jsonify({'error': 'Servono almeno 2 indirizzi', 'km': 0})
        
        # Per ora simuliamo il calcolo - in produzione usare Google Maps API o simile
        # Calcolo basato su distanza approssimativa tra città italiane
        total_km = 0
        
        for i in range(len(addresses) - 1):
            start = addresses[i].lower()
            end = addresses[i + 1].lower()
            
            # Calcolo approssimativo basato su città principali
            segment_km = calculate_approximate_distance(start, end)
            total_km += segment_km
        
        return jsonify({
            'success': True,
            'km': round(total_km, 1),
            'segments': len(addresses) - 1
        })
        
    except Exception as e:
        return jsonify({'error': str(e), 'km': 0})

def calculate_approximate_distance(start_address, end_address):
    """Calcola distanza approssimativa tra due indirizzi italiani"""
    
    # Database semplificato di coordinate delle principali città italiane
    city_coords = {
        'roma': (41.9028, 12.4964),
        'milano': (45.4642, 9.1900),
        'napoli': (40.8518, 14.2681),
        'torino': (45.0703, 7.6869),
        'firenze': (43.7696, 11.2558),
        'bologna': (44.4949, 11.3426),
        'genova': (44.4056, 8.9463),
        'palermo': (38.1157, 13.3613),
        'bari': (41.1171, 16.8719),
        'catania': (37.5079, 15.0830),
        'venezia': (45.4408, 12.3155),
        'verona': (45.4384, 10.9916),
        'messina': (38.1938, 15.5540),
        'padova': (45.4064, 11.8768),
        'trieste': (45.6495, 13.7768),
        'brescia': (45.5416, 10.2118),
        'parma': (44.8015, 10.3279),
        'modena': (44.6471, 10.9252),
        'reggio calabria': (38.1059, 15.6219),
        'perugia': (43.1122, 12.3888)
    }
    
    def extract_city(address):
        """Estrae il nome della città dall'indirizzo"""
        address = address.lower().strip()
        
        # Cerca la città nell'indirizzo
        for city in city_coords.keys():
            if city in address:
                return city
        
        # Se non trova nulla, restituisce None
        return None
    
    def haversine_distance(lat1, lon1, lat2, lon2):
        """Calcola distanza usando formula haversine"""
        import math
        
        # Raggio della Terra in km
        R = 6371
        
        # Converti in radianti
        lat1, lon1, lat2, lon2 = map(math.radians, [lat1, lon1, lat2, lon2])
        
        # Differenze
        dlat = lat2 - lat1
        dlon = lon2 - lon1
        
        # Formula haversine
        a = math.sin(dlat/2)**2 + math.cos(lat1) * math.cos(lat2) * math.sin(dlon/2)**2
        c = 2 * math.asin(math.sqrt(a))
        
        return R * c
    
    # Estrai città dagli indirizzi
    start_city = extract_city(start_address)
    end_city = extract_city(end_address)
    
    # Se entrambe le città sono note, calcola distanza reale
    if start_city and end_city and start_city in city_coords and end_city in city_coords:
        lat1, lon1 = city_coords[start_city]
        lat2, lon2 = city_coords[end_city]
        return haversine_distance(lat1, lon1, lat2, lon2)
    
    # Altrimenti, stima generica basata su lunghezza stringa (molto approssimativa)
    return max(10, len(start_address + end_address) * 2)

@expense_bp.route('/mileage/export')
@login_required
@require_mileage_permission
def mileage_requests_excel():
    """Export Excel richieste rimborsi chilometrici"""
    if not current_user.can_view_mileage_requests():
        flash('Non hai i permessi per esportare le richieste di rimborso chilometrico.', 'warning')
        return redirect(url_for('dashboard.dashboard'))
    
    # Query con filtri sede (with company filter)
    query = filter_by_company(MileageRequest.query).options(
        joinedload(MileageRequest.user),
        joinedload(MileageRequest.approver),
        joinedload(MileageRequest.vehicle)
    )
    
    if not current_user.all_sedi and current_user.sede_id:
        query = query.join(User, MileageRequest.user_id == User.id).filter(User.sede_id == current_user.sede_id)
    
    requests = query.order_by(MileageRequest.created_at.desc()).all()
    
    # Creazione Excel
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from io import BytesIO
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Rimborsi Chilometrici"
    
    # Headers
    headers = ['Utente', 'Data Viaggio', 'Tragitto', 'Km Totali', 'Importo', 'Scopo', 'Stato', 'Approvata da', 'Note', 'Creata il']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Dati
    for row, req in enumerate(requests, 2):
        route_str = ' → '.join(req.route_addresses) if req.route_addresses else 'N/A'
        ws.cell(row=row, column=1, value=req.user.get_full_name())
        ws.cell(row=row, column=2, value=req.travel_date.strftime('%d/%m/%Y'))
        ws.cell(row=row, column=3, value=route_str[:100])  # Tronca se troppo lungo
        ws.cell(row=row, column=4, value=f"{req.total_km:.1f}")
        ws.cell(row=row, column=5, value=f"€{req.total_amount:.2f}")
        ws.cell(row=row, column=6, value=req.purpose[:50] if req.purpose else '')
        ws.cell(row=row, column=7, value=req.status.upper())
        ws.cell(row=row, column=8, value=req.approver.get_full_name() if req.approver else '')
        ws.cell(row=row, column=9, value=req.notes[:100] if req.notes else '')
        ws.cell(row=row, column=10, value=req.created_at.strftime('%d/%m/%Y %H:%M'))
    
    # Auto-adjust columns
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Salva in memoria
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    try:
        from utils import italian_now
        today = italian_now().strftime("%Y%m%d")
    except ImportError:
        today = datetime.now().strftime("%Y%m%d")
    
    filename = f"rimborsi_chilometrici_{today}.xlsx"
    
    response = make_response(output.read())
    response.headers['Content-Disposition'] = f'attachment; filename={filename}'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    
    return response


@expense_bp.route('/mileage/requests/export')
@login_required
@require_mileage_permission
def export_mileage_requests():
    """Esporta le richieste di rimborso chilometrico in Excel"""
    if not (current_user.can_view_mileage_requests() or current_user.can_manage_mileage_requests()):
        flash('Non hai i permessi per esportare le richieste di rimborso chilometrico.', 'warning')
        return redirect(url_for('dashboard.dashboard'))
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from sqlalchemy.orm import joinedload
        
        # Base query (with company filter)
        query = filter_by_company(MileageRequest.query).options(
            joinedload(MileageRequest.user),
            joinedload(MileageRequest.approver),
            joinedload(MileageRequest.vehicle)
        )
        
        # Filtri per sede
        if not current_user.all_sedi and current_user.sede_id:
            query = query.join(User).filter(User.sede_id == current_user.sede_id)
        
        requests = query.order_by(MileageRequest.travel_date.desc()).all()
        
        # Crea workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Rimborsi Chilometrici"
        
        # Header con stile
        headers = [
            'Data Viaggio', 'Utente', 'Percorso', 'KM Totali', 'Veicolo', 
            'Importo (€)', 'Stato', 'Motivazione', 'Note', 'Data Richiesta', 
            'Approvato/Rifiutato da', 'Data Approvazione', 'Commento Approvazione'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Dati
        for row, req in enumerate(requests, 2):
            ws.cell(row=row, column=1, value=req.travel_date.strftime('%d/%m/%Y'))
            ws.cell(row=row, column=2, value=req.user.get_full_name())
            ws.cell(row=row, column=3, value=req.get_route_summary())
            ws.cell(row=row, column=4, value=req.total_km)
            
            # Veicolo
            if req.vehicle:
                vehicle_info = f"{req.vehicle.marca} {req.vehicle.modello}"
            else:
                vehicle_info = req.vehicle_description or "Non specificato"
            ws.cell(row=row, column=5, value=vehicle_info)
            
            ws.cell(row=row, column=6, value=req.total_amount or 0)
            
            # Stato con colore
            status_cell = ws.cell(row=row, column=7, value=req.get_status_display())
            if req.status == 'approved':
                status_cell.fill = PatternFill(start_color="D4F4DD", end_color="D4F4DD", fill_type="solid")
            elif req.status == 'rejected':
                status_cell.fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
            
            ws.cell(row=row, column=8, value=req.purpose)
            ws.cell(row=row, column=9, value=req.notes or '')
            ws.cell(row=row, column=10, value=req.created_at.strftime('%d/%m/%Y %H:%M'))
            
            if req.approver:
                ws.cell(row=row, column=11, value=req.approver.get_full_name())
            if req.approved_at:
                ws.cell(row=row, column=12, value=req.approved_at.strftime('%d/%m/%Y %H:%M'))
            ws.cell(row=row, column=13, value=req.approval_comment or '')
        
        # Auto-fit colonne
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Prepara response
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = make_response(output.read())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename=rimborsi_chilometrici_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        return response
        
    except Exception as e:
        flash(f'Errore durante l\'esportazione: {str(e)}', 'danger')
        return redirect(url_for('expense.mileage_requests'))

# =============================================================================
# ADDITIONAL OVERTIME TYPE ROUTES - Edit and Delete routes for existing overtime types
# =============================================================================

# NOTE: These routes are duplicates of existing routes in this blueprint.
# The existing edit and delete routes may conflict with these.
# Remove duplicate routes and keep these as they have correct URL patterns.

# =============================================================================
# BLUEPRINT REGISTRATION READY
# =============================================================================
# This blueprint is ready to be registered in main.py:
# from blueprints.expense import expense_bp
# app.register_blueprint(expense_bp)
# =============================================================================