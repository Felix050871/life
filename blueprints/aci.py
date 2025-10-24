# =============================================================================
# ACI TABLES BLUEPRINT - Sistema gestione tabelle ACI per rimborsi chilometrici
# =============================================================================
#
# ROUTES INCLUSE:
# 1. /aci_tables (GET/POST) - Visualizza tabelle ACI con filtri
# 2. /aci_tables/upload (GET/POST) - Upload e importazione file Excel
# 3. /aci_tables/create (GET/POST) - Creazione manuale record ACI
# 4. /aci_tables/<record_id>/edit (GET/POST) - Modifica record esistente
# 5. /aci_tables/<record_id>/delete (POST) - Cancellazione singolo record
# 6. /aci_tables/export (GET) - Export Excel tabelle ACI
# 7. /aci_tables/bulk_delete (POST) - Cancellazione in massa per tipologia
# 8. /api/aci/marcas (GET) - API marche filtrate per tipologia
# 9. /api/aci/modelos (GET) - API modelli filtrati per tipologia e marca
#
# Total routes: 9 ACI management routes
# =============================================================================

from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, make_response
from flask_login import login_required, current_user
from datetime import datetime
from functools import wraps
from app import db
from models import ACITable
from forms import ACIFilterForm, ACIUploadForm, ACIRecordForm
from utils_tenant import filter_by_company, set_company_on_create, get_user_company_id

# Create blueprint
aci_bp = Blueprint('aci', __name__, url_prefix='/aci')

# Helper function for admin access
def admin_required(f):
    """Decorator per verificare che l'utente sia Admin o Amministratore"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or current_user.role not in ['Admin', 'Amministratore']:
            flash('Accesso negato. Solo gli amministratori possono accedere a questa sezione.', 'danger')
            return redirect(url_for('dashboard.dashboard'))
        return f(*args, **kwargs)
    
    return decorated_function

# =============================================================================
# MAIN ACI TABLES ROUTES
# =============================================================================

@aci_bp.route("/tables", methods=["GET", "POST"])
@login_required
@admin_required
def tables():
    """Visualizza tabelle ACI con caricamento lazy - record caricati solo dopo filtro"""
    form = ACIFilterForm()
    tables = []
    total_records = filter_by_company(ACITable.query).count()
    
    # LAZY LOADING: carica record solo se è stato applicato un filtro (POST)
    if request.method == "POST" and form.validate_on_submit():
        query = filter_by_company(ACITable.query)
        
        # Applica filtri selezionati
        filters_applied = False
        if form.tipologia.data:
            query = query.filter(ACITable.tipologia == form.tipologia.data)
            filters_applied = True
        if form.marca.data:
            query = query.filter(ACITable.marca == form.marca.data)
            filters_applied = True
        if form.modello.data:
            query = query.filter(ACITable.modello == form.modello.data)
            filters_applied = True
            
        # Carica risultati solo se almeno un filtro è applicato
        if filters_applied:
            tables = query.order_by(ACITable.tipologia, ACITable.marca, ACITable.modello).all()
            import logging
            logging.info(f"ACI Tables: Caricati {len(tables)} record con filtri applicati")
        else:
            # Se nessun filtro è selezionato ma form è stato inviato, mostra messaggio
            flash("⚠️ Seleziona almeno un filtro prima di cercare", "warning")
    
    return render_template("aci_tables.html", 
                         tables=tables, 
                         form=form, 
                         total_records=total_records,
                         show_results=(request.method == "POST"))

@aci_bp.route("/upload", methods=["GET", "POST"])
@login_required
@admin_required
def upload():
    """Upload e importazione file Excel ACI"""
    form = ACIUploadForm()
    
    if form.validate_on_submit():
        file = form.excel_file.data
        tipologia = form.tipologia.data
        
        try:
            # Salva file temporaneo
            import tempfile
            import os
            import pandas as pd
            
            with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
                file.save(tmp_file.name)
                
                # Leggi file Excel con ottimizzazioni per file grandi
                df = pd.read_excel(tmp_file.name, engine='openpyxl', 
                                 usecols=[0, 1, 2],  # Leggi solo colonne A, B, C
                                 dtype={0: 'str', 1: 'str', 2: 'float64'},  # Forza tipi di dato
                                 na_filter=True)
                
                # Pulisci file temporaneo
                os.unlink(tmp_file.name)
            
            # Verifica struttura file Excel - richiede almeno 3 colonne (A, B, C)
            if len(df.columns) < 3:
                flash(f"Errore: Il file Excel deve avere almeno 3 colonne (MARCA, MODELLO, COSTO KM). Trovate {len(df.columns)} colonne.", "danger")
                return render_template("aci_upload.html", form=form)
            
            # Log struttura per debug
            import logging
            logging.info(f"Colonne Excel trovate: {list(df.columns)}")
            logging.info(f"Utilizzo solo colonne A, B, C - ignorando tutte le altre")
            
            # Usa il nome del file come tipologia se non specificata esplicitamente
            if not tipologia.strip():
                # Estrae nome file senza estensione
                import os
                filename = file.filename or "Excel_File"
                tipologia = os.path.splitext(filename)[0]
            
            # Processa dati Excel - SOLO COLONNE A, B, C con ottimizzazione per file grandi
            imported_count = 0
            skipped_count = 0
            batch_size = 100  # Processa a lotti per evitare timeout
            
            # Converti DataFrame in lista per processamento più veloce
            rows_data = []
            for index, row in df.iterrows():
                # UTILIZZO SOLO LE PRIME 3 COLONNE (A, B, C)
                marca = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ""
                modello = str(row.iloc[1]).strip() if pd.notna(row.iloc[1]) else ""
                costo_km = row.iloc[2] if pd.notna(row.iloc[2]) else None
                
                # Skip righe vuote o incomplete
                if not marca or not modello or modello == 'nan' or costo_km is None:
                    skipped_count += 1
                    continue
                    
                rows_data.append((marca, modello, costo_km))
            
            logging.info(f"Processando {len(rows_data)} righe valide in batch da {batch_size}")
            
            # Processa in batch per evitare timeout
            for i in range(0, len(rows_data), batch_size):
                batch = rows_data[i:i + batch_size]
                
                for marca, modello, costo_km in batch:
                    # Verifica duplicati esistenti (con filtro company)
                    existing_query = filter_by_company(ACITable.query).filter_by(
                        tipologia=tipologia,
                        marca=marca,
                        modello=modello
                    )
                    existing = existing_query.first()
                    
                    if existing:
                        # Aggiorna record esistente - SOLO COSTO KM
                        existing.costo_km = float(costo_km)
                        imported_count += 1
                    else:
                        # Crea nuovo record ACI con company_id
                        aci_record = ACITable(
                            tipologia=tipologia,
                            marca=marca,
                            modello=modello,
                            costo_km=float(costo_km)
                        )
                        set_company_on_create(aci_record)
                        db.session.add(aci_record)
                        imported_count += 1
                
                # Commit ogni batch per liberare memoria
                try:
                    db.session.commit()
                    logging.info(f"Batch {i//batch_size + 1} completato: {len(batch)} record")
                except Exception as batch_error:
                    db.session.rollback()
                    logging.error(f"Errore batch {i//batch_size + 1}: {batch_error}")
                    raise batch_error
            
            # Messaggi di feedback dettagliati
            if imported_count > 0:
                flash(f"✅ File Excel importato con successo!", "success")
                flash(f"📊 {imported_count} record processati (nuovi o aggiornati)", "info")
                if skipped_count > 0:
                    flash(f"⏭️ {skipped_count} righe saltate (intestazioni o righe vuote)", "info")
                flash(f"🚫 Importate solo colonne A, B, C (MARCA, MODELLO, COSTO KM) - tutte le altre colonne ignorate", "info")
            else:
                flash("⚠️ Nessun record valido trovato nel file Excel", "warning")
                
            return redirect(url_for("aci.tables"))
            
        except Exception as e:
            db.session.rollback()
            flash(f"Errore durante l'importazione del file: {str(e)}", "danger")
    
    return render_template("aci_upload.html", form=form)

@aci_bp.route("/create", methods=["GET", "POST"])
@login_required
@admin_required
def create():
    """Crea nuovo record ACI manualmente"""
    form = ACIRecordForm()
    
    if form.validate_on_submit():
        try:
            aci_record = ACITable(
                tipologia=form.tipologia.data,
                marca=form.marca.data,
                modello=form.modello.data,
                costo_km=form.costo_km.data
            )
            set_company_on_create(aci_record)
            
            db.session.add(aci_record)
            db.session.commit()
            flash("Record ACI creato con successo!", "success")
            return redirect(url_for("aci.tables"))
            
        except Exception as e:
            db.session.rollback()
            flash(f"Errore durante la creazione del record: {str(e)}", "danger")
    
    return render_template("aci_create.html", form=form)

@aci_bp.route("/<int:record_id>/edit", methods=["GET", "POST"])
@login_required
@admin_required
def edit(record_id):
    """Modifica record ACI esistente"""
    aci_record = filter_by_company(ACITable.query).filter_by(id=record_id).first_or_404()
    form = ACIRecordForm(obj=aci_record)
    
    if form.validate_on_submit():
        try:
            aci_record.tipologia = form.tipologia.data
            aci_record.marca = form.marca.data
            aci_record.modello = form.modello.data
            aci_record.costo_km = form.costo_km.data
            
            db.session.commit()
            flash("Record ACI aggiornato con successo!", "success")
            return redirect(url_for("aci.tables"))
            
        except Exception as e:
            db.session.rollback()
            flash(f"Errore durante l'aggiornamento del record: {str(e)}", "danger")
    
    return render_template("aci_edit.html", form=form, record=aci_record)

@aci_bp.route("/<int:record_id>/delete", methods=["POST"])
@login_required
@admin_required
def delete(record_id):
    """Cancella record ACI"""
    try:
        aci_record = filter_by_company(ACITable.query).filter_by(id=record_id).first_or_404()
        db.session.delete(aci_record)
        db.session.commit()
        flash("Record ACI cancellato con successo!", "success")
        
    except Exception as e:
        db.session.rollback()
        flash(f"Errore durante la cancellazione: {str(e)}", "danger")
    
    return redirect(url_for("aci.tables"))

# =============================================================================
# ACI EXPORT ROUTES
# =============================================================================

@aci_bp.route("/export")
@login_required
@admin_required
def export():
    """Export Excel delle tabelle ACI"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from io import BytesIO
        
        # Recupera tutti i record ACI (filtrati per company)
        records = filter_by_company(ACITable.query).order_by(ACITable.tipologia, ACITable.tipo, ACITable.marca, ACITable.modello).all()
        
        # Crea workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Tabelle ACI"
        
        # Intestazioni
        headers = [
            "ID", "Tipologia", "Tipo", "Marca", "Modello", 
            "Costo KM", "Fringe Benefit 10%", "Fringe Benefit 25%", 
            "Fringe Benefit 30%", "Fringe Benefit 50%", 
            "Creato il", "Aggiornato il"
        ]
        
        # Stile intestazione
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Dati
        for row, record in enumerate(records, 2):
            ws.cell(row=row, column=1, value=record.id)
            ws.cell(row=row, column=2, value=record.tipologia)
            ws.cell(row=row, column=3, value=record.tipo)
            ws.cell(row=row, column=4, value=record.marca)
            ws.cell(row=row, column=5, value=record.modello)
            ws.cell(row=row, column=6, value=float(record.costo_km) if record.costo_km else None)
            ws.cell(row=row, column=7, value=float(record.fringe_benefit_10) if record.fringe_benefit_10 else None)
            ws.cell(row=row, column=8, value=float(record.fringe_benefit_25) if record.fringe_benefit_25 else None)
            ws.cell(row=row, column=9, value=float(record.fringe_benefit_30) if record.fringe_benefit_30 else None)
            ws.cell(row=row, column=10, value=float(record.fringe_benefit_50) if record.fringe_benefit_50 else None)
            ws.cell(row=row, column=11, value=record.created_at.strftime('%d/%m/%Y %H:%M') if record.created_at else "")
            ws.cell(row=row, column=12, value=record.updated_at.strftime('%d/%m/%Y %H:%M') if record.updated_at else "")
        
        # Autofit colonne
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Salva in BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Response
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename=tabelle_aci_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        return response
        
    except Exception as e:
        flash(f"Errore durante l'export: {str(e)}", "danger")
        return redirect(url_for("aci.tables"))

@aci_bp.route("/bulk_delete", methods=["POST"])
@login_required
@admin_required
def bulk_delete():
    """Cancellazione in massa per tipologia"""
    tipologia = request.form.get('tipologia')
    if not tipologia:
        flash("Tipologia non specificata.", "warning")
        return redirect(url_for("aci.tables"))
    
    try:
        # Prima conta i record da cancellare (con filtro company)
        records_to_delete = filter_by_company(ACITable.query).filter_by(tipologia=tipologia).count()
        
        if records_to_delete == 0:
            flash("Nessun record trovato per la tipologia specificata.", "info")
            return redirect(url_for("aci.tables"))
        
        # Esegui la cancellazione (con filtro company)
        deleted_count = filter_by_company(ACITable.query).filter_by(tipologia=tipologia).delete()
        db.session.commit()
        
        flash(f"✅ Cancellazione completata con successo!", "success")
        flash(f"📊 {deleted_count} record eliminati dalla tipologia '{tipologia}'", "info")
        
        import logging
        logging.info(f"Admin {current_user.username} ha cancellato {deleted_count} record ACI tipologia '{tipologia}'")
        
    except Exception as e:
        db.session.rollback()
        flash(f"❌ Errore durante la cancellazione in massa: {str(e)}", "danger")
        import logging
        logging.error(f"Errore cancellazione bulk ACI: {str(e)}")
    
    return redirect(url_for("aci.tables"))

# =============================================================================
# ACI API ROUTES
# =============================================================================

@aci_bp.route("/api/marcas")
@login_required
def api_marcas():
    """API per ottenere le marche filtrate per tipologia"""
    tipologia = request.args.get('tipologia')
    
    try:
        # Query con filtro company
        base_query = filter_by_company(ACITable.query)
        query = db.session.query(ACITable.marca).distinct().filter(ACITable.id.in_(
            db.session.query(ACITable.id).select_from(base_query.subquery())
        ))
        
        if tipologia:
            query = query.filter(ACITable.tipologia == tipologia)
        
        marcas = [row.marca for row in query.order_by(ACITable.marca).all()]
        return jsonify({'success': True, 'marcas': marcas})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@aci_bp.route("/api/modelos")
@login_required
def api_modelos():
    """API per ottenere i modelli filtrati per tipologia e marca con ID e costo km"""
    tipologia = request.args.get('tipologia')
    marca = request.args.get('marca')
    
    try:
        # Query con filtro company - ritorna record completi invece di solo i nomi
        query = filter_by_company(ACITable.query)
        
        if tipologia:
            query = query.filter(ACITable.tipologia == tipologia)
        if marca:
            query = query.filter(ACITable.marca == marca)
        
        # Ritorna record completi con id, modello e costo_km
        vehicles = query.order_by(ACITable.modello).all()
        modelos = [
            {
                'id': v.id,
                'modello': v.modello,
                'costo_km': float(v.costo_km) if v.costo_km else 0.0
            }
            for v in vehicles
        ]
        return jsonify({'success': True, 'modelos': modelos})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500