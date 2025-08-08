# =============================================================================
# EXPORT BLUEPRINT - Sistema esportazione dati
# =============================================================================
#
# ROUTES INCLUSE:
# 1. /shifts/export/excel (GET) - Export turni Excel
# 2. /shifts/export/pdf (GET) - Export turni PDF
# 3. /attendance/export/excel (GET) - Export presenze CSV
# 4. /leave/export/excel (GET) - Export ferie/permessi Excel
# 5. /expense/export/excel (GET) - Export note spese Excel
# 6. /interventions/general/export/excel (GET) - Export interventi generici Excel
# 7. /interventions/reperibilita/export/excel (GET) - Export interventi reperibilità Excel
#
# Total routes: 7 export routes
# =============================================================================

from flask import Blueprint, request, jsonify, render_template, redirect, url_for, flash, make_response
from flask_login import login_required, current_user
from datetime import datetime, date, timedelta
from functools import wraps
from app import db
from models import User, Shift, AttendanceEvent, LeaveRequest, ExpenseReport, ExpenseCategory, Intervention, ReperibilitaIntervention, italian_now
from io import BytesIO, StringIO
import tempfile
import os

# Create blueprint
export_bp = Blueprint('export', __name__, url_prefix='/export')

# =============================================================================
# SHIFTS EXPORT ROUTES
# =============================================================================

@export_bp.route('/shifts/excel')
@login_required
def shifts_excel():
    """Export turni in formato Excel"""
    # Parametri dalla query string
    view_mode = request.args.get('view', 'month')  # month, week, day
    show_my_shifts = request.args.get('my_shifts', 'false') == 'true'
    date_param = request.args.get('date', datetime.now().strftime('%Y-%m-%d'))
    
    try:
        current_date = datetime.strptime(date_param, '%Y-%m-%d').date()
    except:
        current_date = date.today()
    
    # Calcola range di date in base alla vista
    if view_mode == 'day':
        start_date = current_date
        end_date = current_date
        filename = f"turni_{current_date.strftime('%Y-%m-%d')}.xlsx"
    elif view_mode == 'week':
        # Settimana (Lunedì - Domenica)
        days_since_monday = current_date.weekday()
        start_date = current_date - timedelta(days=days_since_monday)
        end_date = start_date + timedelta(days=6)
        filename = f"turni_settimana_{start_date.strftime('%Y-%m-%d')}.xlsx"
    else:  # month
        start_date = current_date.replace(day=1)
        if current_date.month == 12:
            end_date = date(current_date.year + 1, 1, 1) - timedelta(days=1)
        else:
            end_date = date(current_date.year, current_date.month + 1, 1) - timedelta(days=1)
        filename = f"turni_{current_date.strftime('%Y-%m')}.xlsx"
    
    # Query dei turni
    shifts_query = Shift.query.filter(
        Shift.date >= start_date,
        Shift.date <= end_date
    )
    
    # Filtro per "I Miei Turni" se richiesto
    if show_my_shifts:
        shifts_query = shifts_query.filter(Shift.user_id == current_user.id)
        filename = f"miei_{filename}"
    
    shifts = shifts_query.order_by(Shift.date, Shift.start_time).all()
    
    # Crea Excel in memoria usando openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Turni"
    
    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Header
    headers = ['Data', 'Utente', 'Ruolo', 'Orario Inizio', 'Orario Fine', 'Tipo Turno', 'Durata (ore)']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    # Dati
    for row_idx, shift in enumerate(shifts, 2):
        duration = (datetime.combine(date.today(), shift.end_time) - 
                   datetime.combine(date.today(), shift.start_time)).total_seconds() / 3600
        
        row_data = [
            shift.date.strftime('%d/%m/%Y'),
            shift.user.get_full_name(),
            shift.user.role,
            shift.start_time.strftime('%H:%M'),
            shift.end_time.strftime('%H:%M'),
            shift.shift_type,
            f"{duration:.1f}"
        ]
        
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border
            if col in [1, 4, 5]:  # Data e Orari
                cell.alignment = Alignment(horizontal='center')
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to temporary file
    temp_dir = tempfile.mkdtemp()
    excel_path = os.path.join(temp_dir, filename)
    wb.save(excel_path)
    
    # Read file for response
    with open(excel_path, 'rb') as f:
        excel_data = f.read()
    
    # Cleanup
    os.remove(excel_path)
    os.rmdir(temp_dir)
    
    response = make_response(excel_data)
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response

@export_bp.route('/shifts/pdf')
@login_required  
def shifts_pdf():
    """Export calendario turni in formato PDF"""
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    
    # Parametri dalla query string
    view_mode = request.args.get('view', 'month')
    show_my_shifts = request.args.get('my_shifts', 'false') == 'true'
    date_param = request.args.get('date', datetime.now().strftime('%Y-%m-%d'))
    
    try:
        current_date = datetime.strptime(date_param, '%Y-%m-%d').date()
    except:
        current_date = date.today()
    
    # Calcola range di date in base alla vista
    if view_mode == 'day':
        start_date = current_date
        end_date = current_date
        filename = f"calendario_turni_{current_date.strftime('%Y-%m-%d')}.pdf"
        title_period = current_date.strftime('%d/%m/%Y')
    elif view_mode == 'week':
        days_since_monday = current_date.weekday()
        start_date = current_date - timedelta(days=days_since_monday)
        end_date = start_date + timedelta(days=6)
        filename = f"calendario_turni_settimana_{start_date.strftime('%Y-%m-%d')}.pdf"
        title_period = f"Settimana {start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')}"
    else:  # month
        start_date = current_date.replace(day=1)
        if current_date.month == 12:
            end_date = date(current_date.year + 1, 1, 1) - timedelta(days=1)
        else:
            end_date = date(current_date.year, current_date.month + 1, 1) - timedelta(days=1)
        filename = f"calendario_turni_{current_date.strftime('%Y-%m')}.pdf"
        title_period = current_date.strftime('%B %Y').title()
    
    # Query turni
    shifts_query = Shift.query.filter(
        Shift.date >= start_date,
        Shift.date <= end_date
    )
    
    if show_my_shifts:
        shifts_query = shifts_query.filter(Shift.user_id == current_user.id)
        filename = f"miei_{filename}"
    
    shifts = shifts_query.order_by(Shift.date, Shift.start_time).all()
    
    # Crea PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    story = []
    
    # Titolo
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=30,
        alignment=1
    )
    story.append(Paragraph(f"Calendario Turni - {title_period}", title_style))
    
    if show_my_shifts:
        story.append(Paragraph(f"Utente: {current_user.get_full_name()}", styles['Normal']))
        story.append(Spacer(1, 20))
    
    # Prepara dati per tabella
    table_data = [['Data', 'Utente', 'Orario', 'Durata', 'Tipo']]
    
    for shift in shifts:
        duration = (datetime.combine(date.today(), shift.end_time) - 
                   datetime.combine(date.today(), shift.start_time)).total_seconds() / 3600
        
        table_data.append([
            shift.date.strftime('%d/%m'),
            shift.user.get_full_name()[:20],  # Limita lunghezza nome
            f"{shift.start_time.strftime('%H:%M')}-{shift.end_time.strftime('%H:%M')}",
            f"{duration:.1f}h",
            shift.shift_type[:10]  # Limita lunghezza tipo
        ])
    
    if len(table_data) > 1:
        # Crea tabella
        table = Table(table_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(table)
    else:
        story.append(Paragraph("Nessun turno trovato per il periodo selezionato.", styles['Normal']))
    
    # Genera PDF
    doc.build(story)
    buffer.seek(0)
    pdf_data = buffer.getvalue()
    buffer.close()
    
    response = make_response(pdf_data)
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response

# =============================================================================
# ATTENDANCE EXPORT ROUTES
# =============================================================================

@export_bp.route('/attendance/excel')
@login_required  
def attendance_excel():
    """Export presenze in formato CSV"""
    from defusedcsv import csv
    
    # Controllo permessi
    if not current_user.can_access_attendance():
        flash('Non hai i permessi per esportare presenze.', 'danger')
        return redirect(url_for('dashboard.dashboard'))
    
    # Handle team/personal view toggle for PM
    view_mode = request.args.get('view', 'personal')
    if current_user.role == 'Management':
        show_team_data = (view_mode == 'team')
    else:
        show_team_data = False
    
    # Handle date filtering
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')
    
    if not start_date_str or not end_date_str:
        end_date = datetime.now().date()
        start_date = end_date - timedelta(days=30)
    else:
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        except ValueError:
            end_date = datetime.now().date()
            start_date = end_date - timedelta(days=30)
    
    if show_team_data:
        team_users = User.query.filter(
            User.role.in_(['Redattore', 'Sviluppatore', 'Operatore', 'Management', 'Responsabili']),
            User.active.is_(True)
        ).all()
        
        records = []
        for user in team_users:
            user_records = AttendanceEvent.get_events_as_records(user.id, start_date, end_date)
            records.extend(user_records)
    else:
        records = AttendanceEvent.get_events_as_records(current_user.id, start_date, end_date)
    
    records.sort(key=lambda x: x.date, reverse=True)
    
    # Create CSV
    output = StringIO()
    writer = csv.writer(output)
    
    if show_team_data:
        writer.writerow(['Data', 'Utente', 'Ruolo', 'Entrata', 'Pausa Inizio', 'Pausa Fine', 'Uscita', 'Ore Lavorate', 'Note'])
    else:
        writer.writerow(['Data', 'Entrata', 'Pausa Inizio', 'Pausa Fine', 'Uscita', 'Ore Lavorate', 'Note'])
    
    for record in records:
        row = [record.date.strftime('%d/%m/%Y')]
        
        if show_team_data and hasattr(record, 'user') and record.user:
            row.extend([record.user.get_full_name(), record.user.role])
        elif show_team_data:
            row.extend(['--', '--'])
        
        row.extend([
            record.clock_in.strftime('%H:%M') if record.clock_in else '--:--',
            record.break_start.strftime('%H:%M') if record.break_start else '--:--',
            record.break_end.strftime('%H:%M') if record.break_end else '--:--',
            record.clock_out.strftime('%H:%M') if record.clock_out else '--:--',
            f"{record.get_work_hours():.2f}" if record.clock_in and record.clock_out else '0.00',
            record.notes or ''
        ])
        
        writer.writerow(row)
    
    output.seek(0)
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'text/csv; charset=utf-8'
    
    filename = f"presenze_{'team' if show_team_data else 'personali'}_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.csv"
    response.headers['Content-Disposition'] = f'attachment; filename={filename}'
    
    return response

# =============================================================================
# LEAVE REQUESTS EXPORT ROUTES
# =============================================================================

@export_bp.route('/leave/excel')
@login_required
def leave_excel():
    """Export delle richieste di ferie/permessi in formato Excel"""
    if not current_user.can_view_leave_requests() and not current_user.can_request_leave():
        flash('Non hai i permessi per esportare le richieste', 'danger')
        return redirect(url_for('dashboard.dashboard'))
    
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    
    # Determina se l'utente può vedere tutte le richieste o solo le proprie
    can_approve = current_user.can_approve_leave()
    
    if can_approve:
        # Admin può vedere tutte le richieste
        requests = LeaveRequest.query.order_by(LeaveRequest.start_date.desc()).all()
        filename = f"richieste_ferie_permessi_{date.today().strftime('%Y%m%d')}.xlsx"
    else:
        # Utente normale vede solo le proprie
        requests = LeaveRequest.query.filter_by(user_id=current_user.id).order_by(LeaveRequest.start_date.desc()).all()
        filename = f"mie_richieste_ferie_permessi_{date.today().strftime('%Y%m%d')}.xlsx"
    
    # Crea il workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Richieste Ferie e Permessi"
    
    # Definisci gli header
    if can_approve:
        headers = ['Utente', 'Ruolo', 'Periodo', 'Durata', 'Tipo', 'Motivo', 'Stato', 'Data Richiesta', 'Approvato da', 'Data Approvazione']
    else:
        headers = ['Periodo', 'Durata', 'Tipo', 'Motivo', 'Stato', 'Data Richiesta', 'Approvato da', 'Data Approvazione']
    
    # Scrive gli header
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Scrive i dati
    for row_idx, request in enumerate(requests, 2):
        col = 1
        
        if can_approve:
            # Utente
            ws.cell(row=row_idx, column=col, value=request.user.get_full_name())
            col += 1
            
            # Ruolo
            ws.cell(row=row_idx, column=col, value=request.user.role)
            col += 1
        
        # Periodo
        if request.leave_type == 'Permesso' and request.is_time_based():
            periodo = f"{request.start_date.strftime('%d/%m/%Y')} {request.start_time.strftime('%H:%M')}-{request.end_time.strftime('%H:%M')}"
        elif request.start_date != request.end_date:
            periodo = f"{request.start_date.strftime('%d/%m/%Y')} - {request.end_date.strftime('%d/%m/%Y')}"
        else:
            periodo = request.start_date.strftime('%d/%m/%Y')
        ws.cell(row=row_idx, column=col, value=periodo)
        col += 1
        
        # Durata
        if request.leave_type == 'Permesso' and request.is_time_based():
            durata = f"{request.duration_hours}h"
        else:
            durata = f"{request.duration_days} giorni"
        ws.cell(row=row_idx, column=col, value=durata)
        col += 1
        
        # Tipo
        ws.cell(row=row_idx, column=col, value=request.leave_type)
        col += 1
        
        # Motivo
        ws.cell(row=row_idx, column=col, value=request.reason or '-')
        col += 1
        
        # Stato
        status_cell = ws.cell(row=row_idx, column=col, value=request.status)
        if request.status == 'Approved':
            status_cell.fill = PatternFill(start_color="D4F8D4", end_color="D4F8D4", fill_type="solid")
        elif request.status == 'Rejected':
            status_cell.fill = PatternFill(start_color="F8D4D4", end_color="F8D4D4", fill_type="solid")
        elif request.status == 'Pending':
            status_cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        col += 1
        
        # Data Richiesta
        ws.cell(row=row_idx, column=col, value=request.created_at.strftime('%d/%m/%Y %H:%M'))
        col += 1
        
        # Approvato da
        ws.cell(row=row_idx, column=col, value=request.approved_by.get_full_name() if request.approved_by else '-')
        col += 1
        
        # Data Approvazione
        ws.cell(row=row_idx, column=col, value=request.approved_at.strftime('%d/%m/%Y %H:%M') if request.approved_at else '-')
        col += 1
    
    # Ajusta la larghezza delle colonne
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    # Prepara la risposta
    response = make_response()
    
    # Salva in un buffer temporaneo
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    response.data = buffer.getvalue()
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename={filename}'
    
    return response

# =============================================================================
# EXPENSE REPORTS EXPORT ROUTES
# =============================================================================

@export_bp.route('/expense/excel')
@login_required
def expense_excel():
    """Export delle note spese in formato Excel"""
    if not current_user.can_view_expense_reports() and not current_user.can_create_expense_reports():
        flash('Non hai i permessi per esportare le note spese', 'danger')
        return redirect(url_for('dashboard.dashboard'))
    
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from sqlalchemy.orm import joinedload
    
    # Determina se l'utente può vedere tutte le note spese o solo le proprie
    can_manage = current_user.can_view_expense_reports()
    
    if can_manage:
        # Manager può vedere tutte le note spese
        reports = ExpenseReport.query.options(
            joinedload(ExpenseReport.user),
            joinedload(ExpenseReport.category),
            joinedload(ExpenseReport.approved_by)
        ).order_by(ExpenseReport.expense_date.desc()).all()
        filename = f"note_spese_tutte_{date.today().strftime('%Y%m%d')}.xlsx"
    else:
        # Utente normale vede solo le proprie
        reports = ExpenseReport.query.filter_by(user_id=current_user.id).options(
            joinedload(ExpenseReport.category),
            joinedload(ExpenseReport.approved_by)
        ).order_by(ExpenseReport.expense_date.desc()).all()
        filename = f"mie_note_spese_{date.today().strftime('%Y%m%d')}.xlsx"
    
    # Crea il workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Note Spese"
    
    # Definisci gli header
    if can_manage:
        headers = ['Utente', 'Data Spesa', 'Categoria', 'Descrizione', 'Importo', 'Stato', 'Data Richiesta', 'Approvato da', 'Data Approvazione']
    else:
        headers = ['Data Spesa', 'Categoria', 'Descrizione', 'Importo', 'Stato', 'Data Richiesta', 'Approvato da', 'Data Approvazione']
    
    # Scrive gli header
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Scrive i dati
    for row_idx, report in enumerate(reports, 2):
        col = 1
        
        if can_manage:
            # Utente
            ws.cell(row=row_idx, column=col, value=report.user.get_full_name())
            col += 1
        
        # Data Spesa
        ws.cell(row=row_idx, column=col, value=report.expense_date.strftime('%d/%m/%Y'))
        col += 1
        
        # Categoria
        ws.cell(row=row_idx, column=col, value=report.category.name if report.category else '-')
        col += 1
        
        # Descrizione
        ws.cell(row=row_idx, column=col, value=report.description or '-')
        col += 1
        
        # Importo
        ws.cell(row=row_idx, column=col, value=f"€ {report.amount:.2f}")
        col += 1
        
        # Stato
        status_cell = ws.cell(row=row_idx, column=col, value=report.status)
        if report.status == 'Approved':
            status_cell.fill = PatternFill(start_color="D4F8D4", end_color="D4F8D4", fill_type="solid")
        elif report.status == 'Rejected':
            status_cell.fill = PatternFill(start_color="F8D4D4", end_color="F8D4D4", fill_type="solid")
        elif report.status == 'Pending':
            status_cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        col += 1
        
        # Data Richiesta
        ws.cell(row=row_idx, column=col, value=report.created_at.strftime('%d/%m/%Y %H:%M'))
        col += 1
        
        # Approvato da
        ws.cell(row=row_idx, column=col, value=report.approved_by.get_full_name() if report.approved_by else '-')
        col += 1
        
        # Data Approvazione
        ws.cell(row=row_idx, column=col, value=report.approved_at.strftime('%d/%m/%Y %H:%M') if report.approved_at else '-')
        col += 1
    
    # Ajusta la larghezza delle colonne
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    # Prepara la risposta
    response = make_response()
    
    # Salva in un buffer temporaneo
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    response.data = buffer.getvalue()
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename={filename}'
    
    return response

# =============================================================================
# INTERVENTIONS EXPORT ROUTES
# =============================================================================

@export_bp.route('/interventions/general/excel')
@login_required
def general_interventions_excel():
    """Export interventi generici in formato Excel"""
    if current_user.role == 'Admin':
        flash('Accesso non autorizzato.', 'danger')
        return redirect(url_for('dashboard.dashboard'))
    
    from zoneinfo import ZoneInfo
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    
    italy_tz = ZoneInfo('Europe/Rome')
    today = datetime.now(italy_tz).date()
    
    # Gestisci filtri data
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')
    
    if start_date_str and end_date_str:
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        except ValueError:
            start_date = today.replace(day=1)
            end_date = today
    else:
        start_date = today.replace(day=1)
        end_date = today
    
    # Query degli interventi nel periodo
    interventions = Intervention.query.filter(
        Intervention.start_time >= datetime.combine(start_date, datetime.min.time()),
        Intervention.start_time <= datetime.combine(end_date, datetime.max.time())
    ).order_by(Intervention.start_time.desc()).all()
    
    # Filtra per utente se non Manager/Admin
    if current_user.role not in ['Management', 'Admin']:
        interventions = [i for i in interventions if i.user_id == current_user.id]
    
    # Crea Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Interventi Generici"
    
    # Headers
    headers = ['Data', 'Utente', 'Inizio', 'Fine', 'Durata (min)', 'Descrizione']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    
    # Dati
    for row_idx, intervention in enumerate(interventions, 2):
        # Converti timestamps in orario italiano
        start_time_italy = intervention.start_time
        if start_time_italy.tzinfo is None:
            start_time_italy = start_time_italy.replace(tzinfo=ZoneInfo('UTC')).astimezone(italy_tz)
        else:
            start_time_italy = start_time_italy.astimezone(italy_tz)
        
        end_time_italy = None
        duration_minutes = 0
        if intervention.end_time:
            end_time_italy = intervention.end_time
            if end_time_italy.tzinfo is None:
                end_time_italy = end_time_italy.replace(tzinfo=ZoneInfo('UTC')).astimezone(italy_tz)
            else:
                end_time_italy = end_time_italy.astimezone(italy_tz)
            
            duration_minutes = int((end_time_italy - start_time_italy).total_seconds() / 60)
        
        row_data = [
            start_time_italy.strftime('%d/%m/%Y'),
            intervention.user.get_full_name() if intervention.user else 'N/A',
            start_time_italy.strftime('%H:%M'),
            end_time_italy.strftime('%H:%M') if end_time_italy else 'In corso',
            str(duration_minutes) if duration_minutes > 0 else '-',
            intervention.description or '-'
        ]
        
        for col, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col, value=value)
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Prepara la risposta
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    response = make_response(buffer.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    
    filename = f"interventi_generici_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx"
    response.headers['Content-Disposition'] = f'attachment; filename={filename}'
    
    return response

@export_bp.route('/interventions/reperibilita/excel')
@login_required
def reperibilita_interventions_excel():
    """Export interventi reperibilità in formato Excel"""
    if current_user.role == 'Admin':
        flash('Accesso non autorizzato.', 'danger')
        return redirect(url_for('dashboard.dashboard'))
    
    from zoneinfo import ZoneInfo
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    
    italy_tz = ZoneInfo('Europe/Rome')
    today = datetime.now(italy_tz).date()
    
    # Gestisci filtri data
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')
    
    if start_date_str and end_date_str:
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        except ValueError:
            start_date = today.replace(day=1)
            end_date = today
    else:
        start_date = today.replace(day=1)
        end_date = today
    
    # Query degli interventi di reperibilità nel periodo
    interventions = ReperibilitaIntervention.query.filter(
        ReperibilitaIntervention.start_time >= datetime.combine(start_date, datetime.min.time()),
        ReperibilitaIntervention.start_time <= datetime.combine(end_date, datetime.max.time())
    ).order_by(ReperibilitaIntervention.start_time.desc()).all()
    
    # Filtra per utente se non Manager/Admin
    if current_user.role not in ['Management', 'Admin']:
        interventions = [i for i in interventions if i.user_id == current_user.id]
    
    # Crea Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Interventi Reperibilità"
    
    # Headers
    headers = ['Data', 'Utente', 'Inizio', 'Fine', 'Durata (min)', 'Descrizione', 'Tipo']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    
    # Dati
    for row_idx, intervention in enumerate(interventions, 2):
        # Converti timestamps in orario italiano
        start_time_italy = intervention.start_time
        if start_time_italy.tzinfo is None:
            start_time_italy = start_time_italy.replace(tzinfo=ZoneInfo('UTC')).astimezone(italy_tz)
        else:
            start_time_italy = start_time_italy.astimezone(italy_tz)
        
        end_time_italy = None
        duration_minutes = 0
        if intervention.end_time:
            end_time_italy = intervention.end_time
            if end_time_italy.tzinfo is None:
                end_time_italy = end_time_italy.replace(tzinfo=ZoneInfo('UTC')).astimezone(italy_tz)
            else:
                end_time_italy = end_time_italy.astimezone(italy_tz)
            
            duration_minutes = int((end_time_italy - start_time_italy).total_seconds() / 60)
        
        row_data = [
            start_time_italy.strftime('%d/%m/%Y'),
            intervention.user.get_full_name() if intervention.user else 'N/A',
            start_time_italy.strftime('%H:%M'),
            end_time_italy.strftime('%H:%M') if end_time_italy else 'In corso',
            str(duration_minutes) if duration_minutes > 0 else '-',
            intervention.description or '-',
            getattr(intervention, 'intervention_type', 'Reperibilità')
        ]
        
        for col, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col, value=value)
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Prepara la risposta
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    response = make_response(buffer.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    
    filename = f"interventi_reperibilita_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx"
    response.headers['Content-Disposition'] = f'attachment; filename={filename}'
    
    return response