# =============================================================================
# ATTENDANCE BLUEPRINT - Modulo gestione presenze e timbrature
# =============================================================================
#
# ROUTES INCLUSE:
# 1. check_shift_before_clock_in (POST) - Controllo pre-entrata
# 2. clock_in (POST) - Timbratura entrata  
# 3. check_shift_before_clock_out (POST) - Controllo pre-uscita
# 4. clock_out (POST) - Timbratura uscita
# 5. break_start (POST) - Inizio pausa
# 6. break_end (POST) - Fine pausa
# 7. attendance (GET/POST) - Pagina principale presenze
# 8. export_attendance_excel (GET) - Export dati presenze
# 9. api/work_hours/<user_id>/<date> (GET) - API ore lavorate
# 10. quick_attendance/<action> (GET/POST) - Timbratura rapida QR
#
# Total routes: 10
# =============================================================================

from flask import Blueprint, request, jsonify, render_template, redirect, url_for, flash, make_response
from flask_login import login_required, current_user
from datetime import datetime, date, timedelta, time
from functools import wraps
from app import db
from models import User, AttendanceEvent, Shift, Sede, ReperibilitaShift, Intervention, LeaveRequest, WorkSchedule, MonthlyTimesheet, AttendanceType, italian_now
from utils_tenant import get_user_company_id, filter_by_company, set_company_on_create
from io import StringIO
from defusedcsv import csv
from forms import AttendanceForm
import io

# Create blueprint
attendance_bp = Blueprint('attendance', __name__, url_prefix='/attendance')

# Helper functions
def get_current_user_sede(user):
    """Get current user's sede - copy from routes.py"""
    if hasattr(user, 'sedi') and user.sedi:
        return user.sedi[0]  # Return first sede
    elif hasattr(user, 'sede_obj') and user.sede_obj:
        return user.sede_obj  # Return direct sede relationship
    elif hasattr(user, 'sede_id') and user.sede_id:
        # Fallback: get sede by ID
        from models import Sede
        return Sede.query.get(user.sede_id)
    return None

def require_login(f):
    """Decorator to require login for routes - copy from routes.py"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated:
            return redirect(url_for('auth.login'))
        return f(*args, **kwargs)
    return decorated_function

# =============================================================================
# MAIN ATTENDANCE PAGE ROUTE
# =============================================================================

@attendance_bp.route('/', methods=['GET', 'POST'])
@login_required
def attendance():
    """Main attendance page - migrated from routes.py"""
    # Controllo permessi di accesso alle presenze
    view_mode = request.args.get('view', 'personal')
    
    # Controllo specifico per vista sede
    if view_mode == 'sede' and not current_user.can_view_sede_attendance():
        flash('Non hai i permessi per visualizzare le presenze della sede.', 'danger')
        return redirect(url_for('dashboard.dashboard'))
    
    # Controllo generale per accesso alle presenze
    if not current_user.can_access_attendance():
        flash('Non hai i permessi per accedere alla gestione presenze.', 'danger')
        return redirect(url_for('dashboard.dashboard'))
    
    form = AttendanceForm(user=current_user)
    
    # Ottieni stato attuale dell'utente e eventi di oggi (solo se non è Ente o Staff)
    if current_user.role not in ['Ente', 'Staff']:
        user_status, last_event = AttendanceEvent.get_user_status(current_user.id)
        today_events = AttendanceEvent.get_daily_events(current_user.id)
        today_work_hours = AttendanceEvent.get_daily_work_hours(current_user.id)
    else:
        # Ente e Staff non hanno dati personali di presenza
        user_status, last_event = 'out', None
        today_events = []
        today_work_hours = 0
    
    # Blocca POST per utenti Ente e Staff (solo visualizzazione)
    if request.method == 'POST' and form.validate_on_submit() and current_user.role not in ['Ente', 'Staff']:
        # Salva note nell'ultimo evento di oggi o crea un nuovo evento note
        if form.notes.data:
            if today_events and last_event:
                # Aggiorna note dell'ultimo evento
                last_event.notes = form.notes.data
                db.session.commit()
                flash('Note salvate', 'success')
            else:
                # Crea evento note se non ci sono eventi oggi
                from zoneinfo import ZoneInfo
                italy_tz = ZoneInfo('Europe/Rome')
                now = datetime.now(italy_tz)
                
                # Determina sede_id
                sede_id = None
                if current_user.all_sedi and form.sede_id.data:
                    sede_id = form.sede_id.data
                elif current_user.sede_id:
                    sede_id = current_user.sede_id
                
                note_event = AttendanceEvent(
                    user_id=current_user.id,
                    date=now.date(),
                    event_type='clock_in',  # Evento fittizio per salvare le note
                    timestamp=now,
                    sede_id=sede_id,
                    notes=form.notes.data
                )
                # Imposta company_id automaticamente
                set_company_on_create(note_event)
                db.session.add(note_event)
                db.session.commit()
                flash('Note salvate', 'success')
        return redirect(url_for('attendance.attendance'))
    
    # Handle team/personal view toggle for PM, Management, Responsabili, Ente and Staff
    view_mode = request.args.get('view', 'personal')
    if current_user.role in ['Management'] and current_user.can_view_all_attendance():
        # Management può scegliere vista personale o team
        show_team_data = (view_mode == 'team')
    elif current_user.role in ['Ente', 'Staff']:
        # Ente e Staff vedono sempre e solo dati team
        show_team_data = True
        view_mode = 'team'
    elif current_user.role in ['Amministratore']:
        # Amministratore vede sempre dati team con sede
        show_team_data = True
        view_mode = 'sede'
    elif current_user.can_view_sede_attendance() and view_mode == 'sede':
        # Utenti con permesso "Visualizzare Presenze Sede" possono vedere presenze della propria sede
        show_team_data = True
        view_mode = 'sede'
    else:
        # Altri utenti vedono solo dati personali
        show_team_data = False
        view_mode = 'personal'
    
    # Handle date filtering
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')
    
    # Default to current month (from first to last day) if no dates provided
    if not start_date_str or not end_date_str:
        from calendar import monthrange
        today = datetime.now().date()
        # Primo giorno del mese corrente
        start_date = date(today.year, today.month, 1)
        # Ultimo giorno del mese corrente
        last_day_num = monthrange(today.year, today.month)[1]
        end_date = date(today.year, today.month, last_day_num)
    else:
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        except ValueError:
            # Fallback to current month if invalid dates
            from calendar import monthrange
            today = datetime.now().date()
            start_date = date(today.year, today.month, 1)
            last_day_num = monthrange(today.year, today.month)[1]
            end_date = date(today.year, today.month, last_day_num)
    
    if show_team_data:
        # Get team attendance data for PM, Management, Responsabili and Ente
        if current_user.role == 'Staff':
            # Staff vede tutti gli utenti di tutte le sedi (esclusi Admin e Staff), filtrati per company
            team_users = filter_by_company(User.query).filter(
                User.active.is_(True),
                ~User.role.in_(['Admin', 'Staff'])
            ).all()
        elif current_user.role == 'Management':
            # Management vedono solo utenti della propria sede (esclusi Admin e Staff), filtrati per company
            team_users = filter_by_company(User.query).filter(
                User.sede_id == current_user.sede_id,
                User.active.is_(True),
                ~User.role.in_(['Admin', 'Staff'])
            ).all()
        elif view_mode == 'sede' or current_user.role == 'Amministratore':
            # Utenti con permessi visualizzazione sede o amministratori
            if current_user.all_sedi or current_user.role == 'Amministratore':
                # Utenti multi-sede e amministratori vedono tutti gli utenti attivi di tutte le sedi, filtrati per company
                team_users = filter_by_company(User.query).filter(
                    User.active.is_(True),
                    ~User.role.in_(['Admin', 'Staff'])
                ).all()
            elif current_user.sede_id:
                # Utenti sede-specifica vedono solo utenti della propria sede, filtrati per company
                team_users = filter_by_company(User.query).filter(
                    User.sede_id == current_user.sede_id,
                    User.active.is_(True),
                    ~User.role.in_(['Admin', 'Staff'])
                ).all()
            else:
                team_users = []
        else:
            # PM e Ente vedono solo utenti operativi (esclusi Admin e Staff), filtrati per company
            team_users = filter_by_company(User.query).filter(
                User.role.in_(['Redattore', 'Sviluppatore', 'Operatore', 'Management', 'Staff']),
                User.active.is_(True)
            ).all()
        
        old_records = []
        event_records = []
        user_shifts = []
        
        for user in team_users:
            # Get event records for this user
            user_event_records = AttendanceEvent.get_events_as_records(user.id, start_date, end_date)
            event_records.extend(user_event_records)
            
            # Get shifts for this user
            user_user_shifts = Shift.query.filter(
                Shift.user_id == user.id,
                Shift.date >= start_date,
                Shift.date <= end_date
            ).all()
            user_shifts.extend(user_user_shifts)
    else:
        # Ottieni tutti gli eventi come record individuali
        event_records = AttendanceEvent.get_events_as_records(current_user.id, start_date, end_date)
        old_records = []  # Non più necessario
        
        # Get user shifts for shift comparison
        user_shifts = Shift.query.filter(
            Shift.user_id == current_user.id,
            Shift.date >= start_date,
            Shift.date <= end_date
        ).all()
    
    # Create a shift lookup by date and user
    if show_team_data:
        # Per i dati team, crea lookup per utente+data
        shifts_by_user_date = {}
        for shift in user_shifts:
            key = (shift.user_id, shift.date)
            shifts_by_user_date[key] = shift
    else:
        # Per i dati personali, usa la lookup per data
        shifts_by_date = {shift.date: shift for shift in user_shifts}
    
    # Add shift status to event records (both entry and exit indicators)
    for record in event_records:
        # Trova il turno corrispondente
        shift = None
        if show_team_data:
            # Per dati team, cerca per utente+data
            key = (record.user_id, record.date)
            shift = shifts_by_user_date.get(key)
        else:
            # Per dati personali, cerca per data
            shift = shifts_by_date.get(record.date)
        
        # Inizializza indicatori
        record.shift_status = 'normale'
        record.exit_status = 'normale'
        
        if shift:
            from zoneinfo import ZoneInfo
            italy_tz = ZoneInfo('Europe/Rome')
            utc_tz = ZoneInfo('UTC')
            
            # Calcola indicatori di ENTRATA solo se l'utente ha un orario che richiede controlli
            if hasattr(record, 'clock_in') and record.clock_in and record.user.should_check_attendance_timing():
                # Recupera il WorkSchedule della sede per controllare la flessibilità
                work_schedule = None
                if record.user.sede_id:
                    work_schedule = filter_by_company(WorkSchedule.query).filter(
                        WorkSchedule.sede_id == record.user.sede_id,
                        WorkSchedule.active == True
                    ).first()
                
                # Converti il timestamp di clock_in da UTC a orario italiano
                clock_in_time = record.clock_in
                if clock_in_time.tzinfo is None:
                    clock_in_time = clock_in_time.replace(tzinfo=utc_tz)
                clock_in_time_italy = clock_in_time.astimezone(italy_tz)
                
                # Se c'è un WorkSchedule con flessibilità, usa quello
                if work_schedule and work_schedule.start_time_min and work_schedule.start_time_max:
                    # Usa gli orari flessibili
                    start_min_datetime = datetime.combine(record.date, work_schedule.start_time_min).replace(tzinfo=italy_tz)
                    start_max_datetime = datetime.combine(record.date, work_schedule.start_time_max).replace(tzinfo=italy_tz)
                    
                    # Anticipo: prima di start_min - 30min
                    early_limit = start_min_datetime - timedelta(minutes=30)
                    # Ritardo: dopo start_max + 15min
                    late_limit = start_max_datetime + timedelta(minutes=15)
                    
                    if clock_in_time_italy < early_limit:
                        record.shift_status = 'anticipo'
                    elif clock_in_time_italy > late_limit:
                        record.shift_status = 'ritardo'
                    else:
                        record.shift_status = 'normale'
                else:
                    # Fallback: usa l'orario fisso del turno
                    shift_start_datetime = datetime.combine(record.date, shift.start_time)
                    shift_start_datetime = shift_start_datetime.replace(tzinfo=italy_tz)
                    early_limit = shift_start_datetime - timedelta(minutes=30)
                    late_limit = shift_start_datetime + timedelta(minutes=15)
                    
                    if clock_in_time_italy < early_limit:
                        record.shift_status = 'anticipo'
                    elif clock_in_time_italy > late_limit:
                        record.shift_status = 'ritardo'
                    else:
                        record.shift_status = 'normale'
            
            # Calcola indicatori di USCITA solo se l'utente ha un orario che richiede controlli
            if hasattr(record, 'clock_out') and record.clock_out and record.user.should_check_attendance_timing():
                # Recupera il WorkSchedule della sede per controllare la flessibilità
                work_schedule = None
                if record.user.sede_id:
                    work_schedule = filter_by_company(WorkSchedule.query).filter(
                        WorkSchedule.sede_id == record.user.sede_id,
                        WorkSchedule.active == True
                    ).first()
                
                # Converti il timestamp di clock_out da UTC a orario italiano
                clock_out_time = record.clock_out
                if clock_out_time.tzinfo is None:
                    clock_out_time = clock_out_time.replace(tzinfo=utc_tz)
                clock_out_time_italy = clock_out_time.astimezone(italy_tz)
                
                # Se c'è un WorkSchedule con flessibilità, usa quello
                if work_schedule and work_schedule.end_time_min and work_schedule.end_time_max:
                    # Usa gli orari flessibili
                    end_min_datetime = datetime.combine(record.date, work_schedule.end_time_min).replace(tzinfo=italy_tz)
                    end_max_datetime = datetime.combine(record.date, work_schedule.end_time_max).replace(tzinfo=italy_tz)
                    
                    # Uscita anticipata: prima di end_min - 5min
                    early_exit_limit = end_min_datetime - timedelta(minutes=5)
                    # Straordinario: dopo end_max + 10min
                    late_exit_limit = end_max_datetime + timedelta(minutes=10)
                    
                    if clock_out_time_italy < early_exit_limit:
                        record.exit_status = 'anticipo'
                    elif clock_out_time_italy > late_exit_limit:
                        record.exit_status = 'straordinario'
                    else:
                        record.exit_status = 'normale'
                else:
                    # Fallback: usa l'orario fisso del turno
                    shift_end_datetime = datetime.combine(record.date, shift.end_time)
                    shift_end_datetime = shift_end_datetime.replace(tzinfo=italy_tz)
                    early_exit_limit = shift_end_datetime - timedelta(minutes=5)
                    late_exit_limit = shift_end_datetime + timedelta(minutes=10)
                    
                    if clock_out_time_italy < early_exit_limit:
                        record.exit_status = 'anticipo'
                    elif clock_out_time_italy > late_exit_limit:
                        record.exit_status = 'straordinario'
                    else:
                        record.exit_status = 'normale'
    
    # Non ci sono più turni da controllare - rimosso controllo assenze basato su turni
    
    # Aggiungi record per le giornate con ferie/permessi/malattie approvate
    leave_records = []
    
    # Determina gli utenti per cui cercare le richieste di ferie
    if show_team_data:
        # Per vista team, cerca ferie di tutti gli utenti del team
        target_user_ids = [user.id for user in team_users]
    else:
        # Per vista personale, solo l'utente corrente
        target_user_ids = [current_user.id]
    
    if target_user_ids:
        # Cerca richieste di ferie approvate E in attesa nel periodo per tutti gli utenti target
        # Include sia Approved che Pending per mostrare all'utente tutte le assenze
        approved_leaves = LeaveRequest.query.filter(
            LeaveRequest.user_id.in_(target_user_ids),
            LeaveRequest.status.in_(['Approved', 'Pending']),  # Mostra sia validate che in attesa
            LeaveRequest.start_date <= end_date,
            LeaveRequest.end_date >= start_date
        ).all()
        
        for leave in approved_leaves:
            current_date = max(leave.start_date, start_date)
            while current_date <= min(leave.end_date, end_date):
                # Verifica se esiste già un record di presenza per questa data e utente
                existing_record = any(
                    r.date == current_date and getattr(r, 'user_id', None) == leave.user_id 
                    for r in event_records + old_records
                )
                
                if not existing_record:
                    # Crea un record per la giornata di ferie/permesso/malattia
                    class LeaveRecord:
                        def __init__(self, date, leave_type, reason, user_id, leave_request):
                            self.date = date
                            self.break_start = None
                            self.break_end = None
                            self.notes = f"{leave_type}: {reason}" if reason else leave_type
                            self.user_id = user_id
                            self.user = User.query.get(user_id)
                            self.shift_status = leave_type.lower() if leave_type else 'permesso'  # 'ferie', 'permesso', 'malattia'
                            self.exit_status = 'normale'
                            self.leave_type = leave_type
                            self.leave_reason = reason
                            # Aggiungi lo status della richiesta per distinguere approvate da pending
                            self.request_status = leave_request.status if hasattr(leave_request, 'status') else 'Approved'
                            self.is_pending = (self.request_status == 'Pending')
                            
                            # Determina orari in base al tipo di assenza
                            if (leave_type and leave_type.lower() == 'permesso') and hasattr(leave_request, 'start_time') and leave_request.start_time:
                                # Per i permessi usa gli orari specifici della richiesta - converti in datetime italiano
                                from datetime import datetime
                                from zoneinfo import ZoneInfo
                                
                                if isinstance(leave_request.start_time, datetime):
                                    # Se è già datetime, assicurati che sia in fuso italiano
                                    if leave_request.start_time.tzinfo is None:
                                        # Assumi UTC se non ha timezone
                                        utc_time = leave_request.start_time.replace(tzinfo=ZoneInfo('UTC'))
                                        self.clock_in = utc_time.astimezone(ZoneInfo('Europe/Rome'))
                                    else:
                                        self.clock_in = leave_request.start_time.astimezone(ZoneInfo('Europe/Rome'))
                                else:
                                    # Se è time, combinalo con la data corrente in fuso italiano
                                    self.clock_in = datetime.combine(self.date, leave_request.start_time, ZoneInfo('Europe/Rome'))
                                
                                if hasattr(leave_request, 'end_time') and leave_request.end_time:
                                    if isinstance(leave_request.end_time, datetime):
                                        if leave_request.end_time.tzinfo is None:
                                            utc_time = leave_request.end_time.replace(tzinfo=ZoneInfo('UTC'))
                                            self.clock_out = utc_time.astimezone(ZoneInfo('Europe/Rome'))
                                        else:
                                            self.clock_out = leave_request.end_time.astimezone(ZoneInfo('Europe/Rome'))
                                    else:
                                        self.clock_out = datetime.combine(self.date, leave_request.end_time, ZoneInfo('Europe/Rome'))
                                else:
                                    self.clock_out = None
                            else:
                                # Per ferie e malattie usa orari standard di lavoro dell'utente
                                from models import WorkSchedule
                                from datetime import datetime, time
                                user_schedule = None
                                if self.user.work_schedule_id:
                                    user_schedule = WorkSchedule.query.get(self.user.work_schedule_id)
                                
                                if user_schedule:
                                    # Usa orari standard del turno - converti time in datetime italiano
                                    from zoneinfo import ZoneInfo
                                    self.clock_in = datetime.combine(self.date, user_schedule.start_time_min, ZoneInfo('Europe/Rome'))
                                    self.clock_out = datetime.combine(self.date, user_schedule.end_time_max, ZoneInfo('Europe/Rome'))
                                else:
                                    # Fallback a orari generici 9-17 - converti in datetime italiano
                                    from zoneinfo import ZoneInfo
                                    self.clock_in = datetime.combine(self.date, time(9, 0), ZoneInfo('Europe/Rome'))
                                    self.clock_out = datetime.combine(self.date, time(17, 0), ZoneInfo('Europe/Rome'))
                        
                        def get_work_hours(self):
                            return 0  # Nessuna ora lavorata durante ferie/permessi
                        
                        def get_attendance_indicators(self):
                            return {'entry': None, 'exit': None}
                    
                    # Usa la relazione leave_type_obj invece del campo testuale legacy
                    leave_type_name = leave.leave_type_obj.name if leave.leave_type_obj else (leave.leave_type or 'Permesso')
                    leave_records.append(LeaveRecord(
                        date=current_date,
                        leave_type=leave_type_name,
                        reason=leave.reason,
                        user_id=leave.user_id,
                        leave_request=leave
                    ))
                
                current_date += timedelta(days=1)
    
    # Combina tutti i record
    records = []
    records.extend(event_records)
    records.extend(old_records)
    records.extend(leave_records)
    
    # Riordina per data e timestamp decrescente
    def sort_key(record):
        if hasattr(record, 'timestamp') and record.timestamp:
            return (record.date, record.timestamp)
        elif hasattr(record, 'created_at') and record.created_at:
            return (record.date, record.created_at)
        else:
            return (record.date, datetime.min)
    
    records.sort(key=sort_key, reverse=True)
    
    # Organizza i record per sede per utenti multi-sede in modalità team
    records_by_sede = {}
    all_sedi_list = []
    if show_team_data and current_user.all_sedi and view_mode == 'sede':
        from collections import defaultdict
        from models import Sede
        
        # Ottieni tutte le sedi attive
        all_sedi_list = Sede.query.filter_by(active=True).order_by(Sede.name).all()
        records_by_sede = {sede.name: [] for sede in all_sedi_list}
        
        # Aggiungi anche una categoria per utenti senza sede
        records_by_sede['Nessuna Sede'] = []
        
        # Distribuisci i record per sede
        for record in records:
            if hasattr(record, 'user') and record.user:
                sede_name = record.user.sede_obj.name if record.user.sede_obj else 'Nessuna Sede'
                if sede_name in records_by_sede:
                    records_by_sede[sede_name].append(record)
        
        # Rimuovi "Nessuna Sede" se vuota
        if not records_by_sede['Nessuna Sede']:
            del records_by_sede['Nessuna Sede']
    
    # Genera tutti i giorni del periodo per vista personale (Le Mie Presenze)
    all_days = []
    if not show_team_data:
        # Ottieni work schedule dell'utente per determinare se sabati/domeniche sono abilitati
        user_work_schedule = current_user.work_schedule
        works_on_saturday = False
        works_on_sunday = False
        
        if user_work_schedule and user_work_schedule.days_of_week:
            # days_of_week è un array JSON: [0,1,2,3,4] = Lun-Ven, 5=Sabato, 6=Domenica
            works_on_saturday = 5 in user_work_schedule.days_of_week
            works_on_sunday = 6 in user_work_schedule.days_of_week
        
        # Crea dict di lookup per records esistenti (per data)
        records_by_date = {}
        for record in records:
            record_date = record.date
            if record_date not in records_by_date:
                records_by_date[record_date] = []
            records_by_date[record_date].append(record)
        
        # Genera tutti i giorni del periodo
        current_day = start_date
        while current_day <= end_date:
            day_of_week = current_day.weekday()  # 0=Monday, 6=Sunday
            is_saturday = (day_of_week == 5)
            is_sunday = (day_of_week == 6)
            is_weekend = is_saturday or is_sunday
            
            # Determina se il giorno è lavorativo per l'utente
            day_enabled = True
            if is_saturday and not works_on_saturday:
                day_enabled = False
            elif is_sunday and not works_on_sunday:
                day_enabled = False
            
            # Ottieni i record per questo giorno (se esistono)
            day_records = records_by_date.get(current_day, [])
            
            # Crea oggetto per il giorno
            day_info = {
                'date': current_day,
                'is_weekend': is_weekend,
                'is_saturday': is_saturday,
                'is_sunday': is_sunday,
                'day_enabled': day_enabled,
                'records': day_records,
                'has_records': len(day_records) > 0
            }
            
            all_days.append(day_info)
            current_day += timedelta(days=1)
    
    # Verifica stato consolidamento del timesheet
    from models import MonthlyTimesheet, TimesheetReopenRequest
    current_month_timesheet = None
    timesheet_is_consolidated = False
    pending_reopen_request = None
    
    if not show_team_data:
        # Ottieni il timesheet del mese corrente
        current_month_timesheet = MonthlyTimesheet.get_or_create(
            user_id=current_user.id,
            year=start_date.year,
            month=start_date.month,
            company_id=current_user.company_id
        )
        timesheet_is_consolidated = current_month_timesheet.is_consolidated
        
        # Verifica se c'è una richiesta di riapertura pendente
        if timesheet_is_consolidated:
            pending_reopen_request = TimesheetReopenRequest.query.filter_by(
                timesheet_id=current_month_timesheet.id,
                status='Pending'
            ).first()
    
    return render_template('attendance.html', 
                         form=form, 
                         records=records,
                         all_days=all_days,
                         records_by_sede=records_by_sede,
                         all_sedi_list=all_sedi_list,
                         today_date=datetime.now().date(),
                         start_date=start_date.strftime('%Y-%m-%d'),
                         end_date=end_date.strftime('%Y-%m-%d'),
                         user_status=user_status,
                         today_events=today_events,
                         today_work_hours=today_work_hours,
                         view_mode=view_mode,
                         show_team_data=show_team_data,
                         is_multi_sede=current_user.all_sedi,
                         timesheet_is_consolidated=timesheet_is_consolidated,
                         pending_reopen_request=pending_reopen_request,
                         current_month_timesheet=current_month_timesheet)

# =============================================================================
# CLOCK IN/OUT PRE-CHECK ROUTES
# =============================================================================

@attendance_bp.route('/check_shift_before_clock_in', methods=['POST'])
@login_required  
def check_shift_before_clock_in():
    """Check if user can clock-in (no shift validation)"""
    # Check if can perform clock-in action
    if not AttendanceEvent.can_perform_action(current_user.id, 'clock_in'):
        status, last_event = AttendanceEvent.get_user_status(current_user.id)
        if status == 'in':
            return jsonify({
                'success': False,
                'message': 'Sei già presente. Devi prima registrare l\'uscita.',
                'already_clocked': True
            })
        elif status == 'break':
            return jsonify({
                'success': False,
                'message': 'Sei in pausa. Devi prima terminare la pausa.',
                'already_clocked': True
            })
    
    # No shift validation - always allow clock-in
    return jsonify({
        'success': True,
        'needs_confirmation': False
    })

@attendance_bp.route('/check_shift_before_clock_out', methods=['POST'])
@login_required  
def check_shift_before_clock_out():
    """Check if user can clock-out (no shift validation)"""
    # Check if user can perform clock-out
    if not AttendanceEvent.can_perform_action(current_user.id, 'clock_out'):
        return jsonify({
            'success': False,
            'message': 'Non puoi registrare l\'uscita in questo momento.'
        })
    
    # No shift validation - always allow clock-out if status permits
    return jsonify({
        'success': True,
        'needs_confirmation': False
    })

# =============================================================================
# CLOCK IN/OUT MAIN ROUTES
# =============================================================================

@attendance_bp.route('/clock_in', methods=['POST'])
@login_required  
def clock_in():
    if not current_user.can_access_attendance():
        return jsonify({
            'success': False,
            'message': 'Non hai i permessi per accedere alle presenze'
        }), 403

    try:
        # Verifica se l'utente può effettuare il clock-in
        if not AttendanceEvent.can_perform_action(current_user.id, 'clock_in'):
            status, last_event = AttendanceEvent.get_user_status(current_user.id)
            if status == 'in':
                return jsonify({
                    'success': False,
                    'message': 'Sei già presente. Devi prima registrare l\'uscita.'
                })
            elif status == 'break':
                return jsonify({
                    'success': False,
                    'message': 'Sei in pausa. Devi prima terminare la pausa.'
                })

        # Crea nuovo evento di entrata
        attendance_event = AttendanceEvent(
            user_id=current_user.id,
            event_type='clock_in',
            timestamp=italian_now(),
            sede_id=get_current_user_sede(current_user).id if get_current_user_sede(current_user) else None
        )
        # Imposta company_id automaticamente
        set_company_on_create(attendance_event)
        db.session.add(attendance_event)
        db.session.commit()
        
        # Controlla se c'è un intervento attivo per questo utente
        active_intervention = Intervention.query.filter(
            Intervention.user_id == current_user.id,
            Intervention.end_datetime.is_(None)
        ).first()
        
        intervention_info = None
        if active_intervention:
            intervention_info = {
                'id': active_intervention.id,
                'description': active_intervention.description,
                'priority': active_intervention.priority,
                'start_datetime': active_intervention.start_datetime.strftime('%d/%m/%Y %H:%M') if active_intervention.start_datetime else None
            }

        return jsonify({
            'success': True,
            'message': f'Entrata registrata alle {attendance_event.timestamp_italian.strftime("%H:%M")}',
            'timestamp': attendance_event.timestamp_italian.strftime('%H:%M'),
            'active_intervention': intervention_info
        })

    except Exception as e:
        db.session.rollback()
        import logging
        logging.error(f"Errore clock_in per utente {current_user.id}: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Errore nel salvare l\'entrata: {str(e)}'
        }), 500

@attendance_bp.route('/clock_out', methods=['POST'])
@login_required
def clock_out():
    if not current_user.can_access_attendance():
        return jsonify({
            'success': False,
            'message': 'Non hai i permessi per accedere alle presenze'
        }), 403

    try:
        # Verifica se l'utente può effettuare il clock-out
        if not AttendanceEvent.can_perform_action(current_user.id, 'clock_out'):
            return jsonify({
                'success': False,
                'message': 'Non puoi registrare l\'uscita in questo momento.'
            })

        # Crea nuovo evento di uscita
        attendance_event = AttendanceEvent(
            user_id=current_user.id,
            event_type='clock_out',
            timestamp=italian_now(),
            sede_id=get_current_user_sede(current_user).id if get_current_user_sede(current_user) else None
        )
        # Imposta company_id automaticamente
        set_company_on_create(attendance_event)
        db.session.add(attendance_event)
        db.session.commit()

        return jsonify({
            'success': True,
            'message': f'Uscita registrata alle {attendance_event.timestamp_italian.strftime("%H:%M")}',
            'timestamp': attendance_event.timestamp_italian.strftime('%H:%M')
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': 'Errore nel salvare l\'uscita'
        }), 500

# =============================================================================
# BREAK START/END ROUTES
# =============================================================================

@attendance_bp.route('/break_start', methods=['POST'])
@login_required
def break_start():
    if not current_user.can_access_attendance():
        return jsonify({
            'success': False,
            'message': 'Non hai i permessi per accedere alle presenze'
        }), 403

    try:
        # Verifica se l'utente può iniziare la pausa
        if not AttendanceEvent.can_perform_action(current_user.id, 'break_start'):
            return jsonify({
                'success': False,
                'message': 'Non puoi iniziare la pausa in questo momento.'
            })

        # Crea nuovo evento di inizio pausa
        attendance_event = AttendanceEvent(
            user_id=current_user.id,
            event_type='break_start',
            timestamp=italian_now(),
            sede_id=get_current_user_sede(current_user).id if get_current_user_sede(current_user) else None
        )
        # Imposta company_id automaticamente
        set_company_on_create(attendance_event)
        db.session.add(attendance_event)
        db.session.commit()

        return jsonify({
            'success': True,
            'message': f'Inizio pausa registrato alle {attendance_event.timestamp_italian.strftime("%H:%M")}',
            'timestamp': attendance_event.timestamp_italian.strftime('%H:%M')
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': 'Errore nel salvare l\'inizio pausa'
        }), 500

@attendance_bp.route('/break_end', methods=['POST'])
@login_required
def break_end():
    if not current_user.can_access_attendance():
        return jsonify({
            'success': False,
            'message': 'Non hai i permessi per accedere alle presenze'
        }), 403

    try:
        # Verifica se l'utente può terminare la pausa
        if not AttendanceEvent.can_perform_action(current_user.id, 'break_end'):
            return jsonify({
                'success': False,
                'message': 'Non puoi terminare la pausa in questo momento.'
            })

        # Crea nuovo evento di fine pausa
        attendance_event = AttendanceEvent(
            user_id=current_user.id,
            event_type='break_end',
            timestamp=italian_now(),
            sede_id=get_current_user_sede(current_user).id if get_current_user_sede(current_user) else None
        )
        # Imposta company_id automaticamente
        set_company_on_create(attendance_event)
        db.session.add(attendance_event)
        db.session.commit()

        return jsonify({
            'success': True,
            'message': f'Fine pausa registrata alle {attendance_event.timestamp_italian.strftime("%H:%M")}',
            'timestamp': attendance_event.timestamp_italian.strftime('%H:%M')
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': 'Errore nel salvare la fine pausa'
        }), 500

# Funzione duplicata rimossa - ora esiste solo attendance() alla linea ~62

# =============================================================================
# ATTENDANCE EXPORT ROUTES
# =============================================================================

@attendance_bp.route('/export_excel')
@login_required  
def export_attendance_excel():
    """Export presenze in formato CSV"""
    if not current_user.can_view_attendance():
        flash('Non hai i permessi per esportare le presenze.', 'error')
        return redirect(url_for('attendance.attendance'))
    
    # Parametri per l'export
    start_date_str = request.args.get('start_date', (date.today() - timedelta(days=30)).strftime('%Y-%m-%d'))
    end_date_str = request.args.get('end_date', date.today().strftime('%Y-%m-%d'))
    
    try:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    except:
        flash('Date non valide.', 'error')
        return redirect(url_for('attendance.attendance'))

    # Query per ottenere tutti gli eventi nel periodo
    events = filter_by_company(AttendanceEvent.query).filter(
        AttendanceEvent.timestamp >= start_date,
        AttendanceEvent.timestamp <= end_date + timedelta(days=1)
    ).order_by(AttendanceEvent.timestamp).all()

    # Creazione file CSV
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Header
    writer.writerow(['Data', 'Utente', 'Tipo Evento', 'Orario', 'Sede'])
    
    # Dati
    for event in events:
        writer.writerow([
            event.timestamp_italian.strftime('%d/%m/%Y'),
            f"{event.user.first_name} {event.user.last_name}" if event.user else 'N/A',
            event.event_type,
            event.timestamp_italian.strftime('%H:%M:%S'),
            event.sede.name if event.sede else 'N/A'
        ])

    # Preparazione response
    output.seek(0)
    response = make_response(output.getvalue())
    response.headers["Content-Disposition"] = f"attachment; filename=presenze_{start_date_str}_{end_date_str}.csv"
    response.headers["Content-type"] = "text/csv; charset=utf-8"
    
    return response

# =============================================================================
# API ROUTES
# =============================================================================

@attendance_bp.route('/api/work_hours/<int:user_id>/<date_str>')
@login_required
def get_work_hours(user_id, date_str):
    """API endpoint per ottenere le ore lavorate aggiornate"""
    try:
        target_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        work_hours = AttendanceEvent.get_daily_work_hours(user_id, target_date)
        return jsonify({'work_hours': round(work_hours, 1)})
    except Exception as e:
        return jsonify({'work_hours': 0})

# =============================================================================
# QUICK ATTENDANCE (QR) ROUTE
# =============================================================================

@attendance_bp.route('/quick/<action>', methods=['GET', 'POST'])
@require_login
def quick_attendance(action):
    """Gestisce la registrazione rapida di entrata/uscita tramite QR"""
    if action not in ['clock_in', 'clock_out', 'break_start', 'break_end']:
        flash('Azione non valida.', 'error')
        return redirect(url_for('attendance.attendance'))
    
    if not current_user.can_access_attendance():
        flash('Non hai i permessi per accedere alle presenze.', 'error')
        return redirect(url_for('dashboard.dashboard'))

    if request.method == 'POST':
        try:
            # Verifica se l'utente può effettuare l'azione
            if not AttendanceEvent.can_perform_action(current_user.id, action):
                status, last_event = AttendanceEvent.get_user_status(current_user.id)
                if action == 'clock_in' and status == 'in':
                    flash('Sei già presente. Devi prima registrare l\'uscita.', 'warning')
                elif action == 'clock_out' and status == 'out':
                    flash('Non sei ancora presente. Devi prima registrare l\'entrata.', 'warning')
                elif action == 'break_start' and status != 'in':
                    flash('Devi essere presente per iniziare una pausa.', 'warning')
                elif action == 'break_end' and status != 'break':
                    flash('Non sei in pausa.', 'warning')
                else:
                    flash('Non puoi effettuare questa azione al momento.', 'error')
                return redirect(url_for('attendance.attendance'))

            # Crea nuovo evento
            attendance_event = AttendanceEvent(
                user_id=current_user.id,
                event_type=action,
                timestamp=italian_now(),
                sede_id=get_current_user_sede(current_user).id if get_current_user_sede(current_user) else None
            )
            db.session.add(attendance_event)
            db.session.commit()
            
            # Messaggio di successo
            action_messages = {
                'clock_in': 'Entrata registrata',
                'clock_out': 'Uscita registrata', 
                'break_start': 'Inizio pausa registrato',
                'break_end': 'Fine pausa registrata'
            }
            
            flash(f'{action_messages[action]} alle {attendance_event.timestamp_italian.strftime("%H:%M")}', 'success')
            return redirect(url_for('attendance.attendance'))

        except Exception as e:
            db.session.rollback()
            flash('Errore nel registrare l\'evento.', 'error')
            return redirect(url_for('attendance.attendance'))

    # GET request - mostra form di conferma
    action_titles = {
        'clock_in': 'Registra Entrata',
        'clock_out': 'Registra Uscita',
        'break_start': 'Inizia Pausa', 
        'break_end': 'Termina Pausa'
    }
    
    return render_template('quick_attendance.html',
                         action=action,
                         action_title=action_titles.get(action, 'Azione'),
                         current_time=italian_now())

# =============================================================================
# MANUAL TIMESHEET ROUTES
# =============================================================================

@attendance_bp.route('/manual_timesheet', methods=['GET'])
@login_required
def manual_timesheet():
    """Interfaccia per inserimento manuale timesheet mensile"""
    if not current_user.can_access_attendance():
        flash('Non hai i permessi per accedere alle presenze.', 'danger')
        return redirect(url_for('dashboard.dashboard'))
    
    # Ottieni mese e anno dal parametro GET o usa mese corrente
    try:
        year = int(request.args.get('year', datetime.now().year))
        month = int(request.args.get('month', datetime.now().month))
    except ValueError:
        year = datetime.now().year
        month = datetime.now().month
    
    # Verifica che mese sia valido
    if month < 1 or month > 12:
        flash('Mese non valido', 'danger')
        return redirect(url_for('attendance.attendance'))
    
    # Ottieni o crea il timesheet mensile
    company_id = get_user_company_id()
    timesheet = MonthlyTimesheet.get_or_create(current_user.id, year, month, company_id)
    
    # Calcola primo e ultimo giorno del mese
    from calendar import monthrange
    first_day = date(year, month, 1)
    last_day_num = monthrange(year, month)[1]
    last_day = date(year, month, last_day_num)
    
    # Ottieni tutti gli eventi del mese per l'utente
    events_query = filter_by_company(AttendanceEvent.query).filter(
        AttendanceEvent.user_id == current_user.id,
        AttendanceEvent.date >= first_day,
        AttendanceEvent.date <= last_day
    ).order_by(AttendanceEvent.date, AttendanceEvent.timestamp).all()
    
    # Ottieni ferie/permessi approvati E in attesa nel mese
    from models import LeaveRequest
    leaves_query = filter_by_company(LeaveRequest.query).filter(
        LeaveRequest.user_id == current_user.id,
        LeaveRequest.status.in_(['Approved', 'Pending']),  # Include sia approvate che in attesa
        LeaveRequest.start_date <= last_day,
        LeaveRequest.end_date >= first_day
    ).all()
    
    # Crea set di date con ferie/permessi
    leave_dates = set()
    leave_info = {}
    for leave in leaves_query:
        current_date = leave.start_date
        while current_date <= leave.end_date:
            if first_day <= current_date <= last_day:
                leave_dates.add(current_date)
                # Ottieni nome tipo ferie
                leave_type_name = leave.leave_type_obj.name if leave.leave_type_obj else (leave.leave_type or 'Permesso')
                leave_info[current_date] = leave_type_name
            current_date += timedelta(days=1)
    
    # Organizza eventi per giorno
    events_by_day = {}
    for event in events_query:
        day = event.date.day
        if day not in events_by_day:
            events_by_day[day] = {
                'events': [],
                'has_manual': False,
                'has_live': False,
                'entry_type': event.entry_type if event.is_manual else 'standard'
            }
        events_by_day[day]['events'].append(event)
        if event.is_manual:
            events_by_day[day]['has_manual'] = True
            events_by_day[day]['entry_type'] = event.entry_type
        else:
            events_by_day[day]['has_live'] = True
    
    # Crea lista giorni del mese con info
    days_data = []
    for day in range(1, last_day_num + 1):
        day_date = date(year, month, day)
        day_events = events_by_day.get(day, {'events': [], 'has_manual': False, 'has_live': False, 'entry_type': 'standard'})
        
        # Verifica se il giorno ha ferie/permessi
        has_leave = day_date in leave_dates
        leave_type = leave_info.get(day_date, '')
        
        # Trova primo clock_in e ultimo clock_out
        clock_in_time = None
        clock_out_time = None
        
        # Funzione helper per convertire timestamp a orario italiano
        from zoneinfo import ZoneInfo
        italy_tz = ZoneInfo('Europe/Rome')
        
        def to_italian_time_str(timestamp):
            """Converte timestamp a stringa HH:MM in orario italiano"""
            if not timestamp:
                return None
            # Se non ha timezone, assume UTC
            if timestamp.tzinfo is None:
                utc_tz = ZoneInfo('UTC')
                timestamp = timestamp.replace(tzinfo=utc_tz)
            # Converti a orario italiano
            italian_time = timestamp.astimezone(italy_tz)
            return italian_time.strftime('%H:%M')
        
        for event in day_events['events']:
            if event.event_type == 'clock_in' and clock_in_time is None:
                clock_in_time = to_italian_time_str(event.timestamp)
            elif event.event_type == 'clock_out':
                clock_out_time = to_italian_time_str(event.timestamp)
        
        days_data.append({
            'day': day,
            'date': day_date,
            'weekday': day_date.strftime('%a'),
            'clock_in': clock_in_time,
            'clock_out': clock_out_time,
            'has_manual': day_events['has_manual'],
            'has_live': day_events['has_live'],
            'is_weekend': day_date.weekday() >= 5,
            'is_future': day_date > date.today(),
            'has_leave': has_leave,
            'leave_type': leave_type,
            'entry_type': day_events['entry_type']
        })
    
    # Nomi mesi in italiano
    month_names = ['', 'Gennaio', 'Febbraio', 'Marzo', 'Aprile', 'Maggio', 'Giugno',
                   'Luglio', 'Agosto', 'Settembre', 'Ottobre', 'Novembre', 'Dicembre']
    
    return render_template('manual_timesheet.html',
                         timesheet=timesheet,
                         year=year,
                         month=month,
                         month_name=month_names[month],
                         days_data=days_data,
                         can_edit=timesheet.can_edit())

@attendance_bp.route('/manual_timesheet/save', methods=['POST'])
@login_required
def save_manual_timesheet():
    """Salva orari inseriti manualmente (salvataggio progressivo)"""
    if not current_user.can_access_attendance():
        return jsonify({'success': False, 'message': 'Permessi insufficienti'}), 403
    
    try:
        data = request.get_json()
        year = int(data.get('year'))
        month = int(data.get('month'))
        day = int(data.get('day'))
        clock_in_str = data.get('clock_in', '').strip()
        clock_out_str = data.get('clock_out', '').strip()
        entry_type = data.get('entry_type', 'standard')
        
        # Ottieni il timesheet mensile
        company_id = get_user_company_id()
        timesheet = MonthlyTimesheet.get_or_create(current_user.id, year, month, company_id)
        
        # Verifica se può essere modificato
        if not timesheet.can_edit():
            return jsonify({'success': False, 'message': 'Timesheet consolidato, non modificabile'}), 400
        
        # Crea data
        day_date = date(year, month, day)
        
        # Blocca inserimenti futuri
        if day_date > date.today():
            return jsonify({'success': False, 'message': 'Non è possibile inserire orari per giorni futuri'}), 400
        
        # Blocca inserimenti su giorni con ferie/permessi approvati o in attesa
        from models import LeaveRequest
        existing_leave = filter_by_company(LeaveRequest.query).filter(
            LeaveRequest.user_id == current_user.id,
            LeaveRequest.status.in_(['Approved', 'Pending']),  # Include anche richieste in attesa
            LeaveRequest.start_date <= day_date,
            LeaveRequest.end_date >= day_date
        ).first()
        
        if existing_leave:
            leave_type_name = existing_leave.leave_type_obj.name if existing_leave.leave_type_obj else (existing_leave.leave_type or 'Permesso')
            status_text = 'approvato' if existing_leave.status == 'Approved' else 'in attesa di approvazione'
            return jsonify({'success': False, 'message': f'Impossibile inserire orari: giorno con {leave_type_name} {status_text}'}), 400
        
        # Elimina eventi manuali esistenti per questo giorno
        filter_by_company(AttendanceEvent.query).filter(
            AttendanceEvent.user_id == current_user.id,
            AttendanceEvent.date == day_date,
            AttendanceEvent.is_manual == True
        ).delete()
        
        # Se entrambi i campi sono vuoti, elimina e basta
        if not clock_in_str and not clock_out_str:
            db.session.commit()
            return jsonify({'success': True, 'message': 'Orari rimossi'})
        
        # Valida e crea nuovi eventi
        from zoneinfo import ZoneInfo
        italy_tz = ZoneInfo('Europe/Rome')
        clock_in_datetime = None
        clock_out_datetime = None
        
        if clock_in_str:
            try:
                clock_in_time = datetime.strptime(clock_in_str, '%H:%M').time()
                clock_in_datetime = datetime.combine(day_date, clock_in_time)
                clock_in_datetime = clock_in_datetime.replace(tzinfo=italy_tz)
                
                clock_in_event = AttendanceEvent(
                    user_id=current_user.id,
                    date=day_date,
                    event_type='clock_in',
                    timestamp=clock_in_datetime,
                    sede_id=current_user.sede_id,
                    is_manual=True,
                    entry_type=entry_type
                )
                set_company_on_create(clock_in_event)
                db.session.add(clock_in_event)
            except ValueError:
                return jsonify({'success': False, 'message': 'Formato orario entrata non valido'}), 400
        
        if clock_out_str:
            try:
                clock_out_time = datetime.strptime(clock_out_str, '%H:%M').time()
                clock_out_datetime = datetime.combine(day_date, clock_out_time)
                clock_out_datetime = clock_out_datetime.replace(tzinfo=italy_tz)
                
                clock_out_event = AttendanceEvent(
                    user_id=current_user.id,
                    date=day_date,
                    event_type='clock_out',
                    timestamp=clock_out_datetime,
                    sede_id=current_user.sede_id,
                    is_manual=True,
                    entry_type=entry_type
                )
                set_company_on_create(clock_out_event)
                db.session.add(clock_out_event)
            except ValueError:
                return jsonify({'success': False, 'message': 'Formato orario uscita non valido'}), 400
        
        # Pausa automatica: se il turno è > 5 ore, inserisci 1h di pausa
        if clock_in_datetime and clock_out_datetime:
            work_duration = (clock_out_datetime - clock_in_datetime).total_seconds() / 3600  # ore
            
            if work_duration > 5:
                # Prova a posizionare la pausa alle 12:30-13:30 se il turno lo copre
                lunch_start = datetime.combine(day_date, datetime.strptime('12:30', '%H:%M').time()).replace(tzinfo=italy_tz)
                lunch_end = lunch_start + timedelta(hours=1)
                
                # Limiti di sicurezza: almeno 30 min dopo entrata e 30 min prima di uscita
                min_break_start = clock_in_datetime + timedelta(minutes=30)
                max_break_end = clock_out_datetime - timedelta(minutes=30)
                
                # Verifica se la pausa 12:30-13:30 sta completamente dentro il turno E rispetta i limiti
                if (lunch_start >= min_break_start and 
                    lunch_end <= max_break_end):
                    break_start_datetime = lunch_start
                    break_end_datetime = lunch_end
                else:
                    # Altrimenti, posiziona la pausa al centro del turno
                    # Assicurandosi che rimanga dentro [clock_in, clock_out]
                    mid_point = clock_in_datetime + timedelta(hours=work_duration / 2)
                    break_start_datetime = mid_point - timedelta(minutes=30)
                    break_end_datetime = mid_point + timedelta(minutes=30)
                    
                    # Verifica che la pausa rimanga dentro i limiti (già calcolati sopra)
                    if break_start_datetime < min_break_start:
                        break_start_datetime = min_break_start
                        break_end_datetime = break_start_datetime + timedelta(hours=1)
                    elif break_end_datetime > max_break_end:
                        break_end_datetime = max_break_end
                        break_start_datetime = break_end_datetime - timedelta(hours=1)
                
                # Crea evento inizio pausa
                break_start_event = AttendanceEvent(
                    user_id=current_user.id,
                    date=day_date,
                    event_type='break_start',
                    timestamp=break_start_datetime,
                    sede_id=current_user.sede_id,
                    is_manual=True,
                    entry_type=entry_type
                )
                set_company_on_create(break_start_event)
                db.session.add(break_start_event)
                
                # Crea evento fine pausa
                break_end_event = AttendanceEvent(
                    user_id=current_user.id,
                    date=day_date,
                    event_type='break_end',
                    timestamp=break_end_datetime,
                    sede_id=current_user.sede_id,
                    is_manual=True,
                    entry_type=entry_type
                )
                set_company_on_create(break_end_event)
                db.session.add(break_end_event)
        
        # Aggiorna timestamp timesheet
        timesheet.updated_at = italian_now()
        
        db.session.commit()
        return jsonify({'success': True, 'message': 'Orari salvati con successo'})
        
    except Exception as e:
        db.session.rollback()
        import logging
        logging.error(f"Errore salvataggio timesheet manuale: {str(e)}")
        return jsonify({'success': False, 'message': f'Errore: {str(e)}'}), 500

@attendance_bp.route('/manual_timesheet/consolidate', methods=['POST'])
@login_required
def consolidate_manual_timesheet():
    """Consolida il timesheet mensile rendendolo immutabile"""
    if not current_user.can_access_attendance():
        return jsonify({'success': False, 'message': 'Permessi insufficienti'}), 403
    
    try:
        data = request.get_json()
        year = int(data.get('year'))
        month = int(data.get('month'))
        
        # Ottieni il timesheet mensile
        company_id = get_user_company_id()
        timesheet = MonthlyTimesheet.get_or_create(current_user.id, year, month, company_id)
        
        # Consolida
        if timesheet.consolidate(current_user.id):
            return jsonify({'success': True, 'message': 'Timesheet consolidato con successo'})
        else:
            return jsonify({'success': False, 'message': 'Timesheet già consolidato'}), 400
            
    except Exception as e:
        db.session.rollback()
        import logging
        logging.error(f"Errore consolidamento timesheet: {str(e)}")
        return jsonify({'success': False, 'message': f'Errore: {str(e)}'}), 500

@attendance_bp.route('/manual_timesheet/delete_day', methods=['POST'])
@login_required
def delete_manual_timesheet_day():
    """Cancella i dati manuali di un giorno specifico"""
    if not current_user.can_access_attendance():
        return jsonify({'success': False, 'message': 'Permessi insufficienti'}), 403
    
    try:
        data = request.get_json()
        year = int(data.get('year'))
        month = int(data.get('month'))
        day = int(data.get('day'))
        
        # Ottieni il timesheet mensile
        company_id = get_user_company_id()
        timesheet = MonthlyTimesheet.get_or_create(current_user.id, year, month, company_id)
        
        # Verifica se può essere modificato
        if not timesheet.can_edit():
            return jsonify({'success': False, 'message': 'Timesheet consolidato, non modificabile'}), 400
        
        # Crea data
        day_date = date(year, month, day)
        
        # Cancella tutti gli eventi manuali del giorno
        events_to_delete = filter_by_company(AttendanceEvent.query).filter(
            AttendanceEvent.user_id == current_user.id,
            AttendanceEvent.date == day_date,
            AttendanceEvent.is_manual == True
        ).all()
        
        for event in events_to_delete:
            db.session.delete(event)
        
        db.session.commit()
        
        return jsonify({
            'success': True, 
            'message': f'Dati del {day_date.strftime("%d/%m/%Y")} cancellati con successo',
            'deleted_count': len(events_to_delete)
        })
        
    except Exception as e:
        db.session.rollback()
        import logging
        logging.error(f"Errore cancellazione giorno timesheet: {str(e)}")
        return jsonify({'success': False, 'message': f'Errore: {str(e)}'}), 500

@attendance_bp.route('/manual_timesheet/bulk_fill', methods=['POST'])
@login_required
def bulk_fill_timesheet():
    """Compila automaticamente il mese con orari standard della sede"""
    if not current_user.can_access_attendance():
        return jsonify({'success': False, 'message': 'Permessi insufficienti'}), 403
    
    try:
        data = request.get_json()
        year = int(data.get('year'))
        month = int(data.get('month'))
        
        # Ottieni il timesheet mensile
        company_id = get_user_company_id()
        timesheet = MonthlyTimesheet.get_or_create(current_user.id, year, month, company_id)
        
        # Verifica se può essere modificato
        if not timesheet.can_edit():
            return jsonify({'success': False, 'message': 'Timesheet consolidato, non modificabile'}), 400
        
        # Ottieni orario di lavoro assegnato all'utente (company-level global, non più legato alla sede)
        from models import WorkSchedule, LeaveRequest
        work_schedule = current_user.work_schedule if current_user.work_schedule_id else None
        
        if not work_schedule:
            return jsonify({'success': False, 'message': 'Nessun orario di lavoro configurato per il tuo profilo'}), 400
        
        # Calcola range mese
        from calendar import monthrange
        first_day = date(year, month, 1)
        last_day_num = monthrange(year, month)[1]
        last_day = date(year, month, last_day_num)
        
        # Ottieni ferie/permessi approvati e in attesa nel mese
        leaves_query = filter_by_company(LeaveRequest.query).filter(
            LeaveRequest.user_id == current_user.id,
            LeaveRequest.status.in_(['Approved', 'Pending']),  # Include anche richieste in attesa
            LeaveRequest.start_date <= last_day,
            LeaveRequest.end_date >= first_day
        ).all()
        
        # Crea set di date con ferie/permessi
        leave_dates = set()
        for leave in leaves_query:
            current_date = leave.start_date
            while current_date <= leave.end_date:
                if first_day <= current_date <= last_day:
                    leave_dates.add(current_date)
                current_date += timedelta(days=1)
        
        # Ottieni eventi già esistenti
        existing_events = filter_by_company(AttendanceEvent.query).filter(
            AttendanceEvent.user_id == current_user.id,
            AttendanceEvent.date >= first_day,
            AttendanceEvent.date <= last_day
        ).all()
        
        # Raggruppa eventi per data e controlla se ha eventi Live
        from collections import defaultdict
        events_by_date = defaultdict(list)
        dates_with_live = set()
        
        for event in existing_events:
            events_by_date[event.date].append(event)
            if not event.is_manual:  # Evento Live
                dates_with_live.add(event.date)
        
        # Compila automaticamente
        days_filled = 0
        days_replaced = 0
        today = date.today()
        
        # Calcola orari medi quando c'è flessibilità
        # Usa sempre i campi min/max per calcolare la media (i legacy sono deprecati)
        start_min_minutes = work_schedule.start_time_min.hour * 60 + work_schedule.start_time_min.minute
        start_max_minutes = work_schedule.start_time_max.hour * 60 + work_schedule.start_time_max.minute
        avg_start_minutes = (start_min_minutes + start_max_minutes) // 2
        standard_start = time(avg_start_minutes // 60, avg_start_minutes % 60)
        
        end_min_minutes = work_schedule.end_time_min.hour * 60 + work_schedule.end_time_min.minute
        end_max_minutes = work_schedule.end_time_max.hour * 60 + work_schedule.end_time_max.minute
        avg_end_minutes = (end_min_minutes + end_max_minutes) // 2
        standard_end = time(avg_end_minutes // 60, avg_end_minutes % 60)
        
        if not standard_start or not standard_end:
            return jsonify({'success': False, 'message': 'Orari standard non definiti correttamente'}), 400
        
        for day in range(1, last_day_num + 1):
            day_date = date(year, month, day)
            
            # Salta se:
            # - Giorno futuro
            # - Weekend
            # - Ha ferie/permessi
            # - Ha eventi Live (non toccare mai dati Live!)
            # - Non è nel days_of_week della sede (se definito)
            if day_date > today:
                continue
            if day_date.weekday() >= 5:  # Weekend
                continue
            if day_date in leave_dates:
                continue
            if day_date in dates_with_live:
                # Ha eventi Live, non toccare
                continue
            
            # Verifica se il giorno è coperto dal WorkSchedule
            if work_schedule.days_of_week and day_date.weekday() not in work_schedule.days_of_week:
                continue
            
            # Se ha eventi manuali esistenti, cancellali prima
            if day_date in events_by_date:
                for event in events_by_date[day_date]:
                    db.session.delete(event)
                days_replaced += 1
            
            # Crea eventi di entrata e uscita
            from zoneinfo import ZoneInfo
            italy_tz = ZoneInfo('Europe/Rome')
            
            clock_in_datetime = datetime.combine(day_date, standard_start).replace(tzinfo=italy_tz)
            clock_out_datetime = datetime.combine(day_date, standard_end).replace(tzinfo=italy_tz)
            
            # Crea evento entrata
            clock_in_event = AttendanceEvent(
                user_id=current_user.id,
                date=day_date,
                event_type='clock_in',
                timestamp=clock_in_datetime,
                sede_id=current_user.sede_id,
                is_manual=True,
                entry_type='standard'
            )
            set_company_on_create(clock_in_event)
            db.session.add(clock_in_event)
            
            # Crea evento uscita
            clock_out_event = AttendanceEvent(
                user_id=current_user.id,
                date=day_date,
                event_type='clock_out',
                timestamp=clock_out_datetime,
                sede_id=current_user.sede_id,
                is_manual=True,
                entry_type='standard'
            )
            set_company_on_create(clock_out_event)
            db.session.add(clock_out_event)
            
            # Pausa automatica: se il turno è > 5 ore, inserisci 1h di pausa
            work_duration = (clock_out_datetime - clock_in_datetime).total_seconds() / 3600  # ore
            
            if work_duration > 5:
                # Prova a posizionare la pausa alle 12:30-13:30 se il turno lo copre
                lunch_start = datetime.combine(day_date, datetime.strptime('12:30', '%H:%M').time()).replace(tzinfo=italy_tz)
                lunch_end = lunch_start + timedelta(hours=1)
                
                # Limiti di sicurezza: almeno 30 min dopo entrata e 30 min prima di uscita
                min_break_start = clock_in_datetime + timedelta(minutes=30)
                max_break_end = clock_out_datetime - timedelta(minutes=30)
                
                # Verifica se la pausa 12:30-13:30 sta completamente dentro il turno E rispetta i limiti
                if (lunch_start >= min_break_start and 
                    lunch_end <= max_break_end):
                    break_start_datetime = lunch_start
                    break_end_datetime = lunch_end
                else:
                    # Altrimenti, posiziona la pausa al centro del turno
                    # Assicurandosi che rimanga dentro [clock_in, clock_out]
                    mid_point = clock_in_datetime + timedelta(hours=work_duration / 2)
                    break_start_datetime = mid_point - timedelta(minutes=30)
                    break_end_datetime = mid_point + timedelta(minutes=30)
                    
                    # Verifica che la pausa rimanga dentro i limiti (già calcolati sopra)
                    if break_start_datetime < min_break_start:
                        break_start_datetime = min_break_start
                        break_end_datetime = break_start_datetime + timedelta(hours=1)
                    elif break_end_datetime > max_break_end:
                        break_end_datetime = max_break_end
                        break_start_datetime = break_end_datetime - timedelta(hours=1)
                
                # Crea evento inizio pausa
                break_start_event = AttendanceEvent(
                    user_id=current_user.id,
                    date=day_date,
                    event_type='break_start',
                    timestamp=break_start_datetime,
                    sede_id=current_user.sede_id,
                    is_manual=True,
                    entry_type='standard'
                )
                set_company_on_create(break_start_event)
                db.session.add(break_start_event)
                
                # Crea evento fine pausa
                break_end_event = AttendanceEvent(
                    user_id=current_user.id,
                    date=day_date,
                    event_type='break_end',
                    timestamp=break_end_datetime,
                    sede_id=current_user.sede_id,
                    is_manual=True,
                    entry_type='standard'
                )
                set_company_on_create(break_end_event)
                db.session.add(break_end_event)
            
            days_filled += 1
        
        # Aggiorna timestamp timesheet
        timesheet.updated_at = italian_now()
        
        db.session.commit()
        
        if days_filled == 0:
            return jsonify({'success': True, 'message': 'Nessun giorno compilato (tutti i giorni lavorativi sono già coperti o nel futuro)'})
        
        # Messaggio con dettagli sostituzioni
        message = f'{days_filled} giorni compilati con orari standard'
        if days_replaced > 0:
            message += f' ({days_replaced} giorni sostituiti)'
        
        return jsonify({'success': True, 'message': message})
        
    except Exception as e:
        db.session.rollback()
        import logging
        logging.error(f"Errore compilazione massiva timesheet: {str(e)}")
        return jsonify({'success': False, 'message': f'Errore: {str(e)}'}), 500

# =============================================================================
# MANUAL ATTENDANCE ENTRY/EDIT ROUTES
# =============================================================================

@attendance_bp.route('/manual_entry', methods=['POST'])
@login_required
def manual_entry():
    """Inserisce manualmente una presenza (clock_in/out, pause)"""
    try:
        from zoneinfo import ZoneInfo
        italy_tz = ZoneInfo('Europe/Rome')
        
        # Ottieni dati dal form
        entry_date_str = request.form.get('entry_date')
        clock_in_time = request.form.get('clock_in_time')
        clock_out_time = request.form.get('clock_out_time')
        break_start_time = request.form.get('break_start_time')
        break_end_time = request.form.get('break_end_time')
        notes = request.form.get('notes', '').strip()
        
        # Validazione data
        if not entry_date_str:
            flash('Data obbligatoria', 'danger')
            return redirect(url_for('attendance.attendance'))
        
        entry_date = datetime.strptime(entry_date_str, '%Y-%m-%d').date()
        
        # Blocca date future
        if entry_date > date.today():
            flash('Non puoi inserire presenze per date future', 'danger')
            return redirect(url_for('attendance.attendance'))
        
        # Verifica se il timesheet è consolidato
        from models import MonthlyTimesheet
        timesheet = MonthlyTimesheet.get_or_create(
            user_id=current_user.id,
            year=entry_date.year,
            month=entry_date.month,
            company_id=current_user.company_id
        )
        
        if not timesheet.can_edit():
            flash('Il timesheet per questo mese è consolidato. Devi richiedere una riapertura per modificarlo.', 'danger')
            return redirect(url_for('attendance.attendance'))
        
        # Validazione orari
        if not clock_in_time or not clock_out_time:
            flash('Orario di entrata e uscita obbligatori', 'danger')
            return redirect(url_for('attendance.attendance'))
        
        # Crea timestamp in Italian time
        clock_in_dt = datetime.strptime(f"{entry_date_str} {clock_in_time}", '%Y-%m-%d %H:%M')
        clock_in_dt = clock_in_dt.replace(tzinfo=italy_tz)
        
        clock_out_dt = datetime.strptime(f"{entry_date_str} {clock_out_time}", '%Y-%m-%d %H:%M')
        clock_out_dt = clock_out_dt.replace(tzinfo=italy_tz)
        
        # Validazione logica orari
        if clock_out_dt <= clock_in_dt:
            flash('L\'orario di uscita deve essere successivo all\'entrata', 'danger')
            return redirect(url_for('attendance.attendance'))
        
        # Converti a UTC naive per storage
        from datetime import timezone
        clock_in_utc = clock_in_dt.astimezone(timezone.utc).replace(tzinfo=None)
        clock_out_utc = clock_out_dt.astimezone(timezone.utc).replace(tzinfo=None)
        
        # Verifica se esistono già eventi per questo giorno
        existing_events = AttendanceEvent.query.filter_by(
            user_id=current_user.id,
            date=entry_date
        ).all()
        
        # Se ci sono eventi non manuali, blocca inserimento
        if any(not event.is_manual for event in existing_events):
            flash('Esistono già timbrature automatiche per questo giorno. Non puoi inserire dati manuali.', 'danger')
            return redirect(url_for('attendance.attendance'))
        
        # Elimina eventi esistenti manuali per questo giorno (sovrascrittura)
        for event in existing_events:
            if event.is_manual:
                db.session.delete(event)
        
        # Crea evento entrata
        clock_in_event = AttendanceEvent(
            user_id=current_user.id,
            date=entry_date,
            event_type='clock_in',
            timestamp=clock_in_utc,
            sede_id=current_user.sede_id,
            notes=notes,
            is_manual=True,
            entry_type='standard'
        )
        set_company_on_create(clock_in_event)
        db.session.add(clock_in_event)
        
        # Crea evento uscita
        clock_out_event = AttendanceEvent(
            user_id=current_user.id,
            date=entry_date,
            event_type='clock_out',
            timestamp=clock_out_utc,
            sede_id=current_user.sede_id,
            notes=notes,
            is_manual=True,
            entry_type='standard'
        )
        set_company_on_create(clock_out_event)
        db.session.add(clock_out_event)
        
        # Gestione pausa (opzionale)
        if break_start_time and break_end_time:
            break_start_dt = datetime.strptime(f"{entry_date_str} {break_start_time}", '%Y-%m-%d %H:%M')
            break_start_dt = break_start_dt.replace(tzinfo=italy_tz)
            
            break_end_dt = datetime.strptime(f"{entry_date_str} {break_end_time}", '%Y-%m-%d %H:%M')
            break_end_dt = break_end_dt.replace(tzinfo=italy_tz)
            
            # Validazione pausa
            if break_start_dt >= break_end_dt:
                flash('Fine pausa deve essere successiva all\'inizio pausa', 'danger')
                return redirect(url_for('attendance.attendance'))
            
            if break_start_dt <= clock_in_dt or break_end_dt >= clock_out_dt:
                flash('La pausa deve essere compresa tra entrata e uscita', 'danger')
                return redirect(url_for('attendance.attendance'))
            
            # Converti a UTC
            break_start_utc = break_start_dt.astimezone(timezone.utc).replace(tzinfo=None)
            break_end_utc = break_end_dt.astimezone(timezone.utc).replace(tzinfo=None)
            
            # Crea eventi pausa
            break_start_event = AttendanceEvent(
                user_id=current_user.id,
                date=entry_date,
                event_type='break_start',
                timestamp=break_start_utc,
                sede_id=current_user.sede_id,
                is_manual=True,
                entry_type='standard'
            )
            set_company_on_create(break_start_event)
            db.session.add(break_start_event)
            
            break_end_event = AttendanceEvent(
                user_id=current_user.id,
                date=entry_date,
                event_type='break_end',
                timestamp=break_end_utc,
                sede_id=current_user.sede_id,
                is_manual=True,
                entry_type='standard'
            )
            set_company_on_create(break_end_event)
            db.session.add(break_end_event)
        
        db.session.commit()
        flash('Presenza inserita manualmente con successo', 'success')
        return redirect(url_for('attendance.attendance'))
        
    except ValueError as ve:
        db.session.rollback()
        flash(f'Formato orario non valido: {str(ve)}', 'danger')
        return redirect(url_for('attendance.attendance'))
    except Exception as e:
        db.session.rollback()
        import logging
        logging.error(f"Errore inserimento manuale presenza: {str(e)}")
        flash(f'Errore inserimento presenza: {str(e)}', 'danger')
        return redirect(url_for('attendance.attendance'))

@attendance_bp.route('/edit_manual_entry/<date_str>', methods=['GET'])
@login_required
def edit_manual_entry(date_str):
    """Ottieni i dati di una presenza manuale per la modifica"""
    try:
        entry_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        
        # Blocca date future
        if entry_date > date.today():
            return jsonify({'success': False, 'message': 'Non puoi modificare date future'})
        
        # Verifica se il timesheet è consolidato
        from models import MonthlyTimesheet
        timesheet = MonthlyTimesheet.get_or_create(
            user_id=current_user.id,
            year=entry_date.year,
            month=entry_date.month,
            company_id=current_user.company_id
        )
        
        if not timesheet.can_edit():
            return jsonify({'success': False, 'message': 'Il timesheet per questo mese è consolidato. Devi richiedere una riapertura.'})
        
        # Ottieni eventi per questo giorno
        events = AttendanceEvent.query.filter_by(
            user_id=current_user.id,
            date=entry_date
        ).order_by(AttendanceEvent.timestamp).all()
        
        # Verifica che tutti gli eventi siano manuali
        if not events or any(not event.is_manual for event in events):
            return jsonify({'success': False, 'message': 'Questa presenza non è modificabile (timbratura automatica)'})
        
        from zoneinfo import ZoneInfo
        italy_tz = ZoneInfo('Europe/Rome')
        
        # Estrai dati
        data = {
            'date': date_str,
            'clock_in': None,
            'clock_out': None,
            'break_start': None,
            'break_end': None,
            'notes': ''
        }
        
        for event in events:
            # Converti timestamp a Italian time
            if event.timestamp:
                from datetime import timezone
                utc_time = event.timestamp.replace(tzinfo=timezone.utc)
                italian_time = utc_time.astimezone(italy_tz)
                time_str = italian_time.strftime('%H:%M')
                
                if event.event_type == 'clock_in':
                    data['clock_in'] = time_str
                    if event.notes:
                        data['notes'] = event.notes
                elif event.event_type == 'clock_out':
                    data['clock_out'] = time_str
                elif event.event_type == 'break_start':
                    data['break_start'] = time_str
                elif event.event_type == 'break_end':
                    data['break_end'] = time_str
        
        return jsonify({'success': True, 'data': data})
        
    except Exception as e:
        import logging
        logging.error(f"Errore caricamento presenza per modifica: {str(e)}")
        return jsonify({'success': False, 'message': str(e)})

@attendance_bp.route('/delete_manual_entry/<date_str>', methods=['POST'])
@login_required
def delete_manual_entry(date_str):
    """Elimina una presenza inserita manualmente"""
    try:
        entry_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        
        # Blocca date future
        if entry_date > date.today():
            flash('Non puoi eliminare date future', 'danger')
            return redirect(url_for('attendance.attendance'))
        
        # Verifica se il timesheet è consolidato
        from models import MonthlyTimesheet
        timesheet = MonthlyTimesheet.get_or_create(
            user_id=current_user.id,
            year=entry_date.year,
            month=entry_date.month,
            company_id=current_user.company_id
        )
        
        if not timesheet.can_edit():
            flash('Il timesheet per questo mese è consolidato. Devi richiedere una riapertura per modificarlo.', 'danger')
            return redirect(url_for('attendance.attendance'))
        
        # Ottieni eventi per questo giorno
        events = AttendanceEvent.query.filter_by(
            user_id=current_user.id,
            date=entry_date
        ).all()
        
        # Verifica che tutti gli eventi siano manuali
        if not events:
            flash('Nessuna presenza trovata per questa data', 'warning')
            return redirect(url_for('attendance.attendance'))
        
        if any(not event.is_manual for event in events):
            flash('Non puoi eliminare una presenza con timbratura automatica', 'danger')
            return redirect(url_for('attendance.attendance'))
        
        # Elimina tutti gli eventi
        for event in events:
            db.session.delete(event)
        
        db.session.commit()
        flash('Presenza eliminata con successo', 'success')
        return redirect(url_for('attendance.attendance'))
        
    except Exception as e:
        db.session.rollback()
        import logging
        logging.error(f"Errore eliminazione presenza manuale: {str(e)}")
        flash(f'Errore eliminazione presenza: {str(e)}', 'danger')
        return redirect(url_for('attendance.attendance'))

# =============================================================================
# TIMESHEET REOPEN REQUEST ROUTES
# =============================================================================

@attendance_bp.route('/request_timesheet_reopen/<int:year>/<int:month>', methods=['GET', 'POST'])
@login_required
def request_timesheet_reopen(year, month):
    """Richiedi la riapertura di un timesheet consolidato"""
    if request.method == 'POST':
        try:
            from models import MonthlyTimesheet, TimesheetReopenRequest
            
            reason = request.form.get('reason', '').strip()
            if not reason:
                flash('Devi fornire una motivazione per la richiesta', 'danger')
                return redirect(url_for('attendance.my_attendance', year=year, month=month))
            
            # Verifica che il timesheet esista e sia consolidato
            timesheet = MonthlyTimesheet.get_or_create(
                user_id=current_user.id,
                year=year,
                month=month,
                company_id=current_user.company_id
            )
            
            if not timesheet.is_consolidated:
                flash('Il timesheet non è consolidato, non serve richiesta di riapertura', 'info')
                return redirect(url_for('attendance.my_attendance', year=year, month=month))
            
            # Verifica se esiste già una richiesta pendente
            existing_request = TimesheetReopenRequest.query.filter_by(
                timesheet_id=timesheet.id,
                status='Pending'
            ).first()
            
            if existing_request:
                flash('Esiste già una richiesta di riapertura pendente per questo mese', 'warning')
                return redirect(url_for('attendance.my_attendance', year=year, month=month))
            
            # Crea la richiesta
            reopen_request = TimesheetReopenRequest(
                timesheet_id=timesheet.id,
                requested_by=current_user.id,
                reason=reason,
                company_id=current_user.company_id
            )
            set_company_on_create(reopen_request)
            db.session.add(reopen_request)
            db.session.commit()
            
            flash('Richiesta di riapertura inviata con successo. Attendi l\'approvazione.', 'success')
            return redirect(url_for('attendance.my_attendance', year=year, month=month))
            
        except Exception as e:
            db.session.rollback()
            import logging
            logging.error(f"Errore richiesta riapertura timesheet: {str(e)}")
            flash(f'Errore invio richiesta: {str(e)}', 'danger')
            return redirect(url_for('attendance.my_attendance', year=year, month=month))
    
    # GET - mostra form
    from models import MonthlyTimesheet
    timesheet = MonthlyTimesheet.get_or_create(
        user_id=current_user.id,
        year=year,
        month=month,
        company_id=current_user.company_id
    )
    
    if not timesheet.is_consolidated:
        flash('Il timesheet non è consolidato', 'info')
        return redirect(url_for('attendance.attendance'))
    
    return render_template('request_timesheet_reopen.html', 
                         timesheet=timesheet,
                         year=year,
                         month=month)

@attendance_bp.route('/timesheet_reopen_requests', methods=['GET'])
@login_required
def timesheet_reopen_requests():
    """Visualizza le richieste di riapertura timesheet (per responsabili)"""
    from models import TimesheetReopenRequest
    
    # Verifica permessi - SOLO HR e Admin possono gestire richieste di riapertura
    if not (current_user.can_manage_hr_data() or current_user.role in ['Admin', 'Amministratore']):
        flash('Non hai i permessi per visualizzare le richieste di riapertura', 'danger')
        return redirect(url_for('dashboard.dashboard'))
    
    # Ottieni le richieste pendenti filtrate per company
    pending_requests = filter_by_company(TimesheetReopenRequest.query).filter_by(
        status='Pending'
    ).order_by(TimesheetReopenRequest.requested_at.desc()).all()
    
    # Ottieni le richieste completate (approvate/rifiutate)
    completed_requests = filter_by_company(TimesheetReopenRequest.query).filter(
        TimesheetReopenRequest.status.in_(['Approved', 'Rejected'])
    ).order_by(TimesheetReopenRequest.reviewed_at.desc()).limit(50).all()
    
    return render_template('timesheet_reopen_requests.html',
                         pending_requests=pending_requests,
                         completed_requests=completed_requests)

@attendance_bp.route('/approve_timesheet_reopen/<int:request_id>', methods=['POST'])
@login_required
def approve_timesheet_reopen(request_id):
    """Approva una richiesta di riapertura timesheet"""
    try:
        from models import TimesheetReopenRequest
        
        reopen_request = filter_by_company(TimesheetReopenRequest.query).filter_by(id=request_id).first()
        if not reopen_request:
            flash('Richiesta non trovata', 'danger')
            return redirect(url_for('attendance.timesheet_reopen_requests'))
        
        # Verifica permessi
        if not reopen_request.can_approve(current_user):
            flash('Non hai i permessi per approvare questa richiesta', 'danger')
            return redirect(url_for('attendance.timesheet_reopen_requests'))
        
        review_notes = request.form.get('review_notes', '').strip()
        
        if reopen_request.approve(current_user.id, review_notes):
            flash('Richiesta approvata e timesheet riaperto con successo', 'success')
        else:
            flash('Impossibile approvare questa richiesta', 'danger')
        
        return redirect(url_for('attendance.timesheet_reopen_requests'))
        
    except Exception as e:
        db.session.rollback()
        import logging
        logging.error(f"Errore approvazione richiesta riapertura: {str(e)}")
        flash(f'Errore approvazione: {str(e)}', 'danger')
        return redirect(url_for('attendance.timesheet_reopen_requests'))

@attendance_bp.route('/reject_timesheet_reopen/<int:request_id>', methods=['POST'])
@login_required
def reject_timesheet_reopen(request_id):
    """Rifiuta una richiesta di riapertura timesheet"""
    try:
        from models import TimesheetReopenRequest
        
        reopen_request = filter_by_company(TimesheetReopenRequest.query).filter_by(id=request_id).first()
        if not reopen_request:
            flash('Richiesta non trovata', 'danger')
            return redirect(url_for('attendance.timesheet_reopen_requests'))
        
        # Verifica permessi
        if not reopen_request.can_approve(current_user):
            flash('Non hai i permessi per rifiutare questa richiesta', 'danger')
            return redirect(url_for('attendance.timesheet_reopen_requests'))
        
        review_notes = request.form.get('review_notes', '').strip()
        if not review_notes:
            flash('Devi fornire una motivazione per il rifiuto', 'danger')
            return redirect(url_for('attendance.timesheet_reopen_requests'))
        
        if reopen_request.reject(current_user.id, review_notes):
            flash('Richiesta rifiutata', 'success')
        else:
            flash('Impossibile rifiutare questa richiesta', 'danger')
        
        return redirect(url_for('attendance.timesheet_reopen_requests'))
        
    except Exception as e:
        db.session.rollback()
        import logging
        logging.error(f"Errore rifiuto richiesta riapertura: {str(e)}")
        flash(f'Errore rifiuto: {str(e)}', 'danger')
        return redirect(url_for('attendance.timesheet_reopen_requests'))

# =============================================================================
# MY ATTENDANCE - PERSONAL ATTENDANCE MANAGEMENT WITH INLINE EDITING
# =============================================================================

@attendance_bp.route('/my_attendance', methods=['GET'])
@login_required
def my_attendance():
    """Interfaccia personale per gestione presenze con editing inline"""
    if not current_user.can_access_attendance():
        flash('Non hai i permessi per accedere alle presenze.', 'danger')
        return redirect(url_for('dashboard.dashboard'))
    
    # Ottieni mese e anno dal parametro GET o usa mese corrente
    try:
        year = int(request.args.get('year', datetime.now().year))
        month = int(request.args.get('month', datetime.now().month))
    except ValueError:
        year = datetime.now().year
        month = datetime.now().month
    
    # Verifica che mese sia valido
    if month < 1 or month > 12:
        flash('Mese non valido', 'danger')
        return redirect(url_for('attendance.my_attendance'))
    
    # Ottieni o crea il timesheet mensile
    company_id = get_user_company_id()
    timesheet = MonthlyTimesheet.get_or_create(current_user.id, year, month, company_id)
    
    # Verifica se esiste una richiesta di riapertura pendente
    from models import TimesheetReopenRequest
    pending_reopen_request = None
    if timesheet.is_consolidated:
        pending_reopen_request = TimesheetReopenRequest.query.filter_by(
            timesheet_id=timesheet.id,
            status='Pending'
        ).first()
    
    # Calcola primo e ultimo giorno del mese
    from calendar import monthrange
    first_day = date(year, month, 1)
    last_day_num = monthrange(year, month)[1]
    last_day = date(year, month, last_day_num)
    
    # Ottieni tutte le sedi disponibili per l'utente
    user_sedi = []
    if hasattr(current_user, 'sedi') and current_user.sedi:
        # Utente multi-sede
        user_sedi = current_user.sedi
    elif current_user.sede_obj:
        # Utente singola sede
        user_sedi = [current_user.sede_obj]
    
    # Ottieni commesse attive per il periodo del mese visualizzato
    active_commesse = current_user.get_commesse_for_period(first_day, last_day)
    
    # Ottieni tipologie di presenza attive
    attendance_types = filter_by_company(AttendanceType.query).filter_by(active=True).order_by(AttendanceType.is_default.desc(), AttendanceType.name).all()
    default_type = next((t for t in attendance_types if t.is_default), attendance_types[0] if attendance_types else None)
    
    # Ottieni tutti gli eventi del mese per l'utente
    events_query = filter_by_company(AttendanceEvent.query).filter(
        AttendanceEvent.user_id == current_user.id,
        AttendanceEvent.date >= first_day,
        AttendanceEvent.date <= last_day
    ).order_by(AttendanceEvent.date, AttendanceEvent.timestamp).all()
    
    # Ottieni ferie/permessi approvati E in attesa nel mese
    leaves_query = filter_by_company(LeaveRequest.query).filter(
        LeaveRequest.user_id == current_user.id,
        LeaveRequest.status.in_(['Approved', 'Pending']),
        LeaveRequest.start_date <= last_day,
        LeaveRequest.end_date >= first_day
    ).all()
    
    # Crea set di date con ferie/permessi
    leave_dates = set()
    leave_info = {}
    for leave in leaves_query:
        current_date = leave.start_date
        while current_date <= leave.end_date:
            if first_day <= current_date <= last_day:
                leave_dates.add(current_date)
                leave_type_name = leave.leave_type_obj.name if leave.leave_type_obj else (leave.leave_type or 'Permesso')
                leave_info[current_date] = leave_type_name
            current_date += timedelta(days=1)
    
    # Organizza eventi per giorno
    events_by_day = {}
    for event in events_query:
        day = event.date.day
        if day not in events_by_day:
            events_by_day[day] = {
                'events': [],
                'has_manual': False,
                'has_live': False,
                'sede_id': event.sede_id,
                'commessa_id': event.commessa_id
            }
        events_by_day[day]['events'].append(event)
        if event.is_manual:
            events_by_day[day]['has_manual'] = True
            # Mantieni sede e commessa dal record manuale
            if event.sede_id:
                events_by_day[day]['sede_id'] = event.sede_id
            if event.commessa_id:
                events_by_day[day]['commessa_id'] = event.commessa_id
        else:
            events_by_day[day]['has_live'] = True
    
    # Crea lista giorni del mese con info
    from zoneinfo import ZoneInfo
    italy_tz = ZoneInfo('Europe/Rome')
    
    def to_italian_time_str(timestamp):
        """Converte timestamp a stringa HH:MM in orario italiano"""
        if not timestamp:
            return None
        if timestamp.tzinfo is None:
            utc_tz = ZoneInfo('UTC')
            timestamp = timestamp.replace(tzinfo=utc_tz)
        italian_time = timestamp.astimezone(italy_tz)
        return italian_time.strftime('%H:%M')
    
    days_data = []
    for day in range(1, last_day_num + 1):
        day_date = date(year, month, day)
        day_events = events_by_day.get(day, {
            'events': [],
            'has_manual': False,
            'has_live': False,
            'sede_id': current_user.sede_id,
            'commessa_id': None
        })
        
        # Verifica se il giorno ha ferie/permessi
        has_leave = day_date in leave_dates
        leave_type = leave_info.get(day_date, '')
        
        # Trova primo clock_in e ultimo clock_out
        clock_in_time = None
        clock_out_time = None
        break_start_time = None
        break_end_time = None
        attendance_type_id = default_type.id if default_type else None
        
        for event in day_events['events']:
            # Cattura attendance_type_id da qualsiasi evento che lo ha
            if event.attendance_type_id and attendance_type_id == (default_type.id if default_type else None):
                attendance_type_id = event.attendance_type_id
            
            if event.event_type == 'clock_in' and clock_in_time is None:
                clock_in_time = to_italian_time_str(event.timestamp)
            elif event.event_type == 'clock_out':
                clock_out_time = to_italian_time_str(event.timestamp)
            elif event.event_type == 'break_start' and break_start_time is None:
                break_start_time = to_italian_time_str(event.timestamp)
            elif event.event_type == 'break_end':
                break_end_time = to_italian_time_str(event.timestamp)
        
        # Nomi giorni in italiano
        italian_weekdays = ['Lun', 'Mar', 'Mer', 'Gio', 'Ven', 'Sab', 'Dom']
        weekday_italian = italian_weekdays[day_date.weekday()]
        
        # Se l'utente ha solo una commessa, usa quella come default
        default_commessa_id = day_events.get('commessa_id')
        if default_commessa_id is None and len(active_commesse) == 1:
            default_commessa_id = active_commesse[0].id
        
        days_data.append({
            'day': day,
            'date': day_date,
            'weekday': weekday_italian,
            'clock_in': clock_in_time,
            'clock_out': clock_out_time,
            'break_start': break_start_time,
            'break_end': break_end_time,
            'sede_id': day_events.get('sede_id', current_user.sede_id),
            'commessa_id': default_commessa_id,
            'attendance_type_id': attendance_type_id,
            'has_manual': day_events['has_manual'],
            'has_live': day_events['has_live'],
            'is_weekend': day_date.weekday() >= 5,
            'is_future': day_date > date.today(),
            'has_leave': has_leave,
            'leave_type': leave_type
        })
    
    # Nomi mesi in italiano
    month_names = ['', 'Gennaio', 'Febbraio', 'Marzo', 'Aprile', 'Maggio', 'Giugno',
                   'Luglio', 'Agosto', 'Settembre', 'Ottobre', 'Novembre', 'Dicembre']
    
    return render_template('my_attendance.html',
                         timesheet=timesheet,
                         year=year,
                         month=month,
                         month_name=month_names[month],
                         days_data=days_data,
                         user_sedi=user_sedi,
                         active_commesse=active_commesse,
                         attendance_types=attendance_types,
                         can_edit=timesheet.can_edit(),
                         pending_reopen_request=pending_reopen_request)

@attendance_bp.route('/my_attendance/save', methods=['POST'])
@login_required
def save_my_attendance():
    """Salva presenza giornaliera con sede e commessa"""
    if not current_user.can_access_attendance():
        return jsonify({'success': False, 'message': 'Permessi insufficienti'}), 403
    
    try:
        data = request.get_json()
        year = int(data.get('year'))
        month = int(data.get('month'))
        day = int(data.get('day'))
        clock_in_str = data.get('clock_in', '').strip()
        clock_out_str = data.get('clock_out', '').strip()
        break_start_str = data.get('break_start', '').strip()
        break_end_str = data.get('break_end', '').strip()
        sede_id = data.get('sede_id')
        commessa_id = data.get('commessa_id') if data.get('commessa_id') != 'none' else None
        attendance_type_id = data.get('attendance_type_id')
        
        # Valida attendance_type_id: deve appartenere alla company e essere attivo
        if attendance_type_id:
            attendance_type = filter_by_company(AttendanceType.query).filter_by(
                id=int(attendance_type_id),
                active=True
            ).first()
            if not attendance_type:
                return jsonify({'success': False, 'message': 'Tipologia di presenza non valida'}), 400
            attendance_type_id = attendance_type.id
        else:
            # Se non specificato, usa il default della company
            default_type = filter_by_company(AttendanceType.query).filter_by(
                is_default=True,
                active=True
            ).first()
            attendance_type_id = default_type.id if default_type else None
        
        # Ottieni il timesheet mensile
        company_id = get_user_company_id()
        timesheet = MonthlyTimesheet.get_or_create(current_user.id, year, month, company_id)
        
        # Verifica se può essere modificato
        if not timesheet.can_edit():
            return jsonify({'success': False, 'message': 'Timesheet consolidato, non modificabile'}), 400
        
        # Crea data
        day_date = date(year, month, day)
        
        # Blocca inserimenti futuri
        if day_date > date.today():
            return jsonify({'success': False, 'message': 'Non è possibile inserire orari per giorni futuri'}), 400
        
        # Blocca inserimenti su giorni con ferie/permessi
        existing_leave = filter_by_company(LeaveRequest.query).filter(
            LeaveRequest.user_id == current_user.id,
            LeaveRequest.status.in_(['Approved', 'Pending']),
            LeaveRequest.start_date <= day_date,
            LeaveRequest.end_date >= day_date
        ).first()
        
        if existing_leave:
            return jsonify({'success': False, 'message': 'Giorno con ferie/permesso/malattia'}), 400
        
        # Verifica che la sede selezionata sia valida per l'utente
        if sede_id:
            sede_id = int(sede_id)
            user_sede_ids = [s.id for s in (current_user.sedi if hasattr(current_user, 'sedi') and current_user.sedi else [current_user.sede_obj] if current_user.sede_obj else [])]
            if sede_id not in user_sede_ids:
                return jsonify({'success': False, 'message': 'Sede non valida per questo utente'}), 400
        
        # Se non ci sono orari, elimina le presenze manuali esistenti
        if not clock_in_str and not clock_out_str:
            filter_by_company(AttendanceEvent.query).filter_by(
                user_id=current_user.id,
                date=day_date,
                is_manual=True
            ).delete()
            db.session.commit()
            return jsonify({'success': True, 'message': 'Presenza eliminata'})
        
        # Validazione orari
        from datetime import datetime as dt
        if clock_in_str:
            try:
                dt.strptime(clock_in_str, '%H:%M')
            except ValueError:
                return jsonify({'success': False, 'message': 'Ora entrata non valida'}), 400
        
        if clock_out_str:
            try:
                dt.strptime(clock_out_str, '%H:%M')
            except ValueError:
                return jsonify({'success': False, 'message': 'Ora uscita non valida'}), 400
        
        # Elimina eventi manuali esistenti per questo giorno
        filter_by_company(AttendanceEvent.query).filter_by(
            user_id=current_user.id,
            date=day_date,
            is_manual=True
        ).delete()
        
        # Crea nuovi eventi
        from zoneinfo import ZoneInfo
        italy_tz = ZoneInfo('Europe/Rome')
        
        def create_timestamp(time_str):
            """Crea timestamp da stringa HH:MM in orario italiano"""
            hour, minute = map(int, time_str.split(':'))
            dt_italian = datetime(year, month, day, hour, minute, tzinfo=italy_tz)
            # Converti a UTC per storage
            dt_utc = dt_italian.astimezone(ZoneInfo('UTC'))
            return dt_utc.replace(tzinfo=None)  # Rimuovi tzinfo per storage naive
        
        if clock_in_str:
            clock_in_event = AttendanceEvent(
                user_id=current_user.id,
                event_type='clock_in',
                timestamp=create_timestamp(clock_in_str),
                date=day_date,
                is_manual=True,
                sede_id=sede_id,
                commessa_id=commessa_id,
                attendance_type_id=attendance_type_id,
                company_id=company_id
            )
            db.session.add(clock_in_event)
        
        if break_start_str:
            break_start_event = AttendanceEvent(
                user_id=current_user.id,
                event_type='break_start',
                timestamp=create_timestamp(break_start_str),
                date=day_date,
                is_manual=True,
                sede_id=sede_id,
                commessa_id=commessa_id,
                attendance_type_id=attendance_type_id,
                company_id=company_id
            )
            db.session.add(break_start_event)
        
        if break_end_str:
            break_end_event = AttendanceEvent(
                user_id=current_user.id,
                event_type='break_end',
                timestamp=create_timestamp(break_end_str),
                date=day_date,
                is_manual=True,
                sede_id=sede_id,
                commessa_id=commessa_id,
                attendance_type_id=attendance_type_id,
                company_id=company_id
            )
            db.session.add(break_end_event)
        
        if clock_out_str:
            clock_out_event = AttendanceEvent(
                user_id=current_user.id,
                event_type='clock_out',
                timestamp=create_timestamp(clock_out_str),
                date=day_date,
                is_manual=True,
                sede_id=sede_id,
                commessa_id=commessa_id,
                attendance_type_id=attendance_type_id,
                company_id=company_id
            )
            db.session.add(clock_out_event)
        
        db.session.commit()
        return jsonify({'success': True, 'message': 'Presenza salvata'})
        
    except Exception as e:
        db.session.rollback()
        import logging
        logging.error(f"Errore salvataggio presenza: {str(e)}")
        return jsonify({'success': False, 'message': f'Errore: {str(e)}'}), 500

@attendance_bp.route('/my_attendance/delete_day', methods=['POST'])
@login_required
def delete_my_attendance_day():
    """Elimina tutti i dati manuali di un giorno"""
    if not current_user.can_access_attendance():
        return jsonify({'success': False, 'message': 'Permessi insufficienti'}), 403
    
    try:
        data = request.get_json()
        year = int(data.get('year'))
        month = int(data.get('month'))
        day = int(data.get('day'))
        
        # Ottieni il timesheet mensile
        company_id = get_user_company_id()
        timesheet = MonthlyTimesheet.get_or_create(current_user.id, year, month, company_id)
        
        # Verifica se può essere modificato
        if not timesheet.can_edit():
            return jsonify({'success': False, 'message': 'Timesheet consolidato, non modificabile'}), 400
        
        day_date = date(year, month, day)
        
        # Elimina tutti gli eventi manuali del giorno
        deleted_count = filter_by_company(AttendanceEvent.query).filter_by(
            user_id=current_user.id,
            date=day_date,
            is_manual=True
        ).delete()
        
        db.session.commit()
        
        if deleted_count > 0:
            return jsonify({'success': True, 'message': f'Presenza del {day}/{month}/{year} eliminata'})
        else:
            return jsonify({'success': False, 'message': 'Nessun dato manuale da eliminare'}), 400
        
    except Exception as e:
        db.session.rollback()
        import logging
        logging.error(f"Errore eliminazione presenza: {str(e)}")
        return jsonify({'success': False, 'message': f'Errore: {str(e)}'}), 500

@attendance_bp.route('/my_attendance/bulk_fill', methods=['POST'])
@login_required
def bulk_fill_month():
    """Compilazione massiva del mese con orari standard"""
    if not current_user.can_access_attendance():
        return jsonify({'success': False, 'message': 'Permessi insufficienti'}), 403
    
    try:
        data = request.get_json()
        year = int(data.get('year'))
        month = int(data.get('month'))
        sede_id = int(data.get('sede_id'))
        
        # Normalizza commessa_id: converte "none" o valori vuoti a None, altrimenti cast a int
        commessa_id_raw = data.get('commessa_id')
        if commessa_id_raw is None or commessa_id_raw == 'none' or commessa_id_raw == '':
            commessa_id = None
        else:
            commessa_id = int(commessa_id_raw)
        
        clock_in_str = data.get('clock_in')
        clock_out_str = data.get('clock_out')
        break_start_str = data.get('break_start')
        break_end_str = data.get('break_end')
        
        # Ottieni il timesheet mensile
        company_id = get_user_company_id()
        timesheet = MonthlyTimesheet.get_or_create(current_user.id, year, month, company_id)
        
        # Verifica se può essere modificato
        if not timesheet.can_edit():
            return jsonify({'success': False, 'message': 'Timesheet consolidato, non modificabile'}), 400
        
        # Verifica che l'utente abbia accesso alla sede
        user_sedi_ids = []
        if hasattr(current_user, 'sedi') and current_user.sedi:
            user_sedi_ids = [s.id for s in current_user.sedi]
        elif current_user.sede_obj:
            user_sedi_ids = [current_user.sede_obj.id]
        
        if sede_id not in user_sedi_ids:
            return jsonify({'success': False, 'message': 'Non hai accesso a questa sede'}), 403
        
        # Helper per creare timestamp - deve creare datetime completo in orario italiano
        def create_timestamp(day_date, time_str):
            if not time_str:
                return None
            hour, minute = map(int, time_str.split(':'))
            # Crea datetime in orario italiano (come gli altri inserimenti manuali)
            from zoneinfo import ZoneInfo
            italy_tz = ZoneInfo('Europe/Rome')
            naive_dt = datetime.combine(day_date, time(hour, minute))
            # Localizza in timezone italiano
            italian_dt = naive_dt.replace(tzinfo=italy_tz)
            # Converti in UTC per il salvataggio (come fa il resto del sistema)
            utc_tz = ZoneInfo('UTC')
            utc_dt = italian_dt.astimezone(utc_tz)
            # Ritorna naive UTC (il database si aspetta naive UTC)
            return utc_dt.replace(tzinfo=None)
        
        # Determina quanti giorni ha il mese
        import calendar
        _, days_in_month = calendar.monthrange(year, month)
        
        filled_count = 0
        skipped_count = 0
        
        # Itera su tutti i giorni del mese
        for day in range(1, days_in_month + 1):
            day_date = date(year, month, day)
            
            # Salta sabato (5) e domenica (6)
            if day_date.weekday() in [5, 6]:
                continue
            
            # Controlla se ci sono eventi live per quel giorno
            live_events = filter_by_company(AttendanceEvent.query).filter_by(
                user_id=current_user.id,
                date=day_date,
                is_manual=False
            ).first()
            
            if live_events:
                skipped_count += 1
                continue
            
            # Controlla se ci sono eventi straordinari (ferie, permessi, malattie)
            leave_request = filter_by_company(LeaveRequest.query).filter(
                LeaveRequest.user_id == current_user.id,
                LeaveRequest.start_date <= day_date,
                LeaveRequest.end_date >= day_date,
                LeaveRequest.status == 'Approved'
            ).first()
            
            if leave_request:
                skipped_count += 1
                continue
            
            # Controlla se ci sono già dati manuali per quel giorno
            existing_manual = filter_by_company(AttendanceEvent.query).filter_by(
                user_id=current_user.id,
                date=day_date,
                is_manual=True
            ).first()
            
            if existing_manual:
                skipped_count += 1
                continue
            
            # Inserisci i dati per questo giorno
            if clock_in_str:
                clock_in_event = AttendanceEvent(
                    user_id=current_user.id,
                    event_type='clock_in',
                    timestamp=create_timestamp(day_date, clock_in_str),
                    date=day_date,
                    is_manual=True,
                    sede_id=sede_id,
                    commessa_id=commessa_id,
                    company_id=company_id
                )
                db.session.add(clock_in_event)
            
            if break_start_str:
                break_start_event = AttendanceEvent(
                    user_id=current_user.id,
                    event_type='break_start',
                    timestamp=create_timestamp(day_date, break_start_str),
                    date=day_date,
                    is_manual=True,
                    sede_id=sede_id,
                    commessa_id=commessa_id,
                    company_id=company_id
                )
                db.session.add(break_start_event)
            
            if break_end_str:
                break_end_event = AttendanceEvent(
                    user_id=current_user.id,
                    event_type='break_end',
                    timestamp=create_timestamp(day_date, break_end_str),
                    date=day_date,
                    is_manual=True,
                    sede_id=sede_id,
                    commessa_id=commessa_id,
                    company_id=company_id
                )
                db.session.add(break_end_event)
            
            if clock_out_str:
                clock_out_event = AttendanceEvent(
                    user_id=current_user.id,
                    event_type='clock_out',
                    timestamp=create_timestamp(day_date, clock_out_str),
                    date=day_date,
                    is_manual=True,
                    sede_id=sede_id,
                    commessa_id=commessa_id,
                    company_id=company_id
                )
                db.session.add(clock_out_event)
            
            filled_count += 1
        
        db.session.commit()
        
        message = f'Compilazione completata: {filled_count} giorni compilati'
        if skipped_count > 0:
            message += f', {skipped_count} giorni saltati (già compilati o con eventi straordinari)'
        
        return jsonify({'success': True, 'message': message})
        
    except Exception as e:
        db.session.rollback()
        import logging
        logging.error(f"Errore compilazione massiva: {str(e)}")
        return jsonify({'success': False, 'message': f'Errore: {str(e)}'}), 500

@attendance_bp.route('/my_attendance/consolidate', methods=['POST'])
@login_required
def consolidate_timesheet():
    """Consolida il timesheet mensile dopo aver verificato la completezza"""
    if not current_user.can_access_attendance():
        return jsonify({'success': False, 'message': 'Permessi insufficienti'}), 403
    
    try:
        data = request.get_json()
        year = int(data.get('year'))
        month = int(data.get('month'))
        
        # Ottieni il timesheet mensile
        company_id = get_user_company_id()
        timesheet = MonthlyTimesheet.get_or_create(current_user.id, year, month, company_id)
        
        # Verifica se è già consolidato
        if timesheet.is_consolidated:
            return jsonify({'success': False, 'message': 'Timesheet già consolidato'}), 400
        
        # Verifica che tutti i giorni lavorativi siano compilati
        import calendar
        _, days_in_month = calendar.monthrange(year, month)
        
        missing_days = []
        
        for day in range(1, days_in_month + 1):
            day_date = date(year, month, day)
            
            # Salta sabato (5) e domenica (6)
            if day_date.weekday() in [5, 6]:
                continue
            
            # Controlla se ci sono eventi straordinari (ferie, permessi, malattie)
            leave_request = filter_by_company(LeaveRequest.query).filter(
                LeaveRequest.user_id == current_user.id,
                LeaveRequest.start_date <= day_date,
                LeaveRequest.end_date >= day_date,
                LeaveRequest.status == 'Approved'
            ).first()
            
            # Se c'è un leave_request approvato, salta questo giorno
            if leave_request:
                continue
            
            # Controlla se ci sono eventi di presenza per quel giorno (clock_in o clock_out)
            attendance_events = filter_by_company(AttendanceEvent.query).filter(
                AttendanceEvent.user_id == current_user.id,
                AttendanceEvent.date == day_date,
                AttendanceEvent.event_type.in_(['clock_in', 'clock_out'])
            ).count()
            
            if attendance_events < 2:  # Deve avere almeno entrata e uscita
                missing_days.append(day)
        
        if missing_days:
            missing_str = ', '.join(map(str, missing_days))
            return jsonify({
                'success': False, 
                'message': f'Impossibile consolidare: mancano dati per i giorni {missing_str}. Completa tutte le giornate lavorative prima di consolidare.'
            }), 400
        
        # Consolida il timesheet
        timesheet.is_consolidated = True
        timesheet.consolidated_at = datetime.now()
        timesheet.consolidated_by_id = current_user.id
        
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Timesheet consolidato con successo'})
        
    except Exception as e:
        db.session.rollback()
        import logging
        logging.error(f"Errore consolidamento timesheet: {str(e)}")
        return jsonify({'success': False, 'message': f'Errore: {str(e)}'}), 500
# =============================================================================
# TIMESHEETS MANAGEMENT - LISTA E VISUALIZZAZIONE STORICA
# =============================================================================

@attendance_bp.route('/timesheets', methods=['GET'])
@login_required
def timesheets():
    """Visualizza la lista di tutti i timesheets dell'utente"""
    if not current_user.can_access_timesheets():
        flash('Non hai i permessi per accedere ai timesheets', 'danger')
        return redirect(url_for('dashboard.dashboard'))
    
    from models import MonthlyTimesheet
    from datetime import datetime
    
    # Ottieni filtro anno (default: anno corrente)
    try:
        year_filter = int(request.args.get('year', datetime.now().year))
    except ValueError:
        year_filter = datetime.now().year
    
    company_id = get_user_company_id()
    
    # Ottieni tutti i timesheets dell'utente per l'anno selezionato
    timesheets_query = MonthlyTimesheet.query.filter_by(
        user_id=current_user.id,
        company_id=company_id,
        year=year_filter
    ).order_by(MonthlyTimesheet.month.desc()).all()
    
    # Ottieni tutti gli anni disponibili per il filtro
    available_years_query = db.session.query(MonthlyTimesheet.year).filter_by(
        user_id=current_user.id,
        company_id=company_id
    ).distinct().order_by(MonthlyTimesheet.year.desc()).all()
    available_years = [y[0] for y in available_years_query]
    
    # Se non ci sono anni disponibili, usa l'anno corrente
    if not available_years:
        available_years = [datetime.now().year]
    
    # Nomi mesi in italiano
    month_names = ['', 'Gennaio', 'Febbraio', 'Marzo', 'Aprile', 'Maggio', 'Giugno',
                   'Luglio', 'Agosto', 'Settembre', 'Ottobre', 'Novembre', 'Dicembre']
    
    return render_template('timesheets.html',
                         timesheets=timesheets_query,
                         year_filter=year_filter,
                         available_years=available_years,
                         month_names=month_names)

@attendance_bp.route('/timesheets/export/<int:timesheet_id>', methods=['GET'])
@login_required
def export_timesheet(timesheet_id):
    """Esporta un timesheet in formato Excel"""
    if not current_user.can_access_timesheets():
        flash('Non hai i permessi per accedere ai timesheets', 'danger')
        return redirect(url_for('dashboard.dashboard'))
    
    try:
        from models import MonthlyTimesheet, AttendanceEvent
        from datetime import date
        from io import BytesIO
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from flask import send_file
        import calendar
        
        # Ottieni il timesheet
        timesheet = filter_by_company(MonthlyTimesheet.query).filter_by(
            id=timesheet_id,
            user_id=current_user.id
        ).first()
        if not timesheet:
            flash('Timesheet non trovato', 'danger')
            return redirect(url_for('attendance.timesheets'))
        
        # Crea workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Timesheet"
        
        # Stili
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Intestazione
        ws['A1'] = f"Timesheet - {current_user.get_full_name()}"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A2'] = f"Mese: {calendar.month_name[timesheet.month]} {timesheet.year}"
        ws['A3'] = f"Stato: {timesheet.get_status()}"
        
        # Riga vuota
        current_row = 5
        
        # Header tabella
        headers = ['Giorno', 'Data', 'Entrata', 'Uscita', 'Inizio Pausa', 'Fine Pausa', 'Ore Lavorate', 'Sede', 'Commessa']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        
        current_row += 1
        
        # Dati
        month_days = calendar.monthrange(timesheet.year, timesheet.month)[1]
        italian_weekdays = ['Lun', 'Mar', 'Mer', 'Gio', 'Ven', 'Sab', 'Dom']
        
        for day in range(1, month_days + 1):
            day_date = date(timesheet.year, timesheet.month, day)
            weekday = italian_weekdays[day_date.weekday()]
            
            # Ottieni eventi del giorno
            events = filter_by_company(AttendanceEvent.query).filter(
                AttendanceEvent.user_id == current_user.id,
                AttendanceEvent.date == day_date
            ).order_by(AttendanceEvent.timestamp).all()
            
            if events:
                clock_in = None
                clock_out = None
                break_start = None
                break_end = None
                sede_name = ""
                commessa_name = ""
                
                for event in events:
                    if event.event_type == 'clock_in' and clock_in is None:
                        clock_in = event.timestamp.strftime('%H:%M') if event.timestamp else ''
                        if event.sede:
                            sede_name = event.sede.name
                        if event.commessa:
                            commessa_name = f"{event.commessa.titolo} - {event.commessa.cliente}"
                    elif event.event_type == 'clock_out':
                        clock_out = event.timestamp.strftime('%H:%M') if event.timestamp else ''
                    elif event.event_type == 'break_start' and break_start is None:
                        break_start = event.timestamp.strftime('%H:%M') if event.timestamp else ''
                    elif event.event_type == 'break_end':
                        break_end = event.timestamp.strftime('%H:%M') if event.timestamp else ''
                
                # Calcola ore lavorate
                ore_lavorate = ""
                if clock_in and clock_out:
                    try:
                        from datetime import datetime
                        in_time = datetime.strptime(clock_in, '%H:%M')
                        out_time = datetime.strptime(clock_out, '%H:%M')
                        total_minutes = (out_time - in_time).seconds // 60
                        
                        # Sottrai pausa
                        if break_start and break_end:
                            break_start_time = datetime.strptime(break_start, '%H:%M')
                            break_end_time = datetime.strptime(break_end, '%H:%M')
                            break_minutes = (break_end_time - break_start_time).seconds // 60
                            total_minutes -= break_minutes
                        
                        hours = total_minutes // 60
                        minutes = total_minutes % 60
                        ore_lavorate = f"{hours}:{minutes:02d}"
                    except:
                        ore_lavorate = ""
                
                ws.cell(row=current_row, column=1, value=weekday).border = border
                ws.cell(row=current_row, column=2, value=day_date.strftime('%d/%m/%Y')).border = border
                ws.cell(row=current_row, column=3, value=clock_in).border = border
                ws.cell(row=current_row, column=4, value=clock_out).border = border
                ws.cell(row=current_row, column=5, value=break_start).border = border
                ws.cell(row=current_row, column=6, value=break_end).border = border
                ws.cell(row=current_row, column=7, value=ore_lavorate).border = border
                ws.cell(row=current_row, column=8, value=sede_name).border = border
                ws.cell(row=current_row, column=9, value=commessa_name).border = border
                
                current_row += 1
        
        # Adatta larghezza colonne
        for col in range(1, 10):
            ws.column_dimensions[chr(64 + col)].width = 15
        
        # Salva in BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"timesheet_{timesheet.year}_{timesheet.month:02d}_{current_user.username}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        import logging
        logging.error(f"Errore export timesheet: {str(e)}")
        flash(f'Errore durante l\'export: {str(e)}', 'danger')
        return redirect(url_for('attendance.timesheets'))


# =============================================================================
# ATTENDANCE TYPES MANAGEMENT ROUTES
# =============================================================================

@attendance_bp.route('/types')
@login_required
def attendance_types():
    """Visualizza elenco tipologie di presenza"""
    if not current_user.can_manage_attendance():
        flash('Non hai i permessi per gestire le tipologie di presenza', 'danger')
        return redirect(url_for('dashboard.dashboard'))
    
    types = filter_by_company(AttendanceType.query).order_by(AttendanceType.name).all()
    return render_template('attendance_types.html', types=types)


@attendance_bp.route('/types/create', methods=['GET', 'POST'])
@login_required
def create_attendance_type():
    """Crea nuova tipologia di presenza"""
    if not current_user.can_manage_attendance():
        flash('Non hai i permessi per creare tipologie di presenza', 'danger')
        return redirect(url_for('attendance.attendance_types'))
    
    from forms import AttendanceTypeForm
    
    form = AttendanceTypeForm()
    
    if form.validate_on_submit():
        # Se questa è impostata come default, rimuovi il flag da tutte le altre
        if form.is_default.data:
            existing_defaults = filter_by_company(AttendanceType.query).filter_by(is_default=True).all()
            for existing in existing_defaults:
                existing.is_default = False
        
        type_obj = AttendanceType(
            code=form.code.data.strip().upper(),
            name=form.name.data,
            description=form.description.data,
            is_default=form.is_default.data,
            active=form.active.data,
            created_by=current_user.id
        )
        set_company_on_create(type_obj)
        
        db.session.add(type_obj)
        
        try:
            db.session.commit()
            flash('Tipologia di presenza creata con successo', 'success')
            return redirect(url_for('attendance.attendance_types'))
        except:
            db.session.rollback()
            flash('Errore nella creazione della tipologia', 'danger')
    
    return render_template('create_attendance_type.html', form=form)


@attendance_bp.route('/types/<int:type_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_attendance_type(type_id):
    """Modifica tipologia di presenza"""
    if not current_user.can_manage_attendance():
        flash('Non hai i permessi per modificare le tipologie di presenza', 'danger')
        return redirect(url_for('attendance.attendance_types'))
    
    from forms import AttendanceTypeForm
    
    type_obj = filter_by_company(AttendanceType.query).filter_by(id=type_id).first_or_404()
    form = AttendanceTypeForm(obj=type_obj)
    
    if form.validate_on_submit():
        # Se questa è impostata come default, rimuovi il flag da tutte le altre
        if form.is_default.data and not type_obj.is_default:
            existing_defaults = filter_by_company(AttendanceType.query).filter_by(is_default=True).all()
            for existing in existing_defaults:
                existing.is_default = False
        
        type_obj.code = form.code.data.strip().upper()
        type_obj.name = form.name.data
        type_obj.description = form.description.data
        type_obj.is_default = form.is_default.data
        type_obj.active = form.active.data
        
        try:
            db.session.commit()
            flash('Tipologia modificata con successo', 'success')
            return redirect(url_for('attendance.attendance_types'))
        except:
            db.session.rollback()
            flash('Errore nella modifica della tipologia', 'danger')
    
    return render_template('edit_attendance_type.html', form=form, type_obj=type_obj)


@attendance_bp.route('/types/<int:type_id>/delete', methods=['POST'])
@login_required
def delete_attendance_type(type_id):
    """Elimina tipologia di presenza"""
    if not current_user.can_manage_attendance():
        flash('Non hai i permessi per eliminare le tipologie di presenza', 'danger')
        return redirect(url_for('attendance.attendance_types'))
    
    type_obj = filter_by_company(AttendanceType.query).filter_by(id=type_id).first_or_404()
    
    # Verifica se ci sono eventi di presenza associati
    if type_obj.attendance_events and len(type_obj.attendance_events) > 0:
        flash('Non è possibile eliminare una tipologia con eventi di presenza associati', 'warning')
        return redirect(url_for('attendance.attendance_types'))
    
    try:
        name = type_obj.name
        db.session.delete(type_obj)
        db.session.commit()
        flash(f'Tipologia "{name}" eliminata con successo', 'success')
    except:
        db.session.rollback()
        flash('Errore nell\'eliminazione della tipologia', 'danger')
    
    return redirect(url_for('attendance.attendance_types'))
