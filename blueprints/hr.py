# =============================================================================
# HR (HUMAN RESOURCES) BLUEPRINT
# =============================================================================
#
# ROUTES INCLUSE:
# 1. hr_list (GET) - Lista dipendenti con dati HR
# 2. hr_detail (GET/POST) - Visualizza/modifica scheda HR completa
# 3. hr_export (GET) - Export Excel dati HR
#
# Total routes: 3
# =============================================================================

from flask import Blueprint, request, render_template, redirect, url_for, flash, make_response, session
from flask_login import login_required, current_user
from datetime import datetime, date
from functools import wraps
from app import db
from models import User, UserHRData, ACITable, Sede, WorkSchedule
from utils_tenant import filter_by_company, set_company_on_create
from utils_hr import assign_cod_si
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os
from werkzeug.utils import secure_filename

# Create blueprint
hr_bp = Blueprint('hr', __name__, url_prefix='/hr')

# Helper functions
def require_hr_permission(f):
    """Decorator per richiedere permessi HR"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated:
            return redirect(url_for('auth.login'))
        if not current_user.can_access_hr_menu():
            flash('Non hai i permessi per accedere a questa sezione', 'error')
            return redirect(url_for('dashboard.dashboard'))
        return f(*args, **kwargs)
    return decorated_function

def get_expiry_status(expiry_date):
    """
    Calcola lo stato di scadenza in base ai giorni rimanenti
    Returns: dict con 'status' (ok/warning/danger/expired), 'color', 'days_left', 'text'
    """
    if not expiry_date:
        return None
    
    today = date.today()
    days_left = (expiry_date - today).days
    
    if days_left < 0:
        # Scaduta
        return {
            'status': 'expired',
            'color': 'danger',
            'icon': 'fa-times-circle',
            'days_left': days_left,
            'text': f'Scaduta da {abs(days_left)} giorni'
        }
    elif days_left <= 30:
        # < 1 mese - arancione
        return {
            'status': 'urgent',
            'color': 'warning',
            'icon': 'fa-exclamation-triangle',
            'days_left': days_left,
            'text': f'Scade tra {days_left} giorni'
        }
    elif days_left <= 90:
        # < 3 mesi - giallo
        return {
            'status': 'warning',
            'color': 'orange',
            'icon': 'fa-exclamation-circle',
            'days_left': days_left,
            'text': f'Scade tra {days_left} giorni'
        }
    else:
        # > 3 mesi - verde
        return {
            'status': 'ok',
            'color': 'success',
            'icon': 'fa-check-circle',
            'days_left': days_left,
            'text': f'Valida per {days_left} giorni'
        }

def get_worst_expiry_status(hr_data):
    """
    Ottiene lo stato peggiore tra tutte le scadenze di un dipendente
    Returns: dict con info sulla scadenza più critica
    """
    if not hr_data:
        return None
    
    expiries = [
        ('Visita Medica', hr_data.medical_visit_expiry),
        ('Form. Generale', hr_data.training_general_expiry),
        ('RSPP', hr_data.training_rspp_expiry),
        ('RLS', hr_data.training_rls_expiry),
        ('Primo Soccorso', hr_data.training_first_aid_expiry),
        ('Emergenza', hr_data.training_emergency_expiry),
        ('Preposto', hr_data.training_supervisor_expiry),
    ]
    
    worst_status = None
    worst_priority = 999  # ok=3, warning=2, urgent=1, expired=0
    worst_name = None
    
    priority_map = {
        'expired': 0,
        'urgent': 1,
        'warning': 2,
        'ok': 3
    }
    
    for name, expiry_date in expiries:
        status = get_expiry_status(expiry_date)
        if status:
            priority = priority_map.get(status['status'], 999)
            if priority < worst_priority:
                worst_priority = priority
                worst_status = status
                worst_name = name
    
    if worst_status:
        worst_status['name'] = worst_name
    
    return worst_status

# =============================================================================
# HR ROUTES
# =============================================================================

@hr_bp.route('/', methods=['GET', 'POST'])
@login_required
@require_hr_permission
def hr_list():
    """Lista dipendenti con dati HR con filtri personalizzabili"""
    
    # Gestione POST - salva configurazione filtri e mostra lista filtrata
    if request.method == 'POST':
        action = request.form.get('action')
        
        # Salva campi selezionati in session
        selected_fields = request.form.getlist('fields')
        session['hr_export_fields'] = selected_fields if selected_fields else None
        
        # Salva filtri in session
        filters = {}
        for key in request.form.keys():
            if key.startswith('filter_') and request.form.get(key):
                filter_name = key.replace('filter_', '')
                filters[filter_name] = request.form.get(key)
        
        session['hr_export_filters'] = filters if filters else None
        
        # Se azione è export, reindirizza all'export
        if action == 'export':
            return redirect(url_for('hr.hr_export'))
        
        # Altrimenti mostra lista filtrata (non fare redirect, gestisci tutto nel POST)
        if filters:
            flash(f'Filtri applicati: {len(filters)} attivi', 'success')
        active_filters = filters if filters else None
    else:
        # GET: azzera sempre i filtri (nuova visita alla pagina)
        session.pop('hr_export_filters', None)
        session.pop('hr_export_fields', None)
        active_filters = None
    
    # Ottieni tutti gli utenti con dati HR (esclusi admin di sistema e ruolo Admin/Amministratore)
    if current_user.can_view_hr_data() or current_user.can_manage_hr_data():
        # HR Manager vedono tutti i dipendenti della company (esclusi admin)
        users_query = filter_by_company(User.query).filter_by(active=True).filter(
            User.role != 'Admin',
            User.role != 'ADMIN',
            User.role != 'Amministratore',
            User.is_system_admin == False
        ).order_by(User.last_name, User.first_name)
    else:
        # Utenti normali vedono solo i propri dati (con filtro company), esclusi se admin
        users_query = filter_by_company(User.query).filter_by(id=current_user.id).filter(
            User.role != 'Admin',
            User.role != 'ADMIN',
            User.role != 'Amministratore',
            User.is_system_admin == False
        )
    
    users = users_query.all()
    
    # Crea dizionario con dati HR per ogni utente
    users_data = []
    for user in users:
        hr_data = user.hr_data
        age = hr_data.get_age() if hr_data else None
        
        # Verifica se il contratto è attivo
        is_contract_active = hr_data.is_contract_active() if hr_data else False
        
        # Calcola stato scadenze
        expiry_status = get_worst_expiry_status(hr_data)
        
        # Calcola giorni rimanenti contratto (per TD e Distacco)
        # SOLO per contratti ATTIVI, altrimenti non è una criticità da gestire
        days_until_contract_end = None
        contract_expiring = False
        if hr_data and hr_data.contract_end_date and is_contract_active:
            days_until_contract_end = hr_data.days_until_contract_end()
            # Contratto in scadenza se è attivo E mancano 60 giorni o meno (ma almeno 1 giorno)
            contract_expiring = days_until_contract_end is not None and 0 < days_until_contract_end <= 60
        
        # Determina se ha certificazioni scadute o in scadenza
        # SOLO per contratti ATTIVI
        has_expired_certs = False
        has_expiring_certs = False
        if expiry_status and is_contract_active:
            if expiry_status['status'] == 'expired':
                has_expired_certs = True
            elif expiry_status['status'] in ['urgent', 'warning']:
                has_expiring_certs = True
        
        users_data.append({
            'user': user,
            'hr_data': hr_data,  # Pass the actual object so template can access sede relationship
            'has_data': hr_data is not None,
            'matricola': hr_data.matricola if hr_data else '-',
            'contract_type': hr_data.contract_type if hr_data else '-',
            'hire_date': hr_data.hire_date.strftime('%d/%m/%Y') if hr_data and hr_data.hire_date else '-',
            'contract_status': 'Attivo' if is_contract_active else 'Non attivo' if hr_data else '-',
            'is_probation': hr_data.is_probation_period() if hr_data else False,
            'gender': hr_data.gender if hr_data else None,
            'age': age,
            'birth_city': hr_data.birth_city if hr_data else None,
            'expiry_status': expiry_status,
            'days_until_contract_end': days_until_contract_end,
            'contract_expiring': contract_expiring,
            'has_expired_certs': has_expired_certs,
            'has_expiring_certs': has_expiring_certs,
        })
    
    # Applica filtri personalizzati PRIMA di calcolare le statistiche
    if active_filters:
        filtered_data = []
        for data in users_data:
            user = data['user']
            hr_data = data['hr_data']
            
            # Se ci sono filtri attivi ma l'utente non ha dati HR, escludilo
            if not hr_data:
                continue
            
            matches = True
            
            # Filtro matricola
            if 'matricola' in active_filters:
                if not hr_data.matricola or active_filters['matricola'].lower() not in hr_data.matricola.lower():
                    matches = False
            
            # Filtro data assunzione (filtra per anno)
            if 'hire_date' in active_filters:
                filter_date_str = active_filters['hire_date']
                try:
                    from datetime import datetime
                    filter_date = datetime.strptime(filter_date_str, '%Y-%m-%d').date()
                    if not hr_data.hire_date or hr_data.hire_date.year != filter_date.year:
                        matches = False
                except:
                    pass  # Se parsing fallisce, ignora il filtro
            
            # Filtro tipo contratto
            if 'contract_type' in active_filters:
                if not hr_data.contract_type or hr_data.contract_type != active_filters['contract_type']:
                    matches = False
            
            # Filtro CCNL
            if 'ccnl' in active_filters:
                if not hr_data.ccnl or active_filters['ccnl'].lower() not in hr_data.ccnl.lower():
                    matches = False
            
            # Filtro mansione
            if 'mansione' in active_filters:
                if not hr_data.mansione or active_filters['mansione'].lower() not in hr_data.mansione.lower():
                    matches = False
            
            # Filtro qualifica
            if 'qualifica' in active_filters:
                if not hr_data.qualifica or active_filters['qualifica'].lower() not in hr_data.qualifica.lower():
                    matches = False
            
            # Filtro sede assunzione
            if 'sede_assunzione' in active_filters:
                if not hr_data.sede_id or str(hr_data.sede_id) != str(active_filters['sede_assunzione']):
                    matches = False
            
            # Filtro genere
            if 'gender' in active_filters:
                if not hr_data.gender or hr_data.gender != active_filters['gender']:
                    matches = False
            
            # Filtro città
            if 'city' in active_filters:
                if not hr_data.city or active_filters['city'].lower() not in hr_data.city.lower():
                    matches = False
            
            if matches:
                filtered_data.append(data)
        
        users_data = filtered_data
    
    # Ricalcola statistiche sui dati filtrati - SOLO CONTRATTI ATTIVI
    # I dipendenti con contratti non attivi sono solo storico, non vanno conteggiati
    active_employees_data = [d for d in users_data if d['contract_status'] == 'Attivo']
    
    total_employees = len(active_employees_data)
    with_hr_data = sum(1 for d in active_employees_data if d['has_data'])
    active_contracts = len(active_employees_data)  # Tutti sono attivi per definizione
    in_probation = sum(1 for d in active_employees_data if d['is_probation'])
    
    female_count = sum(1 for d in active_employees_data if d['gender'] == 'F')
    male_count = sum(1 for d in active_employees_data if d['gender'] == 'M')
    under_36 = sum(1 for d in active_employees_data if d['age'] is not None and d['age'] < 36)
    
    tempo_indeterminato = sum(1 for d in active_employees_data if d['contract_type'] == 'Tempo Indeterminato')
    tempo_determinato = sum(1 for d in active_employees_data if d['contract_type'] == 'Tempo Determinato')
    
    # Statistiche criticità (già filtrate per contratti attivi nella logica precedente)
    contracts_expiring = sum(1 for d in active_employees_data if d['contract_expiring'])
    expired_certifications = sum(1 for d in active_employees_data if d['has_expired_certs'])
    expiring_certifications = sum(1 for d in active_employees_data if d['has_expiring_certs'])
    
    statistics = {
        'total_employees': total_employees,
        'with_hr_data': with_hr_data,
        'active_contracts': active_contracts,
        'in_probation': in_probation,
        'female_count': female_count,
        'male_count': male_count,
        'under_36': under_36,
        'tempo_indeterminato': tempo_indeterminato,
        'tempo_determinato': tempo_determinato,
        'contracts_expiring': contracts_expiring,
        'expired_certifications': expired_certifications,
        'expiring_certifications': expiring_certifications,
    }
    
    # Ottieni lista sedi per filtri
    all_sedi = filter_by_company(Sede.query).order_by(Sede.name).all()
    
    return render_template('hr_list.html', 
                         users_data=users_data, 
                         statistics=statistics,
                         all_sedi=all_sedi,
                         sedi=all_sedi,  # Backward compatibility
                         active_filters=active_filters)


@hr_bp.route('/reset-filters')
@login_required
@require_hr_permission
def hr_reset_filters():
    """Resetta tutti i filtri HR attivi"""
    session.pop('hr_export_filters', None)
    session.pop('hr_export_fields', None)
    flash('Filtri resettati con successo', 'success')
    return redirect(url_for('hr.hr_list'))


@hr_bp.route('/detail/<int:user_id>', methods=['GET', 'POST'])
@login_required
@require_hr_permission
def hr_detail(user_id):
    """Visualizza/modifica scheda HR completa di un utente"""
    
    # Ottieni l'utente
    user = filter_by_company(User.query).filter_by(id=user_id).first_or_404()
    
    # Controllo permessi
    can_edit = current_user.can_manage_hr_data()
    can_view_all = current_user.can_view_hr_data() or current_user.can_manage_hr_data()
    
    # Se non può vedere tutti i dati, può vedere solo i propri
    if not can_view_all and user.id != current_user.id:
        flash('Non hai i permessi per visualizzare questi dati', 'error')
        return redirect(url_for('hr.hr_list'))
    
    # Ottieni record HR esistente
    hr_data = user.hr_data
    
    # Crea record HR solo se l'utente ha permessi di modifica
    if not hr_data and can_edit:
        hr_data = UserHRData(user_id=user.id, company_id=user.company_id)
        set_company_on_create(hr_data)
        db.session.add(hr_data)
        
        # Assegna COD SI auto-incrementato
        try:
            assign_cod_si(hr_data)
        except Exception as e:
            db.session.rollback()
            flash(f'Errore durante l\'assegnazione del COD SI: {str(e)}', 'error')
            return redirect(url_for('hr.hr_list'))
        
        db.session.commit()
    elif not hr_data:
        # Per utenti read-only, crea oggetto vuoto non persistito
        # in modo che il template possa accedere agli attributi senza errori
        hr_data = UserHRData(user_id=user.id, company_id=user.company_id)
    
    if request.method == 'POST' and can_edit:
        try:
            # Aggiorna dati anagrafici
            hr_data.matricola = request.form.get('matricola', '').strip() or None
            hr_data.codice_fiscale = request.form.get('codice_fiscale', '').strip().upper() or None
            hr_data.gender = request.form.get('gender', '') or None
            
            # Data nascita
            birth_date_str = request.form.get('birth_date', '').strip()
            if birth_date_str:
                hr_data.birth_date = datetime.strptime(birth_date_str, '%Y-%m-%d').date()
            else:
                hr_data.birth_date = None
            
            hr_data.birth_city = request.form.get('birth_city', '').strip() or None
            hr_data.birth_province = request.form.get('birth_province', '').strip().upper() or None
            hr_data.birth_country = request.form.get('birth_country', '').strip() or 'Italia'
            
            # Residenza e contatti
            hr_data.address = request.form.get('address', '').strip() or None
            hr_data.city = request.form.get('city', '').strip() or None
            hr_data.province = request.form.get('province', '').strip().upper() or None
            hr_data.postal_code = request.form.get('postal_code', '').strip() or None
            hr_data.country = request.form.get('country', '').strip() or 'Italia'
            hr_data.alternative_domicile = request.form.get('alternative_domicile', '').strip() or None
            hr_data.phone = request.form.get('phone', '').strip() or None
            hr_data.law_104_benefits = request.form.get('law_104_benefits') == 'on'
            hr_data.personal_email = request.form.get('personal_email', '').strip() or None
            
            # Dati contrattuali
            hr_data.contract_type = request.form.get('contract_type', '').strip() or None
            hr_data.distacco_supplier = request.form.get('distacco_supplier', '').strip() or None
            hr_data.consulente_vat = request.form.get('consulente_vat', '').strip() or None
            hr_data.nome_fornitore = request.form.get('nome_fornitore', '').strip() or None
            hr_data.partita_iva_fornitore = request.form.get('partita_iva_fornitore', '').strip() or None
            
            hire_date_str = request.form.get('hire_date', '').strip()
            if hire_date_str:
                hr_data.hire_date = datetime.strptime(hire_date_str, '%Y-%m-%d').date()
            else:
                hr_data.hire_date = None
            
            contract_start_str = request.form.get('contract_start_date', '').strip()
            if contract_start_str:
                hr_data.contract_start_date = datetime.strptime(contract_start_str, '%Y-%m-%d').date()
            else:
                hr_data.contract_start_date = None
            
            contract_end_str = request.form.get('contract_end_date', '').strip()
            if contract_end_str:
                hr_data.contract_end_date = datetime.strptime(contract_end_str, '%Y-%m-%d').date()
            else:
                hr_data.contract_end_date = None
            
            probation_end_str = request.form.get('probation_end_date', '').strip()
            if probation_end_str:
                hr_data.probation_end_date = datetime.strptime(probation_end_str, '%Y-%m-%d').date()
            else:
                hr_data.probation_end_date = None
            
            hr_data.ccnl = request.form.get('ccnl', '').strip() or None
            hr_data.ccnl_level = request.form.get('ccnl_level', '').strip() or None
            hr_data.mansione = request.form.get('mansione', '').strip() or None
            hr_data.qualifica = request.form.get('qualifica', '').strip() or None
            hr_data.rischio_inail = request.form.get('rischio_inail', '').strip() or None
            hr_data.tipo_assunzione = request.form.get('tipo_assunzione', '').strip() or None
            hr_data.ticket_restaurant = request.form.get('ticket_restaurant') == 'on'
            hr_data.other_notes = request.form.get('other_notes', '').strip() or None
            
            work_hours_str = request.form.get('work_hours_week', '').strip()
            if work_hours_str:
                hr_data.work_hours_week = float(work_hours_str.replace(',', '.'))
            else:
                hr_data.work_hours_week = None
            
            # Orario (FT/PT)
            hr_data.working_time_type = request.form.get('working_time_type', '').strip() or None
            
            part_time_percentage_str = request.form.get('part_time_percentage', '').strip()
            if part_time_percentage_str:
                hr_data.part_time_percentage = float(part_time_percentage_str.replace(',', '.'))
            else:
                hr_data.part_time_percentage = None
            
            hr_data.part_time_type = request.form.get('part_time_type', '').strip() or None
            
            superminimo_str = request.form.get('superminimo', '').strip()
            if superminimo_str:
                hr_data.superminimo = float(superminimo_str.replace(',', '.'))
            else:
                hr_data.superminimo = None
            
            rimborsi_str = request.form.get('rimborsi_diarie', '').strip()
            if rimborsi_str:
                hr_data.rimborsi_diarie = float(rimborsi_str.replace(',', '.'))
            else:
                hr_data.rimborsi_diarie = None
            
            # Dati economici
            gross_salary_str = request.form.get('gross_salary', '').strip()
            if gross_salary_str:
                hr_data.gross_salary = float(gross_salary_str.replace(',', '.'))
            else:
                hr_data.gross_salary = None
            
            net_salary_str = request.form.get('net_salary', '').strip()
            if net_salary_str:
                hr_data.net_salary = float(net_salary_str.replace(',', '.'))
            else:
                hr_data.net_salary = None
            
            hr_data.iban = request.form.get('iban', '').strip().upper() or None
            hr_data.payment_method = request.form.get('payment_method', '').strip() or None
            
            # Documenti
            hr_data.id_card_type = request.form.get('id_card_type', '').strip() or None
            hr_data.id_card_number = request.form.get('id_card_number', '').strip() or None
            
            id_issue_str = request.form.get('id_card_issue_date', '').strip()
            if id_issue_str:
                hr_data.id_card_issue_date = datetime.strptime(id_issue_str, '%Y-%m-%d').date()
            else:
                hr_data.id_card_issue_date = None
            
            id_expiry_str = request.form.get('id_card_expiry', '').strip()
            if id_expiry_str:
                hr_data.id_card_expiry = datetime.strptime(id_expiry_str, '%Y-%m-%d').date()
            else:
                hr_data.id_card_expiry = None
            
            hr_data.id_card_issued_by = request.form.get('id_card_issued_by', '').strip() or None
            hr_data.passport_number = request.form.get('passport_number', '').strip() or None
            
            passport_expiry_str = request.form.get('passport_expiry', '').strip()
            if passport_expiry_str:
                hr_data.passport_expiry = datetime.strptime(passport_expiry_str, '%Y-%m-%d').date()
            else:
                hr_data.passport_expiry = None
            
            # Contatto emergenza
            hr_data.emergency_contact_name = request.form.get('emergency_contact_name', '').strip() or None
            hr_data.emergency_contact_phone = request.form.get('emergency_contact_phone', '').strip() or None
            hr_data.emergency_contact_relation = request.form.get('emergency_contact_relation', '').strip() or None
            
            # Formazione
            hr_data.education_level = request.form.get('education_level', '').strip() or None
            hr_data.education_field = request.form.get('education_field', '').strip() or None
            
            # Altri dati
            hr_data.marital_status = request.form.get('marital_status', '').strip() or None
            
            dependents_str = request.form.get('dependents_number', '').strip()
            if dependents_str:
                hr_data.dependents_number = int(dependents_str)
            else:
                hr_data.dependents_number = None
            
            hr_data.disability = request.form.get('disability') == 'on'
            
            disability_perc_str = request.form.get('disability_percentage', '').strip()
            if disability_perc_str and hr_data.disability:
                hr_data.disability_percentage = int(disability_perc_str)
            else:
                hr_data.disability_percentage = None
            
            # Note
            hr_data.notes = request.form.get('notes', '').strip() or None
            
            # Benefit aziendali
            meal_vouchers_str = request.form.get('meal_vouchers_value', '').strip()
            if meal_vouchers_str:
                hr_data.meal_vouchers_value = float(meal_vouchers_str.replace(',', '.'))
            else:
                hr_data.meal_vouchers_value = None
            
            hr_data.fuel_card = request.form.get('fuel_card') == 'on'
            
            # Patente
            hr_data.driver_license_number = request.form.get('driver_license_number', '').strip() or None
            hr_data.driver_license_type = request.form.get('driver_license_type', '').strip() or None
            
            driver_license_expiry_str = request.form.get('driver_license_expiry', '').strip()
            if driver_license_expiry_str:
                hr_data.driver_license_expiry = datetime.strptime(driver_license_expiry_str, '%Y-%m-%d').date()
            else:
                hr_data.driver_license_expiry = None
            
            # Dati operativi - con validazione multi-tenant
            # Sede di assunzione (campo amministrativo HR)
            sede_id_str = request.form.get('sede_id', '').strip()
            if sede_id_str:
                sede_id = int(sede_id_str)
                # Verifica che la sede appartenga alla company dell'utente
                sede = filter_by_company(Sede.query).filter_by(id=sede_id).first()
                if sede:
                    hr_data.sede_id = sede_id
                else:
                    flash('Sede non valida per questa azienda', 'warning')
                    hr_data.sede_id = None
            else:
                hr_data.sede_id = None
            
            # Accesso a tutte le sedi
            hr_data.all_sedi = request.form.get('all_sedi') == 'on'
            
            # Orario di lavoro assegnato
            work_schedule_id_str = request.form.get('work_schedule_id', '').strip()
            if work_schedule_id_str:
                work_schedule_id = int(work_schedule_id_str)
                # Verifica che l'orario appartenga alla company dell'utente
                from models import WorkSchedule
                schedule = filter_by_company(WorkSchedule.query).filter_by(id=work_schedule_id).first()
                if schedule:
                    hr_data.work_schedule_id = work_schedule_id
                else:
                    flash('Orario di lavoro non valido per questa azienda', 'warning')
                    hr_data.work_schedule_id = None
            else:
                hr_data.work_schedule_id = None
            
            aci_vehicle_id_str = request.form.get('aci_vehicle_id', '').strip()
            if aci_vehicle_id_str:
                aci_vehicle_id = int(aci_vehicle_id_str)
                # ACI table è globale, non serve validazione multi-tenant
                hr_data.aci_vehicle_id = aci_vehicle_id
            else:
                hr_data.aci_vehicle_id = None
            
            # Gestione upload libretto di circolazione
            delete_vehicle_doc = request.form.get('delete_vehicle_doc') == '1'
            if delete_vehicle_doc and hr_data.vehicle_registration_document:
                # Elimina il file esistente
                old_file_path = os.path.join('static', 'uploads', 'vehicle_docs', hr_data.vehicle_registration_document.split('/')[-1])
                if os.path.exists(old_file_path):
                    os.remove(old_file_path)
                hr_data.vehicle_registration_document = None
                flash('Libretto di circolazione eliminato con successo', 'success')
            
            if 'vehicle_registration_document' in request.files:
                file = request.files['vehicle_registration_document']
                if file and file.filename:
                    # Verifica dimensione file (max 5MB)
                    file.seek(0, os.SEEK_END)
                    file_size = file.tell()
                    file.seek(0)
                    
                    if file_size > 5 * 1024 * 1024:  # 5MB
                        flash('Il file è troppo grande. Dimensione massima: 5MB', 'error')
                    else:
                        # Verifica estensione
                        allowed_extensions = {'pdf', 'jpg', 'jpeg', 'png'}
                        filename = secure_filename(file.filename)
                        file_ext = filename.rsplit('.', 1)[1].lower() if '.' in filename else ''
                        
                        if file_ext in allowed_extensions:
                            # Crea directory se non esiste
                            upload_dir = os.path.join('static', 'uploads', 'vehicle_docs')
                            os.makedirs(upload_dir, exist_ok=True)
                            
                            # Genera nome univoco per il file
                            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                            unique_filename = f"libretto_{user.id}_{timestamp}.{file_ext}"
                            file_path = os.path.join(upload_dir, unique_filename)
                            
                            # Elimina il file precedente se esiste
                            if hr_data.vehicle_registration_document:
                                old_file_path = os.path.join('static', 'uploads', 'vehicle_docs', hr_data.vehicle_registration_document.split('/')[-1])
                                if os.path.exists(old_file_path):
                                    os.remove(old_file_path)
                            
                            # Salva il nuovo file
                            file.save(file_path)
                            hr_data.vehicle_registration_document = f"/uploads/vehicle_docs/{unique_filename}"
                            flash('Libretto di circolazione caricato con successo', 'success')
                        else:
                            flash('Formato file non supportato. Usa PDF, JPG o PNG', 'error')
            
            hr_data.overtime_enabled = request.form.get('overtime_enabled') == 'on'
            hr_data.overtime_type = request.form.get('overtime_type', '').strip() or None
            
            banca_ore_limite_str = request.form.get('banca_ore_limite_max', '').strip()
            if banca_ore_limite_str:
                hr_data.banca_ore_limite_max = float(banca_ore_limite_str.replace(',', '.'))
            else:
                hr_data.banca_ore_limite_max = None
            
            banca_ore_periodo_str = request.form.get('banca_ore_periodo_mesi', '').strip()
            if banca_ore_periodo_str:
                hr_data.banca_ore_periodo_mesi = int(banca_ore_periodo_str)
            else:
                hr_data.banca_ore_periodo_mesi = None
            
            # Requisiti e sicurezza
            hr_data.minimum_requirements = request.form.get('minimum_requirements', '').strip() or None
            
            # Visita medica
            medical_visit_date_str = request.form.get('medical_visit_date', '').strip()
            if medical_visit_date_str:
                hr_data.medical_visit_date = datetime.strptime(medical_visit_date_str, '%Y-%m-%d').date()
            else:
                hr_data.medical_visit_date = None
            
            medical_visit_expiry_str = request.form.get('medical_visit_expiry', '').strip()
            if medical_visit_expiry_str:
                hr_data.medical_visit_expiry = datetime.strptime(medical_visit_expiry_str, '%Y-%m-%d').date()
            else:
                hr_data.medical_visit_expiry = None
            
            # Formazioni
            # Formazione generale
            training_general_date_str = request.form.get('training_general_date', '').strip()
            if training_general_date_str:
                hr_data.training_general_date = datetime.strptime(training_general_date_str, '%Y-%m-%d').date()
            else:
                hr_data.training_general_date = None
            
            training_general_expiry_str = request.form.get('training_general_expiry', '').strip()
            if training_general_expiry_str:
                hr_data.training_general_expiry = datetime.strptime(training_general_expiry_str, '%Y-%m-%d').date()
            else:
                hr_data.training_general_expiry = None
            
            # Formazione RSPP
            training_rspp_date_str = request.form.get('training_rspp_date', '').strip()
            if training_rspp_date_str:
                hr_data.training_rspp_date = datetime.strptime(training_rspp_date_str, '%Y-%m-%d').date()
            else:
                hr_data.training_rspp_date = None
            
            training_rspp_expiry_str = request.form.get('training_rspp_expiry', '').strip()
            if training_rspp_expiry_str:
                hr_data.training_rspp_expiry = datetime.strptime(training_rspp_expiry_str, '%Y-%m-%d').date()
            else:
                hr_data.training_rspp_expiry = None
            
            # Formazione RLS
            training_rls_date_str = request.form.get('training_rls_date', '').strip()
            if training_rls_date_str:
                hr_data.training_rls_date = datetime.strptime(training_rls_date_str, '%Y-%m-%d').date()
            else:
                hr_data.training_rls_date = None
            
            training_rls_expiry_str = request.form.get('training_rls_expiry', '').strip()
            if training_rls_expiry_str:
                hr_data.training_rls_expiry = datetime.strptime(training_rls_expiry_str, '%Y-%m-%d').date()
            else:
                hr_data.training_rls_expiry = None
            
            # Formazione primo soccorso
            training_first_aid_date_str = request.form.get('training_first_aid_date', '').strip()
            if training_first_aid_date_str:
                hr_data.training_first_aid_date = datetime.strptime(training_first_aid_date_str, '%Y-%m-%d').date()
            else:
                hr_data.training_first_aid_date = None
            
            training_first_aid_expiry_str = request.form.get('training_first_aid_expiry', '').strip()
            if training_first_aid_expiry_str:
                hr_data.training_first_aid_expiry = datetime.strptime(training_first_aid_expiry_str, '%Y-%m-%d').date()
            else:
                hr_data.training_first_aid_expiry = None
            
            # Formazione emergenza
            training_emergency_date_str = request.form.get('training_emergency_date', '').strip()
            if training_emergency_date_str:
                hr_data.training_emergency_date = datetime.strptime(training_emergency_date_str, '%Y-%m-%d').date()
            else:
                hr_data.training_emergency_date = None
            
            training_emergency_expiry_str = request.form.get('training_emergency_expiry', '').strip()
            if training_emergency_expiry_str:
                hr_data.training_emergency_expiry = datetime.strptime(training_emergency_expiry_str, '%Y-%m-%d').date()
            else:
                hr_data.training_emergency_expiry = None
            
            # Formazione preposto
            training_supervisor_date_str = request.form.get('training_supervisor_date', '').strip()
            if training_supervisor_date_str:
                hr_data.training_supervisor_date = datetime.strptime(training_supervisor_date_str, '%Y-%m-%d').date()
            else:
                hr_data.training_supervisor_date = None
            
            training_supervisor_expiry_str = request.form.get('training_supervisor_expiry', '').strip()
            if training_supervisor_expiry_str:
                hr_data.training_supervisor_expiry = datetime.strptime(training_supervisor_expiry_str, '%Y-%m-%d').date()
            else:
                hr_data.training_supervisor_expiry = None
            
            hr_data.updated_at = datetime.now()
            db.session.commit()
            
            flash('Dati HR aggiornati con successo', 'success')
            return redirect(url_for('hr.hr_detail', user_id=user.id))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Errore nel salvataggio: {str(e)}', 'error')
    
    # Carica lista veicoli ACI per il dropdown
    aci_vehicles = ACITable.query.order_by(ACITable.tipologia, ACITable.marca, ACITable.modello).all()
    
    # Carica lista mansioni attive per il dropdown
    from models import Mansione
    mansioni = filter_by_company(Mansione.query).filter_by(active=True).order_by(Mansione.nome).all()
    
    # Carica lista sedi per il dropdown
    sedi = filter_by_company(Sede.query).order_by(Sede.name).all()
    
    # Carica lista orari di lavoro per il dropdown
    from models import WorkSchedule
    work_schedules = filter_by_company(WorkSchedule.query).filter_by(active=True).order_by(WorkSchedule.name).all()
    
    return render_template('hr_detail.html', 
                         user=user, 
                         hr_data=hr_data,
                         can_edit=can_edit,
                         aci_vehicles=aci_vehicles,
                         mansioni=mansioni,
                         sedi=sedi,
                         work_schedules=work_schedules)


@hr_bp.route('/export')
@login_required
@require_hr_permission
def hr_export():
    """Export Excel con dati HR personalizzato (campi e filtri selezionabili)"""
    
    if not (current_user.can_view_hr_data() or current_user.can_manage_hr_data()):
        flash('Non hai i permessi per esportare i dati HR', 'error')
        return redirect(url_for('hr.hr_list'))
    
    # Ottieni campi selezionati e filtri da session
    selected_fields = session.get('hr_export_fields', None)
    active_filters = session.get('hr_export_filters', {})
    
    # Se non ci sono campi selezionati, usa tutti i campi (comportamento predefinito)
    if not selected_fields:
        # Export completo di default
        selected_fields = None
    
    # Ottieni tutti gli utenti con dati HR (escludi admin)
    users_query = filter_by_company(User.query).filter_by(active=True).filter(
        User.role != 'Admin',
        User.role != 'ADMIN',
        User.role != 'Amministratore',
        User.is_system_admin == False
    ).order_by(User.last_name, User.first_name)
    
    users = users_query.all()
    
    # Applica filtri personalizzati
    if active_filters:
        filtered_users = []
        for user in users:
            hr_data = user.hr_data
            
            # Se ci sono filtri attivi ma l'utente non ha dati HR, escludilo
            if not hr_data:
                continue
            
            matches = True
            
            # Applica gli stessi filtri usati nella lista
            if 'matricola' in active_filters:
                if not hr_data.matricola or active_filters['matricola'].lower() not in hr_data.matricola.lower():
                    matches = False
            
            # Filtro data assunzione (filtra per anno)
            if 'hire_date' in active_filters:
                filter_date_str = active_filters['hire_date']
                try:
                    from datetime import datetime
                    filter_date = datetime.strptime(filter_date_str, '%Y-%m-%d').date()
                    if not hr_data.hire_date or hr_data.hire_date.year != filter_date.year:
                        matches = False
                except:
                    pass  # Se parsing fallisce, ignora il filtro
            
            if 'contract_type' in active_filters:
                if not hr_data.contract_type or hr_data.contract_type != active_filters['contract_type']:
                    matches = False
            
            if 'ccnl' in active_filters:
                if not hr_data.ccnl or active_filters['ccnl'].lower() not in hr_data.ccnl.lower():
                    matches = False
            
            if 'mansione' in active_filters:
                if not hr_data.mansione or active_filters['mansione'].lower() not in hr_data.mansione.lower():
                    matches = False
            
            if 'qualifica' in active_filters:
                if not hr_data.qualifica or active_filters['qualifica'].lower() not in hr_data.qualifica.lower():
                    matches = False
            
            if 'sede_assunzione' in active_filters:
                if not hr_data.sede_id or str(hr_data.sede_id) != str(active_filters['sede_assunzione']):
                    matches = False
            
            if 'gender' in active_filters:
                if not hr_data.gender or hr_data.gender != active_filters['gender']:
                    matches = False
            
            if 'city' in active_filters:
                if not hr_data.city or active_filters['city'].lower() not in hr_data.city.lower():
                    matches = False
            
            if matches:
                filtered_users.append(user)
        
        users = filtered_users
    
    # Crea workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Dati HR"
    
    # Styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Headers - Seguendo struttura file Excel: DATI CONTRATTUALI | ANAGRAFICA | VISITE E FORMAZIONE
    headers = [
        # DATI CONTRATTUALI
        'COD SI', 'Cognome', 'Nome',  'Data Assunzione', 'Data Fine Rapporto',
        'Tipologia Contratto', 'Fornitore', 'P.IVA', 'Fornitore Nome', 'Fornitore P.IVA', 'CCNL', 'Mansione', 'Qualifica', 'Livello', 
        'Orario (h/sett.)', 'Tipologia Orario', '% Part Time', 'Tipo Part Time',
        'Sede di Assunzione', 'Accesso Tutte Sedi', 'Orario Lavoro', 'Superminimo/SM Assorbibile', 'Rimborsi/Diarie',
        'Ticket', 'Rischio INAIL', 'Tipo di Assunzione', 'Altro/Note',
        # ANAGRAFICA RISORSA
        'Titolo di Studio', 'Luogo di Nascita', 'Data di Nascita', 'Età', 'Sesso', 'C.F.',
        'Comune di Residenza', 'Indirizzo Residenza', 'Domicilio Alternativo', 'CAP',
        'Recapito Telefonico', 'Fruizione Permessi L. 104/92', 'E-mail Personale', 'E-mail Aziendale',
        'Stato Civile', 'Familiari a Carico', 'Contatto Emergenza', 'Tel. Emergenza',
        'Patente Numero', 'Patente Tipo', 'Patente Scadenza',
        'Veicolo ACI', 'Straordinari', 'Tipologia Straordinario', 'Limite Ore', 'Periodo Mesi',
        # VISITE E FORMAZIONE
        'Possesso Requisiti Minimi',
        'Visita Medica - Data', 'Visita Medica - Scadenza',
        'Formazione Generale - Data', 'Formazione Generale - Scadenza',
        'Formazione RSPP - Data', 'Formazione RSPP - Scadenza',
        'Formazione RLS - Data', 'Formazione RLS - Scadenza',
        'Formazione Primo Soccorso - Data', 'Formazione Primo Soccorso - Scadenza',
        'Formazione Emergenza - Data', 'Formazione Emergenza - Scadenza',
        'Formazione Preposto - Data', 'Formazione Preposto - Scadenza'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    # Dati
    for row_idx, user in enumerate(users, 2):
        hr_data = user.hr_data
        
        row_data = [
            # DATI CONTRATTUALI
            hr_data.cod_si or hr_data.matricola if hr_data else '',
            user.last_name,
            user.first_name,
            hr_data.hire_date.strftime('%d/%m/%Y') if hr_data and hr_data.hire_date else '',
            hr_data.contract_end_date.strftime('%d/%m/%Y') if hr_data and hr_data.contract_end_date else '',
            hr_data.contract_type if hr_data else '',
            hr_data.distacco_supplier if hr_data else '',
            hr_data.consulente_vat if hr_data else '',
            hr_data.nome_fornitore if hr_data else '',
            hr_data.partita_iva_fornitore if hr_data else '',
            hr_data.ccnl if hr_data else '',
            hr_data.mansione if hr_data else '',
            hr_data.qualifica if hr_data else '',
            hr_data.ccnl_level if hr_data else '',
            str(hr_data.work_hours_week).replace('.', ',') if hr_data and hr_data.work_hours_week else '',
            hr_data.working_time_type if hr_data else '',
            str(hr_data.part_time_percentage).replace('.', ',') if hr_data and hr_data.part_time_percentage else '',
            hr_data.part_time_type if hr_data else '',
            hr_data.sede.name if hr_data and hr_data.sede else '',
            'Sì' if hr_data and hr_data.all_sedi else 'No' if hr_data else '',
            hr_data.work_schedule.name if hr_data and hr_data.work_schedule else '',
            str(hr_data.superminimo).replace('.', ',') if hr_data and hr_data.superminimo else '',
            str(hr_data.rimborsi_diarie).replace('.', ',') if hr_data and hr_data.rimborsi_diarie else '',
            'Sì' if hr_data and hr_data.ticket_restaurant else 'No' if hr_data else '',
            hr_data.rischio_inail if hr_data else '',
            hr_data.tipo_assunzione if hr_data else '',
            hr_data.other_notes if hr_data else '',
            # ANAGRAFICA RISORSA
            hr_data.education_level if hr_data else '',
            hr_data.birth_city if hr_data else '',
            hr_data.birth_date.strftime('%d/%m/%Y') if hr_data and hr_data.birth_date else '',
            hr_data.get_age() if hr_data else '',
            hr_data.gender if hr_data else '',
            hr_data.codice_fiscale if hr_data else '',
            hr_data.city if hr_data else '',
            hr_data.address if hr_data else '',
            hr_data.alternative_domicile if hr_data else '',
            hr_data.postal_code if hr_data else '',
            hr_data.phone if hr_data else '',
            'Sì' if hr_data and hr_data.law_104_benefits else 'No' if hr_data else '',
            hr_data.personal_email if hr_data else '',
            user.email,
            hr_data.marital_status if hr_data else '',
            hr_data.dependents_number if hr_data and hr_data.dependents_number else '',
            hr_data.emergency_contact_name if hr_data else '',
            hr_data.emergency_contact_phone if hr_data else '',
            hr_data.driver_license_number if hr_data else '',
            hr_data.driver_license_type if hr_data else '',
            hr_data.driver_license_expiry.strftime('%d/%m/%Y') if hr_data and hr_data.driver_license_expiry else '',
            f"{hr_data.aci_vehicle.marca} {hr_data.aci_vehicle.modello} ({hr_data.aci_vehicle.tipologia})" if hr_data and hr_data.aci_vehicle else '',
            'Sì' if hr_data and hr_data.overtime_enabled else 'No' if hr_data else '',
            hr_data.overtime_type if hr_data and hr_data.overtime_type else '',
            str(hr_data.banca_ore_limite_max).replace('.', ',') if hr_data and hr_data.banca_ore_limite_max else '',
            str(hr_data.banca_ore_periodo_mesi) if hr_data and hr_data.banca_ore_periodo_mesi else '',
            # VISITE E FORMAZIONE
            hr_data.minimum_requirements if hr_data else '',
            hr_data.medical_visit_date.strftime('%d/%m/%Y') if hr_data and hr_data.medical_visit_date else '',
            hr_data.medical_visit_expiry.strftime('%d/%m/%Y') if hr_data and hr_data.medical_visit_expiry else '',
            hr_data.training_general_date.strftime('%d/%m/%Y') if hr_data and hr_data.training_general_date else '',
            hr_data.training_general_expiry.strftime('%d/%m/%Y') if hr_data and hr_data.training_general_expiry else '',
            hr_data.training_rspp_date.strftime('%d/%m/%Y') if hr_data and hr_data.training_rspp_date else '',
            hr_data.training_rspp_expiry.strftime('%d/%m/%Y') if hr_data and hr_data.training_rspp_expiry else '',
            hr_data.training_rls_date.strftime('%d/%m/%Y') if hr_data and hr_data.training_rls_date else '',
            hr_data.training_rls_expiry.strftime('%d/%m/%Y') if hr_data and hr_data.training_rls_expiry else '',
            hr_data.training_first_aid_date.strftime('%d/%m/%Y') if hr_data and hr_data.training_first_aid_date else '',
            hr_data.training_first_aid_expiry.strftime('%d/%m/%Y') if hr_data and hr_data.training_first_aid_expiry else '',
            hr_data.training_emergency_date.strftime('%d/%m/%Y') if hr_data and hr_data.training_emergency_date else '',
            hr_data.training_emergency_expiry.strftime('%d/%m/%Y') if hr_data and hr_data.training_emergency_expiry else '',
            hr_data.training_supervisor_date.strftime('%d/%m/%Y') if hr_data and hr_data.training_supervisor_date else '',
            hr_data.training_supervisor_expiry.strftime('%d/%m/%Y') if hr_data and hr_data.training_supervisor_expiry else ''
        ]
        
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    # Generate response
    response = make_response(excel_file.getvalue())
    response.headers["Content-Disposition"] = f"attachment; filename=dati_hr_{date.today().strftime('%Y%m%d')}.xlsx"
    response.headers["Content-type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    
    return response
