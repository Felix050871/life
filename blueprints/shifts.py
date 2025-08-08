# =============================================================================
# SHIFTS BLUEPRINT - Modulo gestione turni e scheduling
# =============================================================================
#
# ROUTES INCLUSE:
# 1. turni_automatici (GET) - Creazione automatica turni da template
# 2. api/get_shifts_for_template/<template_id> (GET) - API turni template
# 3. visualizza_turni (GET) - Visualizzazione turni
# 4. genera_turni_da_template (POST) - Generazione turni da template
# 5. create_shift (POST) - Creazione singolo turno
# 6. generate_shifts (POST) - Generazione multipla turni
# 7. regenerate_template/<template_id> (POST) - Rigenerazione template
# 8. delete_template/<template_id> (POST) - Eliminazione template
# 9. view_template/<template_id> (GET) - Visualizzazione template
#
# Total routes: 9+ shift management routes
# =============================================================================

from flask import Blueprint, request, jsonify, render_template, redirect, url_for, flash, make_response
from flask_login import login_required, current_user
from datetime import datetime, date, timedelta
from functools import wraps
from app import db, csrf
from models import User, Shift, Sede, PresidioCoverageTemplate, PresidioCoverage, AttendanceEvent, ReperibilitaShift, italian_now
from forms import EditShiftForm
from collections import defaultdict
from io import BytesIO
import json
import tempfile
import os

# Create blueprint
shifts_bp = Blueprint('shifts', __name__, url_prefix='/shifts')

# Helper functions
def require_login(f):
    """Decorator to require login for routes"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated:
            return redirect(url_for('auth.login'))
        return f(*args, **kwargs)
    return decorated_function

# =============================================================================
# SHIFT MANAGEMENT ROUTES
# =============================================================================

@shifts_bp.route('/turni_automatici')
@login_required
def turni_automatici():
    """Sistema nuovo: Creazione automatica turni da template presidio"""
    if not (current_user.can_manage_shifts() or current_user.can_view_shifts()):
        flash('Non hai i permessi per accedere ai turni', 'danger')
        return redirect(url_for('dashboard'))
    
    # Ottieni template presidio attivi per creazione turni
    from models import PresidioCoverageTemplate
    presidio_templates = PresidioCoverageTemplate.query.filter_by(active=True).order_by(
        PresidioCoverageTemplate.start_date.desc()
    ).all()
    
    # Ottieni template selezionato se presente
    template_id = request.args.get('template_id')
    selected_template = None
    turni_per_settimana = {}
    settimane_stats = {}
    shifts = []
    
    if template_id:
        try:
            template_id = int(template_id)
            selected_template = PresidioCoverageTemplate.query.get(template_id)
            
            if selected_template:
                # Ottieni turni del template selezionato
                from collections import defaultdict
                from datetime import timedelta, date
                
                # Filtra turni per il periodo del template
                accessible_sedi = current_user.get_turni_sedi()
                if accessible_sedi:
                    shifts = Shift.query.join(User, Shift.user_id == User.id).filter(
                        User.sede_id.in_([sede.id for sede in accessible_sedi]),
                        Shift.date >= selected_template.start_date,
                        Shift.date <= selected_template.end_date,
                        Shift.shift_type == 'presidio'
                    ).order_by(Shift.date.asc(), Shift.start_time.asc()).all()
                else:
                    shifts = []
                
                # Raggruppa turni per settimana e calcola statistiche
                turni_per_settimana = defaultdict(list)
                settimane_stats = defaultdict(lambda: {
                    'inizio': None,
                    'fine': None,
                    'turni_count': 0,
                    'ore_totali': 0.0,
                    'unique_users': set(),
                    'giorni_coperti': set()
                })
                
                for shift in shifts:
                    # Calcola inizio settimana (lunedì)
                    settimana_inizio = shift.date - timedelta(days=shift.date.weekday())
                    settimana_key = settimana_inizio.strftime('%Y-%m-%d')
                    
                    turni_per_settimana[settimana_key].append(shift)
                    
                    # Aggiorna statistiche settimana
                    stats = settimane_stats[settimana_key]
                    if not stats['inizio']:
                        stats['inizio'] = settimana_inizio
                        stats['fine'] = settimana_inizio + timedelta(days=6)
                    
                    stats['turni_count'] = stats['turni_count'] + 1
                    if isinstance(stats['unique_users'], set):
                        stats['unique_users'].add(shift.user_id)
                    if isinstance(stats['giorni_coperti'], set):
                        stats['giorni_coperti'].add(shift.date.isoformat())
                    
                    # Calcola ore totali
                    if shift.start_time and shift.end_time:
                        start_dt = datetime.combine(shift.date, shift.start_time)
                        end_dt = datetime.combine(shift.date, shift.end_time)
                        if end_dt < start_dt:  # Turno notturno
                            end_dt += timedelta(days=1)
                        ore = (end_dt - start_dt).total_seconds() / 3600
                        if isinstance(stats['ore_totali'], (int, float)):
                            stats['ore_totali'] = stats['ore_totali'] + ore
                
                # Converti set in count per JSON serialization
                for settimana_key, stats in settimane_stats.items():
                    if isinstance(stats['unique_users'], set):
                        stats['unique_users'] = len(stats['unique_users'])
                
                # Ordina settimane per data crescente
                settimane_stats = dict(sorted(settimane_stats.items(), 
                                            key=lambda x: x[1]['inizio'] if x[1]['inizio'] else date.min))
                turni_per_settimana = dict(sorted(turni_per_settimana.items()))
        except (ValueError, TypeError):
            pass
    
    # Ottieni utenti disponibili per creazione turni raggruppati per ruolo
    from collections import defaultdict
    users_by_role = defaultdict(list)
    available_users = User.query.filter(
        User.active.is_(True)
    ).all()
    
    for user in available_users:
        if hasattr(user, 'role') and user.role:
            users_by_role[user.role].append(user)
    
    from datetime import date, timedelta
    return render_template('turni_automatici.html', 
                         presidio_templates=presidio_templates,
                         selected_template=selected_template,
                         turni_per_settimana=turni_per_settimana,
                         settimane_stats=settimane_stats,
                         users_by_role=dict(users_by_role),
                         shifts=shifts,
                         today=date.today(),
                         timedelta=timedelta,
                         can_manage_shifts=current_user.can_manage_shifts())

@shifts_bp.route('/api/get_shifts_for_template/<int:template_id>')
@login_required
def get_shifts_for_template_api(template_id):
    """API per ottenere turni di un template specifico"""
    try:
        from models import PresidioCoverageTemplate
        from collections import defaultdict
        from datetime import timedelta, date
        
        template = PresidioCoverageTemplate.query.get_or_404(template_id)
        
        # Ottieni turni del template
        accessible_sedi = current_user.get_turni_sedi()
        if accessible_sedi:
            shifts = Shift.query.join(User, Shift.user_id == User.id).filter(
                User.sede_id.in_([sede.id for sede in accessible_sedi]),
                Shift.date >= template.start_date,
                Shift.date <= template.end_date,
                Shift.shift_type == 'presidio'
            ).order_by(Shift.date.asc(), Shift.start_time.asc()).all()
        else:
            shifts = []
        
        # Raggruppa per settimana
        weeks_data = []
        turni_per_settimana = defaultdict(list)
        
        for shift in shifts:
            settimana_inizio = shift.date - timedelta(days=shift.date.weekday())
            settimana_key = settimana_inizio.strftime('%Y-%m-%d')
            turni_per_settimana[settimana_key].append(shift)
        
        # Converti in formato per frontend
        for settimana_key in sorted(turni_per_settimana.keys()):
            week_shifts = turni_per_settimana[settimana_key]
            settimana_inizio = datetime.strptime(settimana_key, '%Y-%m-%d').date()
            
            week_data = {
                'week_start': settimana_key,
                'week_end': (settimana_inizio + timedelta(days=6)).strftime('%Y-%m-%d'),
                'shifts_count': len(week_shifts),
                'shifts': []
            }
            
            for shift in week_shifts:
                shift_data = {
                    'id': shift.id,
                    'date': shift.date.strftime('%Y-%m-%d'),
                    'start_time': shift.start_time.strftime('%H:%M') if shift.start_time else '',
                    'end_time': shift.end_time.strftime('%H:%M') if shift.end_time else '',
                    'user_name': shift.user.nome + ' ' + shift.user.cognome if shift.user else '',
                    'role': shift.user.role if shift.user else ''
                }
                week_data['shifts'].append(shift_data)
            
            weeks_data.append(week_data)
        
        return jsonify({
            'success': True,
            'template_name': template.name,
            'weeks_data': weeks_data
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@shifts_bp.route('/visualizza_turni')
@login_required
def visualizza_turni():
    """Visualizzazione turni per mese/settimana"""
    if not (current_user.can_manage_shifts() or current_user.can_view_shifts()):
        flash('Non hai i permessi per visualizzare i turni', 'danger')
        return redirect(url_for('dashboard'))
    
    # Ottieni parametri filtro
    view_mode = request.args.get('view', 'month')  # month, week
    selected_date = request.args.get('date')
    sede_filter = request.args.get('sede')
    
    # Parse data selezionata
    if selected_date:
        try:
            target_date = datetime.strptime(selected_date, '%Y-%m-%d').date()
        except:
            target_date = date.today()
    else:
        target_date = date.today()
    
    # Calcola range di date in base alla modalità
    if view_mode == 'week':
        # Vista settimanale: dal lunedì alla domenica
        start_date = target_date - timedelta(days=target_date.weekday())
        end_date = start_date + timedelta(days=6)
    else:
        # Vista mensile
        start_date = target_date.replace(day=1)
        if target_date.month == 12:
            end_date = date(target_date.year + 1, 1, 1) - timedelta(days=1)
        else:
            end_date = target_date.replace(month=target_date.month + 1, day=1) - timedelta(days=1)
    
    # Query turni nel range
    query = Shift.query.filter(
        Shift.date >= start_date,
        Shift.date <= end_date
    )
    
    # Filtro per sede se specificato
    if sede_filter and sede_filter != 'all':
        try:
            sede_id = int(sede_filter)
            query = query.join(User, Shift.user_id == User.id).filter(User.sede_id == sede_id)
        except:
            pass
    
    # Controllo permessi sede
    accessible_sedi = current_user.get_turni_sedi()
    if accessible_sedi:
        sede_ids = [sede.id for sede in accessible_sedi]
        query = query.join(User, Shift.user_id == User.id).filter(User.sede_id.in_(sede_ids))
    
    shifts = query.order_by(Shift.date.asc(), Shift.start_time.asc()).all()
    
    # Ottieni sedi per filtro
    if accessible_sedi:
        available_sedi = accessible_sedi
    else:
        available_sedi = Sede.query.filter_by(active=True).all()
    
    # Raggruppa turni per giorno
    shifts_by_date = defaultdict(list)
    for shift in shifts:
        shifts_by_date[shift.date].append(shift)
    
    return render_template('shifts.html',
                         shifts=shifts,
                         shifts_by_date=dict(shifts_by_date),
                         view_mode=view_mode,
                         target_date=target_date,
                         start_date=start_date,
                         end_date=end_date,
                         available_sedi=available_sedi,
                         selected_sede=sede_filter,
                         can_manage_shifts=current_user.can_manage_shifts())

@shifts_bp.route('/genera_turni_da_template', methods=['POST'])
@login_required
def genera_turni_da_template():
    """Genera turni automaticamente da template presidio selezionato"""
    if not current_user.can_manage_shifts():
        return jsonify({'success': False, 'message': 'Non hai i permessi per generare turni'}), 403
    
    try:
        template_id = request.form.get('template_id')
        if not template_id:
            return jsonify({'success': False, 'message': 'Template ID richiesto'}), 400
        
        template = PresidioCoverageTemplate.query.get_or_404(int(template_id))
        
        # Import logica generazione turni da utils
        from utils import generate_shifts_from_template
        result = generate_shifts_from_template(template)
        
        if result['success']:
            return jsonify({
                'success': True,
                'message': f'Generati {result["generated_count"]} turni per {template.name}',
                'generated_count': result['generated_count']
            })
        else:
            return jsonify({'success': False, 'message': result.get('message', 'Errore nella generazione turni')}), 400
            
    except Exception as e:
        return jsonify({'success': False, 'message': f'Errore: {str(e)}'}), 500

@shifts_bp.route('/create_shift', methods=['POST'])
@login_required 
def create_shift():
    """Crea un singolo turno manualmente"""
    if not current_user.can_manage_shifts():
        return jsonify({'success': False, 'message': 'Non hai i permessi per creare turni'}), 403
        
    try:
        user_id = request.form.get('user_id')
        date_str = request.form.get('date')
        start_time_str = request.form.get('start_time')
        end_time_str = request.form.get('end_time')
        shift_type = request.form.get('shift_type', 'presidio')
        
        if not all([user_id, date_str, start_time_str, end_time_str]):
            return jsonify({'success': False, 'message': 'Tutti i campi sono richiesti'}), 400
        
        # Parse date e time
        shift_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        start_time = datetime.strptime(start_time_str, '%H:%M').time()
        end_time = datetime.strptime(end_time_str, '%H:%M').time()
        
        user = User.query.get(int(user_id))
        if not user or not user.active:
            return jsonify({'success': False, 'message': 'Utente non trovato o non attivo'}), 400
        
        # Controlla sovrapposizioni
        existing_shifts = Shift.query.filter(Shift.user_id == int(user_id), Shift.date == shift_date).all()
        for existing in existing_shifts:
            if existing.start_time and existing.end_time:
                if (start_time < existing.end_time and end_time > existing.start_time):
                    return jsonify({'success': False, 'message': f'Sovrapposizione con turno esistente'}), 400
        
        # Crea nuovo turno
        new_shift = Shift(
            user_id=int(user_id),
            date=shift_date,
            start_time=start_time,
            end_time=end_time,
            shift_type=shift_type,
            created_by_user_id=current_user.id
        )
        
        db.session.add(new_shift)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'Turno creato per {user.nome} {user.cognome}',
            'shift_id': new_shift.id
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Errore nella creazione turno: {str(e)}'}), 500

# =============================================================================
# ADDITIONAL SHIFT MANAGEMENT ROUTES
# =============================================================================

@shifts_bp.route('/edit/<int:shift_id>', methods=['GET', 'POST'])
@login_required
def edit_shift(shift_id):
    """Edit existing shift"""
    if not current_user.can_access_turni():
        flash('Non hai i permessi per modificare turni', 'danger')
        return redirect(url_for('dashboard.dashboard'))
    
    shift = Shift.query.get_or_404(shift_id)
    
    # Check if shift is in the future or today
    if shift.date < date.today():
        flash('Non è possibile modificare turni passati', 'warning')
        return redirect(url_for('dashboard.dashboard'))
    
    # Verifica permessi sulla sede (se non admin)
    if current_user.role != 'Admin':
        if not current_user.sede_obj or current_user.sede_obj.id != shift.user.sede_id:
            flash('Non hai i permessi per modificare turni per questa sede', 'danger')
            return redirect(url_for('dashboard.dashboard'))
        # Verifica che la sede sia di tipo "Turni" per utenti non-admin
        if not current_user.sede_obj.is_turni_mode():
            flash('La modifica turni è disponibile solo per sedi di tipo "Turni"', 'warning')
            return redirect(url_for('dashboard.dashboard'))
    
    form = EditShiftForm()
    
    # Get available users for assignment (only from the same sede as the shift)
    users = User.query.filter(
        User.role.in_(['Management', 'Redattore', 'Sviluppatore', 'Operatore']),
        User.active.is_(True),
        User.sede_id == shift.user.sede_id
    ).order_by(User.first_name, User.last_name).all()
    
    # Popola le scelte del form con gli utenti disponibili
    form.user_id.choices = [(user.id, f"{user.get_full_name()} - {user.role}") for user in users]
    
    if form.validate_on_submit():
        try:
            # Verifica sovrapposizioni con il nuovo orario e utente
            overlapping_shift = Shift.query.filter(
                Shift.user_id == form.user_id.data,
                Shift.date == shift.date,
                Shift.id != shift.id,
                # Controlla sovrapposizione oraria
                db.or_(
                    db.and_(Shift.start_time <= form.start_time.data, Shift.end_time > form.start_time.data),
                    db.and_(Shift.start_time < form.end_time.data, Shift.end_time >= form.end_time.data),
                    db.and_(Shift.start_time >= form.start_time.data, Shift.end_time <= form.end_time.data)
                )
            ).first()
            
            if overlapping_shift:
                flash(f'Sovrapposizione rilevata: l\'utente selezionato ha già un turno dalle {overlapping_shift.start_time.strftime("%H:%M")} alle {overlapping_shift.end_time.strftime("%H:%M")}', 'warning')
            else:
                # Salva i valori originali per il messaggio
                old_user = shift.user.get_full_name()
                old_time = f"{shift.start_time.strftime('%H:%M')} - {shift.end_time.strftime('%H:%M')}"
                
                # Aggiorna il turno
                shift.user_id = form.user_id.data
                shift.start_time = form.start_time.data
                shift.end_time = form.end_time.data
                
                db.session.commit()
                
                new_user = User.query.get(form.user_id.data)
                new_time = f"{form.start_time.data.strftime('%H:%M')} - {form.end_time.data.strftime('%H:%M')}"
                
                flash(f'Turno modificato con successo: {old_user} ({old_time}) → {new_user.get_full_name()} ({new_time})', 'success')
                
                # Redirect back to the referring page or dashboard
                return redirect(request.referrer or url_for('dashboard.dashboard'))
                
        except Exception as e:
            db.session.rollback()
            flash(f'Errore durante la modifica del turno: {str(e)}', 'danger')
    
    # Pre-popola il form con i dati esistenti
    if request.method == 'GET':
        form.user_id.data = shift.user_id
        form.start_time.data = shift.start_time
        form.end_time.data = shift.end_time
    
    return render_template('edit_shift.html', shift=shift, users=users, form=form)

@shifts_bp.route('/team')
@login_required
def team_shifts():
    """Team shifts view with weekly navigation"""
    # Solo PM può vedere i turni del team
    if not current_user.can_view_shifts():
        flash('Non hai i permessi per accedere a questa funzionalità.', 'danger')
        return redirect(url_for('dashboard.dashboard'))
    
    # Gestisci navigazione settimanale
    date_param = request.args.get('date')
    if date_param:
        try:
            target_date = datetime.strptime(date_param, '%Y-%m-%d').date()
        except:
            target_date = date.today()
    else:
        target_date = date.today()
    
    # Calcola l'inizio e la fine della settimana
    week_start = target_date - timedelta(days=target_date.weekday())
    week_end = week_start + timedelta(days=6)
    
    # Calcola settimana precedente e successiva per navigazione
    prev_week = week_start - timedelta(days=7)
    next_week = week_start + timedelta(days=7)
    
    # Prendi tutti i turni della settimana
    weekly_shifts = Shift.query.filter(
        Shift.date >= week_start,
        Shift.date <= week_end
    ).order_by(Shift.date, Shift.start_time).all()
    
    # Prendi tutti i turni di reperibilità della settimana
    weekly_reperibilita = ReperibilitaShift.query.filter(
        ReperibilitaShift.date >= week_start,
        ReperibilitaShift.date <= week_end
    ).order_by(ReperibilitaShift.date, ReperibilitaShift.start_time).all()
    
    # Crea un dizionario dei turni di reperibilità per data
    reperibilita_by_date = {}
    for rep_shift in weekly_reperibilita:
        if rep_shift.date not in reperibilita_by_date:
            reperibilita_by_date[rep_shift.date] = []
        reperibilita_by_date[rep_shift.date].append(rep_shift)
    
    # Raggruppa i turni per giorno e aggiungi informazioni di presenza
    shifts_by_day = {}
    for shift in weekly_shifts:
        if shift.date not in shifts_by_day:
            shifts_by_day[shift.date] = []
        
        # Calcola lo stato di presenza per ogni turno
        shift.presence_status = calculate_shift_presence(shift)
        shifts_by_day[shift.date].append(shift)
    
    return render_template('team_shifts.html', 
                         shifts_by_day=shifts_by_day,
                         reperibilita_by_date=reperibilita_by_date,
                         week_start=week_start,
                         week_end=week_end,
                         prev_week=prev_week,
                         next_week=next_week,
                         target_date=target_date)

@shifts_bp.route('/change-user/<int:shift_id>', methods=['POST'])
@login_required
@csrf.exempt
def change_shift_user(shift_id):
    """Change user assigned to a shift (API endpoint)"""
    if current_user.role not in ['Management']:
        return jsonify({'success': False, 'message': 'Non hai i permessi per modificare i turni.'})
    
    try:
        new_user_id = request.json.get('user_id')
        if not new_user_id:
            return jsonify({'success': False, 'message': 'ID utente mancante.'})
        
        shift = Shift.query.get_or_404(shift_id)
        old_user = shift.user
        new_user = User.query.get_or_404(new_user_id)
        
        # Verifica che il nuovo utente sia attivo
        if not new_user.active:
            return jsonify({'success': False, 'message': 'L\'utente selezionato non è attivo.'})
        
        # Verifica sovrapposizioni per il nuovo utente
        overlapping_shift = Shift.query.filter(
            Shift.user_id == new_user_id,
            Shift.date == shift.date,
            Shift.id != shift_id,
            # Controlla sovrapposizione oraria
            db.or_(
                db.and_(Shift.start_time <= shift.start_time, Shift.end_time > shift.start_time),
                db.and_(Shift.start_time < shift.end_time, Shift.end_time >= shift.end_time),
                db.and_(Shift.start_time >= shift.start_time, Shift.end_time <= shift.end_time)
            )
        ).first()
        
        if overlapping_shift:
            return jsonify({
                'success': False, 
                'message': f'{new_user.get_full_name()} ha già un turno sovrapposto dalle {overlapping_shift.start_time.strftime("%H:%M")} alle {overlapping_shift.end_time.strftime("%H:%M")}.'
            })
        
        # Aggiorna il turno
        shift.user_id = new_user_id
        db.session.commit()
        
        return jsonify({
            'success': True, 
            'message': f'Turno trasferito da {old_user.get_full_name()} a {new_user.get_full_name()}.',
            'new_user_name': new_user.get_full_name(),
            'new_user_role': new_user.role
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Errore: {str(e)}'})

@shifts_bp.route('/delete/<int:shift_id>', methods=['POST'])
@login_required
def delete_shift(shift_id):
    """Delete a single shift"""
    if not current_user.can_access_turni():
        flash('Non hai i permessi per eliminare turni', 'danger')
        return redirect(url_for('dashboard.dashboard'))
    
    shift = Shift.query.get_or_404(shift_id)
    
    # Verifica che il turno non sia nel passato
    if shift.date < date.today():
        flash('Non è possibile eliminare turni passati', 'warning')
        return redirect(request.referrer or url_for('dashboard.dashboard'))
    
    # Verifica permessi sulla sede (se non admin)
    if current_user.role != 'Admin':
        if not current_user.sede_obj or current_user.sede_obj.id != shift.user.sede_id:
            flash('Non hai i permessi per eliminare turni per questa sede', 'danger')
            return redirect(request.referrer or url_for('dashboard.dashboard'))
        # Verifica che la sede sia di tipo "Turni" per utenti non-admin
        if not current_user.sede_obj.is_turni_mode():
            flash('La modifica turni è disponibile solo per sedi di tipo "Turni"', 'warning')
            return redirect(request.referrer or url_for('dashboard.dashboard'))
    
    try:
        # Salva info per messaggio di conferma
        user_name = shift.user.get_full_name()
        shift_date = shift.date.strftime('%d/%m/%Y')
        shift_time = f"{shift.start_time.strftime('%H:%M')} - {shift.end_time.strftime('%H:%M')}"
        
        db.session.delete(shift)
        db.session.commit()
        
        flash(f'Turno eliminato: {user_name} - {shift_date} ({shift_time})', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Errore durante l\'eliminazione del turno: {str(e)}', 'danger')
    
    return redirect(request.referrer or url_for('dashboard.dashboard'))

# =============================================================================
# HELPER FUNCTIONS
# =============================================================================

def calculate_shift_presence(shift):
    """Calculate presence status for a specific shift"""
    # Combina data e orari del turno per creare datetime completi
    shift_start = datetime.combine(shift.date, shift.start_time)
    shift_end = datetime.combine(shift.date, shift.end_time)
    
    # Prendi tutti gli eventi di presenza dell'utente per quella data
    events = AttendanceEvent.query.filter(
        AttendanceEvent.user_id == shift.user_id,
        AttendanceEvent.date == shift.date
    ).order_by(AttendanceEvent.timestamp).all()
    
    if not events:
        return {'status': 'absent', 'description': 'Nessuna timbratura'}
    
    # Analizza gli eventi per determinare lo stato
    # Implementazione semplificata - può essere estesa secondo necessità
    clock_in_events = [e for e in events if e.action == 'clock_in']
    clock_out_events = [e for e in events if e.action == 'clock_out']
    
    if clock_in_events and clock_out_events:
        return {'status': 'present', 'description': 'Presente'}
    elif clock_in_events:
        return {'status': 'partial', 'description': 'Entrata registrata'}
    else:
        return {'status': 'absent', 'description': 'Assente'}

# =============================================================================
# EXPORT ROUTES
# =============================================================================

@shifts_bp.route('/export/excel')
@login_required
def export_shifts_excel():
    """Export turni in formato Excel"""
    # Parametri dalla query string
    view_mode = request.args.get('view', 'month')  # month, week, day
    show_my_shifts = request.args.get('my_shifts', 'false') == 'true'
    date_param = request.args.get('date', datetime.now().strftime('%Y-%m-%d'))
    
    try:
        current_date = datetime.strptime(date_param, '%Y-%m-%d').date()
    except:
        current_date = date.today()
    
    # Calcola range di date in base alla vista
    if view_mode == 'day':
        start_date = current_date
        end_date = current_date
        filename = f"turni_{current_date.strftime('%Y-%m-%d')}.xlsx"
    elif view_mode == 'week':
        # Settimana (Lunedì - Domenica)
        days_since_monday = current_date.weekday()
        start_date = current_date - timedelta(days=days_since_monday)
        end_date = start_date + timedelta(days=6)
        filename = f"turni_settimana_{start_date.strftime('%Y-%m-%d')}.xlsx"
    else:  # month
        start_date = current_date.replace(day=1)
        if current_date.month == 12:
            end_date = date(current_date.year + 1, 1, 1) - timedelta(days=1)
        else:
            end_date = date(current_date.year, current_date.month + 1, 1) - timedelta(days=1)
        filename = f"turni_{current_date.strftime('%Y-%m')}.xlsx"
    
    # Query dei turni
    shifts_query = Shift.query.filter(
        Shift.date >= start_date,
        Shift.date <= end_date
    )
    
    # Filtro per "I Miei Turni" se richiesto
    if show_my_shifts:
        shifts_query = shifts_query.filter(Shift.user_id == current_user.id)
        filename = f"miei_{filename}"
    
    shifts = shifts_query.order_by(Shift.date, Shift.start_time).all()
    
    # Crea Excel in memoria usando openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Turni"
    
    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Header
    headers = ['Data', 'Utente', 'Ruolo', 'Orario Inizio', 'Orario Fine', 'Tipo Turno', 'Durata (ore)']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    # Dati
    for row_idx, shift in enumerate(shifts, 2):
        duration = (datetime.combine(date.today(), shift.end_time) - 
                   datetime.combine(date.today(), shift.start_time)).total_seconds() / 3600
        
        row_data = [
            shift.date.strftime('%d/%m/%Y'),
            shift.user.get_full_name(),
            shift.user.role,
            shift.start_time.strftime('%H:%M'),
            shift.end_time.strftime('%H:%M'),
            shift.shift_type,
            f"{duration:.1f}"
        ]
        
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border
            if col in [1, 4, 5]:  # Data e Orari
                cell.alignment = Alignment(horizontal='center')
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response

@shifts_bp.route('/export/pdf')
@login_required
def export_shifts_pdf():
    """Export turni in formato PDF"""
    # Parametri dalla query string
    view_mode = request.args.get('view', 'month')
    show_my_shifts = request.args.get('my_shifts', 'false') == 'true'
    date_param = request.args.get('date', datetime.now().strftime('%Y-%m-%d'))
    
    try:
        current_date = datetime.strptime(date_param, '%Y-%m-%d').date()
    except:
        current_date = date.today()
    
    # Calcola range di date e titolo
    if view_mode == 'day':
        start_date = current_date
        end_date = current_date
        title = f"Turni del {current_date.strftime('%d/%m/%Y')}"
        filename = f"turni_{current_date.strftime('%Y-%m-%d')}.pdf"
    elif view_mode == 'week':
        days_since_monday = current_date.weekday()
        start_date = current_date - timedelta(days=days_since_monday)
        end_date = start_date + timedelta(days=6)
        title = f"Turni Settimana {start_date.strftime('%d/%m')} - {end_date.strftime('%d/%m/%Y')}"
        filename = f"turni_settimana_{start_date.strftime('%Y-%m-%d')}.pdf"
    else:
        start_date = current_date.replace(day=1)
        if current_date.month == 12:
            end_date = date(current_date.year + 1, 1, 1) - timedelta(days=1)
        else:
            end_date = date(current_date.year, current_date.month + 1, 1) - timedelta(days=1)
        title = f"Turni {current_date.strftime('%B %Y').title()}"
        filename = f"turni_{current_date.strftime('%Y-%m')}.pdf"
    
    # Query dei turni
    shifts_query = Shift.query.filter(
        Shift.date >= start_date,
        Shift.date <= end_date
    )
    
    if show_my_shifts:
        shifts_query = shifts_query.filter(Shift.user_id == current_user.id)
        title = f"I Miei {title}"
        filename = f"miei_{filename}"
    
    shifts = shifts_query.order_by(Shift.date, Shift.start_time).all()
    
    # Crea PDF in memoria
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=1*inch, bottomMargin=1*inch)
    story = []
    
    # Stili
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=30,
        alignment=1  # Center
    )
    
    # Titolo
    story.append(Paragraph(title, title_style))
    story.append(Paragraph(f"Generato il {datetime.now().strftime('%d/%m/%Y alle %H:%M')}", styles['Normal']))
    story.append(Spacer(1, 20))
    
    if not shifts:
        story.append(Paragraph("Nessun turno trovato per il periodo selezionato.", styles['Normal']))
    else:
        # Raggruppa turni per data
        shifts_by_date = {}
        for shift in shifts:
            if shift.date not in shifts_by_date:
                shifts_by_date[shift.date] = []
            shifts_by_date[shift.date].append(shift)
        
        # Genera calendario giorno per giorno
        for shift_date in sorted(shifts_by_date.keys()):
            day_shifts = shifts_by_date[shift_date]
            
            # Header giorno
            weekday_name = ['Lunedì', 'Martedì', 'Mercoledì', 'Giovedì', 'Venerdì', 'Sabato', 'Domenica'][shift_date.weekday()]
            day_header = f"{weekday_name} {shift_date.strftime('%d/%m/%Y')}"
            
            story.append(Paragraph(day_header, styles['Heading2']))
            
            # Tabella turni
            data = [['Utente', 'Ruolo', 'Orario', 'Durata']]
            
            for shift in day_shifts:
                duration = (datetime.combine(date.today(), shift.end_time) -
                          datetime.combine(date.today(), shift.start_time)).total_seconds() / 3600
                
                data.append([
                    shift.user.get_full_name(),
                    shift.user.role,
                    f"{shift.start_time.strftime('%H:%M')} - {shift.end_time.strftime('%H:%M')}",
                    f"{duration:.1f}h"
                ])
            
            table = Table(data, colWidths=[2.5*inch, 1.5*inch, 1.5*inch, 1*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(table)
            story.append(Spacer(1, 20))
    
    doc.build(story)
    
    # Crea response
    buffer.seek(0)
    response = make_response(buffer.getvalue())
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response

# =============================================================================
# ADMIN COVERAGE MANAGEMENT ROUTES
# =============================================================================

@shifts_bp.route('/admin/coverage/create', methods=['POST'])
@login_required
def create_shift_coverage():
    """Crea nuova copertura turni con supporto numerosità ruoli"""
    if not current_user.can_access_turni():
        flash('Non hai i permessi per creare coperture turni', 'danger')
        return redirect(url_for('dashboard.dashboard'))
    
    try:
        # Ottieni dati dal form
        sede_id = request.form.get('sede_id')
        start_date = request.form.get('start_date')
        end_date = request.form.get('end_date')
        base_start_time = request.form.get('base_start_time')
        base_end_time = request.form.get('base_end_time')
        description = request.form.get('description', '')
        
        # Giorni selezionati
        days_of_week = request.form.getlist('days_of_week')
        
        # Estrai ruoli selezionati e numerosità
        roles_dict = {}
        # Cerca i campi role_count_ per determinare ruoli e numerosità
        for key in request.form.keys():
            if key.startswith('role_count_'):
                role_name = key.replace('role_count_', '')
                # Verifica se il checkbox corrispondente esiste ed è selezionato
                checkbox_found = False
                for checkbox_key in request.form.keys():
                    if checkbox_key.startswith('role_') and request.form.get(checkbox_key) == role_name:
                        checkbox_found = True
                        break
                
                if checkbox_found:
                    count = int(request.form.get(key, 1))
                    if count > 0:
                        roles_dict[role_name] = count
        
        if not roles_dict:
            flash('Devi selezionare almeno un ruolo con numerosità', 'danger')
            return redirect(url_for('manage_turni'))
        
        # Converti date
        start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
        
        # Converti orari  
        start_time_obj = datetime.strptime(base_start_time, '%H:%M').time()
        end_time_obj = datetime.strptime(base_end_time, '%H:%M').time()
        
        success_count = 0
        
        # Crea copertura per ogni giorno selezionato
        for day_str in days_of_week:
            day_of_week = int(day_str)
            
            # Controlla orari personalizzati per questo giorno
            custom_start_key = f'start_time_{day_of_week}'
            custom_end_key = f'end_time_{day_of_week}'
            
            day_start_time = start_time_obj
            day_end_time = end_time_obj
            
            if custom_start_key in request.form and custom_end_key in request.form:
                custom_start = request.form.get(custom_start_key)
                custom_end = request.form.get(custom_end_key)
                if custom_start and custom_end:
                    day_start_time = datetime.strptime(custom_start, '%H:%M').time()
                    day_end_time = datetime.strptime(custom_end, '%H:%M').time()
            
            # Crea la copertura
            coverage = PresidioCoverage(
                day_of_week=day_of_week,
                start_time=day_start_time,
                end_time=day_end_time,
                description=description,
                active=True,
                start_date=start_date_obj,
                end_date=end_date_obj,
                created_by=current_user.id
            )
            
            # Imposta ruoli con numerosità
            coverage.set_required_roles_dict(roles_dict)
            
            db.session.add(coverage)
            success_count += 1
        
        db.session.commit()
        
        if success_count > 0:
            total_resources = sum(roles_dict.values())
            flash(f'Copertura creata con successo per {success_count} giorni! Total risorse richieste per turno: {total_resources}', 'success')
        else:
            flash('Nessuna nuova copertura creata', 'warning')
            
    except Exception as e:
        db.session.rollback()
        flash(f'Errore durante la creazione della copertura: {str(e)}', 'danger')
    
    return redirect(url_for('manage_turni'))

@shifts_bp.route('/admin/coperture')
@login_required
def view_turni_coverage():
    """Visualizza le coperture create per una sede specifica"""
    if not current_user.can_access_turni():
        flash('Non hai i permessi per visualizzare le coperture', 'danger')
        return redirect(url_for('dashboard.dashboard'))
    
    sede_id = request.args.get('sede', type=int)
    if not sede_id:
        # Se l'utente non è Admin, usa la sua sede per default
        if current_user.role != 'Admin' and current_user.sede_obj and current_user.sede_obj.is_turni_mode():
            sede_id = current_user.sede_obj.id
        else:
            flash('ID sede non specificato. Seleziona una sede dalla pagina Gestione Turni.', 'warning')
            return redirect(url_for('manage_turni'))
    
    sede = Sede.query.get_or_404(sede_id)
    
    # Verifica permessi sulla sede - supporta utenti multi-sede
    if not current_user.can_manage_shifts() and not current_user.can_view_shifts():
        flash('Non hai i permessi per visualizzare le coperture', 'danger')
        return redirect(url_for('dashboard.dashboard'))
    
    # Per utenti non-admin, verifica accesso alla sede specifica
    if not current_user.all_sedi and current_user.sede_obj and current_user.sede_obj.id != sede_id:
        flash('Non hai accesso a questa sede specifica', 'danger')
        return redirect(url_for('manage_turni'))
    
    if not sede.is_turni_mode():
        flash('La sede selezionata non è configurata per la modalità turni', 'warning')
        return redirect(url_for('manage_turni'))
    
    # Ottieni le coperture create per questa sede
    coperture = PresidioCoverage.query.filter_by(active=True).order_by(
        PresidioCoverage.start_date.desc(),
        PresidioCoverage.day_of_week,
        PresidioCoverage.start_time
    ).all()
    
    # Raggruppa coperture per periodo di validità (evita duplicati)
    coperture_grouped = {}
    coperture_ids_seen = set()
    for copertura in coperture:
        # Evita duplicati
        if copertura.id in coperture_ids_seen:
            continue
        coperture_ids_seen.add(copertura.id)
        
        period_key = f"{copertura.start_date.strftime('%Y-%m-%d')} - {copertura.end_date.strftime('%Y-%m-%d')}"
        if period_key not in coperture_grouped:
            coperture_grouped[period_key] = {
                'start_date': copertura.start_date,
                'end_date': copertura.end_date,
                'coperture': [],
                'active_status': copertura.active and copertura.end_date >= date.today()
            }
        coperture_grouped[period_key]['coperture'].append(copertura)
    
    # Statistiche
    total_coperture = len(coperture)
    active_coperture = len([c for c in coperture if c.is_valid_for_date(date.today())])
    
    return render_template('view_turni_coverage.html',
                         sede=sede,
                         coperture_grouped=coperture_grouped,
                         total_coperture=total_coperture,
                         active_coperture=active_coperture,
                         today=datetime.now().date(),
                         is_admin=(current_user.role == 'Admin'))

@shifts_bp.route('/admin/genera-da-coperture')
@login_required
def generate_turni_from_coverage():
    """Pagina per generare turni basati sulle coperture create"""
    if not current_user.can_access_turni():
        flash('Non hai i permessi per generare turni', 'danger')
        return redirect(url_for('dashboard.dashboard'))
    
    sede_id = request.args.get('sede', type=int)
    if not sede_id:
        # Se l'utente non è Admin, usa la sua sede per default
        if current_user.role != 'Admin' and current_user.sede_obj and current_user.sede_obj.is_turni_mode():
            sede_id = current_user.sede_obj.id
        else:
            flash('ID sede non specificato. Seleziona una sede dalla pagina Genera Turni.', 'warning')
            return redirect(url_for('generate_turnazioni'))
    
    sede = Sede.query.get_or_404(sede_id)
    
    # Verifica permessi sulla sede - supporta utenti multi-sede  
    if not current_user.can_manage_shifts():
        flash('Non hai i permessi per generare turni', 'danger')
        return redirect(url_for('dashboard.dashboard'))
        
    # Per utenti non-admin, verifica accesso alla sede specifica
    if not current_user.all_sedi and current_user.sede_obj and current_user.sede_obj.id != sede_id:
        flash('Non hai accesso a questa sede specifica', 'danger')
        return redirect(url_for('generate_turnazioni'))
    
    if not sede.is_turni_mode():
        flash('La sede selezionata non è configurata per la modalità turni', 'warning')
        return redirect(url_for('generate_turnazioni'))
    
    # Ottieni le coperture attive per questa sede
    coperture = PresidioCoverage.query.filter_by(active=True).order_by(
        PresidioCoverage.start_date.desc(),
        PresidioCoverage.day_of_week,
        PresidioCoverage.start_time
    ).all()
    
    # Raggruppa coperture per periodo di validità (evita duplicati con ID univoci)
    coperture_grouped = {}
    coperture_ids_seen = set()
    for copertura in coperture:
        # Evita duplicati
        if copertura.id in coperture_ids_seen:
            continue
        coperture_ids_seen.add(copertura.id)
        
        period_key = f"{copertura.start_date.strftime('%Y-%m-%d')} - {copertura.end_date.strftime('%Y-%m-%d')}"
        if period_key not in coperture_grouped:
            coperture_grouped[period_key] = {
                'start_date': copertura.start_date,
                'end_date': copertura.end_date,
                'coperture': [],
                'active_status': copertura.active and copertura.end_date >= date.today(),
                'period_id': f"{copertura.start_date.strftime('%Y%m%d')}-{copertura.end_date.strftime('%Y%m%d')}"
            }
        coperture_grouped[period_key]['coperture'].append(copertura)
    
    # Statistiche
    total_coperture = len(coperture)
    active_coperture = len([c for c in coperture if c.is_valid_for_date(date.today())])
    
    return render_template('generate_turni_from_coverage.html',
                         sede=sede,
                         coperture_grouped=coperture_grouped,
                         total_coperture=total_coperture,
                         active_coperture=active_coperture,
                         today=datetime.now().date(),
                         is_admin=(current_user.role == 'Admin'))

@shifts_bp.route('/admin/process-generate-from-coverage', methods=['POST'])
@login_required
def process_generate_turni_from_coverage():
    """Processa la generazione dei turni basata sulle coperture"""
    if not current_user.can_access_turni():
        flash('Non hai i permessi per generare turni', 'danger')
        return redirect(url_for('dashboard.dashboard'))
    
    sede_id = request.form.get('sede_id', type=int)
    coverage_period_id = request.form.get('coverage_period_id')
    use_coverage_dates = 'use_coverage_dates' in request.form
    replace_existing = 'replace_existing' in request.form
    confirm_overwrite = 'confirm_overwrite' in request.form
    
    if not sede_id or not coverage_period_id or coverage_period_id.strip() == '':
        flash(f'Dati mancanti per la generazione turni (sede_id: {sede_id}, coverage_period_id: \'{coverage_period_id}\')', 'danger')
        return redirect(url_for('generate_turnazioni'))
    
    sede = Sede.query.get_or_404(sede_id)
    
    # Verifica permessi - supporta utenti multi-sede
    if not current_user.can_manage_shifts():
        flash('Non hai i permessi per generare turni', 'danger')
        return redirect(url_for('dashboard.dashboard'))
        
    # Per utenti non-admin, verifica accesso alla sede specifica
    if not current_user.all_sedi and current_user.sede_obj and current_user.sede_obj.id != sede_id:
        flash('Non hai accesso per generare turni per questa sede', 'danger')
        return redirect(url_for('generate_turnazioni'))
    
    try:
        # Decodifica period_id per ottenere le date della copertura
        start_str, end_str = coverage_period_id.split('-')
        start_date = datetime.strptime(start_str, '%Y%m%d').date()
        end_date = datetime.strptime(end_str, '%Y%m%d').date()
        
        # Trova coperture
        coperture = PresidioCoverage.query.filter(
            PresidioCoverage.start_date <= end_date,
            PresidioCoverage.end_date >= start_date,
            PresidioCoverage.active == True
        ).all()
        
        if not coperture:
            flash('Nessuna copertura trovata per il periodo specificato', 'warning')
            return redirect(url_for('generate_turnazioni'))
        
        # Controlla se esistono già turni nel periodo
        existing_shifts = Shift.query.join(User, Shift.user_id == User.id).filter(
            User.sede_id == sede_id,
            Shift.date >= start_date,
            Shift.date <= end_date
        ).all()
        
        # Se esistono turni e non confermato, chiedi conferma
        if existing_shifts and not replace_existing and not confirm_overwrite:
            turni_count = len(existing_shifts)
            date_range = f"{start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')}"
            
            return render_template('confirm_overwrite_shifts.html',
                                 sede=sede,
                                 period_id=coverage_period_id,
                                 start_date=start_date,
                                 end_date=end_date,
                                 date_range=date_range,
                                 existing_shifts_count=turni_count,
                                 use_coverage_dates=use_coverage_dates,
                                 replace_existing=replace_existing)
        
        # Genera turni
        turni_creati = 0
        turni_sostituiti = 0
        current_date = start_date
        
        while current_date <= end_date:
            day_of_week = current_date.weekday()
            coperture_giorno = [c for c in coperture if c.day_of_week == day_of_week and c.is_valid_for_date(current_date)]
            
            for copertura in coperture_giorno:
                existing_shift = Shift.query.filter_by(
                    date=current_date,
                    start_time=copertura.start_time,
                    end_time=copertura.end_time
                ).first()
                
                if existing_shift and not replace_existing:
                    continue
                elif existing_shift and replace_existing:
                    db.session.delete(existing_shift)
                    turni_sostituiti += 1
                
                required_roles_dict = copertura.get_required_roles_dict()
                
                for role, count_needed in required_roles_dict.items():
                    available_users = User.query.filter(
                        User.sede_id == sede_id,
                        User.active == True,
                        User.role == role
                    ).all()
                    
                    if len(available_users) >= count_needed:
                        for i in range(count_needed):
                            user_index = (current_date.day + copertura.id + i) % len(available_users)
                            assigned_user = available_users[user_index]
                            
                            new_shift = Shift(
                                user_id=assigned_user.id,
                                date=current_date,
                                start_time=copertura.start_time,
                                end_time=copertura.end_time,
                                shift_type='Normale',
                                created_by=current_user.id
                            )
                            db.session.add(new_shift)
                            turni_creati += 1
            
            current_date += timedelta(days=1)
        
        db.session.commit()
        
        if turni_creati > 0 or turni_sostituiti > 0:
            message_parts = []
            if turni_creati > 0:
                message_parts.append(f'{turni_creati} turni creati')
            if turni_sostituiti > 0:
                message_parts.append(f'{turni_sostituiti} turni sostituiti')
            
            flash(f'Generazione completata! {" e ".join(message_parts)} per {sede.name} dal {start_date.strftime("%d/%m/%Y")} al {end_date.strftime("%d/%m/%Y")}', 'success')
        else:
            flash(f'Nessun turno generato - potrebbero già esistere turni per il periodo o non ci sono utenti disponibili', 'warning')
        
    except (ValueError, AttributeError):
        flash('ID periodo non valido', 'danger')
        return redirect(url_for('generate_turnazioni'))
    
    return redirect(url_for('generate_turnazioni'))

@shifts_bp.route('/admin/visualizza-generati')
@login_required
def view_generated_shifts():
    """Visualizza i turni generati per una specifica copertura"""
    if not current_user.can_access_turni():
        flash('Non hai i permessi per visualizzare i turni', 'danger')
        return redirect(url_for('dashboard.dashboard'))
    
    sede_id = request.args.get('sede', type=int)
    period_id = request.args.get('period') or request.args.get('coverage_period')
    
    if not all([sede_id, period_id]):
        flash('Parametri mancanti per la visualizzazione turni', 'danger')
        return redirect(url_for('generate_turnazioni'))
    
    sede = Sede.query.get_or_404(sede_id)
    
    # Verifica permessi sulla sede - supporta utenti multi-sede
    if not current_user.can_view_shifts() and not current_user.can_manage_shifts():
        flash('Non hai i permessi per visualizzare i turni', 'danger')
        return redirect(url_for('dashboard.dashboard'))
        
    # Per utenti non-admin, verifica accesso alla sede specifica
    if not current_user.all_sedi and current_user.sede_obj and current_user.sede_obj.id != sede_id:
        flash('Non hai accesso per visualizzare i turni di questa sede', 'danger')
        return redirect(url_for('generate_turnazioni'))
    
    # Decodifica period_id per ottenere le date
    try:
        start_str, end_str = period_id.split('-')
        coverage_start_date = datetime.strptime(start_str, '%Y%m%d').date()
        coverage_end_date = datetime.strptime(end_str, '%Y%m%d').date()
    except (ValueError, AttributeError):
        flash('Periodo non valido specificato', 'danger')
        return redirect(url_for('generate_turnazioni'))
    
    # Ottieni i turni generati nel periodo delle coperture
    shifts = Shift.query.filter(
        Shift.user.has(sede_id=sede_id),
        Shift.date >= coverage_start_date,
        Shift.date <= coverage_end_date
    ).order_by(Shift.date, Shift.start_time).all()
    
    # Raggruppa turni per data
    shifts_by_date = {}
    for shift in shifts:
        date_str = shift.date.strftime('%Y-%m-%d')
        if date_str not in shifts_by_date:
            shifts_by_date[date_str] = {
                'date': shift.date,
                'date_display': shift.date.strftime('%d/%m/%Y'),
                'day_name': ['Lunedì', 'Martedì', 'Mercoledì', 'Giovedì', 'Venerdì', 'Sabato', 'Domenica'][shift.date.weekday()],
                'shifts': []
            }
        shifts_by_date[date_str]['shifts'].append(shift)
    
    # Ottieni le coperture di riferimento per il confronto
    reference_coverages = PresidioCoverage.query.filter(
        PresidioCoverage.start_date <= coverage_end_date,
        PresidioCoverage.end_date >= coverage_start_date,
        PresidioCoverage.active == True
    ).all()
    
    total_shifts = len(shifts)
    dates_with_shifts = len(shifts_by_date)
    period_days = (coverage_end_date - coverage_start_date).days + 1
    
    # Calcola utenti unici coinvolti
    unique_users = set()
    for shift in shifts:
        if shift.user:
            unique_users.add(shift.user.id)
    unique_users_count = len(unique_users)
    
    return render_template('view_generated_shifts.html',
                         sede=sede,
                         coverage_start_date=coverage_start_date,
                         coverage_end_date=coverage_end_date,
                         shifts_by_date=shifts_by_date,
                         reference_coverages=reference_coverages,
                         total_shifts=total_shifts,
                         dates_with_shifts=dates_with_shifts,
                         period_days=period_days,
                         unique_users_count=unique_users_count,
                         today=datetime.now().date(),
                         is_admin=(current_user.role == 'Admin'))

@shifts_bp.route('/admin/regenerate-from-coverage', methods=['POST'])
@login_required
def regenerate_turni_from_coverage():
    """Rigenera i turni eliminando quelli esistenti da oggi in poi e creandone di nuovi"""
    if not current_user.can_access_turni():
        flash('Non hai i permessi per rigenerare turni', 'danger')
        return redirect(url_for('dashboard.dashboard'))
    
    sede_id = request.form.get('sede_id', type=int)
    coverage_period_id = request.form.get('coverage_period_id')
    
    if not all([sede_id, coverage_period_id]):
        flash('Dati mancanti per la rigenerazione turni', 'danger')
        return redirect(url_for('generate_turnazioni'))
    
    sede = Sede.query.get_or_404(sede_id)
    
    # Verifica permessi
    if current_user.role != 'Admin':
        if not current_user.sede_obj or current_user.sede_obj.id != sede_id:
            flash('Non hai i permessi per rigenerare turni per questa sede', 'danger')
            return redirect(url_for('generate_turnazioni'))
    
    try:
        # Decodifica period_id per ottenere le date della copertura
        start_str, end_str = coverage_period_id.split('-')
        start_date = datetime.strptime(start_str, '%Y%m%d').date()
        end_date = datetime.strptime(end_str, '%Y%m%d').date()
        
        # Data di inizio per l'eliminazione (da oggi in poi)
        today = date.today()
        delete_from_date = max(start_date, today)
        
        # Elimina turni esistenti da oggi in poi
        shifts_to_delete = Shift.query.join(User, Shift.user_id == User.id).filter(
            User.sede_id == sede_id,
            Shift.date >= delete_from_date,
            Shift.date <= end_date
        ).all()
        
        deleted_count = len(shifts_to_delete)
        for shift in shifts_to_delete:
            db.session.delete(shift)
        
        db.session.commit()
        
        # Rigenera turni per tutto il periodo originale
        coperture = PresidioCoverage.query.filter(
            PresidioCoverage.start_date <= end_date,
            PresidioCoverage.end_date >= start_date,
            PresidioCoverage.active == True
        ).all()
        
        if not coperture:
            flash('Nessuna copertura trovata per il periodo specificato', 'warning')
            return redirect(url_for('generate_turnazioni'))
        
        # Genera i nuovi turni
        new_shifts_count = 0
        current_date = start_date
        
        while current_date <= end_date:
            day_of_week = current_date.weekday()
            
            day_coverages = [c for c in coperture if 
                           c.start_date <= current_date <= c.end_date and
                           c.day_of_week == day_of_week]
            
            for coverage in day_coverages:
                available_users = User.query.filter(
                    User.sede_id == sede_id,
                    User.active == True,
                    User.role.in_(['Operatore', 'Sviluppatore', 'Redattore', 'Management'])
                ).all()
                
                roles_dict = coverage.get_required_roles_dict()
                total_required_staff = sum(roles_dict.values()) if roles_dict else 1
                
                if available_users and total_required_staff > 0:
                    selected_users = available_users[:total_required_staff]
                    
                    for user in selected_users:
                        new_shift = Shift(
                            user_id=user.id,
                            date=current_date,
                            start_time=coverage.start_time,
                            end_time=coverage.end_time,
                            shift_type='Turno',
                            created_by=current_user.id
                        )
                        db.session.add(new_shift)
                        new_shifts_count += 1
            
            current_date += timedelta(days=1)
        
        db.session.commit()
        
        # Messaggio di successo
        if deleted_count > 0:
            flash(f'Turni rigenerati con successo! Eliminati {deleted_count} turni esistenti, creati {new_shifts_count} nuovi turni.', 'success')
        else:
            flash(f'Turni generati con successo! Creati {new_shifts_count} nuovi turni.', 'success')
        
        return redirect(url_for('shifts.view_generated_shifts', sede=sede_id, period=coverage_period_id))
        
    except Exception as e:
        db.session.rollback()
        flash(f'Errore durante la rigenerazione turni: {str(e)}', 'danger')
        return redirect(url_for('generate_turnazioni'))